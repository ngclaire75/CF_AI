"""CF_AI Security Dashboard — Flask web application."""
from __future__ import annotations
import ipaddress
import json as _json
import os

# Load .env from project root before anything else reads os.environ
try:
    from dotenv import load_dotenv as _load_dotenv
    import pathlib as _pl
    _load_dotenv(_pl.Path(__file__).parent.parent / '.env', override=False)
except ImportError:
    pass
import datetime as _datetime
import hashlib as _hashlib
import re
import sys
import time as _time
import shutil as _shutil
import ssl as _ssl
import subprocess as _subprocess
import threading as _threading
import urllib.error as _up_err
import urllib.parse as _up_parse
import urllib.request as _up_req
import uuid as _uuid
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent.parent))

from flask import Flask, render_template, jsonify, abort, request, Response, stream_with_context, redirect, session, url_for, flash, send_file, make_response
from functools import wraps
from werkzeug.security import generate_password_hash, check_password_hash
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import dashboard.db as db
from dashboard.remediations import REMEDIATIONS

db.init_db()

_BROWSER_UA = ('Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
               'AppleWebKit/537.36 (KHTML, like Gecko) '
               'Chrome/124.0.0.0 Safari/537.36')

try:
    import requests as _requests
    import urllib3 as _urllib3
    _urllib3.disable_warnings(_urllib3.exceptions.InsecureRequestWarning)
    _HAS_REQUESTS = True
except ImportError:
    _HAS_REQUESTS = False

try:
    import cloudscraper as _cloudscraper
    _HAS_CLOUDSCRAPER = True
except ImportError:
    _HAS_CLOUDSCRAPER = False

_SCRAPER_API_KEY = os.environ.get('SCRAPER_API_KEY', '').strip()

# ── GeoIP lookup cache (uses ip-api.com, free, no key, 45 req/min) ─────────
_geoip_cache: dict = {}

def _geoip_detail(ip: str) -> dict:
    """Return {country, country_code, lat, lon} for an IP. Cached."""
    _blank = {'country': '', 'country_code': '', 'lat': 0.0, 'lon': 0.0}
    if not ip or ip in ('127.0.0.1', '::1', 'localhost', '0.0.0.0', ''):
        return _blank
    if ip in _geoip_cache:
        return _geoip_cache[ip]
    try:
        import urllib.request as _ur2
        with _ur2.urlopen(
            f'http://ip-api.com/json/{ip}?fields=country,countryCode,lat,lon',
            timeout=3
        ) as _r2:
            _d2 = _json.loads(_r2.read())
            result = {
                'country':      _d2.get('country', ''),
                'country_code': _d2.get('countryCode', ''),
                'lat':          float(_d2.get('lat', 0)),
                'lon':          float(_d2.get('lon', 0)),
            }
    except Exception:
        result = _blank.copy()
    _geoip_cache[ip] = result
    return result


# ── Auto-remediation rules ────────────────────────────────────────────────────
_REMED_RULES = [
    {'name': 'brute_force_block',     'event_types': ['login_failed'],
     'threshold': 10, 'window_min': 5,   'action': 'block_ip',
     'severity': 'HIGH',     'description': 'Block IPs with >10 failed logins in 5 min'},
    {'name': 'sql_injection_block',   'event_types': ['sql_injection'],
     'threshold': 1,  'window_min': 60,  'action': 'block_ip',
     'severity': 'CRITICAL', 'description': 'Immediately block SQL injection sources'},
    {'name': 'xss_block',             'event_types': ['xss_attempt'],
     'threshold': 3,  'window_min': 10,  'action': 'block_ip',
     'severity': 'HIGH',     'description': 'Block IPs with ≥3 XSS attempts in 10 min'},
    {'name': 'scanner_block',         'event_types': ['port_scan', 'vuln_scan'],
     'threshold': 5,  'window_min': 2,   'action': 'block_ip',
     'severity': 'MEDIUM',   'description': 'Block aggressive scanners (≥5 probes in 2 min)'},
    {'name': 'credential_stuffing',   'event_types': ['login_failed'],
     'threshold': 50, 'window_min': 60,  'action': 'block_ip',
     'severity': 'CRITICAL', 'description': 'Block credential-stuffing sources (≥50 failures/hour)'},
    {'name': 'critical_vuln_incident','event_types': ['vulnerability_detected'],
     'threshold': 1,  'window_min': 1440,'action': 'create_incident',
     'severity': 'CRITICAL', 'description': 'Open incident for every critical vulnerability'},
]


def _run_remediation(event_id: int, event: dict) -> None:
    """Check rules against the new event; trigger actions when thresholds are met."""
    ip        = event.get('ip_address', '')
    ev_type   = event.get('event_type', '')
    target    = event.get('target', '')

    for rule in _REMED_RULES:
        if ev_type not in rule['event_types']:
            continue
        count = db.count_events_by_ip(ip, rule['event_types'], rule['window_min'])
        if count < rule['threshold']:
            continue

        action_id = db.log_remediation(
            trigger_event_id=event_id, rule_name=rule['name'],
            action_type=rule['action'], target=ip or target,
            parameters=_json.dumps({'rule': rule['name'], 'count': count, 'ip': ip}),
            status='running', auto_triggered=True,
        )

        try:
            if rule['action'] == 'block_ip' and ip and not db.is_ip_blocked(ip):
                _auto_block_ip(ip, rule['name'], event.get('country', ''), action_id)
            elif rule['action'] == 'create_incident':
                db.create_incident(
                    title=f"[Auto] {event.get('description','Security event')}",
                    description=f"Rule: {rule['description']}\nEvent ID: {event_id}\nTarget: {target}",
                    severity=rule['severity'],
                    target=target,
                )
                db.update_remediation(action_id, 'success', 'Incident created')
        except Exception as _e:
            db.update_remediation(action_id, 'failed', str(_e)[:200])


def _auto_block_ip(ip: str, rule_name: str, country: str, action_id: int) -> None:
    """Block an IP via Cloudflare Firewall Rules (requires Zone WAF:Edit permission)."""
    def _strip_env_prefix(v):
        return v.split('=', 1)[-1].strip() if '=' in v else v.strip()

    cf_token = _strip_env_prefix(os.environ.get('CF_API_TOKEN', '').strip())
    if not cf_token:
        db.add_blocked_ip(ip, country, rule_name)
        db.update_remediation(action_id, 'success', f'Blocked locally (no CF token)')
        return

    # Find a zone to apply the rule to (use first available zone)
    try:
        zr = _json.loads(_cf_request('/zones?per_page=1', cf_token, timeout=10)[1])
        zones = (zr.get('result') or [])
        zone_id = zones[0]['id'] if zones else ''
    except Exception:
        zone_id = ''

    cf_rule_id = ''
    if zone_id and _HAS_REQUESTS:
        url  = f'https://api.cloudflare.com/client/v4/zones/{zone_id}/firewall/rules'
        hdrs = {'Authorization': f'Bearer {cf_token}', 'Content-Type': 'application/json'}
        try:
            r = _requests.post(url, headers=hdrs, json=[{
                'filter': {'expression': f'(ip.src eq {ip})'},
                'action': 'block',
                'description': f'CF_AI auto-block: {rule_name}',
            }], timeout=15, verify=True)
            data = r.json()
            if r.status_code == 200 and data.get('success'):
                cf_rule_id = (data.get('result') or [{}])[0].get('id', '')
        except Exception:
            pass

    db.add_blocked_ip(ip, country, rule_name, cf_rule_id, zone_id)
    db.update_remediation(action_id, 'success',
        f'IP {ip} blocked' + (f' via CF rule {cf_rule_id}' if cf_rule_id else ' locally'))


def _wp_request(url: str, method: str = 'GET', headers: dict | None = None,
                body: bytes | None = None, timeout: int = 20) -> tuple[int, str]:
    """HTTP request with progressive bypass for Cloudflare/WAF-protected sites.

    Layer 0: ScraperAPI — routes through residential IPs, bypasses VPS/datacenter blocks
    Layer 1: requests library — verify=False, full browser headers
    Layer 2: cloudscraper — solves Cloudflare JS challenges
    Layer 3: curl -sk — handles edge cases (cert issues, header quirks)

    Returns (status_code, body_text).  status_code=0 means total failure.
    The error string is always non-empty on failure so callers can surface it.
    """
    hdrs = {
        'User-Agent':      _BROWSER_UA,
        'Accept':          'application/json, */*',
        'Accept-Language': 'en-US,en;q=0.9',
        'Cache-Control':   'no-cache',
    }
    if headers:
        hdrs.update(headers)

    last_err = 'unknown error'

    # Layer 0 — ScraperAPI (residential IPs — bypasses VPS/datacenter Cloudflare blocks)
    # Skip for cookie-auth requests: nonce is bound to the session established by _wp_cookie_auth
    if _SCRAPER_API_KEY and _HAS_REQUESTS and 'Cookie' not in hdrs:
        try:
            scraper_url = (
                f'http://api.scraperapi.com/?api_key={_SCRAPER_API_KEY}'
                f'&url={_up_parse.quote(url, safe="")}&render=false'
            )
            resp = _requests.request(
                method, scraper_url,
                headers=hdrs,
                data=body,
                verify=False,
                timeout=timeout + 10,
            )
            return resp.status_code, resp.text
        except Exception as e:
            last_err = str(e)

    # Layer 1 — requests with SSL verification disabled
    if _HAS_REQUESTS:
        try:
            resp = _requests.request(
                method, url,
                headers=hdrs,
                data=body,
                verify=False,
                timeout=timeout,
                allow_redirects=True,
            )
            return resp.status_code, resp.text
        except Exception as e:
            last_err = str(e)

    # Layer 2 — cloudscraper (solves Cloudflare JS challenge v1/v2)
    if _HAS_CLOUDSCRAPER:
        try:
            cs = _cloudscraper.create_scraper(
                browser={'browser': 'chrome', 'platform': 'linux', 'mobile': False}
            )
            resp = cs.request(method, url, headers=hdrs, data=body,
                              verify=False, timeout=timeout)
            return resp.status_code, resp.text
        except Exception as e:
            last_err = str(e)

    # Layer 3 — curl -sk (handles remaining edge cases)
    if _shutil.which('curl'):
        # Write body to temp file to avoid stdin issues on Linux/Windows
        tmp_path = None
        try:
            if body:
                import tempfile
                with tempfile.NamedTemporaryFile(delete=False, suffix='.json') as tf:
                    tf.write(body)
                    tmp_path = tf.name
            cmd = [
                'curl', '-sk', '-L',
                '-w', '\n__CF_STATUS__:%{http_code}',
                '--max-time', str(timeout),
                '--connect-timeout', '10',
                '-X', method,
            ]
            for k, v in hdrs.items():
                cmd += ['-H', f'{k}: {v}']
            if tmp_path:
                cmd += ['--data-binary', f'@{tmp_path}']
            cmd.append(url)
            res = _subprocess.run(cmd, capture_output=True, timeout=timeout + 10)
            text = res.stdout.decode('utf-8', errors='replace')
            if '__CF_STATUS__:' in text:
                body_part, stat = text.rsplit('\n__CF_STATUS__:', 1)
                code = int(stat.strip() or '0')
                if code > 0:
                    return code, body_part
                last_err = f'curl exited with HTTP 0 — site may be blocking VPS/datacenter IPs'
            else:
                stderr = res.stderr.decode('utf-8', errors='replace')[:200]
                last_err = f'curl no response — {stderr or "connection refused or timed out"}'
        except Exception as e:
            last_err = str(e)
        finally:
            if tmp_path:
                try:
                    import os; os.unlink(tmp_path)
                except Exception:
                    pass
    else:
        last_err = 'curl not found; requests unavailable'

    return 0, last_err


def _wp_cookie_auth(site_url: str, username: str, password: str) -> tuple[str | None, str | None]:
    """Login via wp-login.php and return (nonce, cookie_header) for REST API cookie auth.

    Works with regular WordPress admin passwords (not just Application Passwords).
    Uses ScraperAPI URL mode (residential IPs) — proxy mode fails Cloudflare JS challenge.
    Returns (nonce, cookie_header) on success, (None, None) on failure.
    """
    if not _HAS_REQUESTS:
        return None, None
    import re as _rmod

    def _sa_req(target, method='GET', data=None, extra_hdrs=None):
        sa = (f'http://api.scraperapi.com/?api_key={_SCRAPER_API_KEY}'
              f'&url={_up_parse.quote(target, safe="")}')
        h = {'User-Agent': _BROWSER_UA}
        if extra_hdrs:
            h.update(extra_hdrs)
        if method == 'POST':
            h.setdefault('Content-Type', 'application/x-www-form-urlencoded')
            return _requests.post(sa, data=data, headers=h, verify=False, timeout=30)
        return _requests.get(sa, headers=h, verify=False, timeout=20)

    def _try_scraper():
        if not _SCRAPER_API_KEY:
            return None, None
        try:
            login_r = _sa_req(
                f'{site_url}/wp-login.php', method='POST',
                data={'log': username, 'pwd': password,
                      'wp-submit': 'Log In', 'redirect_to': '/wp-admin/',
                      'testcookie': '1'},
                extra_hdrs={'Cookie': 'wordpress_test_cookie=WP Cookie check'},
            )
            # Collect cookies — ScraperAPI may return them in response.cookies or Set-Cookie headers
            cookies = {k: v for k, v in login_r.cookies.items()}
            for raw in login_r.headers.getlist('set-cookie') if hasattr(login_r.headers, 'getlist') \
                    else [login_r.headers.get('set-cookie', '')]:
                nv = raw.split(';')[0].strip()
                if '=' in nv:
                    n, v2 = nv.split('=', 1)
                    cookies[n.strip()] = v2.strip()
            if not any('wordpress_logged_in' in k for k in cookies):
                return None, None
            cookie_hdr = '; '.join(f'{k}={v}' for k, v in cookies.items())
            admin_r = _sa_req(f'{site_url}/wp-admin/', extra_hdrs={'Cookie': cookie_hdr})
            nonce = ''
            m = _rmod.search(r'"nonce"\s*:\s*"([a-f0-9]{10})"', admin_r.text)
            if m:
                nonce = m.group(1)
            return nonce, cookie_hdr
        except Exception:
            return None, None

    def _try_direct():
        try:
            sess = _requests.Session()
            sess.verify = False
            sess.headers.update({'User-Agent': _BROWSER_UA})
            sess.post(
                f'{site_url}/wp-login.php',
                data={'log': username, 'pwd': password,
                      'wp-submit': 'Log In', 'redirect_to': '/wp-admin/',
                      'testcookie': '1'},
                cookies={'wordpress_test_cookie': 'WP Cookie check'},
                allow_redirects=True, timeout=25,
            )
            if not any('wordpress_logged_in' in k for k in sess.cookies.keys()):
                return None, None
            admin_r = sess.get(f'{site_url}/wp-admin/', allow_redirects=True, timeout=15)
            nonce = ''
            m = _rmod.search(r'"nonce"\s*:\s*"([a-f0-9]{10})"', admin_r.text)
            if m:
                nonce = m.group(1)
            cookie_hdr = '; '.join(f'{k}={v}' for k, v in sess.cookies.items())
            return nonce, cookie_hdr
        except Exception:
            return None, None

    nonce, ck = _try_scraper()
    if nonce is not None:
        return nonce, ck
    nonce, ck = _try_direct()
    if nonce is not None:
        return nonce, ck

    # ScraperAPI can't preserve session cookies (it's a scraper, not a session proxy).
    # Fall back to XML-RPC to at least verify whether the credentials are correct.
    if _wp_xmlrpc_verify(site_url, username, password):
        return '__xmlrpc_verified__', ''  # credentials OK but no cookie session possible
    return None, None


def _wp_xmlrpc_verify(site_url: str, username: str, password: str) -> bool:
    """Verify WordPress credentials via XML-RPC (accepts regular admin passwords)."""
    if not _HAS_REQUESTS:
        return False
    import xml.sax.saxutils as _sax
    payload = (
        '<?xml version="1.0"?><methodCall>'
        '<methodName>wp.getProfile</methodName><params>'
        '<param><value><int>1</int></value></param>'
        f'<param><value><string>{_sax.escape(username)}</string></value></param>'
        f'<param><value><string>{_sax.escape(password)}</string></value></param>'
        '</params></methodCall>'
    ).encode()
    xmlrpc_url = f'{site_url}/xmlrpc.php'
    hdrs = {'Content-Type': 'text/xml', 'User-Agent': _BROWSER_UA}
    try:
        if _SCRAPER_API_KEY and _HAS_REQUESTS:
            sa = (f'http://api.scraperapi.com/?api_key={_SCRAPER_API_KEY}'
                  f'&url={_up_parse.quote(xmlrpc_url, safe="")}')
            r = _requests.post(sa, data=payload, headers=hdrs, verify=False, timeout=25)
        else:
            r = _requests.post(xmlrpc_url, data=payload, headers=hdrs, verify=False, timeout=20)
        return (r.status_code == 200
                and '<fault>' not in r.text
                and '<methodResponse>' in r.text
                and 'user_login' in r.text)
    except Exception:
        return False


app = Flask(__name__, template_folder='templates')
app.secret_key = os.environ.get('CFAI_SECRET_KEY', 'cfai-dev-secret-change-in-prod-2026')

# ── SMTP / email config ───────────────────────────────────────────────────────
_SMTP_USER = os.environ.get('SMTP_USER', '')
_SMTP_PASS = os.environ.get('SMTP_PASS', '')
_SMTP_HOST = os.environ.get('SMTP_HOST', 'smtp.gmail.com')
_SMTP_PORT = int(os.environ.get('SMTP_PORT', '465'))
_BASE_URL  = os.environ.get('CFAI_BASE_URL', 'http://localhost:8889')

# ── App version (git short hash — auto-updates on every commit) ───────────────
_GIT_ROOT = os.path.dirname(os.path.abspath(__file__))
_CC_PREFIX = re.compile(
    r'^(feat|fix|chore|refactor|docs|style|test|perf|build|ci|revert)(\(.+?\))?[!]?:\s*',
    re.IGNORECASE,
)

def _git_run(*args) -> str:
    try:
        r = _subprocess.run(
            ['git'] + list(args),
            capture_output=True, text=True, timeout=5, cwd=_GIT_ROOT,
        )
        return r.stdout.strip()
    except Exception:
        return ''

def _get_git_commit_count() -> int:
    try:
        return int(_git_run('rev-list', '--count', 'HEAD') or '0')
    except ValueError:
        return 0

def _get_git_version() -> str:
    n = _get_git_commit_count()
    return f'1.0.{n}' if n else '1.0.0'

def _get_git_build_date() -> str:
    v = _git_run('log', '-1', '--format=%cd', '--date=format:%d %b %Y')
    return v if v else _datetime.datetime.now().strftime('%d %b %Y')

def _clean_commit_msg(msg: str) -> str:
    msg = _CC_PREFIX.sub('', msg).strip()
    return msg[:1].upper() + msg[1:] if msg else msg

def _get_git_changelog(n: int = 12) -> list[dict]:
    total = _get_git_commit_count()
    raw = _git_run('log', f'-{n}', '--format=%s')
    entries = []
    for i, line in enumerate(raw.splitlines()):
        line = line.strip()
        if not line:
            continue
        entries.append({
            'version': f'1.0.{total - i}',
            'msg': _clean_commit_msg(line),
        })
    return entries

_APP_VERSION    = _get_git_version()
_APP_BUILD_DATE = _get_git_build_date()
_APP_CHANGELOG  = _get_git_changelog()

# ── Privacy policy change detection ──────────────────────────────────────────
_TEMPLATE_PATH    = os.path.join(os.path.dirname(__file__), 'templates', 'index.html')
_POLICY_HASH_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'policy_hash.txt')

# ── Shared email HTML wrapper (light + dark mode, mobile-safe) ────────────────
def _email_html(subject: str, body: str) -> str:
    year = _datetime.datetime.now().year
    return f"""<!DOCTYPE html>
<html lang="en" xmlns="http://www.w3.org/1999/xhtml">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<meta name="color-scheme" content="light dark">
<meta name="supported-color-schemes" content="light dark">
<title>{subject}</title>
<style>
  :root{{color-scheme:light dark;}}
  body,html{{margin:0;padding:0;width:100%;background-color:#eff6ff;}}
  @media(prefers-color-scheme:dark){{
    body,html{{background-color:#0c1120!important;}}
    .eout{{background-color:#0c1120!important;}}
    .ecard{{background-color:#1a2035!important;border-color:#1e3a8a!important;}}
    .ebody{{background-color:#1a2035!important;}}
    .efoot{{background-color:#111827!important;border-color:#1e3a8a!important;}}
    .eh1{{color:#e2e8f0!important;}}
    .ep{{color:#94a3b8!important;}}
    .elink{{color:#60a5fa!important;}}
    .efp{{color:#4b5563!important;}}
  }}
  @media only screen and(max-width:600px){{
    .ecard{{border-radius:8px!important;}}
    .ebody{{padding:22px 18px!important;}}
    .efoot{{padding:12px 18px!important;}}
  }}
</style>
</head>
<body style="margin:0;padding:0;background-color:#eff6ff;">
<table class="eout" width="100%" cellpadding="0" cellspacing="0" border="0"
  style="background-color:#eff6ff;width:100%;">
<tr><td align="center" style="padding:40px 16px;">
  <table class="ecard" width="100%" cellpadding="0" cellspacing="0" border="0"
    style="max-width:480px;background-color:#ffffff;border:1px solid #bfdbfe;
           border-radius:14px;overflow:hidden;width:100%;">
    <tr>
      <td style="background:linear-gradient(135deg,#1e3a8a 0%,#1d4ed8 100%);padding:24px 32px;">
        <div style="font-size:22px;font-weight:800;color:#ffffff;letter-spacing:-.5px;
                    font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">CyberINK</div>
        <div style="font-size:10px;color:#93c5fd;letter-spacing:.8px;text-transform:uppercase;
                    margin-top:4px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">Security Intelligence</div>
      </td>
    </tr>
    <tr>
      <td class="ebody" style="padding:32px;background-color:#ffffff;
                               font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
        {body}
      </td>
    </tr>
    <tr>
      <td class="efoot" style="background-color:#f8faff;border-top:1px solid #bfdbfe;padding:14px 32px;">
        <p class="efp" style="font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;
                               color:#64748b;font-size:11px;margin:0;line-height:1.6;">
          &copy; {year} CyberINK Security Intelligence &mdash; Automated message, do not reply.
        </p>
      </td>
    </tr>
  </table>
</td></tr>
</table>
</body>
</html>"""


def _send_verification_email(to_email: str, token: str) -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        verify_url = f'{_BASE_URL}/verify/{token}'
        subject = 'Verify your CyberINK account'
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 12px;">
            Verify your email address
          </p>
          <p class="ep" style="color:#475569;font-size:13px;line-height:1.65;margin:0 0 28px;">
            Thanks for signing up to CyberINK. Click the button below to verify your
            email address and activate your account.<br><br>
            This link expires in <strong>24&nbsp;hours</strong>.
          </p>
          <table cellpadding="0" cellspacing="0" border="0"><tr><td>
            <a href="{verify_url}"
              style="display:inline-block;background:#2563eb;color:#ffffff;text-decoration:none;
                     padding:13px 32px;border-radius:8px;font-size:14px;font-weight:700;
                     letter-spacing:.2px;font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
              Verify my account &rarr;
            </a>
          </td></tr></table>
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:28px;line-height:1.7;">
            If you didn't create a CyberINK account, you can safely ignore this email.<br>
            Or copy this link into your browser:<br>
            <a href="{verify_url}" class="elink"
              style="color:#2563eb;word-break:break-all;font-size:11px;">{verify_url}</a>
          </p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK <{_SMTP_USER}>'
        msg['To']      = to_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, to_email, msg.as_string())
        return True
    except Exception:
        return False

# ── User store ────────────────────────────────────────────────────────────────
_USERS_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'users.json')

_DEFAULT_ADMIN = 'admin'
_DEFAULT_ADMIN_PASS = 'admin123'

def _load_users() -> dict:
    os.makedirs(os.path.dirname(_USERS_FILE), exist_ok=True)
    if not os.path.exists(_USERS_FILE):
        default = {_DEFAULT_ADMIN: {
            'password': generate_password_hash(_DEFAULT_ADMIN_PASS),
            'role': 'admin', 'email': '', 'verified': True, 'verification_token': None,
        }}
        with open(_USERS_FILE, 'w') as f:
            _json.dump(default, f, indent=2)
        return default
    with open(_USERS_FILE) as f:
        users = _json.load(f)
    # Default page list for non-admin users
    _DEFAULT_USER_PAGES = [
        'dashboard', 'chatbot', 'pluginlogs', 'logexplorer', 'inventories', 'network', 'syslog', 'grc',
    ]
    # Migrate older entries + always enforce admin account
    changed = False
    for uname, u in users.items():
        for field, default_val in [('verified', True), ('email', ''), ('verification_token', None)]:
            if field not in u:
                u[field] = default_val
                changed = True
        # Migrate: add allowed_pages for non-admin users who don't have it yet
        if u.get('role') != 'admin' and 'allowed_pages' not in u:
            u['allowed_pages'] = _DEFAULT_USER_PAGES[:]
            changed = True
        # Migrate: add plan field — admins default to pro, users to basic
        if 'plan' not in u:
            u['plan'] = 'pro' if u.get('role') == 'admin' else 'basic'
            changed = True
    # Ensure default admin always exists with admin role
    if _DEFAULT_ADMIN not in users:
        users[_DEFAULT_ADMIN] = {
            'password': generate_password_hash(_DEFAULT_ADMIN_PASS),
            'role': 'admin', 'email': '', 'verified': True, 'verification_token': None,
        }
        changed = True
    elif users[_DEFAULT_ADMIN].get('role') != 'admin':
        users[_DEFAULT_ADMIN]['role'] = 'admin'
        changed = True
    if changed:
        with open(_USERS_FILE, 'w') as f:
            _json.dump(users, f, indent=2)
    return users

def _save_users(users: dict) -> None:
    # Never allow demoting the default admin
    if _DEFAULT_ADMIN in users:
        users[_DEFAULT_ADMIN]['role'] = 'admin'
    os.makedirs(os.path.dirname(_USERS_FILE), exist_ok=True)
    with open(_USERS_FILE, 'w') as f:
        _json.dump(users, f, indent=2)

def _find_user_by_identifier(identifier: str, users: dict):
    """Return (key, user_dict) by username key or email address."""
    if '@' in identifier:
        for k, v in users.items():
            if v.get('email', '').lower() == identifier.lower():
                return k, v
        return None, None
    return identifier, users.get(identifier)

# ── Privacy policy change detection + notification ────────────────────────────
def _get_privacy_policy_hash() -> str:
    try:
        with open(_TEMPLATE_PATH, 'r', encoding='utf-8', errors='ignore') as f:
            content = f.read()
        start = content.find('<!-- ══ PRIVACY POLICY PAGE')
        end   = content.find('<!-- ══', start + 1) if start != -1 else -1
        section = content[start:end] if (start != -1 and end != -1) else content
        return _hashlib.sha256(section.encode()).hexdigest()[:20]
    except Exception:
        return ''

def _send_policy_update_email(to_email: str, version: str) -> None:
    if not _SMTP_USER or not _SMTP_PASS:
        return
    try:
        updated = _datetime.datetime.now().strftime('%d %B %Y')
        subject = 'CyberINK Privacy Policy Updated'
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 12px;">
            Privacy Policy Updated
          </p>
          <p class="ep" style="color:#475569;font-size:13px;line-height:1.65;margin:0 0 16px;">
            The CyberINK Privacy Policy has been updated on <strong>{updated}</strong>
            (platform version <code style="background:#eff6ff;padding:2px 6px;border-radius:4px;
            font-size:12px;color:#1d4ed8;border:1px solid #bfdbfe;">v{version}</code>).
          </p>
          <p class="ep" style="color:#475569;font-size:13px;line-height:1.65;margin:0 0 24px;">
            We recommend reviewing the updated policy to understand how your data is handled.
            You can view the full Privacy Policy by logging in to the CyberINK platform and
            navigating to <strong>Pricing &rarr; Privacy Policy</strong>.
          </p>
          <table cellpadding="0" cellspacing="0" border="0"><tr><td>
            <a href="{_BASE_URL}"
              style="display:inline-block;background:#2563eb;color:#ffffff;text-decoration:none;
                     padding:11px 28px;border-radius:8px;font-size:13px;font-weight:700;
                     font-family:-apple-system,BlinkMacSystemFont,'Segoe UI',sans-serif;">
              View Platform &rarr;
            </a>
          </td></tr></table>
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:28px;line-height:1.7;">
            This notification was sent because your account has administrator access to CyberINK.<br>
            If you have questions, contact <a href="mailto:ngclaire75@gmail.com" class="elink"
              style="color:#2563eb;">ngclaire75@gmail.com</a>.
          </p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK <{_SMTP_USER}>'
        msg['To']      = to_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, to_email, msg.as_string())
    except Exception:
        pass

def _check_privacy_policy_update() -> None:
    if not _SMTP_USER or not _SMTP_PASS:
        return
    current_hash = _get_privacy_policy_hash()
    if not current_hash:
        return
    try:
        stored_hash = ''
        os.makedirs(os.path.dirname(os.path.abspath(_POLICY_HASH_FILE)), exist_ok=True)
        if os.path.exists(_POLICY_HASH_FILE):
            with open(_POLICY_HASH_FILE, 'r') as f:
                stored_hash = f.read().strip()
        with open(_POLICY_HASH_FILE, 'w') as f:
            f.write(current_hash)
        if not stored_hash or stored_hash == current_hash:
            return  # first run or no change
        users = _load_users()
        admin_emails = [
            u.get('email', '').strip()
            for u in users.values()
            if u.get('role') == 'admin' and u.get('email', '').strip()
        ]
        for email in admin_emails:
            _threading.Thread(
                target=_send_policy_update_email,
                args=(email, _APP_VERSION),
                daemon=True
            ).start()
    except Exception:
        pass

_policy_check_done = False

@app.before_request
def _startup_policy_check():
    global _policy_check_done
    if not _policy_check_done:
        _policy_check_done = True
        _threading.Thread(target=_check_privacy_policy_update, daemon=True).start()

# ── Global auth enforcement ───────────────────────────────────────────────────
_PUBLIC_PATHS = ('/login', '/signup', '/verify/', '/logout',
                 '/api/syslog/hec', '/api/syslog/ingest', '/api/events/ingest',
                 '/api/payment/notification')

# ── Midtrans configuration ────────────────────────────────────────────────────
_MIDTRANS_SERVER_KEY     = os.environ.get('MIDTRANS_SERVER_KEY', '').strip()
_MIDTRANS_CLIENT_KEY     = os.environ.get('MIDTRANS_CLIENT_KEY', '').strip()
_MIDTRANS_IS_PRODUCTION  = os.environ.get('MIDTRANS_IS_PRODUCTION', '0').strip() == '1'

def _midtrans_api_base() -> str:
    return 'https://app.midtrans.com' if _MIDTRANS_IS_PRODUCTION else 'https://app.sandbox.midtrans.com'

def _midtrans_snap_js_url() -> str:
    base = 'https://app.midtrans.com' if _MIDTRANS_IS_PRODUCTION else 'https://app.sandbox.midtrans.com'
    return f'{base}/snap/snap.js'

def _midtrans_snap_create(order_id: str, amount: int, customer: dict, item_name: str, plan_type: str) -> dict:
    import base64 as _b64
    url  = f'{_midtrans_api_base()}/snap/v1/transactions'
    auth = _b64.b64encode(f'{_MIDTRANS_SERVER_KEY}:'.encode()).decode()
    payload: dict = {
        'transaction_details': {'order_id': order_id, 'gross_amount': int(amount)},
    }
    email = customer.get('email', '').strip()
    name  = (customer.get('name') or customer.get('username') or '').strip()
    if email or name:
        cust: dict = {}
        if name:  cust['first_name'] = name
        if email: cust['email']      = email
        payload['customer_details'] = cust
    payload['item_details'] = [
        {'id': plan_type[:50], 'price': int(amount), 'quantity': 1, 'name': item_name[:50]}
    ]
    resp = _requests.post(url, json=payload, headers={
        'Authorization': f'Basic {auth}',
        'Content-Type':  'application/json',
        'Accept':        'application/json',
    }, timeout=15)
    if not resp.ok:
        try:
            detail = resp.json()
        except Exception:
            detail = resp.text
        raise RuntimeError(f'Midtrans {resp.status_code}: {detail}')
    return resp.json()

def _midtrans_verify_signature(order_id: str, status_code: str, gross_amount: str) -> str:
    import hashlib
    raw = f'{order_id}{status_code}{gross_amount}{_MIDTRANS_SERVER_KEY}'
    return hashlib.sha512(raw.encode()).hexdigest()

@app.before_request
def _enforce_auth():
    if any(request.path.startswith(p) for p in _PUBLIC_PATHS):
        return None
    if request.path.startswith('/static/'):
        return None
    if not session.get('user'):
        if request.path.startswith('/api/'):
            return jsonify({'error': 'Not authenticated', 'redirect': '/login'}), 401
        return redirect(url_for('login_page'))
    # Enforce role-based API access
    u = session['user']
    if request.path.startswith('/api/admin/') and u.get('role') != 'admin':
        return jsonify({'error': 'Admin access required'}), 403

# ── Auth routes ───────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
def login_page():
    if session.get('user'):
        return redirect(url_for('index'))
    error   = None
    success = request.args.get('verified') == '1'
    if request.method == 'POST':
        identifier = (request.form.get('username') or '').strip()
        password   = request.form.get('password') or ''
        users      = _load_users()
        key, user  = _find_user_by_identifier(identifier, users)
        if user and check_password_hash(user['password'], password):
            if not user.get('verified', True):
                error = 'Please verify your email before logging in. Check your inbox for the verification link.'
            else:
                session['user'] = {'username': key, 'role': user['role'], 'email': user.get('email', ''),
                                   'country': user.get('country', ''), 'currency_code': user.get('currency_code', 'USD')}
                return redirect(url_for('index'))
        else:
            error = 'Invalid username/email or password.'
    return render_template('login.html', error=error, success='Account verified! You can now sign in.' if success else None)

@app.route('/signup', methods=['GET', 'POST'])
def signup_page():
    if session.get('user'):
        return redirect(url_for('index'))
    error        = None
    pending_email = None
    if request.method == 'POST':
        identifier = (request.form.get('username') or '').strip()
        password   = request.form.get('password') or ''
        confirm    = request.form.get('confirm') or ''
        is_email   = '@' in identifier and '.' in identifier.split('@', 1)[-1]
        key        = identifier.lower() if is_email else identifier
        if not identifier or not password:
            error = 'Username/email and password are required.'
        elif password != confirm:
            error = 'Passwords do not match.'
        elif len(password) < 6:
            error = 'Password must be at least 6 characters.'
        elif key == _DEFAULT_ADMIN:
            error = 'That username is reserved.'
        else:
            users = _load_users()
            email_taken = any(v.get('email', '').lower() == identifier.lower() for v in users.values())
            if key in users or email_taken:
                error = 'That username or email is already registered.'
            else:
                token    = str(_uuid.uuid4())
                verified = not is_email
                users[key] = {
                    'password': generate_password_hash(password),
                    'role': 'user',
                    'email': identifier if is_email else '',
                    'verified': verified,
                    'verification_token': token if is_email else None,
                }
                _save_users(users)
                if is_email:
                    sent = _send_verification_email(identifier, token)
                    if sent:
                        pending_email = identifier
                    else:
                        # SMTP not configured — auto-verify so user isn't stuck
                        users[key]['verified'] = True
                        users[key]['verification_token'] = None
                        _save_users(users)
                        session['user'] = {'username': key, 'role': 'user', 'email': identifier,
                                           'country': '', 'currency_code': 'USD'}
                        return redirect(url_for('index'))
                else:
                    session['user'] = {'username': key, 'role': 'user', 'email': '',
                                       'country': '', 'currency_code': 'USD'}
                    return redirect(url_for('index'))
    return render_template('signup.html', error=error, pending_email=pending_email)

@app.route('/verify/<token>')
def verify_email(token):
    users = _load_users()
    for uname, user in users.items():
        if user.get('verification_token') == token:
            user['verified'] = True
            user['verification_token'] = None
            _save_users(users)
            return redirect(url_for('login_page') + '?verified=1')
    return render_template('login.html', error='This verification link is invalid or has already been used.', success=None)

@app.route('/resend-verification', methods=['POST'])
def resend_verification():
    email = (request.form.get('email') or '').strip().lower()
    users = _load_users()
    key, user = _find_user_by_identifier(email, users)
    if user and not user.get('verified', True):
        token = str(_uuid.uuid4())
        user['verification_token'] = token
        _save_users(users)
        _send_verification_email(email, token)
    return render_template('login.html', error=None,
                           success='If that email exists and is unverified, a new link has been sent.')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login_page'))

# ── RBAC decorators ───────────────────────────────────────────────────────────
def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get('user'):
            return redirect(url_for('login_page'))
        return f(*args, **kwargs)
    return decorated


def _cu() -> dict:
    """Return current session user dict."""
    return session.get('user', {'username': 'admin', 'role': 'admin'})

def _cu_username() -> str:
    """Return current user's username."""
    return _cu().get('username', 'admin')

def _cu_filter():
    """Return username for DB scoping — None means admin (sees all scans)."""
    u = _cu()
    return None if u.get('role') == 'admin' else u.get('username', '')

def _admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        u = session.get('user')
        if not u:
            return jsonify({'error': 'Not authenticated'}), 401
        if u.get('role') != 'admin':
            return jsonify({'error': 'Admin access required'}), 403
        return f(*args, **kwargs)
    return decorated

# ── Admin user-management API ─────────────────────────────────────────────────
@app.route('/api/admin/users', methods=['GET'])
@_admin_required
def admin_list_users():
    users = _load_users()
    return jsonify({'users': [
        {
            'username':      k,
            'role':          v['role'],
            'email':         v.get('email', ''),
            'verified':      v.get('verified', True),
            'allowed_pages': v.get('allowed_pages'),  # None for admins = unrestricted
            'plan':          v.get('plan', 'basic'),
        }
        for k, v in users.items()
    ]})

@app.route('/api/admin/users', methods=['POST'])
@_admin_required
def admin_create_user():
    data     = request.get_json() or {}
    username = (data.get('username') or '').strip()
    password = data.get('password') or ''
    role     = data.get('role', 'user')
    if not username or not password:
        return jsonify({'error': 'Username and password are required'}), 400
    if len(password) < 6:
        return jsonify({'error': 'Password must be at least 6 characters'}), 400
    if role not in ('admin', 'user'):
        return jsonify({'error': 'Invalid role'}), 400
    if username == _DEFAULT_ADMIN:
        return jsonify({'error': 'That username is reserved'}), 400
    users = _load_users()
    if username in users:
        return jsonify({'error': 'Username already taken'}), 409
    users[username] = {
        'password': generate_password_hash(password),
        'role': role, 'email': '', 'verified': True, 'verification_token': None,
    }
    _save_users(users)
    return jsonify({'ok': True})

@app.route('/api/admin/users/<username>/role', methods=['POST'])
@_admin_required
def admin_set_role(username):
    data     = request.get_json() or {}
    new_role = data.get('role')
    if new_role not in ('admin', 'user'):
        return jsonify({'error': 'Invalid role'}), 400
    if username == _DEFAULT_ADMIN:
        return jsonify({'error': 'The default admin role cannot be changed'}), 400
    users = _load_users()
    if username not in users:
        return jsonify({'error': 'User not found'}), 404
    if username == session['user']['username'] and new_role != 'admin':
        return jsonify({'error': 'Cannot demote yourself'}), 400
    users[username]['role'] = new_role
    _save_users(users)
    return jsonify({'ok': True})

@app.route('/api/admin/users/<username>/plan', methods=['POST'])
@_admin_required
def admin_set_plan(username):
    data     = request.get_json() or {}
    new_plan = data.get('plan')
    if new_plan not in ('basic', 'pro'):
        return jsonify({'error': 'Plan must be basic or pro'}), 400
    users = _load_users()
    if username not in users:
        return jsonify({'error': 'User not found'}), 404
    users[username]['plan'] = new_plan
    _save_users(users)
    return jsonify({'ok': True})

@app.route('/api/admin/users/<username>/pages', methods=['GET'])
@_admin_required
def admin_get_user_pages(username):
    users = _load_users()
    if username not in users:
        return jsonify({'error': 'User not found'}), 404
    u = users[username]
    return jsonify({'allowed_pages': u.get('allowed_pages')})  # None = admin / unrestricted


@app.route('/api/admin/users/<username>/pages', methods=['POST'])
@_admin_required
def admin_set_user_pages(username):
    users = _load_users()
    if username not in users:
        return jsonify({'error': 'User not found'}), 404
    data  = request.get_json(silent=True) or {}
    pages = data.get('allowed_pages')
    if pages is not None and not isinstance(pages, list):
        return jsonify({'error': 'allowed_pages must be a list'}), 400
    users[username]['allowed_pages'] = pages
    _save_users(users)
    return jsonify({'ok': True})


@app.route('/api/admin/users/<username>/verify', methods=['POST'])
@_admin_required
def admin_verify_user(username):
    users = _load_users()
    if username not in users:
        return jsonify({'error': 'User not found'}), 404
    users[username]['verified'] = True
    users[username]['verification_token'] = None
    _save_users(users)
    return jsonify({'ok': True})

@app.route('/api/admin/users/<username>', methods=['DELETE'])
@_admin_required
def admin_delete_user(username):
    if username == _DEFAULT_ADMIN:
        return jsonify({'error': 'The default admin account cannot be deleted'}), 400
    if username == session['user']['username']:
        return jsonify({'error': 'Cannot delete yourself'}), 400
    users = _load_users()
    if username not in users:
        return jsonify({'error': 'User not found'}), 404
    del users[username]
    _save_users(users)
    return jsonify({'ok': True})

# ── Account self-service ──────────────────────────────────────────────────────
@app.route('/api/account/update', methods=['POST'])
def account_update():
    if 'user' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    data             = request.get_json(force=True) or {}
    action           = data.get('action', '')
    current_username = session['user']['username']
    users            = _load_users()
    if current_username not in users:
        return jsonify({'error': 'User not found'}), 404
    user = users[current_username]

    if action == 'username':
        new_un = (data.get('username') or '').strip()
        if not new_un:
            return jsonify({'error': 'Username is required'}), 400
        if new_un == current_username:
            return jsonify({'error': 'That is already your username'}), 400
        if new_un.lower() == _DEFAULT_ADMIN.lower():
            return jsonify({'error': 'That username is reserved'}), 400
        if new_un in users:
            return jsonify({'error': 'Username already taken'}), 400
        users[new_un] = users.pop(current_username)
        _save_users(users)
        session['user']['username'] = new_un
        return jsonify({'ok': True, 'message': 'Username updated. You are now logged in as ' + new_un + '.'})

    elif action == 'email':
        new_email = (data.get('email') or '').strip().lower()
        if not new_email or '@' not in new_email or '.' not in new_email.split('@', 1)[-1]:
            return jsonify({'error': 'A valid email address is required'}), 400
        for k, u in users.items():
            if k != current_username and u.get('email', '').lower() == new_email:
                return jsonify({'error': 'Email address already in use by another account'}), 400
        user['email'] = new_email
        _save_users(users)
        session['user']['email'] = new_email
        return jsonify({'ok': True, 'message': 'Email address updated successfully.'})

    elif action == 'password':
        cur_pass     = data.get('current_password', '')
        new_pass     = data.get('new_password', '')
        confirm_pass = data.get('confirm_password', '')
        if not cur_pass or not new_pass:
            return jsonify({'error': 'Current password and new password are required'}), 400
        if not check_password_hash(user['password'], cur_pass):
            return jsonify({'error': 'Current password is incorrect'}), 400
        if new_pass != confirm_pass:
            return jsonify({'error': 'New passwords do not match'}), 400
        if len(new_pass) < 6:
            return jsonify({'error': 'Password must be at least 6 characters'}), 400
        if cur_pass == new_pass:
            return jsonify({'error': 'New password must differ from the current one'}), 400
        user['password'] = generate_password_hash(new_pass)
        _save_users(users)
        return jsonify({'ok': True, 'message': 'Password updated successfully. Use your new password next time you sign in.'})

    elif action == 'country':
        country       = (data.get('country') or '').strip()
        currency_code = (data.get('currency_code') or '').strip().upper()
        if not country:
            return jsonify({'error': 'Country is required'}), 400
        if not currency_code:
            return jsonify({'error': 'Currency is required'}), 400
        user['country']       = country
        user['currency_code'] = currency_code
        _save_users(users)
        session['user']['country']       = country
        session['user']['currency_code'] = currency_code
        session.modified = True
        return jsonify({'ok': True, 'message': f'Country set to {country} ({currency_code}).'})

    return jsonify({'error': 'Invalid action'}), 400


# ── Customer Service / Contact ────────────────────────────────────────────────
_SUPPORT_EMAIL = 'darynnclaire88@gmail.com'

def _send_contact_email(category: str, title: str, full_name: str,
                        from_email: str, message: str, username: str) -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        subject = f'[CyberINK Support] {category} — {title}'
        safe_msg = message.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 16px;">
            New Support Request
          </p>
          <table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:20px;">
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:110px;">Category</td>
              <td style="padding:7px 0;color:#2563eb;">{category}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Title</td>
              <td style="padding:7px 0;color:#2563eb;">{title}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Full Name</td>
              <td style="padding:7px 0;color:#2563eb;">{full_name}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Email</td>
              <td style="padding:7px 0;color:#2563eb;"><a href="mailto:{from_email}" style="color:#2563eb;">{from_email}</a></td>
            </tr>
            <tr>
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Account</td>
              <td style="padding:7px 0;color:#2563eb;">{username}</td>
            </tr>
          </table>
          <div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;
                      letter-spacing:.4px;margin-bottom:8px;">Message</div>
          <div style="background:#f8faff;border:1px solid #bfdbfe;border-radius:8px;
                      padding:14px 16px;font-size:13px;color:#1e3a8a;line-height:1.65;">
            {safe_msg}
          </div>
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
            Reply directly to this email to respond to the sender at
            <a href="mailto:{from_email}" class="elink" style="color:#2563eb;">{from_email}</a>.
          </p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Support <{_SMTP_USER}>'
        msg['To']      = _SUPPORT_EMAIL
        msg['Reply-To'] = from_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, _SUPPORT_EMAIL, msg.as_string())
        return True
    except Exception:
        return False


def _send_contact_ack_email(category: str, title: str, full_name: str,
                              to_email: str, message: str) -> None:
    if not _SMTP_USER or not _SMTP_PASS:
        return
    try:
        import datetime as _dt
        submitted_at = _dt.datetime.utcnow().strftime('%d %B %Y, %H:%M UTC')
        safe_msg = message.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')
        subject = f'CyberINK — We received your message: {title}'
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 8px;">
            Message Received
          </p>
          <p class="ep" style="color:#3b82f6;font-size:13px;margin:0 0 20px;line-height:1.6;">
            Hello {full_name}, thank you for reaching out to CyberINK Security. We have successfully received your message and it has been forwarded to our Data Protection Officer (DPO). You can expect a reply within <strong style="color:#1e3a8a;">3 business days</strong>.
          </p>

          <div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;letter-spacing:.4px;margin-bottom:8px;">Your Submission</div>
          <table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:20px;">
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:120px;">Category</td>
              <td style="padding:7px 0;color:#2563eb;">{category}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Subject</td>
              <td style="padding:7px 0;color:#2563eb;">{title}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Full Name</td>
              <td style="padding:7px 0;color:#2563eb;">{full_name}</td>
            </tr>
            <tr>
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Submitted</td>
              <td style="padding:7px 0;color:#64748b;">{submitted_at}</td>
            </tr>
          </table>

          <div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;letter-spacing:.4px;margin-bottom:8px;">Your Message</div>
          <div style="background:#f8faff;border:1px solid #bfdbfe;border-radius:8px;padding:14px 16px;font-size:13px;color:#1e3a8a;line-height:1.65;margin-bottom:20px;">
            {safe_msg}
          </div>

          <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;padding:14px 16px;font-size:12px;color:#1e3a8a;line-height:1.7;">
            <strong>What happens next?</strong><br>
            Our DPO will review your enquiry and respond to this email address within <strong>3 business days</strong>. If your matter is urgent, you may also reach us directly at
            <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.
          </div>

          <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
            Please keep this email as a record of your submission. Do not reply to this message — our team will contact you directly from <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.
          </p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Security <{_SMTP_USER}>'
        msg['To']      = to_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, to_email, msg.as_string())
    except Exception:
        pass


@app.route('/api/contact', methods=['POST'])
def contact_submit():
    if 'user' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    data      = request.get_json(force=True) or {}
    category  = (data.get('category')  or '').strip()
    title     = (data.get('title')     or '').strip()
    full_name = (data.get('full_name') or '').strip()
    email     = (data.get('email')     or '').strip()
    message   = (data.get('message')   or '').strip()
    if not all([category, title, full_name, email, message]):
        return jsonify({'error': 'All fields are required'}), 400
    if '@' not in email or '.' not in email.split('@', 1)[-1]:
        return jsonify({'error': 'A valid email address is required'}), 400
    if len(message) < 10:
        return jsonify({'error': 'Message must be at least 10 characters'}), 400
    ok = _send_contact_email(category, title, full_name, email, message, session['user']['username'])
    if not ok:
        return jsonify({'error': 'Failed to send message. Please try again later or contact support directly.'}), 500
    _send_contact_ack_email(category, title, full_name, email, message)
    return jsonify({'ok': True, 'message': 'Your message has been sent. We will respond within 3 business days.'})


# ── Temporary password ───────────────────────────────────────────────────────
import string as _string
import secrets as _secrets

def _gen_temp_password(length: int = 12) -> str:
    alphabet = _string.ascii_letters + _string.digits
    return ''.join(_secrets.choice(alphabet) for _ in range(length))

def _send_temp_password_email(to_email: str, username: str, temp_pass: str) -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        subject = 'CyberINK — Your Temporary Password'
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 12px;">
            Temporary Password Request
          </p>
          <p class="ep" style="color:#475569;font-size:13px;line-height:1.65;margin:0 0 20px;">
            A temporary password has been generated for your CyberINK account
            <strong style="color:#1e3a8a;">{username}</strong>.
            Use it to fill in the <strong>Current Password</strong> field on the
            Account Settings page, then set a new password immediately.
          </p>
          <div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;
                      padding:16px 20px;margin-bottom:20px;text-align:center;">
            <div style="font-size:10px;font-weight:700;color:#60a5fa;text-transform:uppercase;
                        letter-spacing:.5px;margin-bottom:6px;">Temporary Password</div>
            <div style="font-family:monospace;font-size:22px;font-weight:800;
                        color:#1e3a8a;letter-spacing:2px;">{temp_pass}</div>
          </div>
          <p class="ep" style="color:#475569;font-size:13px;line-height:1.65;margin:0 0 8px;">
            Steps to regain access:
          </p>
          <ol style="font-size:13px;color:#475569;line-height:1.8;margin:0 0 20px;padding-left:20px;">
            <li>Go to <strong>Account Settings</strong> in the CyberINK sidebar.</li>
            <li>Under <strong>Change Password</strong>, enter the temporary password above in the <strong>Current Password</strong> field.</li>
            <li>Enter and confirm your new password, then click <strong>Update Password</strong>.</li>
          </ol>
          <p class="ep" style="color:#64748b;font-size:11px;margin:0;line-height:1.6;">
            If you did not request this, please contact support immediately at
            <a href="mailto:{_SUPPORT_EMAIL}" class="elink" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.
            Your previous password has already been replaced by this temporary one.
          </p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK <{_SMTP_USER}>'
        msg['To']      = to_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, to_email, msg.as_string())
        return True
    except Exception:
        return False


@app.route('/api/account/temp-password', methods=['POST'])
def account_temp_password():
    if 'user' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    current_username = session['user']['username']
    users = _load_users()
    if current_username not in users:
        return jsonify({'error': 'User not found'}), 404
    user  = users[current_username]
    email = user.get('email', '').strip()
    if not email:
        return jsonify({'error': 'No email address is registered on this account. '
                                 'Please contact an administrator to reset your password.'}), 400
    temp_pass = _gen_temp_password()
    user['password'] = generate_password_hash(temp_pass)
    _save_users(users)
    ok = _send_temp_password_email(email, current_username, temp_pass)
    if not ok:
        return jsonify({'error': 'Email could not be sent. SMTP may not be configured. '
                                 'Please contact an administrator.'}), 500
    return jsonify({'ok': True,
                    'message': f'A temporary password has been sent to {email}. '
                               'Use it in the Current Password field, then set a new password immediately.'})


# ── In-memory scan job store (Connect Your Website feature) ──────────────────
_scan_jobs: dict = {}

# ── Pentest engagement background jobs ────────────────────────────────────────
_jobs: dict = {}

_JOB_DIR = os.path.join(os.path.dirname(__file__), '..', 'data', 'jobs')

def _job_persist(job_id: str, job: dict) -> None:
    """Write terminal job state to disk so polls survive server restarts."""
    try:
        os.makedirs(_JOB_DIR, exist_ok=True)
        path = os.path.join(_JOB_DIR, f'{job_id}.json')
        with open(path, 'w') as fh:
            _json.dump({
                'status':  job.get('status', 'done'),
                'scan_id': job.get('scan_id'),
                'error':   job.get('error'),
                'domain':  job.get('domain', ''),
                'target':  job.get('target', ''),
            }, fh)
    except Exception:
        pass

def _job_load(job_id: str) -> dict | None:
    """Try to load a completed job from disk (fallback after server restart)."""
    try:
        path = os.path.join(_JOB_DIR, f'{job_id}.json')
        if os.path.exists(path):
            with open(path) as fh:
                return _json.load(fh)
    except Exception:
        pass
    return None

# ── Plugin inventory parser ───────────────────────────────────────────────────

def _parse_and_save_plugins(scan_id: int, target: str, output: str, username: str = '') -> int:
    """Extract plugins/packages from scan output and upsert them into the DB."""
    found: dict = {}

    # WordPress wp plugin list table: | slug | active | 1.2.3 |
    for m in re.finditer(
            r'\|\s*([\w][\w-]{1,60})\s*\|\s*(active|inactive|must-use|drop-in)\s*\|\s*([\d.]+)\s*\|',
            output, re.I):
        slug, status, ver = m.group(1).strip(), m.group(2).strip().lower(), m.group(3).strip()
        key = slug.lower()
        found[key] = {'name': slug, 'version': ver, 'plugin_type': 'WordPress Plugin',
                      'status': status, 'vulnerable': 0}

    # WordPress REST API JSON blobs: "slug":"name","version":"x.y"
    for m in re.finditer(r'"slug"\s*:\s*"([^"]{2,60})"[^}]{0,200}?"version"\s*:\s*"([^"]+)"',
                         output, re.I | re.S):
        name, ver = m.group(1).strip(), m.group(2).strip()
        key = name.lower()
        if key not in found:
            found[key] = {'name': name, 'version': ver, 'plugin_type': 'WordPress Plugin',
                          'status': 'active', 'vulnerable': 0}

    # wp-content/plugins/slug paths
    for m in re.finditer(r'wp-content/plugins/([\w-]{3,60})', output, re.I):
        slug = m.group(1).strip()
        key = slug.lower()
        if key not in found:
            found[key] = {'name': slug, 'version': '', 'plugin_type': 'WordPress Plugin',
                          'status': 'active', 'vulnerable': 0}

    # WP-LOG plugin activation/deactivation entries
    for m in re.finditer(
            r'WP-LOG[^\n]*(?:activated|deactivated|installed)\s+(?:plugin\s+)?[:\s]+"?([^"|\n]{3,60}?)"?\s*\|',
            output, re.I):
        name = m.group(1).strip().rstrip('.')
        key = name.lower()
        if key not in found:
            found[key] = {'name': name, 'version': '', 'plugin_type': 'WordPress Plugin',
                          'status': 'active', 'vulnerable': 0}

    # wp_security_scan output: [ACTIVE] Plugin Name v1.0.0 — slug/slug.php
    for m in re.finditer(
            r'\[(ACTIVE|INACTIVE|MUST-USE|DROP-IN)\]\s+(.+?)\s+v([\d][0-9.]*)',
            output, re.I):
        status = m.group(1).lower()
        name   = m.group(2).strip().rstrip(' —-')
        ver    = m.group(3).strip()
        key    = name.lower()
        if key not in found and 2 < len(name) < 80:
            found[key] = {'name': name, 'version': ver, 'plugin_type': 'WordPress Plugin',
                          'status': status, 'vulnerable': 0}

    # Active theme line: "Active theme: ThemeName v1.0"
    for m in re.finditer(r'active theme:\s+(.+?)\s+v([\d][0-9.]*)', output, re.I):
        name = m.group(1).strip()
        ver  = m.group(2).strip()
        key  = name.lower()
        if key not in found and len(name) < 80:
            found[key] = {'name': name, 'version': ver, 'plugin_type': 'WordPress Theme',
                          'status': 'active', 'vulnerable': 0}

    # Generic plugin lines: "Plugin: name v1.0" or "Detected plugin: name 1.0"
    for m in re.finditer(
            r'(?:detected\s+)?plugin[:\s]+([a-zA-Z][\w ._-]{2,50}?)\s+v?([\d]+\.[\d.]+)',
            output, re.I):
        name, ver = m.group(1).strip(), m.group(2).strip()
        key = name.lower()
        if key not in found and len(name) < 60:
            found[key] = {'name': name, 'version': ver, 'plugin_type': 'Plugin',
                          'status': 'active', 'vulnerable': 0}

    # npm packages: "package-name@version" patterns in output
    for m in re.finditer(r'(?:^|\s)([@\w][\w/-]{1,50})@([\d]+\.[\d.]+)', output, re.M):
        name, ver = m.group(1).strip(), m.group(2).strip()
        if name.startswith('@') or '/' in name:
            continue  # skip scoped packages
        key = name.lower()
        if key not in found:
            found[key] = {'name': name, 'version': ver, 'plugin_type': 'npm Package',
                          'status': 'active', 'vulnerable': 0}

    # Mark known-vulnerable plugins
    vuln_ctx = re.findall(
        r'(?:CVE-\d{4}-\d{4,}|vulnerable|exploit|critical\s+vuln)[^\n]{0,120}',
        output, re.I)
    for ctx in vuln_ctx:
        for key in list(found.keys()):
            if key in ctx.lower():
                found[key]['vulnerable'] = 1

    # Persist to DB
    for data in found.values():
        try:
            db.upsert_plugin(
                target=target, name=data['name'], version=data['version'],
                plugin_type=data['plugin_type'], status=data['status'],
                vulnerable=data['vulnerable'], scan_id=scan_id, username=username,
            )
        except Exception:
            pass

    return len(found)

# ── IP geolocation (ip-api.com, free, no key required) ───────────────────────
_geo_cache: dict[str, str] = {}

def _geoip(ip_or_url: str) -> str:
    """Return 'Country (City)' for a real IP or hostname. Returns '' on failure."""
    raw = (ip_or_url or '').strip()
    if not raw or raw in ('-', '--', ''):
        return ''
    # Extract hostname from URL
    if raw.startswith('http'):
        raw = _up_parse.urlparse(raw).netloc or raw
    ip = raw.split(':')[0].strip()
    if not ip:
        return ''
    if ip in _geo_cache:
        return _geo_cache[ip]
    # Skip private/reserved IPs
    try:
        addr = ipaddress.ip_address(ip)
        if addr.is_private or addr.is_loopback or addr.is_reserved:
            _geo_cache[ip] = ''
            return ''
    except ValueError:
        pass  # hostname — proceed with lookup
    try:
        url = f'http://ip-api.com/json/{_up_parse.quote(ip)}?fields=status,country,city'
        req = _up_req.Request(url, headers={'User-Agent': 'CF_AI/1.0'})
        with _up_req.urlopen(req, timeout=5) as r:
            data = _json.loads(r.read().decode())
        if data.get('status') == 'success':
            country = data.get('country', '')
            city    = data.get('city', '')
            result  = f'{country} ({city})' if city else country
            _geo_cache[ip] = result
            return result
    except Exception:
        pass
    _geo_cache[ip] = ''
    return ''


def _build_cred_block(site_type: str, creds: dict, domain: str) -> str:
    """Return an agent instruction block for authenticated scanning."""
    if not site_type or site_type == 'none':
        return ''

    wp_user  = creds.get('wp_user', '')
    wp_pass  = creds.get('wp_pass', '')
    wp_app   = creds.get('wp_app_pass', '')
    cp_user  = creds.get('cpanel_user', '')
    ssh_host = creds.get('ssh_host', '') or domain
    ssh_user = creds.get('ssh_user', 'root')
    ssh_pass = creds.get('ssh_pass', '')
    ssh_port = creds.get('ssh_port', '22') or '22'
    ftp_host = creds.get('ftp_host', '') or domain
    ftp_user = creds.get('ftp_user', '')
    ftp_port = creds.get('ftp_port', '') or ('22' if site_type == 'sftp' else '21')

    hdr = (
        '\n\n══════════════ AUTHENTICATED SCAN ══════════════\n'
        'Credentials provided. You MUST use them for every check.\n'
        'NEVER print passwords in output — write [REDACTED] instead.\n'
        '════════════════════════════════════════════════\n\n'
    )

    if site_type == 'wordpress' and (wp_user or wp_pass):
        # Pre-compute conditional strings — avoids nested f-strings inside f-string expressions
        app_status  = '(provided — use for REST API)' if wp_app else '(not provided — cookie auth only)'
        rest_users  = (f'curl -s -u "{wp_user}:{wp_app}" '
                       f'"https://{domain}/wp-json/wp/v2/users?context=edit&per_page=100"'
                       if wp_app else
                       f'curl -s -b /tmp/wp_auth.txt "https://{domain}/wp-json/wp/v2/users?context=edit"')
        rest_plugins = (f'curl -s -u "{wp_user}:{wp_app}" "https://{domain}/wp-json/wp/v2/plugins"'
                        if wp_app else
                        '# app password required for /wp-json/wp/v2/plugins — skipping')
        return (hdr
            + 'WORDPRESS ADMIN CREDENTIALS\n'
            + f'  Username : {wp_user}\n'
            + f'  App Pass : {app_status}\n\n'
            + 'Run ALL of these checks in order:\n\n'
            + '1. Login and capture session cookie (required for most checks below):\n'
            + f'   curl -s -L -c /tmp/wp_auth.txt -b /tmp/wp_auth.txt \\\n'
            + f'     -d "log={wp_user}&pwd=$WP_PASSWORD&wp-submit=Log+In&redirect_to=%2Fwp-admin%2F&testcookie=1" \\\n'
            + '     -H "Cookie: wordpress_test_cookie=WP+Cookie+check" \\\n'
            + f'     "https://{domain}/wp-login.php" -w "%{{http_code}}" -o /tmp/wp_login_resp.html\n'
            + '   grep -iP "error|invalid|incorrect" /tmp/wp_login_resp.html | head -5\n\n'
            + '2. Enumerate all WordPress users (admin view — reveals roles):\n'
            + f'   curl -s -b /tmp/wp_auth.txt "https://{domain}/wp-admin/users.php" \\\n'
            + "     | grep -oP '(?<=user-login\">)[^<]+'\n\n"
            + '3. Installed plugins and versions (identify outdated/vulnerable):\n'
            + f'   curl -s -b /tmp/wp_auth.txt "https://{domain}/wp-admin/plugins.php" \\\n'
            + "     | grep -oP '(?<=<strong>)[^<]+|(?<=Version )[0-9.]+'\n\n"
            + '4. WordPress core version and debug/security settings:\n'
            + f'   curl -s -b /tmp/wp_auth.txt "https://{domain}/wp-admin/about.php" | grep -oP "(?<=Version )[\\d.]+"\n'
            + f'   curl -s -b /tmp/wp_auth.txt "https://{domain}/wp-admin/options-general.php" \\\n'
            + '     | grep -iP "debug|ssl_force|login_lockout|two.factor|recaptcha"\n\n'
            + '5. REST API with real auth (lists all users including admin):\n'
            + f'   {rest_users}\n'
            + f'   {rest_plugins}\n\n'
            + '6. Admin AJAX — test for unauthenticated fallback:\n'
            + f'   curl -s -b /tmp/wp_auth.txt -d "action=heartbeat" "https://{domain}/wp-admin/admin-ajax.php"\n\n'
            + '7. File editor check (should be disabled — enables RCE):\n'
            + f'   curl -s -b /tmp/wp_auth.txt "https://{domain}/wp-admin/theme-editor.php" \\\n'
            + '     | grep -iP "disabled|not allowed|higher level"\n\n'
            + '8. XML-RPC authenticated call (test for DDoS amplification):\n'
            + "   curl -s -d '<?xml version=\"1.0\"?><methodCall><methodName>system.listMethods"
            + f'</methodName></methodCall>\' "https://{domain}/xmlrpc.php"'
            + " | grep -oP '(?<=string>)[^<]+' | head -20\n\n"
            + '9. WP_DEBUG log and error exposure:\n'
            + f'   curl -s "https://{domain}/wp-content/debug.log" | head -30\n'
            + f'   curl -s "https://{domain}/?debug=1" | grep -iP "fatal|error|warning|deprecated"\n'
        )

    if site_type == 'cpanel' and cp_user:
        cp = f'curl -sk -u "{cp_user}:$CPANEL_PASSWORD"'
        api = f'https://{domain}:2083/execute'
        # python3 snippet using % to avoid brace conflicts with f-string
        py_filter = (
            "python3 -c \"import sys,json; d=json.load(sys.stdin); "
            "[print(f['file']) for f in d.get('data',{}).get('files',[]) "
            "if any(f['file'].endswith(e) for e in ['.env','.sql','.zip','.bak','.tar'])]\""
        )
        return (hdr
            + 'CPANEL CREDENTIALS\n'
            + f'  Username : {cp_user}\n'
            + f'  API base : https://{domain}:2083  (try :2082 for HTTP)\n\n'
            + 'Run ALL cPanel UAPI checks:\n\n'
            + f'1. PHP versions:\n   {cp} "{api}/LangPHP/php_get_vhost_versions"\n'
            + f'   {cp} "{api}/LangPHP/php_get_installed_versions"\n\n'
            + f'2. SSL certificate:\n   {cp} "{api}/SSL/fetch_best_for_domain?domain={domain}"\n\n'
            + f'3. All domains/subdomains:\n   {cp} "{api}/DomainInfo/domains_data?format=json"\n'
            + f'   {cp} "{api}/SubDomain/listsubdomains"\n\n'
            + f'4. Email accounts:\n   {cp} "{api}/Email/list_pops"\n\n'
            + f'5. MySQL databases and users:\n   {cp} "{api}/Mysql/list_databases"\n'
            + f'   {cp} "{api}/Mysql/list_users"\n\n'
            + f'6. Cron jobs:\n   {cp} "{api}/Cron/list_cron"\n\n'
            + f'7. Files in public_html (find sensitive files):\n'
            + f'   {cp} "{api}/Fileman/list_files?path=/public_html&show_hidden=1" | {py_filter}\n\n'
            + f'8. ModSecurity status:\n   {cp} "{api}/ModSecurity/has_modsec_installed"\n\n'
            + f'9. Hotlink protection:\n   {cp} "{api}/Hotlink/get_status"\n\n'
            + f'10. .htaccess security rules:\n    curl -s "https://{domain}/.htaccess" | head -50\n'
        )

    if site_type == 'ssh' and ssh_user and (ssh_pass or creds.get('ssh_key')):
        if creds.get('ssh_key'):
            key_setup = ('Setup SSH key first:\n'
                         '  python3 -c "import os; open(\'/tmp/cf_id_rsa\',\'w\').write(os.environ[\'SSH_KEY\']); '
                         'os.chmod(\'/tmp/cf_id_rsa\', 0o600)"\n')
            sc = (f'ssh -o StrictHostKeyChecking=no -o ConnectTimeout=10 '
                  f'-i /tmp/cf_id_rsa -p {ssh_port} {ssh_user}@{ssh_host}')
        else:
            key_setup = 'Install sshpass if missing: apt-get install -y sshpass 2>/dev/null\n'
            sc = (f'sshpass -e ssh -o StrictHostKeyChecking=no -o ConnectTimeout=10 '
                  f'-p {ssh_port} {ssh_user}@{ssh_host}')
        # awk command — written as plain string, no f-string brace conflicts
        awk_users  = "awk -F: '$7!~/nologin|false|sync/{print $1,$7}' /etc/passwd"
        awk_bruteforce = "awk '{print $11}' | sort | uniq -c | sort -rn | head -10"
        return (hdr
            + 'SSH CREDENTIALS\n'
            + f'  Host     : {ssh_host}:{ssh_port}\n'
            + f'  Username : {ssh_user}\n'
            + '  Password : $SSHPASS (in environment — do not print)\n\n'
            + key_setup + '\n'
            + f'SSH prefix: {sc} "<cmd>"\n\n'
            + 'Run ALL server-side checks:\n\n'
            + f'1. Web/PHP/OS versions:\n'
            + f'   {sc} "php -v 2>&1|head -1; nginx -v 2>&1; apache2 -v 2>&1|head -1; lsb_release -d 2>/dev/null"\n\n'
            + f'2. Sensitive config files (hardcoded credentials):\n'
            + f'   {sc} "cat /var/www/html/.env 2>/dev/null|grep -vP \'^#|^$\'|head -30"\n'
            + f'   {sc} "cat /var/www/html/wp-config.php 2>/dev/null|grep -P \'DB_|AUTH_KEY|SECRET\'|head -20"\n\n'
            + f'3. World-writable PHP files (malware injection risk):\n'
            + f'   {sc} "find /var/www/html -perm -0002 -name \'*.php\' 2>/dev/null|head -20"\n\n'
            + f'4. Backup/dump files in webroot (data exposure):\n'
            + f'   {sc} "find /var/www/html -name \'*.sql\' -o -name \'*.zip\' -o -name \'*.bak\' 2>/dev/null|head -20"\n\n'
            + f'5. Open ports and services:\n'
            + f'   {sc} "ss -tlnp 2>/dev/null|head -25"\n\n'
            + f'6. Users with shell access:\n'
            + f'   {sc} "{awk_users}"\n\n'
            + f'7. Sudo rules (overly permissive = high risk):\n'
            + f'   {sc} "sudo -l 2>/dev/null|head -20"\n\n'
            + f'8. Brute force evidence (failed SSH logins):\n'
            + f'   {sc} "grep \'Failed password\' /var/log/auth.log 2>/dev/null|tail -30|{awk_bruteforce}"\n\n'
            + f'9. Crontabs (check for backdoors):\n'
            + f'   {sc} "crontab -l 2>/dev/null; ls /etc/cron* 2>/dev/null"\n\n'
            + f'10. SSL certificate expiry:\n'
            + f'    {sc} "openssl s_client -connect {domain}:443 -servername {domain} </dev/null 2>/dev/null'
            + '     | openssl x509 -noout -dates 2>/dev/null"\n\n'
            + f'11. Firewall rules:\n'
            + f'    {sc} "ufw status 2>/dev/null; iptables -L INPUT -n 2>/dev/null|head -20"\n'
        )

    if site_type == 'mysql' and creds.get('db_host') and creds.get('db_name'):
        db_host   = creds.get('db_host', '')
        db_port   = creds.get('db_port', '3306') or '3306'
        db_name   = creds.get('db_name', '')
        db_user   = creds.get('db_user', '')
        db_prefix = creds.get('db_prefix', 'wp_') or 'wp_'
        mc = (f'mysql -h "{db_host}" -P {db_port} -u "{db_user}" -p"$DB_PASSWORD" '
              f'--connect-timeout=10 "{db_name}"')
        return (hdr
            + 'MYSQL / DATABASE CREDENTIALS\n'
            + f'  Host     : {db_host}:{db_port}\n'
            + f'  Database : {db_name}\n'
            + f'  User     : {db_user}\n'
            + f'  Prefix   : {db_prefix}\n'
            + '  Password : $DB_PASSWORD (in environment — do not print)\n\n'
            + 'Run ALL MySQL security checks:\n\n'
            + f'1. WordPress users (admins + last login):\n'
            + f'   {mc} -e "SELECT user_login,user_email,user_registered FROM {db_prefix}users ORDER BY user_registered DESC LIMIT 30;"\n\n'
            + f'2. WordPress options (siteurl, admin email, active plugins):\n'
            + f'   {mc} -e "SELECT option_name,option_value FROM {db_prefix}options WHERE option_name IN (\'siteurl\',\'admin_email\',\'blogname\',\'active_plugins\',\'blogdescription\');"\n\n'
            + f'3. Recent Simple History activity log (login, plugin changes):\n'
            + f'   {mc} -e "SELECT date,initiator,action,object_type,object_name FROM {db_prefix}simple_history ORDER BY id DESC LIMIT 50;" 2>/dev/null || '
            + f'{mc} -e "SELECT created_at,user_login,action FROM {db_prefix}wsal_occurrences oc JOIN {db_prefix}wsal_metadata m ON oc.id=m.occurrence_id WHERE m.name=\'username\' ORDER BY oc.created_at DESC LIMIT 50;" 2>/dev/null\n\n'
            + f'4. Wordfence login log (brute force evidence):\n'
            + f'   {mc} -e "SELECT ctime,IP,username,hitcount,blockedHits FROM {db_prefix}wflogins ORDER BY ctime DESC LIMIT 30;" 2>/dev/null\n\n'
            + f'5. Wordfence blocked IPs:\n'
            + f'   {mc} -e "SELECT IP,blockedTime,reason FROM {db_prefix}wfblockediplog ORDER BY blockedTime DESC LIMIT 20;" 2>/dev/null\n\n'
            + f'6. User capabilities (check for hidden admins):\n'
            + f'   {mc} -e "SELECT user_id,meta_value FROM {db_prefix}usermeta WHERE meta_key=\'{db_prefix}capabilities\' AND meta_value LIKE \'%administrator%\';"\n\n'
            + f'7. Recent user registrations (last 30 days):\n'
            + f'   {mc} -e "SELECT user_login,user_email,user_registered FROM {db_prefix}users WHERE user_registered > DATE_SUB(NOW(),INTERVAL 30 DAY) ORDER BY user_registered DESC;"\n\n'
            + f'8. Active plugins list from DB:\n'
            + f'   {mc} -e "SELECT option_value FROM {db_prefix}options WHERE option_name=\'active_plugins\';" | tr \',\' \'\\n\' | grep -oP \'[a-z0-9-]+/[a-z0-9-]+\\.php\'\n\n'
            + f'9. Scheduled events (check for malicious crons):\n'
            + f'   {mc} -e "SELECT option_value FROM {db_prefix}options WHERE option_name=\'cron\';" | python3 -c "import sys,json; d=json.loads(sys.stdin.read()); [print(k,list(v.keys())) for k,v in d.items() if k!=\'version\']" 2>/dev/null\n'
        )

    if site_type == 'sftp' and ftp_user and ftp_host:
        proto  = 'sftp' if site_type == 'sftp' else 'ftp'
        cp_ftp = f'curl -sk --user "{ftp_user}:$FTP_PASSWORD"'
        base   = f'{proto}://{ftp_host}:{ftp_port}/public_html'
        sens_files = '.env wp-config.php config.php database.php settings.php .htpasswd'
        bak_files  = 'backup.sql backup.zip site.sql dump.sql site-backup.tar.gz'
        return (hdr
            + 'SFTP/FTP CREDENTIALS\n'
            + f'  Host     : {ftp_host}:{ftp_port}\n'
            + f'  Username : {ftp_user}\n'
            + '  Password : $FTP_PASSWORD (in environment)\n\n'
            + 'Run ALL file system checks:\n\n'
            + f'1. List webroot:\n   {cp_ftp} "{base}/" 2>&1|head -50\n\n'
            + f'2. Sensitive files (check each for 200/non-000 response):\n'
            + '   for f in ' + sens_files + '; do\n'
            + f'     code=$({cp_ftp} "{base}/$f" -o /tmp/ftp_f -w "%{{http_code}}" 2>&1)\n'
            + '     [ "$code" != "000" ] && echo "FOUND $f ($code)" && head -20 /tmp/ftp_f\n'
            + '   done\n\n'
            + f'3. Backup/dump files:\n'
            + '   for b in ' + bak_files + '; do\n'
            + f'     echo -n "$b: "; {cp_ftp} -o /dev/null -w "%{{http_code}}" "{base}/$b" 2>&1\n'
            + '     echo\n'
            + '   done\n\n'
            + f'4. Exposed .git directory:\n'
            + f'   curl -s "https://{domain}/.git/HEAD"|head -5\n'
            + f'   curl -s "https://{domain}/.git/config"|head -20\n\n'
            + f'5. PHP config:\n'
            + f'   {cp_ftp} "{base}/php.ini" 2>&1|grep -iP "disable_functions|open_basedir|expose_php"\n\n'
            + f'6. .htaccess security rules:\n'
            + f'   {cp_ftp} "{base}/.htaccess" 2>&1|head -40\n\n'
            + f'7. Uploads directory — check for PHP shells:\n'
            + f'   {cp_ftp} "{base}/wp-content/uploads/" 2>&1|grep -iP "\\.php|\\.phtml|\\.php5"\n'
        )

    if site_type == 'drupal' and creds.get('drupal_user'):
        d_url  = (creds.get('drupal_url') or f'https://{domain}').rstrip('/')
        d_user = creds.get('drupal_user', '')
        return (hdr
            + 'DRUPAL CREDENTIALS\n'
            + f'  URL      : {d_url}\n'
            + f'  Username : {d_user}\n'
            + '  Password : $DRUPAL_PASSWORD (in environment)\n\n'
            + 'Run ALL Drupal security checks:\n\n'
            + f'1. Drupal version:\n'
            + f'   curl -s "{d_url}/CHANGELOG.txt"|head -5\n'
            + f'   curl -s "{d_url}/core/CHANGELOG.txt"|head -5\n\n'
            + f'2. Obtain session token:\n'
            + f'   curl -s -c /tmp/drupal_auth.txt -b /tmp/drupal_auth.txt \\\n'
            + f'     -d "name={d_user}&pass=$DRUPAL_PASSWORD&form_id=user_login_form&op=Log+in" \\\n'
            + f'     "{d_url}/user/login"\n\n'
            + f'3. JSON:API — list users (requires admin):\n'
            + f'   curl -s -b /tmp/drupal_auth.txt -H "Accept: application/vnd.api+json" \\\n'
            + f'     "{d_url}/jsonapi/user/user?page[limit]=50" | python3 -m json.tool 2>/dev/null | grep -P "name|mail|status|roles"\n\n'
            + f'4. Installed modules and versions:\n'
            + f'   curl -s -b /tmp/drupal_auth.txt "{d_url}/admin/reports/updates" \\\n'
            + '     | grep -oP \'(?<=<td>)[A-Za-z ]+(?=</td>)|(?<=version">)[^<]+\' | head -60\n\n'
            + f'5. Available security updates:\n'
            + f'   curl -s -b /tmp/drupal_auth.txt "{d_url}/admin/reports/updates/update" \\\n'
            + '     | grep -iP "security update|critical|warning"\n\n'
            + f'6. Status report (critical config issues):\n'
            + f'   curl -s -b /tmp/drupal_auth.txt "{d_url}/admin/reports/status" \\\n'
            + '     | grep -iP "error|warning|critical|vulnerable|insecure"\n\n'
            + f'7. Public file exposure:\n'
            + f'   curl -s "{d_url}/sites/default/settings.php" | head -10\n'
            + f'   curl -s "{d_url}/.git/HEAD" | head -3\n\n'
            + f'8. PHP filter module (critical RCE risk if enabled):\n'
            + f'   curl -s -b /tmp/drupal_auth.txt "{d_url}/admin/modules" \\\n'
            + '     | grep -iP "php filter|php_filter"\n'
        )

    if site_type == 'joomla' and creds.get('joomla_user'):
        j_user  = creds.get('joomla_user', '')
        j_token = creds.get('joomla_token', '')
        auth_hdr = f'-H "X-Joomla-Token: {j_token}"' if j_token else '-b /tmp/joomla_auth.txt'
        return (hdr
            + 'JOOMLA CREDENTIALS\n'
            + f'  URL      : https://{domain}\n'
            + f'  Username : {j_user}\n'
            + ('  API Token: (provided)\n' if j_token else '  Password : $JOOMLA_PASSWORD (in environment)\n')
            + '\nRun ALL Joomla security checks:\n\n'
            + ('' if j_token else
               f'1. Authenticate (form login):\n'
               + f'   JTOKEN=$(curl -sc /tmp/joomla_auth.txt -s "https://{domain}/administrator/index.php" \\\n'
               + '     | grep -oP \'(?<=name="[0-9a-f]{32}" value=")[^"]+\' | head -1)\n'
               + f'   curl -sb /tmp/joomla_auth.txt -s \\\n'
               + f'     -d "username={j_user}&passwd=$JOOMLA_PASSWORD&option=com_login&task=login&return=aW5kZXgucGhw&${{JTOKEN}}=1" \\\n'
               + f'     "https://{domain}/administrator/index.php" -o /dev/null\n\n')
            + f'2. Joomla version:\n'
            + f'   curl -s "https://{domain}/administrator/manifests/files/joomla.xml" | grep -oP "(?<=<version>)[^<]+"\n\n'
            + f'3. Installed extensions (API — Joomla 4+):\n'
            + f'   curl -s {auth_hdr} "https://{domain}/api/index.php/v1/extensions?page[limit]=100" \\\n'
            + '     | python3 -m json.tool 2>/dev/null | grep -P "name|version|enabled|type" | head -80\n\n'
            + f'4. Users list (API):\n'
            + f'   curl -s {auth_hdr} "https://{domain}/api/index.php/v1/users?page[limit]=50" \\\n'
            + '     | python3 -m json.tool 2>/dev/null | grep -P "name|username|email|groups" | head -60\n\n'
            + f'5. Global config — check debug/error display:\n'
            + f'   curl -s {auth_hdr} "https://{domain}/api/index.php/v1/config/application" \\\n'
            + '     | python3 -m json.tool 2>/dev/null | grep -P "debug|error|ssl|force_ssl|session" | head -30\n\n'
            + f'6. Exposed sensitive files:\n'
            + f'   curl -s "https://{domain}/configuration.php" | head -5\n'
            + f'   curl -s "https://{domain}/.git/HEAD" | head -3\n'
            + f'   curl -s "https://{domain}/administrator/logs/" | head -20\n'
        )

    if site_type == 'generic' and (creds.get('gen_user') or creds.get('gen_api_key')):
        g_url       = (creds.get('gen_url') or f'https://{domain}').rstrip('/')
        g_user      = creds.get('gen_user', '')
        g_api_key   = creds.get('gen_api_key', '')
        g_framework = creds.get('gen_framework', 'auto')
        api_hdr     = f'-H "Authorization: Bearer {g_api_key}"' if g_api_key else '-b /tmp/app_auth.txt'
        return (hdr
            + f'GENERIC WEB APP CREDENTIALS ({g_framework.upper()})\n'
            + f'  Login URL : {g_url}\n'
            + f'  Username  : {g_user}\n'
            + ('  API Key   : (provided)\n' if g_api_key else '  Password  : $GEN_PASSWORD (in environment)\n')
            + '\nRun ALL application security checks:\n\n'
            + ('' if g_api_key else
               f'1. Authenticate (form login):\n'
               + f'   curl -sc /tmp/app_auth.txt -s -X POST \\\n'
               + f'     -d "username={g_user}&password=$GEN_PASSWORD" \\\n'
               + f'     "{g_url}" -L -o /tmp/app_login.html -w "%{{http_code}}"\n'
               + f'   curl -sc /tmp/app_auth.txt -s -X POST \\\n'
               + f'     -H "Content-Type: application/json" \\\n'
               + f'     -d \'{{"username":"{g_user}","password":"$GEN_PASSWORD"}}\' \\\n'
               + f'     "{g_url}" -L -w "%{{http_code}}"\n\n')
            + f'2. Security HTTP headers:\n'
            + f'   curl -sI "https://{domain}/" | grep -iP "x-frame|x-content|strict-transport|content-security|referrer|permissions"\n\n'
            + f'3. Common sensitive file exposure:\n'
            + '   for p in .env .env.local .env.production config/database.yml config/secrets.yml '
            + 'storage/logs/laravel.log var/log/prod.log .git/HEAD; do\n'
            + f'     code=$(curl -so /dev/null -w "%{{http_code}}" "https://{domain}/$p")\n'
            + '     [ "$code" = "200" ] && echo "EXPOSED: $p"\n'
            + '   done\n\n'
            + f'4. API endpoint discovery:\n'
            + f'   curl -s {api_hdr} "https://{domain}/api/" | python3 -m json.tool 2>/dev/null | head -40\n'
            + f'   curl -s {api_hdr} "https://{domain}/api/v1/" | python3 -m json.tool 2>/dev/null | head -40\n\n'
            + f'5. Authentication bypass tests:\n'
            + f'   curl -s "https://{domain}/admin" -o /dev/null -w "%{{http_code}} %{{url_effective}}\\n"\n'
            + f'   curl -s "https://{domain}/dashboard" -o /dev/null -w "%{{http_code}} %{{url_effective}}\\n"\n'
            + f'   curl -s -H "X-Forwarded-For: 127.0.0.1" "https://{domain}/admin" -o /dev/null -w "%{{http_code}}\\n"\n\n'
            + f'6. CORS misconfiguration:\n'
            + f'   curl -sI -H "Origin: https://evil.com" "https://{domain}/api/" \\\n'
            + '     | grep -i "access-control"\n\n'
            + (f'7. Framework-specific checks ({g_framework}):\n'
               + (f'   curl -s "https://{domain}/_debugbar/open" | head -5\n'
                  f'   curl -s "https://{domain}/telescope/api/requests" {api_hdr} | python3 -m json.tool 2>/dev/null | head -20\n'
                  if g_framework == 'laravel' else
                  f'   curl -s "https://{domain}/__debug__/" | head -20\n'
                  f'   curl -s "https://{domain}/admin/" -o /dev/null -w "%{{http_code}}\\n"\n'
                  if g_framework == 'django' else
                  f'   curl -s "https://{domain}/admin" | grep -iP "flask|werkzeug|jinja"\n'
                  if g_framework == 'flask' else
                  f'   curl -sI "https://{domain}/" | grep -iP "x-powered-by|server|x-rails"\n')
               + '\n')
        )

    if site_type == 'plesk' and creds.get('plesk_host') and creds.get('plesk_user'):
        pk_host = creds.get('plesk_host', '') or domain
        pk_user = creds.get('plesk_user', '')
        pk_port = creds.get('plesk_port', '8443') or '8443'
        pk_api  = f'https://{pk_host}:{pk_port}/api/v2'
        pk_hdr  = f'-u "{pk_user}:$PLESK_PASSWORD" -H "Content-Type: application/json"'
        return (hdr
            + 'PLESK CREDENTIALS\n'
            + f'  Host     : {pk_host}:{pk_port}\n'
            + f'  Username : {pk_user}\n'
            + '  Password : $PLESK_PASSWORD (in environment)\n\n'
            + 'Run ALL Plesk security checks:\n\n'
            + f'1. List all domains/sites:\n'
            + f'   curl -sk {pk_hdr} "{pk_api}/domains" | python3 -m json.tool 2>/dev/null | grep -P "name|status|hosting"\n\n'
            + f'2. SSL certificates:\n'
            + f'   curl -sk {pk_hdr} "{pk_api}/certificates" | python3 -m json.tool 2>/dev/null | grep -P "name|validTo|status"\n\n'
            + f'3. PHP handler versions:\n'
            + f'   curl -sk {pk_hdr} "{pk_api}/domains" | python3 -c "import sys,json;d=json.load(sys.stdin);[print(x.get(\'name\'),x.get(\'phpHandlerId\',\'?\')) for x in d.get(\'data\',[])if isinstance(x,dict)]" 2>/dev/null\n\n'
            + f'4. Mail accounts:\n'
            + f'   curl -sk {pk_hdr} "{pk_api}/mail" | python3 -m json.tool 2>/dev/null | grep -P "name|domain|enabled" | head -40\n\n'
            + f'5. Server info and version:\n'
            + f'   curl -sk {pk_hdr} "{pk_api}/server/info" | python3 -m json.tool 2>/dev/null\n\n'
            + f'6. Database servers:\n'
            + f'   curl -sk {pk_hdr} "{pk_api}/dbservers" | python3 -m json.tool 2>/dev/null | grep -P "type|host|status"\n'
        )

    if site_type == 'postgresql' and creds.get('pg_host') and creds.get('pg_name'):
        pg_host = creds.get('pg_host', '')
        pg_port = creds.get('pg_port', '5432') or '5432'
        pg_name = creds.get('pg_name', '')
        pg_user = creds.get('pg_user', 'postgres')
        pg_ssl  = creds.get('pg_ssl', 'prefer')
        pg_cmd  = f'PGPASSWORD="$PG_PASSWORD" psql -h "{pg_host}" -p {pg_port} -U "{pg_user}" -d "{pg_name}" --no-password -c'
        return (hdr
            + 'POSTGRESQL CREDENTIALS\n'
            + f'  Host     : {pg_host}:{pg_port}\n'
            + f'  Database : {pg_name}\n'
            + f'  User     : {pg_user}\n'
            + f'  SSL Mode : {pg_ssl}\n'
            + '  Password : $PG_PASSWORD (in environment)\n\n'
            + 'Install client if missing: apt-get install -y postgresql-client 2>/dev/null\n\n'
            + 'Run ALL PostgreSQL security checks:\n\n'
            + f'1. Server version:\n   {pg_cmd} "SELECT version();"\n\n'
            + f'2. All users and roles:\n   {pg_cmd} "SELECT usename,usesuper,usecreatedb,usecreaterole,valuntil FROM pg_user ORDER BY usesuper DESC;"\n\n'
            + f'3. Superuser accounts (minimize these):\n   {pg_cmd} "SELECT rolname FROM pg_roles WHERE rolsuper=true;"\n\n'
            + f'4. Databases and sizes:\n   {pg_cmd} "SELECT datname,pg_size_pretty(pg_database_size(datname)),datistemplate FROM pg_database;"\n\n'
            + f'5. Public schema privileges (should not be world-writable):\n   {pg_cmd} "SELECT nspname,nspacl FROM pg_namespace WHERE nspname=\'public\';"\n\n'
            + f'6. SSL enforcement:\n   {pg_cmd} "SHOW ssl;"\n   {pg_cmd} "SELECT name,setting FROM pg_settings WHERE name IN (\'ssl\',\'ssl_ca_file\',\'password_encryption\');"\n\n'
            + f'7. Installed extensions (check for dangerous ones):\n   {pg_cmd} "SELECT name,default_version,installed_version FROM pg_available_extensions WHERE installed_version IS NOT NULL;"\n\n'
            + f'8. Active connections and IPs:\n   {pg_cmd} "SELECT datname,usename,client_addr,state,query_start FROM pg_stat_activity WHERE state!=\'idle\' LIMIT 30;"\n\n'
            + f'9. Tables with sensitive-sounding names:\n   {pg_cmd} "SELECT table_schema,table_name FROM information_schema.tables WHERE table_name ~* \'user|password|token|secret|credit|payment\' AND table_schema NOT IN (\'pg_catalog\',\'information_schema\') LIMIT 30;"\n'
        )

    if site_type == 'mongodb' and (creds.get('mongo_uri') or creds.get('mongo_host')):
        m_uri  = creds.get('mongo_uri', '')
        m_host = creds.get('mongo_host', 'localhost')
        m_port = creds.get('mongo_port', '27017') or '27017'
        m_db   = creds.get('mongo_db', '')
        m_user = creds.get('mongo_user', '')
        conn   = m_uri if m_uri else (
            f'mongodb://{m_user}:$MONGO_PASSWORD@{m_host}:{m_port}/{m_db}?authSource=admin'
            if m_user else f'mongodb://{m_host}:{m_port}/{m_db}'
        )
        mc = f'mongosh "{conn}" --quiet --eval'
        return (hdr
            + 'MONGODB CREDENTIALS\n'
            + f'  Connection: {conn.replace(m_user, "[user]") if m_user else conn}\n'
            + '  Password  : $MONGO_PASSWORD (in environment)\n\n'
            + 'Install client if missing: apt-get install -y mongodb-mongosh 2>/dev/null\n\n'
            + 'Run ALL MongoDB security checks:\n\n'
            + f'1. Server version and build info:\n   {mc} "db.adminCommand({{buildInfo:1}}).version"\n\n'
            + f'2. List all databases:\n   {mc} "db.adminCommand({{listDatabases:1}}).databases.map(d=>d.name+\' (\'+d.sizeOnDisk+\' bytes)\')"\n\n'
            + f'3. All users and roles:\n   {mc} "db.adminCommand({{usersInfo:1}}).users.map(u=>u.user+\':\'+JSON.stringify(u.roles))"\n\n'
            + f'4. Auth status (CRITICAL — empty means no auth required):\n   {mc} "db.adminCommand({{connectionStatus:1}}).authInfo"\n\n'
            + f'5. Collections in target DB:\n   {mc} "use {m_db or "admin"}; db.getCollectionNames()"\n\n'
            + f'6. Check for exposed credentials in collections:\n   {mc} "use {m_db or "admin"}; db.getCollectionNames().forEach(c=>{{var s=db[c].findOne({{$or:[{{password:{{$exists:true}}}},{{token:{{$exists:true}}}},{{api_key:{{$exists:true}}}}]}});if(s)print(c,Object.keys(s));}})"\n\n'
            + f'7. Active operations:\n   {mc} "db.adminCommand({{currentOp:1}}).inprog.map(o=>{{return {{op:o.op,ns:o.ns,secs:o.secs_running}};}})"\n\n'
            + f'8. Replica set / auth config:\n   {mc} "db.adminCommand({{getCmdLineOpts:1}}).parsed"\n'
        )

    if site_type == 'firebase' and creds.get('fb_project'):
        fb_proj = creds.get('fb_project', '')
        fb_db   = creds.get('fb_db_url', '') or f'https://{fb_proj}-default-rtdb.firebaseio.com'
        fb_key  = creds.get('fb_api_key', '')
        return (hdr
            + 'FIREBASE / FIRESTORE CREDENTIALS\n'
            + f'  Project ID   : {fb_proj}\n'
            + f'  Database URL : {fb_db}\n'
            + ('  API Key      : (provided)\n' if fb_key else '')
            + ('  Service Acct : $FB_JSON (in environment)\n' if creds.get('fb_json') else '')
            + '\nRun ALL Firebase security checks:\n\n'
            + f'1. Realtime Database unauthenticated read (CRITICAL if 200):\n'
            + f'   curl -s "{fb_db}/.json?shallow=true" | python3 -m json.tool 2>/dev/null | head -30\n'
            + f'   echo "HTTP code:"; curl -so /dev/null -w "%{{http_code}}" "{fb_db}/.json"\n\n'
            + f'2. Unauthenticated write test (CRITICAL if 200/204):\n'
            + f'   curl -s -X PUT -H "Content-Type:application/json" -d \'{{"cfai_probe":true}}\' \\\n'
            + f'     "{fb_db}/cfai_security_probe.json" -w "\\nHTTP: %{{http_code}}"\n'
            + f'   # Clean up probe if written:\n'
            + f'   curl -s -X DELETE "{fb_db}/cfai_security_probe.json"\n\n'
            + f'3. Storage bucket public access:\n'
            + f'   curl -s "https://firebasestorage.googleapis.com/v0/b/{fb_proj}.appspot.com/o" \\\n'
            + '     | python3 -m json.tool 2>/dev/null | head -20\n\n'
            + (f'4. Firestore security rules (requires API key):\n'
               + f'   curl -s "https://firestore.googleapis.com/v1/projects/{fb_proj}/databases/(default)/documents" \\\n'
               + f'     -H "x-goog-api-key: {fb_key}" | python3 -m json.tool 2>/dev/null | head -30\n\n'
               if fb_key else '')
            + f'5. Auth providers and settings (requires service account):\n'
            + f'   # Run: firebase auth:export /tmp/fb_users.json --project {fb_proj}\n'
            + f'   # Then: head -30 /tmp/fb_users.json\n\n'
            + f'6. Hosting configuration exposure:\n'
            + f'   curl -s "https://{fb_proj}.web.app/__/firebase/init.json" | python3 -m json.tool 2>/dev/null\n'
            + f'   curl -s "https://{fb_proj}.web.app/__/firebase/init.js" | head -20\n'
        )

    if site_type == 'redis' and creds.get('redis_host'):
        r_host = creds.get('redis_host', 'localhost')
        r_port = creds.get('redis_port', '6379') or '6379'
        r_db   = creds.get('redis_db', '0') or '0'
        auth   = '-a "$REDIS_PASSWORD"' if creds.get('redis_pass') else ''
        rc     = f'redis-cli -h {r_host} -p {r_port} {auth} -n {r_db}'
        return (hdr
            + 'REDIS CREDENTIALS\n'
            + f'  Host     : {r_host}:{r_port}\n'
            + f'  DB Index : {r_db}\n'
            + ('  Password : $REDIS_PASSWORD (in environment)\n' if creds.get('redis_pass') else '  Password : NONE (testing for unauthenticated access)\n')
            + '\nInstall client if missing: apt-get install -y redis-tools 2>/dev/null\n\n'
            + 'Run ALL Redis security checks:\n\n'
            + f'1. Unauthenticated access test:\n   redis-cli -h {r_host} -p {r_port} PING\n\n'
            + f'2. Server info:\n   {rc} INFO server | grep -P "redis_version|os|tcp_port|config_file"\n\n'
            + f'3. Protected mode (should be ON):\n   {rc} CONFIG GET protected-mode\n\n'
            + f'4. Bind address (danger if 0.0.0.0):\n   {rc} CONFIG GET bind\n\n'
            + f'5. All keys (SCAN — avoid KEYS in production):\n   {rc} SCAN 0 COUNT 100 | head -30\n\n'
            + f'6. Session/token data exposure:\n'
            + f'   {rc} SCAN 0 MATCH "*session*" COUNT 50\n'
            + f'   {rc} SCAN 0 MATCH "*token*" COUNT 50\n'
            + f'   {rc} SCAN 0 MATCH "*password*" COUNT 50\n\n'
            + f'7. Dangerous command availability:\n'
            + f'   {rc} CONFIG GET maxmemory\n'
            + f'   {rc} CONFIG REWRITE 2>&1\n'
            + f'   {rc} DEBUG SLEEP 0 2>&1\n\n'
            + f'8. ACL users (Redis 6+):\n   {rc} ACL LIST 2>/dev/null | head -20\n\n'
            + f'9. Client connections:\n   {rc} CLIENT LIST | head -20\n'
        )

    if site_type == 'mssql' and creds.get('mssql_host') and creds.get('mssql_db'):
        ms_host = creds.get('mssql_host', '')
        ms_port = creds.get('mssql_port', '1433') or '1433'
        ms_db   = creds.get('mssql_db', '')
        ms_user = creds.get('mssql_user', 'sa')
        ms_auth = creds.get('mssql_auth', 'sql')
        sq = (f'sqlcmd -S "{ms_host},{ms_port}" -d "{ms_db}" -U "{ms_user}" '
              f'-P "$MSSQL_PASSWORD" -Q')
        return (hdr
            + 'MICROSOFT SQL SERVER CREDENTIALS\n'
            + f'  Host     : {ms_host}:{ms_port}\n'
            + f'  Database : {ms_db}\n'
            + f'  User     : {ms_user}\n'
            + f'  Auth     : {ms_auth}\n'
            + '  Password : $MSSQL_PASSWORD (in environment)\n\n'
            + 'Install client if missing: apt-get install -y mssql-tools 2>/dev/null\n\n'
            + 'Run ALL SQL Server security checks:\n\n'
            + f'1. Server version:\n   {sq} "SELECT @@VERSION"\n\n'
            + f'2. All logins and roles:\n   {sq} "SELECT name,type_desc,is_disabled,is_policy_checked FROM sys.sql_logins ORDER BY name"\n\n'
            + f'3. Sysadmin accounts (minimize — sa should be renamed/disabled):\n   {sq} "SELECT name FROM sys.sql_logins WHERE IS_SRVROLEMEMBER(\'sysadmin\',name)=1"\n\n'
            + f'4. xp_cmdshell status (CRITICAL if enabled — allows OS command exec):\n   {sq} "EXEC sp_configure \'xp_cmdshell\'"\n\n'
            + f'5. Linked servers (lateral movement risk):\n   {sq} "SELECT name,data_source,provider FROM sys.servers WHERE is_linked=1"\n\n'
            + f'6. Database encryption (TDE):\n   {sq} "SELECT name,is_encrypted FROM sys.databases"\n\n'
            + f'7. Audit specifications:\n   {sq} "SELECT name,is_state_enabled FROM sys.server_audits"\n\n'
            + f'8. Tables with sensitive data:\n'
            + f'   {sq} "SELECT TABLE_SCHEMA,TABLE_NAME FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME LIKE \'%user%\' OR TABLE_NAME LIKE \'%password%\' OR TABLE_NAME LIKE \'%token%\' OR TABLE_NAME LIKE \'%credit%\'"\n\n'
            + f'9. Recent failed logins (brute force):\n   {sq} "SELECT TOP 30 EventTime,LoginName,ClientHost FROM sys.event_log WHERE EventClass=\'AUDIT_LOGIN_FAILED\' ORDER BY EventTime DESC" 2>/dev/null\n'
        )

    return ''


def _run_background_scan(job_id: str, target: str, agent_type: str,
                          model: str, site_type: str, creds: dict, username: str = ''):
    """Run a WSTG agent in a background thread and stream chunks to _scan_jobs."""
    job = _scan_jobs[job_id]

    # Pre-initialise so the except handler can always reference these safely
    parts:  list = []
    tools:  list = [0]
    t0           = _time.time()
    domain       = ''
    model_used   = model or ''

    # Set credential env vars so agents can reference $WP_PASSWORD, $SSHPASS, etc.
    env_restore: dict = {}
    cred_env = {
        'WP_USER':         creds.get('wp_user', ''),
        'WP_APP_PASSWORD': creds.get('wp_app_pass', ''),
        'WP_PASSWORD':     creds.get('wp_pass', ''),
        'CPANEL_PASSWORD': creds.get('cpanel_pass', ''),
        'SSHPASS':         creds.get('ssh_pass', ''),
        'FTP_PASSWORD':    creds.get('ftp_pass', ''),
        'SSH_KEY':         creds.get('ssh_key', ''),
        'DB_HOST':         creds.get('db_host', ''),
        'DB_PORT':         creds.get('db_port', '3306'),
        'DB_NAME':         creds.get('db_name', ''),
        'DB_USER':         creds.get('db_user', ''),
        'DB_PASSWORD':     creds.get('db_pass', ''),
        'DB_PREFIX':       creds.get('db_prefix', 'wp_'),
        'DRUPAL_PASSWORD': creds.get('drupal_pass', ''),
        'JOOMLA_PASSWORD': creds.get('joomla_pass', ''),
        'GEN_PASSWORD':    creds.get('gen_pass', ''),
        'PLESK_PASSWORD':  creds.get('plesk_pass', ''),
        'PG_PASSWORD':     creds.get('pg_pass', ''),
        'MONGO_PASSWORD':  creds.get('mongo_pass', ''),
        'FB_JSON':         creds.get('fb_json', ''),
        'REDIS_PASSWORD':  creds.get('redis_pass', ''),
        'MSSQL_PASSWORD':  creds.get('mssql_pass', ''),
    }
    for k, v in cred_env.items():
        if v:
            env_restore[k] = os.environ.get(k, '')
            os.environ[k] = v

    try:
        import dataclasses as dc
        from urllib.parse import urlparse as _up
        from agents.wstg_agents import WSTG_REGISTRY
        from sdk.agents import Runner
        from sdk import tracing

        # Normalise domain (strip scheme + path)
        _url = target if '://' in target else 'https://' + target
        _p   = _up(_url)
        domain = (_p.netloc or _p.path.split('/')[0]).rstrip('/')
        job['domain'] = domain

        # Build authenticated-scan instructions and announce cred type in terminal
        cred_block = _build_cred_block(site_type, creds, domain)
        if cred_block:
            job['chunks'].append({'k': 'txt', 'd': f'[AUTH] {site_type.upper()} credentials loaded — authenticated scan enabled'})

        def _ot(t):
            if job.get('aborted'):
                raise RuntimeError('Scan aborted by user')
            parts.append(t)
            job['chunks'].append({'k': 'txt', 'd': t})
        def _oo(n, a):  tools[0] += 1;     job['chunks'].append({'k': 'tool', 'n': n, 'a': str(a)[:200]})
        def _or(n, r, e):
            r_full = str(r)
            # Capture tool results containing WP-LOG lines or from MCP tools so
            # extract_wp_logs() and the Plugin Logs modal can parse them after save.
            if ('WP-LOG' in r_full or n in ('wp_security_scan', 'wp_api_call')) and r_full.strip():
                parts.append(f'[TOOL:{n}]\n{r_full}')
            job['chunks'].append({'k': 'res', 'n': n, 'r': r_full[:300], 'e': bool(e)})

        if agent_type == 'pentest':
            from agents.pentest import run_full_pentest
            t0 = _time.time()
            run_full_pentest(domain, model=model or None, on_text=_ot, on_tool=_oo, on_result=_or,
                             cred_block=cred_block or None,
                             is_aborted=lambda: bool(job.get('aborted')))
            model_used = model or ''
        elif agent_type in ('ctf', 'ot', 'enum'):
            from agents.special_agents import SPECIAL_REGISTRY as _SREG
            base = _SREG.get(agent_type)
            if base is None:
                job.update({'status': 'error', 'error': f'Unknown special agent: {agent_type}'}); return
            _s_instr = base.instructions.replace('{target}', domain)
            if cred_block:
                _s_instr += cred_block
            agent    = dc.replace(base, instructions=_s_instr)
            if model:
                agent = dc.replace(agent, model=model)
            elif cred_block:
                # Credentials supplied — use Claude for reliable execution
                _cm = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
                agent = dc.replace(agent, model=_cm)
            model_used = getattr(agent, 'model', model) or ''
            _label = {'ctf': 'CTF Solver', 'ot': 'OT/ICS Security', 'enum': 'API Enumeration'}.get(agent_type, agent_type.upper())
            t0 = _time.time()
            with tracing.span(f'dashboard:{agent_type}') as span:
                span.set_attribute('cfai.target', domain)
                Runner.run(agent, f'Begin {_label} on {domain}.',
                           on_text=_ot, on_tool=_oo, on_result=_or)
        elif agent_type in ('recon-passive', 'breach', 'subdomain', 'tech-detect', 'supply-chain'):
            from dashboard.passive_agents import PASSIVE_RUNNERS
            runner = PASSIVE_RUNNERS.get(agent_type)
            if runner is None:
                job.update({'status': 'error', 'error': f'Unknown passive agent: {agent_type}'}); return
            t0 = _time.time()
            runner(domain, _ot)
        else:
            base = WSTG_REGISTRY.get(agent_type)
            if base is None:
                job.update({'status': 'error', 'error': f'Unknown agent type: {agent_type}'}); return

            # Inject credential instructions after the base instructions
            base_instructions = base.instructions.replace('{domain}', domain) + cred_block
            agent = dc.replace(base, instructions=base_instructions)
            if model:
                agent = dc.replace(agent, model=model)
            model_used = getattr(agent, 'model', model) or ''

            # Any agent + credentials → Claude model for reliable execution.
            # WordPress site type or WP creds → also inject MCP tools for all agents.
            # APIT → always gets MCP tools regardless of site type.
            _wp_creds  = (creds.get('wp_user') or creds.get('wp_pass') or creds.get('wp_app_pass'))
            _has_creds = bool(
                _wp_creds
                or creds.get('cpanel_user') or creds.get('cpanel_pass')
                or creds.get('ssh_user')    or creds.get('ssh_pass')
                or creds.get('ftp_user')    or creds.get('ftp_pass')
                or creds.get('db_host')     or creds.get('db_name')
            )
            _needs_mcp = (site_type == 'wordpress') or bool(_wp_creds) or (agent_type == 'apit')

            if _needs_mcp:
                from tools.wordpress_mcp import wp_api_call, wp_security_scan
                mcp_block = (
                    '\n\n══════════════ MCP DIRECT CONNECTION ══════════════\n'
                    'You have wp_security_scan and wp_api_call tools connected via MCP.\n\n'
                    'STEP 0 — ALWAYS DO THIS FIRST:\n'
                    f'  1. Call wp_security_scan(site_url="https://{domain}")\n'
                    '     Runs a full WordPress security audit and emits WP-LOG entries.\n'
                    '  2. CRITICAL: Copy ALL lines starting with "WP-LOG |" from the tool\n'
                    '     result VERBATIM into your response — DO NOT generate WP-LOG lines\n'
                    '     yourself, only echo what the tool returned.\n'
                    f'  3. Call wp_api_call(site_url="https://{domain}", endpoint="/wp-json/wp/v2/users")\n'
                    '     and other REST endpoints for deeper investigation.\n'
                    '  4. If no WP-LOG entries were returned by the tool, instruct the user\n'
                    '     to install the free "WP Activity Log" plugin (wp-security-audit-log)\n'
                    '     and re-run the scan to get real admin login events with IP addresses.\n'
                    '     If admin credentials are provided, attempt installation via WP-CLI:\n'
                    f'     generic_linux_command("wp plugin install wp-security-audit-log --activate --path=/var/www/html --url=https://{domain}")\n'
                    'Auth is handled automatically: Basic Auth → Cookie+Nonce → public.\n'
                    '═══════════════════════════════════════════════════\n\n'
                )
                _existing = {getattr(t, '__name__', '') for t in agent.tools}
                new_tools = [t for t in [wp_api_call, wp_security_scan]
                             if getattr(t, '__name__', '') not in _existing] + list(agent.tools)
                _claude_model = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
                agent = dc.replace(
                    agent,
                    model=_claude_model,
                    instructions=base_instructions + mcp_block,
                    tools=new_tools,
                )
                model_used = _claude_model
            elif _has_creds:
                # Non-WordPress credentials (SSH / cPanel / SFTP) — switch to Claude so
                # the agent reliably executes the credential block instructions.
                _claude_model = os.environ.get('ANTHROPIC_MODEL', 'claude-sonnet-4-6')
                agent = dc.replace(agent, model=_claude_model,
                                   instructions=base_instructions)
                model_used = _claude_model

            t0 = _time.time()
            with tracing.span(f'dashboard:{agent_type}') as span:
                span.set_attribute('cfai.target', domain)
                Runner.run(agent, f'Run all WSTG-{agent_type.upper()} checks on {domain}.',
                           on_text=_ot, on_tool=_oo, on_result=_or)

        elapsed  = _time.time() - t0
        output   = '\n\n'.join(parts)
        was_aborted = job.get('aborted', False)
        db_status   = 'interrupted' if was_aborted else 'ok'

        scan_id = db.save_scan(
            target=domain, agent_type=agent_type,
            model=model_used, status=db_status,
            latency_s=round(elapsed, 2),
            tool_count=tools[0], output=output,
            username=username,
        )
        final_status = 'interrupted' if was_aborted else 'done'
        if was_aborted:
            job['chunks'].append({'k': 'txt', 'd': '\n[Scan stopped — findings logged to dashboard]\n'})
        job['chunks'].append({'k': 'saved', 'id': scan_id})
        job.update({'status': final_status, 'elapsed': round(elapsed, 2),
                    'tool_count': tools[0], 'scan_id': scan_id})
        _parse_and_save_plugins(scan_id, domain, output, username=username)
        _job_persist(job_id, job)

    except Exception as exc:
        import traceback as _tb
        tb = _tb.format_exc()[-1200:]
        job['chunks'].append({'k': 'txt', 'd': f'\n[ERROR] {exc}\n{tb}'})
        # Save partial output — variables are always defined because they were pre-initialised
        try:
            scan_id = db.save_scan(
                target=domain or target or '',
                agent_type=agent_type,
                model=model_used,
                status='error',
                latency_s=round(_time.time() - t0, 2),
                tool_count=tools[0],
                output='\n\n'.join(parts) or f'[ERROR] {exc}',
                username=username,
            )
            job['chunks'].append({'k': 'saved', 'id': scan_id})
            job.update({'status': 'error', 'error': str(exc), 'trace': tb, 'scan_id': scan_id})
            _parse_and_save_plugins(scan_id, domain or target or '', '\n\n'.join(parts), username=username)
        except Exception:
            job.update({'status': 'error', 'error': str(exc), 'trace': tb})
        _job_persist(job_id, job)
    finally:
        for k, orig in env_restore.items():
            if orig: os.environ[k] = orig
            else: os.environ.pop(k, None)


# Optional shared secret for the remote-save API endpoint.
# Set CFAI_API_KEY in the VPS .env to protect POST /api/scan.
# If unset, the endpoint accepts any request (fine on a private VPS).
_API_KEY = os.environ.get('CFAI_API_KEY', '')


# ── Risk classification ────────────────────────────────────────────────────────

_HIGH_KW = [
    'critical', 'high severity', 'exploit', 'vulnerable', 'vulnerability',
    'remote code execution', 'rce', 'sql injection', 'sqli', 'xss',
    'cross-site scripting', 'authentication bypass', 'privilege escalation',
    'unauthorized access', 'exposed credentials', 'leaked secret',
    'path traversal', 'directory traversal', 'file inclusion',
    'command injection', 'deserialization', 'idor', 'insecure direct object',
    'broken access control', 'account takeover',
]
_MED_KW = [
    'medium', 'moderate', 'information disclosure', 'outdated version',
    'weak cipher', 'misconfigured', 'missing security header',
    'deprecated', 'insecure cookie', 'cors misconfiguration',
    'open redirect', 'clickjacking', 'csrf', 'open port',
    'server version', 'default credentials',
]
_LOW_KW = [
    'low severity', 'informational', 'best practice', 'minor',
    'consider enabling', 'consider disabling', 'suggestion',
]

_ACTION_RE = re.compile(
    r'\b(update|upgrade|patch|disable|enable|restrict|remove|add|fix|'
    r'configure|implement|enforce|rotate|revoke|harden|review|audit|'
    r'change|replace|block|sanitize|validate|encrypt)\b',
    re.I,
)

_REC_HEADERS = (
    'recommendation', 'action item', 'remediation', 'next step',
    'action plan', 'suggested fix', 'what to do', 'mitigation',
    'to fix', 'to remediate',
)

# Lines containing these phrases describe attempts or failed checks — skip them
_NEGATION = re.compile(
    r'\b(no\b|not\b|failed|unsuccessful|did not|does not|returned no|'
    r'found no|no evidence|could not|unable to|attempting|will attempt|'
    r'will try|will now|testing for|checking for|i will|let me|'
    r'next i |next,|explore potential|no result|no data|no output|'
    r'empty response|no vuln|not vuln|not found|not detect|not appear)\b',
    re.I,
)


def risk_level(text: str) -> str:
    """Derive risk only from lines that confirm a finding, skipping attempt/failure lines."""
    lines = text.splitlines()
    for kw_list, label in ((_HIGH_KW, 'HIGH'), (_MED_KW, 'MEDIUM'), (_LOW_KW, 'LOW')):
        for line in lines:
            if _NEGATION.search(line):
                continue
            if any(k in line.lower() for k in kw_list):
                return label
    return 'INFO'


def rec_risk(text: str) -> str:
    """Score a single recommendation line independently — avoids scan-wide HIGH bleeding."""
    if not text:
        return 'INFO'
    tl = text.lower()
    for kw_list, label in ((_HIGH_KW, 'HIGH'), (_MED_KW, 'MEDIUM'), (_LOW_KW, 'LOW')):
        if any(k in tl for k in kw_list):
            return label
    return 'MEDIUM'  # actionable but no severity keyword → treat as medium


def _strip_md(text: str) -> str:
    text = re.sub(r'\*\*([^*\n]+)\*\*', r'\1', text)
    text = re.sub(r'\*([^*\n]+)\*', r'\1', text)
    return text.strip()


# Recommendations that describe scanner execution problems, not real vulnerabilities
_SCANNER_NOISE = re.compile(
    r'\b(check connectivity|internet connectivity|assessment platform|'
    r'ensure proper internet|rerun with|retry with|adjust.*header|'
    r'passive recon|historical.*data source|alternative means|'
    r'manual inspection|firewall.*rule.*site|rate limit.*rule|'
    r'scan.*block|blocked.*scan|increase.*timeout|reduce.*thread|'
    r'try.*different.*approach|diagnostic|next step.*scan|'
    r'consider passive|proper internet|platform.*connect|'
    r'connectivity.*assess|assess.*connect|'
    # Agent execution errors — not real security findings
    r'syntax error.*execution|preventing complete execution|'
    r'reliability.*assessment.*syntax|syntax review.*re-execution|re-execution|'
    r'review.*correct.*python|correct.*script syntax|script syntax|'
    r'alternative reconnaissance method|correct.*operational endpoint.*alternative|'
    r'confirm correct url.*alternative)\b',
    re.I,
)


_LOG_LINE_RE = re.compile(
    r'^(WP-LOG|WP-PLUGIN-CHANGE|WP-LOG-STATUS|WP-USER|WP-USER-ENUM|WP-USER-CONFIRMED|'
    r'FAILED_LOGIN|TOOL:|#\s*Log\s+sync)\s*[\|:]',
    re.I,
)


def extract_recs(text: str) -> list[str]:
    recs, in_sec = [], False
    for raw in text.splitlines():
        line = raw.strip()
        if not line:
            in_sec = False
            continue
        # Skip structured activity-log lines — they are not security recommendations
        if _LOG_LINE_RE.match(line):
            in_sec = False
            continue
        if _NEGATION.search(line) or _SCANNER_NOISE.search(line):
            in_sec = False
            continue
        clean = _strip_md(re.sub(r'^[-*+\d.)#]+\s*', '', line))
        clean_lower = clean.lower()
        if any(h in clean_lower for h in _REC_HEADERS):
            in_sec = True
            if ':' in clean:
                after = clean.split(':', 1)[1].strip()
                if len(after) > 12 and not _SCANNER_NOISE.search(after):
                    recs.append(after)
                    in_sec = False
            continue
        is_bullet = line.startswith(('-', '*', '+')) or re.match(r'^\d+[.)]\s', line)
        if in_sec and is_bullet:
            item = _strip_md(re.sub(r'^[-*+\d.)]+\s*', '', line).strip())
            if len(item) > 12:
                recs.append(item)
        elif _ACTION_RE.search(clean) and 25 < len(clean) < 300:
            recs.append(clean)
    seen, out = set(), []
    for r in recs:
        k = r[:50].lower()
        if k not in seen:
            seen.add(k)
            out.append(r)
    return out[:12]


_WP_LOG_RE = re.compile(
    r'^WP-LOG\s*\|\s*([^|\n]+?)\s*\|\s*([^|\n]+?)\s*\|\s*([^|\n]+?)\s*\|\s*([^|\n]+?)\s*\|\s*(HIGH|MEDIUM|LOW|INFO)\s*$',
    re.I | re.MULTILINE,
)
_WP_LOG_STATUS_RE = re.compile(
    r'^WP-LOG-STATUS\s*\|\s*(\w+)\s*\|\s*([^|\n]+?)\s*\|',
    re.I | re.MULTILINE,
)

# Fallback patterns: parse structured agent output lines into WP-LOG-style entries
# for scans that predate the WP-LOG emission updates.
_WP_LOG_FALLBACK = [
    (re.compile(r'^EXPOSED_FILE \| (\S+) \| (/\S+)', re.M),
     lambda m: ('CF_AI', f'Exposed sensitive file: {m.group(2)} (HTTP {m.group(1)})', '-', 'HIGH')),
    (re.compile(r'^CREDS_FOUND_XMLRPC \| ([^\|\n]+?) \|', re.M),
     lambda m: (m.group(1).strip(), 'WordPress credentials verified via XML-RPC brute-force', '-', 'HIGH')),
    (re.compile(r'^CREDS_FOUND_FORM \| ([^\|\n]+?) \|', re.M),
     lambda m: (m.group(1).strip(), 'WordPress credentials verified via login form', '-', 'HIGH')),
    (re.compile(r'^APP_PASS_CREATED(?:_COOKIE)? \| ([^\|\n]+?) \|', re.M),
     lambda m: (m.group(1).strip(), 'Application Password created by CF_AI scanner', '-', 'HIGH')),
    (re.compile(r'^WP-USER-CONFIRMED \| ([^\|\n]+?) \|', re.M),
     lambda m: (m.group(1).strip(), 'WordPress username confirmed via login error oracle', '-', 'MEDIUM')),
    (re.compile(r'^WP-USER \| (\S+) \| (\S+) \|', re.M),
     lambda m: (m.group(2), f'WordPress user enumerated via REST API (id={m.group(1)})', '-', 'MEDIUM')),
    (re.compile(r'^WP-USER-ENUM \| (\S+) \| (\S+) \|', re.M),
     lambda m: (m.group(2), 'WordPress user enumerated via author redirect', '-', 'MEDIUM')),
    (re.compile(r'^FOUND_DB_USER:\s*(\S+)', re.M),
     lambda m: ('CF_AI', f'Database username exposed in config file: {m.group(1)}', '-', 'HIGH')),
]


def _wp_log_fallback(output: str) -> list:
    """Parse structured agent lines into WP-LOG entries (fallback for older scans)."""
    entries = []
    seen = set()
    for pat, builder in _WP_LOG_FALLBACK:
        for m in pat.finditer(output):
            user, event, ip, risk = builder(m)
            key = (user, event)
            if key not in seen:
                seen.add(key)
                entries.append({'timestamp': '--', 'user': user, 'event': event, 'ip': ip, 'risk': risk})
    return entries


# Users that indicate scanner-generated entries, not real admin logins
_SCANNER_USERS = {'system', 'cf_ai', 'cf_ai-mcp', 'cf_ai_mcp', 'scanner'}


def extract_wp_logs(output: str) -> dict:
    """Parse WP-LOG lines from agent output. Returns {entries, status}.

    Scanner-generated entries (user == 'system', 'CF_AI', etc.) are excluded —
    Plugin Logs shows only real WordPress user activity.
    """
    entries = []
    for m in _WP_LOG_RE.finditer(output):
        user = m.group(2).strip()
        if user.lower() in _SCANNER_USERS:
            continue
        entries.append({
            'timestamp': m.group(1).strip(),
            'user':      user,
            'event':     m.group(3).strip(),
            'ip':        m.group(4).strip(),
            'risk':      m.group(5).strip().upper(),
        })
    # Fallback: structured agent-emitted lines (CF_AI scanner patterns)
    fallback = [e for e in _wp_log_fallback(output)
                if e.get('user', '').lower() not in _SCANNER_USERS]
    entries.extend(fallback)
    status_match = _WP_LOG_STATUS_RE.search(output)
    status_code = status_match.group(1) if status_match else ('found' if entries else 'none')
    status_msg  = status_match.group(2).strip() if status_match else ''
    return {'entries': entries, 'status': status_code, 'status_msg': status_msg}


# Narrower negation for remediation matching — only skips lines that say
# a check *passed* or the agent is *about to* test something.
# Deliberately does NOT include \bno\b / \bnot\b because "No X-Frame-Options
# header" is a real finding, not a negation.
_REM_NEGATION = re.compile(
    r'\b(no vulnerability|not vulnerable|not affected|properly configured|'
    r'no issue found|no problem|correctly set|header present|header found|'
    r'attempting|will attempt|will try|will now|testing for|checking for|'
    r'i will test|let me test|i will check|checking if)\b',
    re.I,
)

# Maps lowercase fix-key substrings → internal stack identifier
_FIX_STACK_KEYS: dict[str, str] = {
    'nginx':          'nginx',
    'apache':         'apache',
    '.htaccess':      'apache',
    'php':            'php',
    'wordpress':      'wp',
    'wp-cli':         'wp',
    'functions.php':  'wp',
    'wp-config':      'wp',
}

# Smart remediation scoring — extract structured signals from scan output
_CVE_RE  = re.compile(r'\bCVE-\d{4}-\d{4,7}\b', re.I)
_CVSS_RE = re.compile(r'\b(?:cvss\s*(?:score)?|severity\s*score)[:\s=]+([0-9]+(?:\.[0-9]+)?)\b', re.I)

# Phrases that confirm a finding is a *real* problem (not a test step or passing check)
_CONFIRM_RE = re.compile(
    r'\b(missing|absent|not\s+set|not\s+found|not\s+present|not\s+configured|'
    r'no\s+(?:csp|hsts|csp|x-frame|nosniff|referrer)|'
    r'confirmed|detected|vulnerable|exploitable|exposed|accessible|'
    r'returns?\s*200|http\s*200|status\s*(?:code\s*)?200|injection\s+(?:found|success)|'
    r'xss\s+(?:found|confirmed)|rce\s+(?:found|confirmed)|sqli\s+(?:found|confirmed)|'
    r'header\s+(?:is\s+)?missing|no\s+header|header\s+not|without\s+(?:the\s+)?header)\b',
    re.I,
)

_SEV_WEIGHT = {'CRITICAL': 40, 'HIGH': 30, 'MEDIUM': 20, 'LOW': 10}

# ── Technology version extraction (shared with cve.py) ────────────────────────
_TECH_VER_PATTERNS: list[tuple] = [
    (re.compile(r'\bApache[/ ]([\d]+\.[\d]+(?:\.[\d]+)?)\b', re.I),      'Apache HTTP Server', 'server'),
    (re.compile(r'\bnginx[/ ]([\d]+\.[\d]+(?:\.[\d]+)?)\b',  re.I),      'Nginx',              'server'),
    (re.compile(r'\bPHP[/ ]([\d]+\.[\d]+(?:\.[\d]+)?)\b',    re.I),      'PHP',                'php'),
    (re.compile(r'\bWordPress[/ ]?([\d]+\.[\d]+(?:\.[\d]+)?)\b', re.I),  'WordPress',          'cms'),
    (re.compile(r'\bDrupal[/ ]?([\d]+\.[\d]+(?:\.[\d]+)?)\b', re.I),     'Drupal',             'cms'),
    (re.compile(r'\bJoomla[! ]?([\d]+\.[\d]+(?:\.[\d]+)?)\b', re.I),     'Joomla',             'cms'),
    (re.compile(r'\bOpenSSL[/ ]([\d]+\.[\d]+[^\s]{0,10})\b', re.I),      'OpenSSL',            'ssl'),
    (re.compile(r'\bMySQL[/ ]?([\d]+\.[\d]+(?:\.[\d]+)?)\b', re.I),      'MySQL',              'db'),
    (re.compile(r'\bPostgreSQL[/ ]?([\d]+\.[\d]+)\b', re.I),             'PostgreSQL',         'db'),
    (re.compile(r'\bLaravel[/ ]?([\d]+\.[\d]+(?:\.[\d]+)?)\b', re.I),    'Laravel',            'framework'),
    (re.compile(r'\bDjango[/ ]?([\d]+\.[\d]+(?:\.[\d]+)?)\b', re.I),     'Django',             'framework'),
    (re.compile(r'\bNode\.js[/ ]?([\d]+\.[\d]+(?:\.[\d]+)?)\b', re.I),   'Node.js',            'runtime'),
]

# PHP end-of-life mapping: major.minor → (EOL date, status label)
_PHP_EOL: dict[str, tuple[str, str]] = {
    '5':   ('2018-12-31', 'CRITICAL — EOL since 2018, zero security patches'),
    '7.0': ('2018-12-03', 'CRITICAL — EOL since Dec 2018'),
    '7.1': ('2019-12-01', 'CRITICAL — EOL since Dec 2019'),
    '7.2': ('2020-11-30', 'CRITICAL — EOL since Nov 2020'),
    '7.3': ('2021-12-06', 'HIGH — EOL since Dec 2021'),
    '7.4': ('2022-11-28', 'HIGH — EOL since Nov 2022'),
    '8.0': ('2023-11-26', 'HIGH — EOL since Nov 2023'),
    '8.1': ('2024-12-31', 'MEDIUM — EOL Dec 2024'),
}

# Co-occurrence boosts: if both IDs are present in the same scan, boost the named ID.
# Format: ({id_a, id_b}, boost_target_id, extra_score)
_COOCCURRENCE: list[tuple] = [
    ({'xss-reflected',            'missing-csp'},           'missing-csp',              25),
    ({'xss-reflected',            'missing-xcto'},          'missing-xcto',             10),
    ({'sql-injection-confirmed',  'server-version-disclosed'}, 'server-version-disclosed', 15),
    ({'sql-injection-confirmed',  'insecure-cookies'},      'insecure-cookies',          20),
    ({'sql-injection-confirmed',  'directory-listing'},     'directory-listing',         10),
    ({'php-code-execution',       'php-file-upload'},        'php-file-upload',          25),
    ({'php-code-execution',       'sensitive-data-exposed'}, 'sensitive-data-exposed',   20),
    ({'weak-tls',                 'sensitive-data-exposed'}, 'weak-tls',                 20),
    ({'missing-hsts',             'weak-tls'},               'missing-hsts',             15),
    ({'cors-wildcard',            'insecure-cookies'},       'cors-wildcard',            15),
    ({'cors-wildcard',            'sql-injection-confirmed'},'cors-wildcard',            10),
    ({'exposed-env-git',          'sensitive-data-exposed'}, 'exposed-env-git',          20),
    ({'wp-login-no-ratelimit',    'wp-admin-exposed'},       'wp-login-no-ratelimit',    15),
    ({'open-redirect',            'xss-reflected'},          'open-redirect',            15),
    ({'broken-access-control',    'sql-injection-confirmed'},'broken-access-control',    15),
    ({'unpatched-cms',            'sql-injection-confirmed'},'unpatched-cms',            20),
    ({'unpatched-cms',            'xss-reflected'},          'unpatched-cms',            15),
    ({'open-port-service',        'sql-injection-confirmed'},'open-port-service',        20),
    ({'missing-csp',              'cors-wildcard'},          'missing-csp',              10),
]

# Regex for open port extraction from nmap-style output
_PORT_RE = re.compile(r'\b(\d{1,5})/(?:tcp|udp)\s+(?:open|filtered)\b', re.I)
# Regex for sensitive path mentions in findings
_FINDING_PATH_RE = re.compile(
    r'\b(/(?:wp-content|wp-admin|wp-includes|uploads|backup|config|admin|api|'
    r'\.env|\.git|phpinfo|storage|logs?)[^\s"\'<>]{0,80})\b')


def _extract_site_context(text: str, target: str) -> dict:
    """Extract concrete site-specific details from scan output.

    Returns a context dict used to personalise remediation descriptions and fix code.
    """
    ctx: dict = {
        'target': target, 'server': None, 'php': None, 'cms': None,
        'db': None, 'ssl': None, 'framework': None, 'runtime': None,
        'versions': {}, 'cves': [], 'open_ports': [], 'specific_paths': [],
        'php_eol': None,
    }
    for pat, name, category in _TECH_VER_PATTERNS:
        m = pat.search(text)
        if m:
            version = m.group(1)
            ctx['versions'][name] = version
            if ctx.get(category) is None:
                ctx[category] = f'{name}/{version}'

    # PHP end-of-life check
    php_ver = ctx['versions'].get('PHP', '')
    if php_ver:
        major_minor = '.'.join(php_ver.split('.')[:2])
        major = php_ver.split('.')[0]
        for key in (php_ver, major_minor, major):
            if key in _PHP_EOL:
                ctx['php_eol'] = (php_ver, *_PHP_EOL[key])
                break

    ctx['cves']         = list(dict.fromkeys(m.upper() for m in _CVE_RE.findall(text)))[:8]
    ctx['open_ports']   = sorted({int(p) for p in _PORT_RE.findall(text)})[:15]
    ctx['specific_paths'] = list(dict.fromkeys(_FINDING_PATH_RE.findall(text)))[:8]
    return ctx


def _extract_evidence_lines(patterns: list[str], pos_lines: list[str],
                             n: int = 3) -> list[str]:
    """Return up to n actual scan lines that triggered this remediation.

    These are shown verbatim in the UI as 'Evidence found in your scan'.
    """
    hits: list[str] = []
    seen: set[str] = set()
    for line in pos_lines:
        stripped = line.strip()
        if len(stripped) < 12 or len(stripped) > 220:
            continue
        if any(p in line for p in patterns):
            key = stripped[:80].lower()
            if key not in seen:
                seen.add(key)
                hits.append(stripped)
            if len(hits) >= n:
                break
    return hits


def _urgency_tier(score: int, confirmed: bool) -> tuple[str, str]:
    """Map score → (urgency_label, urgency_class).

    urgency_class is used as a CSS class suffix on the frontend.
    """
    if score >= 65 or (confirmed and score >= 50):
        return ('Fix within 24 hours', 'critical')
    if score >= 45 or (confirmed and score >= 30):
        return ('Fix within 7 days', 'high')
    if score >= 28:
        return ('Fix within 30 days', 'medium')
    return ('Best practice / future sprint', 'low')


def _enrich_remediation(rem: dict, site_ctx: dict, evidence: list[str]) -> dict:
    """Inject site-specific context and evidence into a matched remediation.

    - Personalises description with actual versions, EOL warnings, CVE IDs
    - Substitutes real paths/versions into fix code
    - Adds evidence quotes from the actual scan output
    - Computes urgency tier
    """
    rem = dict(rem)
    desc_parts: list[str] = [rem['description']]

    # PHP EOL warning
    if site_ctx.get('php_eol'):
        php_ver, eol_date, eol_msg = site_ctx['php_eol']
        if 'php' in rem.get('id', '').lower() or 'php' in rem['description'].lower():
            desc_parts.append(
                f'Your site is running PHP {php_ver} ({eol_msg} on {eol_date}). '
                'Upgrade to PHP 8.2+ immediately to receive security patches.'
            )

    # Server version context
    if site_ctx.get('server'):
        sn = site_ctx['server'].lower()
        if ('nginx' in sn and 'nginx' in ''.join(rem.get('fixes', {}).keys()).lower()) or \
           ('apache' in sn and 'apache' in ''.join(rem.get('fixes', {}).keys()).lower()):
            desc_parts.append(f'Detected on your server: {site_ctx["server"]}.')

    # CVE context
    rel_cves = rem.get('cves', [])
    if rel_cves:
        desc_parts.append(
            f'CVE(s) found in this scan: {", ".join(rel_cves[:4])}. '
            'Click "Run CVE Analysis" to get full descriptions and CVSS scores.'
        )

    # Specific paths
    if site_ctx.get('specific_paths'):
        path_str = ', '.join(site_ctx['specific_paths'][:3])
        if any(p in rem.get('id', '') for p in ('env', 'git', 'dir', 'server', 'version')):
            desc_parts.append(f'Paths detected in scan output: {path_str}')

    rem['description'] = ' '.join(desc_parts)

    # Evidence from actual scan output
    if evidence:
        rem['evidence'] = evidence

    # Urgency tier
    score      = rem.get('priority_score', 10)
    confirmed  = rem.get('confirmed', False)
    urgency, uc = _urgency_tier(score, confirmed)
    rem['urgency']       = urgency
    rem['urgency_class'] = uc

    # Substitute real values into fix code
    tgt = site_ctx.get('target', '')
    fixes: dict[str, str] = {}
    for k, v in rem.get('fixes', {}).items():
        if tgt:
            v = v.replace('yourdomain.com', tgt)
        v = v.replace('YOUR.OFFICE.IP.HERE', '[your office IP]')
        php_ver_str = site_ctx['versions'].get('PHP', '')
        if php_ver_str and 'php' in k.lower():
            major_minor = '.'.join(php_ver_str.split('.')[:2])
            v = v.replace('php-fpm.sock', f'php{major_minor}-fpm.sock')
        fixes[k] = v
    rem['fixes'] = fixes
    return rem


# Negation patterns specific to stack detection — a line saying
# "checking for xmlrpc.php" or "xmlrpc.php returned 404" is NOT evidence of WordPress.
_STACK_NEG = re.compile(
    r'\b(checking|testing|looking for|scanning for|probing|not found|'
    r'returns?\s*404|http\s*404|status\s*404|not present|not installed|'
    r'not detected|not running|no evidence|not a wordpress|not wordpress|'
    r'no wordpress|does not appear|does not use|is not running)\b',
    re.I,
)


def _detect_stacks(text: str) -> set[str]:
    """Detect server / CMS stacks referenced in agent output.

    Only fires on lines with *positive* evidence — lines that look like
    "checking for X" or "X returned 404" are skipped via _STACK_NEG.
    """
    found: set[str] = set()
    for line in text.splitlines():
        if _STACK_NEG.search(line) or _NEGATION.search(line):
            continue
        ll = line.lower()
        # WordPress: require definitive positive signals, not just any mention
        if any(k in ll for k in ('wp-content/', 'wp-admin/', 'wp-login.php',
                                  '/wp-json/', 'wordpress version', 'woocommerce',
                                  '/wp-includes/', 'wp-config.php', 'xmlrpc.php')):
            found.add('wp')
        if 'nginx' in ll:
            found.add('nginx')
        if 'apache' in ll or '.htaccess' in ll:
            found.add('apache')
        if 'php' in ll:
            found.add('php')
    return found


def _filter_fixes(fixes: dict, detected: set[str]) -> dict:
    """Return only the fix entries relevant to detected stacks.

    Fix keys that belong to an undetected stack are hidden; generic keys
    (certbot, bash, manual, general) are always shown.
    Falls back to all fixes when nothing was detected.
    """
    if not detected:
        return fixes
    filtered = {}
    for key, code in fixes.items():
        kl = key.lower()
        fix_stack = next(
            (sid for pattern, sid in _FIX_STACK_KEYS.items() if pattern in kl),
            None,
        )
        is_generic = fix_stack is None or any(
            g in kl for g in ('bash', 'manual', 'general', 'certbot', 'waf')
        )
        if is_generic or fix_stack in detected:
            filtered[key] = code
    return filtered if filtered else fixes


def _score_remediation(rem: dict, pos_lines: list[str],
                        cves_by_line: list[tuple[str, set[str]]],
                        cvss_max: float) -> tuple[int, bool, list[str]]:
    """Compute priority score for a matched remediation.

    Returns (score, confirmed, cve_ids_relevant).
    Higher score → shown first in priority actions list.
    """
    base     = _SEV_WEIGHT.get(rem.get('severity', 'LOW'), 10)
    patterns = rem['patterns']

    # Lines in the scan output that actually matched this remediation
    hit_lines = [l for l in pos_lines if any(p in l for p in patterns)]

    # Confirmation: at least one hit line also has positive-finding language
    confirmed = any(_CONFIRM_RE.search(l) for l in hit_lines)
    if confirmed:
        base += 15

    # CVE bonus: CVEs that appear ON the same line as a pattern hit
    relevant_cves: set[str] = set()
    for line, line_cves in cves_by_line:
        if any(p in line for p in patterns) and line_cves:
            relevant_cves |= line_cves
    base += min(len(relevant_cves) * 5, 20)

    # CVSS score bonus (global max from scan output)
    if cvss_max >= 9.0:
        base += 12
    elif cvss_max >= 7.0:
        base += 7
    elif cvss_max >= 4.0:
        base += 3

    # More confirming lines → higher confidence signal
    base += min(len(hit_lines), 5) * 2

    return base, confirmed, sorted(relevant_cves)


def _extract_http_headers_from_output(text: str) -> dict[str, str]:
    """Pull HTTP response header name→value pairs mentioned in scan output.

    Handles lines like:
      Server: nginx/1.18.0
      X-Powered-By: PHP/7.4.3
      > strict-transport-security: max-age=31536000
    """
    hdr_re = re.compile(
        r'^\s*>?\s*([a-zA-Z][-a-zA-Z0-9]{2,40})\s*:\s*(.{1,200})$', re.M)
    headers = {}
    for m in hdr_re.finditer(text):
        name  = m.group(1).strip().lower()
        value = m.group(2).strip()
        # Avoid picking up line-noise (e.g. "URL: https://...")
        if name in ('url', 'host', 'connection', 'date', 'user-agent',
                    'accept', 'accept-encoding', 'content-length', 'transfer-encoding',
                    'keep-alive', 'vary', 'cache-control'):
            continue
        if name not in headers:
            headers[name] = value
    return headers


# ── Scan tool output parsers ──────────────────────────────────────────────────
# Each parser takes raw scan output and returns a list of synthetic remediation
# dicts that feed directly into the match_remediations pipeline.  Entries are
# pre-scored so structured tool findings always outrank generic pattern matches.

_NIKTO_ITEM_RE = re.compile(
    r'\+\s+([^:]+):\s+(.+)', re.M)
_NIKTO_CVE_RE  = re.compile(r'CVE-\d{4}-\d{4,}', re.I)
_NIKTO_OSVDB_RE = re.compile(r'OSVDB-\d+', re.I)

def _parse_nikto_output(text: str) -> list[dict]:
    """Extract structured findings from Nikto scan output.

    Nikto lines look like:
        + Server: Apache/2.4.29
        + OSVDB-3268: /admin/: Directory indexing found.
        + CVE-2017-7679: mod_mime buffer overread.
        + X-Frame-Options header is not present.
    """
    if '+ Target IP' not in text and 'Nikto' not in text:
        return []

    synthetic: list[dict] = []
    # Mapping: Nikto message fragment → remediation template id to boost
    _nikto_hint_map = [
        ('x-frame-options',         'missing-x-frame-options',  55),
        ('x-content-type-options',  'missing-xcto',             50),
        ('strict-transport-security','missing-hsts',             60),
        ('content-security-policy', 'missing-csp',              55),
        ('httponly',                'insecure-cookies',          55),
        ('samesite',                'insecure-cookies',          50),
        ('directory indexing',      'directory-listing',         70),
        ('directory listing',       'directory-listing',         70),
        ('robots.txt',              'info-disclosure',           40),
        ('server banner',           'info-disclosure',           45),
        ('x-powered-by',            'info-disclosure',           45),
        ('etag',                    'info-disclosure',           40),
        ('ssl',                     'ssl-expired',               50),
        ('tls 1.0',                 'tls-old-versions',          65),
        ('tls 1.1',                 'tls-old-versions',          65),
        ('sslv2',                   'tls-old-versions',          70),
        ('sslv3',                   'tls-old-versions',          70),
        ('put method',              'http-put-delete-methods',   75),
        ('delete method',           'http-put-delete-methods',   75),
        ('trace method',            'http-trace-method',         65),
        ('phpinfo',                 'info-disclosure',           70),
        ('sql',                     'sql-injection',             60),
        ('xss',                     'xss-reflected',             65),
        ('/wp-login',               'wp-admin-exposed',          70),
        ('/wp-content',             'wp-outdated-plugin',        45),
        ('php/5.',                  'php-outdated',              80),
        ('php/7.0',                 'php-outdated',              75),
        ('php/7.1',                 'php-outdated',              75),
        ('php/7.2',                 'php-outdated',              75),
        ('php/7.3',                 'php-outdated',              70),
        ('apache/2.2',              'server-version-exposed',    65),
        ('iis/6.',                  'server-version-exposed',    70),
        ('iis/7.',                  'server-version-exposed',    65),
        ('default files',           'default-credentials',       60),
        ('default page',            'info-disclosure',           45),
        ('backup',                  'info-disclosure',           55),
        ('.bak',                    'info-disclosure',           60),
        ('.old',                    'info-disclosure',           55),
        ('/.git',                   'info-disclosure',           75),
        ('/.svn',                   'info-disclosure',           70),
        ('cross-site',              'xss-reflected',             65),
        ('clickjack',               'missing-x-frame-options',   65),
        ('basic auth',              'http-basic-auth-unencrypted', 65),
        ('no auth',                 'default-credentials',       60),
        ('weak cipher',             'tls-weak-cipher',           70),
        ('rc4',                     'tls-weak-cipher',           75),
    ]

    for m in _NIKTO_ITEM_RE.finditer(text):
        label = m.group(1).strip()
        detail = m.group(2).strip()
        combined = (label + ' ' + detail).lower()
        cves = [c.upper() for c in _NIKTO_CVE_RE.findall(label + ' ' + detail)]

        for hint, rem_id, score in _nikto_hint_map:
            if hint in combined:
                # Find the base template and clone it with boosted score
                base = next((r for r in REMEDIATIONS if r['id'] == rem_id), None)
                if base:
                    entry = dict(base)
                    entry['priority_score'] = score + (10 if cves else 0)
                    entry['confirmed']      = True
                    entry['cves']           = cves
                    entry['evidence']       = [f'[Nikto] {label}: {detail[:120]}']
                    entry['urgency'], entry['urgency_class'] = _urgency_tier(
                        entry['priority_score'], True)
                    synthetic.append(entry)
                break

    return synthetic


_WPSCAN_VULN_RE = re.compile(
    r'\|\s+(?:Title|Name):\s+(.+?)(?:\n|\r)', re.M)
_WPSCAN_CVE_RE  = re.compile(r'CVE-\d{4}-\d{4,}', re.I)
_WPSCAN_FIXED_IN_RE = re.compile(r'Fixed\s+in\s*:\s*([\d.]+)', re.I)
_WPSCAN_CVSS_RE = re.compile(r'(?:CVSS|Score)\s*[:\-=]\s*([\d.]+)', re.I)

def _parse_wpscan_output(text: str) -> list[dict]:
    """Extract structured findings from WPScan output.

    WPScan blocks look like:
     | Title: WordPress 5.8 - SQL injection via WP_Query
     |  - CVE: CVE-2022-21661
     |  - CVSS Score: 9.8
     |  - Fixed in: 5.8.3
    """
    if '[+] URL' not in text and 'wpscan' not in text.lower() and 'WordPress' not in text:
        return []

    synthetic: list[dict] = []

    # Detect plugin-specific vulnerabilities (WPScan lists them per plugin)
    plugin_vuln_re = re.compile(
        r'\[!\]\s+(.+?)\s*\n.*?(?:CVE|vulnerability|vuln|inject|xss|sqli)', re.I | re.S)

    # Map WPScan finding titles to remediation IDs
    _wp_hint_map = [
        ('sql',             'sql-injection-confirmed', 90),
        ('xss',             'xss-reflected',           80),
        ('csrf',            'csrf-missing',             75),
        ('rce',             'rce-possible',             95),
        ('file inclusion',  'rce-possible',             90),
        ('file upload',     'unsafe-file-upload',       85),
        ('path traversal',  'path-traversal',           80),
        ('ssrf',            'ssrf',                     80),
        ('xxe',             'xxe',                      80),
        ('ssti',            'ssti',                     80),
        ('privilege',       'missing-authz',            75),
        ('authentication',  'wp-user-enum',             65),
        ('enumeration',     'wp-user-enum',             70),
        ('user enumerat',   'wp-user-enum',             75),
        ('outdated',        'wp-outdated-plugin',        65),
        ('update',          'wp-outdated-plugin',        60),
        ('admin',           'wp-admin-exposed',         70),
        ('debug',           'wp-debug-mode',            65),
        ('xmlrpc',          'wp-xmlrpc-enabled',        70),
        ('backup',          'info-disclosure',          65),
        ('redirect',        'open-redirect',            65),
        ('log4',            'log4shell',                95),
        ('deserialization', 'rce-possible',             90),
        ('open redirect',   'open-redirect',            65),
        ('plugin',          'wp-outdated-plugin',       55),
        ('theme',           'wp-outdated-plugin',       50),
        ('default',         'wp-default-content',       55),
    ]

    seen_ids: set[str] = set()
    for m in _WPSCAN_VULN_RE.finditer(text):
        title = m.group(1).strip()
        title_l = title.lower()

        # Grab CVEs and CVSS from the same block (~10 lines after title)
        block_start = m.end()
        block = text[block_start:block_start + 500]
        cves  = [c.upper() for c in _WPSCAN_CVE_RE.findall(title + ' ' + block)]
        cvss_vals = [float(x) for x in _WPSCAN_CVSS_RE.findall(block) if float(x) <= 10.0]
        cvss  = max(cvss_vals) if cvss_vals else 0.0
        fixed = (_WPSCAN_FIXED_IN_RE.search(block) or ['', ''])[0]
        if hasattr(fixed, 'group'):
            fixed = fixed.group(1)

        for hint, rem_id, base_score in _wp_hint_map:
            if hint in title_l:
                dedup_key = rem_id + ('|'.join(cves) if cves else title[:40])
                if dedup_key in seen_ids:
                    continue
                seen_ids.add(dedup_key)
                base = next((r for r in REMEDIATIONS if r['id'] == rem_id), None)
                if base:
                    entry = dict(base)
                    cvss_boost = int(cvss * 3) if cvss else 0
                    entry['priority_score'] = base_score + cvss_boost + (15 if cves else 0)
                    entry['confirmed']      = True
                    entry['cves']           = cves
                    ev = f'[WPScan] {title}'
                    if fixed:
                        ev += f' — fix available in v{fixed}'
                    entry['evidence']       = [ev[:140]]
                    entry['urgency'], entry['urgency_class'] = _urgency_tier(
                        entry['priority_score'], True)
                    synthetic.append(entry)
                break

    return synthetic


_NMAP_PORT_RE = re.compile(
    r'(\d+)/(tcp|udp)\s+(open|filtered)\s+(\S+)(?:\s+(.+))?', re.M)
_NMAP_CVE_RE  = re.compile(r'CVE-\d{4}-\d{4,}', re.I)

def _parse_nmap_services(text: str) -> list[dict]:
    """Extract service/port findings from nmap output.

    Lines like:
        22/tcp   open  ssh     OpenSSH 7.4 (protocol 2.0)
        3306/tcp open  mysql   MySQL 5.5.62
        6379/tcp open  redis   Redis key-value store
    """
    if 'Nmap scan report' not in text and 'nmap' not in text.lower():
        return []

    # Map (port or service keyword) → remediation id and base score
    _port_service_map = [
        (21,   'ftp',           'info-disclosure',              55),
        (22,   'ssh',           'server-version-exposed',       40),
        (23,   'telnet',        'info-disclosure',              80),
        (25,   'smtp',          'missing-spf',                  55),
        (53,   'dns',           'dns-zone-transfer',            60),
        (80,   'http',          'missing-hsts',                 50),
        (443,  'https',         'ssl-expired',                  40),
        (445,  'smb',           'info-disclosure',              75),
        (1433, 'mssql',         'info-disclosure',              70),
        (3306, 'mysql',         'info-disclosure',              65),
        (3389, 'rdp',           'info-disclosure',              70),
        (5432, 'postgresql',    'info-disclosure',              60),
        (6379, 'redis',         'redis-exposed',                85),
        (8080, 'http-proxy',    'info-disclosure',              50),
        (8443, 'https-alt',     'ssl-expired',                  45),
        (9200, 'elasticsearch', 'elasticsearch-exposed',        85),
        (9300, 'elasticsearch', 'elasticsearch-exposed',        80),
        (27017,'mongodb',       'mongodb-exposed',              85),
        (5984, 'couchdb',       'info-disclosure',              70),
        (11211,'memcache',      'info-disclosure',              70),
        (2181, 'zookeeper',     'info-disclosure',              65),
        (4848, 'glassfish',     'default-credentials',          70),
        (8161, 'activemq',      'default-credentials',          70),
        (9090, 'websm',         'info-disclosure',              60),
    ]
    _service_keyword_map = [
        ('telnet',          'info-disclosure',              80),
        ('ftp',             'info-disclosure',              60),
        ('redis',           'redis-exposed',                85),
        ('elasticsearch',   'elasticsearch-exposed',        85),
        ('mongodb',         'mongodb-exposed',              85),
        ('mysql',           'info-disclosure',              65),
        ('postgres',        'info-disclosure',              60),
        ('smtp',            'missing-spf',                  55),
        ('smb',             'info-disclosure',              75),
        ('rdp',             'info-disclosure',              70),
        ('vnc',             'info-disclosure',              75),
        ('memcache',        'info-disclosure',              70),
        ('rpcbind',         'info-disclosure',              60),
        ('nfs',             'info-disclosure',              65),
        ('snmp',            'info-disclosure',              70),
        ('irc',             'info-disclosure',              50),
        ('ajp',             'info-disclosure',              70),
        ('jmx',             'default-credentials',          70),
        ('jboss',           'default-credentials',          75),
        ('tomcat',          'default-credentials',          70),
        ('jenkins',         'default-credentials',          75),
        ('docker',          'info-disclosure',              75),
        ('kubernetes',      'info-disclosure',              70),
        ('zookeeper',       'info-disclosure',              65),
        ('cassandra',       'info-disclosure',              65),
    ]

    synthetic: list[dict] = []
    seen: set[str] = set()

    for m in _NMAP_PORT_RE.finditer(text):
        port    = int(m.group(1))
        state   = m.group(3)
        service = (m.group(4) or '').lower()
        banner  = (m.group(5) or '').strip()

        if state not in ('open', 'filtered'):
            continue

        rem_id = None
        score  = 0

        # Check port first
        for p, svc_hint, rid, sc in _port_service_map:
            if port == p:
                rem_id, score = rid, sc
                break

        # Then service keyword
        if not rem_id:
            for kw, rid, sc in _service_keyword_map:
                if kw in service or kw in banner.lower():
                    rem_id, score = rid, sc
                    break

        if not rem_id:
            continue

        key = f'{rem_id}:{port}'
        if key in seen:
            continue
        seen.add(key)

        base = next((r for r in REMEDIATIONS if r['id'] == rem_id), None)
        if not base:
            continue

        entry = dict(base)
        entry['priority_score'] = score
        entry['confirmed']      = True
        entry['cves']           = []
        ev_banner = f' ({banner})' if banner else ''
        entry['evidence']       = [f'[Nmap] {port}/tcp open {service}{ev_banner}']
        entry['urgency'], entry['urgency_class'] = _urgency_tier(score, True)
        synthetic.append(entry)

    return synthetic


_NUCLEI_FINDING_RE = re.compile(
    r'\[(\S+)\]\s+\[(\w+)\]\s+\[(\w+)\]\s+(\S+)(?:\s+\[(.+?)\])?', re.M)

def _parse_nuclei_output(text: str) -> list[dict]:
    """Extract findings from Nuclei output.

    Nuclei lines look like:
        [template-id] [protocol] [severity] https://example.com [matcher-name]
        [CVE-2021-44228] [http] [critical] https://target.com/path
        [wordpress-user-enum] [http] [medium] https://target.com
    """
    if '[http]' not in text and '[dns]' not in text and 'nuclei' not in text.lower():
        return []

    _sev_to_score = {'critical': 95, 'high': 80, 'medium': 60, 'low': 40, 'info': 25}

    # Map nuclei template-id keywords → remediation ids
    _nuclei_hint_map = [
        ('cve-2021-44228',      'log4shell',                   100),
        ('cve-2021-45046',      'log4shell',                   100),
        ('log4j',               'log4shell',                   100),
        ('log4shell',           'log4shell',                   100),
        ('sqli',                'sql-injection-confirmed',      90),
        ('sql-injection',       'sql-injection-confirmed',      90),
        ('xss',                 'xss-reflected',                80),
        ('csrf',                'csrf-missing',                 75),
        ('ssrf',                'ssrf',                         85),
        ('xxe',                 'xxe',                          80),
        ('ssti',                'ssti',                         85),
        ('rce',                 'rce-possible',                 95),
        ('lfi',                 'path-traversal',               80),
        ('path-traversal',      'path-traversal',               80),
        ('open-redirect',       'open-redirect',                65),
        ('default-login',       'default-credentials',          80),
        ('default-password',    'default-credentials',          80),
        ('weak-password',       'weak-password-policy',         70),
        ('exposed-panel',       'info-disclosure',              65),
        ('exposure',            'info-disclosure',              60),
        ('hsts',                'missing-hsts',                 60),
        ('x-frame',             'missing-x-frame-options',      55),
        ('csp',                 'missing-csp',                  55),
        ('xcto',                'missing-xcto',                 50),
        ('cors',                'info-disclosure',              60),
        ('cors-misconfig',      'info-disclosure',              70),
        ('tls',                 'ssl-expired',                  55),
        ('ssl',                 'ssl-expired',                  55),
        ('wp-user-enum',        'wp-user-enum',                 75),
        ('wordpress',           'wp-admin-exposed',             60),
        ('wp-xmlrpc',           'wp-xmlrpc-enabled',            70),
        ('wp-debug',            'wp-debug-mode',                65),
        ('redis-unauth',        'redis-exposed',                90),
        ('elasticsearch',       'elasticsearch-exposed',        85),
        ('mongodb-unauth',      'mongodb-exposed',              90),
        ('graphql',             'graphql-introspection',        70),
        ('swagger',             'swagger-exposed',              65),
        ('jwt',                 'jwt-weak',                     75),
        ('api-key',             'api-key-exposed',              80),
        ('directory-listing',   'directory-listing',            70),
        ('git-exposed',         'info-disclosure',              80),
        ('env-exposure',        'info-disclosure',              85),
        ('phpinfo',             'info-disclosure',              75),
        ('backup-files',        'info-disclosure',              65),
        ('prototype-pollution', 'prototype-pollution',          70),
        ('sri',                 'sri-missing',                  55),
        ('nosql',               'nosql-injection',              80),
        ('header-injection',    'xss-reflected',                70),
        ('email-header',        'missing-spf',                  60),
        ('spf',                 'missing-spf',                  65),
        ('dmarc',               'missing-dmarc',                65),
    ]

    synthetic: list[dict] = []
    seen: set[str] = set()

    for m in _NUCLEI_FINDING_RE.finditer(text):
        template_id = m.group(1).lower()
        severity    = m.group(3).lower()
        extra       = (m.group(5) or '').lower()
        url         = m.group(4)

        base_score  = _sev_to_score.get(severity, 40)

        # Check for direct CVE template
        cves = [c.upper() for c in _NIKTO_CVE_RE.findall(template_id)]

        rem_id = None
        score  = base_score

        for hint, rid, hint_score in _nuclei_hint_map:
            if hint in template_id or hint in extra:
                rem_id = rid
                score  = max(base_score, hint_score)
                break

        if not rem_id and cves:
            # CVE without a mapped template → use generic info-disclosure or sql/xss
            rem_id = 'info-disclosure'
            score  = base_score + 15

        if not rem_id:
            continue

        key = f'{rem_id}:{template_id}'
        if key in seen:
            continue
        seen.add(key)

        base = next((r for r in REMEDIATIONS if r['id'] == rem_id), None)
        if not base:
            continue

        entry = dict(base)
        entry['priority_score'] = score
        entry['confirmed']      = True
        entry['cves']           = cves
        entry['evidence']       = [
            f'[Nuclei/{severity.upper()}] {template_id} @ {url[:80]}'
        ]
        entry['urgency'], entry['urgency_class'] = _urgency_tier(score, True)
        synthetic.append(entry)

    return synthetic


def match_remediations(text: str, target: str = '') -> list[dict]:
    """Return remediations dynamically scored, enriched, and personalised to this scan.

    Pipeline:
    1. Stack detection via text signals + HTTP header values from output
    2. CVE extraction and per-line attribution
    3. CVSS extraction for score boosting
    4. Per-remediation scoring (severity + confirmation + CVE + evidence count)
    5. Co-occurrence boosting (related vulnerabilities amplify each other)
    6. Site context extraction (server versions, PHP EOL, CVEs, open ports)
    7. Evidence line extraction (actual scan output lines that triggered each match)
    8. Enrichment (_enrich_remediation injects context + urgency tier into each item)
    9. Scan tool parser results (Nikto/WPScan/Nmap/Nuclei) merged at top
    10. Final sort: confirmed first, highest score first
    """
    raw_lines = text.splitlines()
    pos_lines = [l.lower() for l in raw_lines if not _REM_NEGATION.search(l)]
    pos_text  = ' '.join(pos_lines)

    # ── Stack detection enriched with HTTP header values ─────────────────────
    detected  = _detect_stacks(text)
    http_hdrs = _extract_http_headers_from_output(text)
    if 'nginx'    in (http_hdrs.get('server')       or '').lower():  detected.add('nginx')
    if 'apache'   in (http_hdrs.get('server')       or '').lower():  detected.add('apache')
    if 'php'      in (http_hdrs.get('x-powered-by') or '').lower():  detected.add('php')
    if 'wordpress'in (http_hdrs.get('x-powered-by') or '').lower():  detected.add('wp')

    # ── Per-line CVE extraction for attribution ───────────────────────────────
    cves_by_line: list[tuple[str, set[str]]] = []
    for raw_l in raw_lines:
        if _REM_NEGATION.search(raw_l):
            cves_by_line.append(('', set()))
            continue
        lc   = raw_l.lower()
        cvs  = {m.upper() for m in _CVE_RE.findall(raw_l)}
        cves_by_line.append((lc, cvs))

    # Max CVSS extracted from scan output
    cvss_vals = [float(m) for m in _CVSS_RE.findall(text) if float(m) <= 10.0]
    cvss_max  = max(cvss_vals) if cvss_vals else 0.0

    # ── Site context (versions, EOL, ports, specific paths) ───────────────────
    site_ctx = _extract_site_context(text, target)

    # ── Scan tool structured parsers (Nikto / WPScan / Nmap / Nuclei) ──────────
    # These run first and produce pre-confirmed entries with exact evidence lines.
    # They will be merged with pattern-matched results below, keeping the highest
    # score for any id that appears in both sources.
    tool_entries: dict[str, dict] = {}
    for tool_result in (
        _parse_nikto_output(text),
        _parse_wpscan_output(text),
        _parse_nmap_services(text),
        _parse_nuclei_output(text),
    ):
        for entry in tool_result:
            eid = entry['id']
            if eid not in tool_entries or entry['priority_score'] > tool_entries[eid]['priority_score']:
                tool_entries[eid] = entry

    # ── Initial matching pass ──────────────────────────────────────────────────
    matched: list[dict] = []
    matched_ids: set[str] = set()

    for rem in REMEDIATIONS:
        if not any(p in pos_text for p in rem['patterns']):
            continue
        if rem['id'].startswith('wp-') and 'wp' not in detected:
            continue

        score, confirmed, rel_cves = _score_remediation(
            rem, pos_lines, cves_by_line, cvss_max)

        fixes = _filter_fixes(rem['fixes'], detected)

        evidence = _extract_evidence_lines(rem['patterns'], pos_lines, n=4)

        # Merge with any tool-parser entry for the same id
        if rem['id'] in tool_entries:
            tool_e = tool_entries.pop(rem['id'])
            score    = max(score, tool_e['priority_score'])
            confirmed = confirmed or tool_e.get('confirmed', False)
            rel_cves  = list(dict.fromkeys(rel_cves + tool_e.get('cves', [])))
            evidence  = (tool_e.get('evidence', []) + evidence)[:5]

        matched.append({
            **rem,
            'fixes':          fixes,
            'priority_score': score,
            'confirmed':      confirmed,
            'cves':           rel_cves,
            'evidence':       evidence,
        })
        matched_ids.add(rem['id'])

    # Append any tool-parser entries whose template id wasn't in REMEDIATIONS patterns
    for entry in tool_entries.values():
        if entry['id'] not in matched_ids:
            matched.append(entry)
            matched_ids.add(entry['id'])

    # ── Co-occurrence boost pass ───────────────────────────────────────────────
    for id_pair, boost_id, boost_amount in _COOCCURRENCE:
        if id_pair.issubset(matched_ids):
            for rem in matched:
                if rem['id'] == boost_id:
                    rem['priority_score'] += boost_amount
                    if not rem.get('confirmed'):
                        rem['confirmed'] = True   # co-occurring vulns confirm each other
                    break

    # ── Enrich each matched item with site context + urgency tier ─────────────
    matched = [_enrich_remediation(rem, site_ctx, rem.get('evidence', []))
               for rem in matched]

    # ── Final sort: confirmed first, then by priority score descending ─────────
    matched.sort(key=lambda r: (r['confirmed'], r['priority_score']), reverse=True)
    return matched


_AGENT_LABELS = {
    'info':    'Information Gathering',
    'conf':    'Configuration Review',
    'athn':    'Authentication Testing',
    'athz':    'Authorization Testing',
    'sess':    'Session Management',
    'inpv':    'Input Validation',
    'cryp':    'Cryptography Review',
    'clnt':    'Client-Side Testing',
    'apit':    'API Security Testing',
    'js':      'JavaScript Analysis',
    'idnt':    'Identity Management',
    'ctf':     'CTF / Challenge',
    'ot':      'OT/ICS Security',
    'enum':    'API Enumeration',
    'pentest': 'Full Penetration Test',
    'recon-passive': 'Passive Recon',
    'breach':        'Breach & Credential Check',
    'subdomain':     'Subdomain Enumeration',
    'tech-detect':   'Tech Stack Detection',
    'supply-chain':  'Supply Chain / JS Audit',
    'recon':   'Reconnaissance',
    'analyst': 'Security Analysis',
    'exploit': 'Exploit Development',
}


def agent_label(a: str) -> str:
    return _AGENT_LABELS.get((a or '').lower(), (a or '').upper())


def _norm_target(raw: str) -> str:
    """Strip scheme and path — keep only lowercase host[:port]."""
    t = (raw or '').replace('https://', '').replace('http://', '')
    return t.split('/')[0].split('?')[0].rstrip('.').lower()


def enrich(scan: dict) -> dict:
    scan = dict(scan)
    scan['target']       = _norm_target(scan.get('target', ''))
    out  = scan.get('output', '') or ''
    scan['risk']         = risk_level(out)
    scan['agent_label']  = agent_label(scan.get('agent_type', ''))
    scan['recs']         = extract_recs(out)
    remediations         = match_remediations(out, target=scan['target'])
    username             = scan.get('username', '')

    # ── Age escalation — persist findings history and boost long-open items ────
    for rem in remediations:
        try:
            db.update_finding_history(scan['target'], rem['id'], username)
            age_days = db.get_finding_age_days(scan['target'], rem['id'], username)
            if age_days >= 90:
                rem['priority_score'] = rem.get('priority_score', 0) + 30
                rem['age_label'] = f'Open {age_days}d — escalated'
                rem['urgency'], rem['urgency_class'] = _urgency_tier(
                    rem['priority_score'], rem.get('confirmed', False))
            elif age_days >= 60:
                rem['priority_score'] = rem.get('priority_score', 0) + 20
                rem['age_label'] = f'Open {age_days}d'
                rem['urgency'], rem['urgency_class'] = _urgency_tier(
                    rem['priority_score'], rem.get('confirmed', False))
            elif age_days >= 30:
                rem['priority_score'] = rem.get('priority_score', 0) + 10
                rem['age_label'] = f'Open {age_days}d'
            else:
                rem['age_label'] = ''
        except Exception:
            rem['age_label'] = ''

    scan['remediations'] = remediations
    scan['preview']      = out[:400].replace('\n', ' ')
    dt = scan.get('created_at', '') or ''
    scan['display_date'] = dt[:16].replace('T', ' ')
    return scan


# ── Routes ────────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    ctx = _build_template_context()
    user = session.get('user', {'username': 'admin', 'role': 'admin'})
    ctx['current_user'] = user
    # Inject per-user page permissions (None for admins = sees everything)
    if user.get('role') == 'admin':
        ctx['user_allowed_pages'] = None
        ctx['user_plan'] = 'pro'
    else:
        users = _load_users()
        u = users.get(user['username'], {})
        ctx['user_allowed_pages'] = u.get('allowed_pages', [
            'dashboard', 'chatbot', 'pluginlogs', 'logexplorer', 'inventories', 'network',
        ])
        ctx['user_plan'] = u.get('plan', 'basic')
        # Always read fresh profile fields from users.json so template reflects saved changes
        user = dict(user)
        user['country']       = u.get('country', user.get('country', ''))
        user['currency_code'] = u.get('currency_code', user.get('currency_code', 'USD'))
        user['email']         = u.get('email', user.get('email', ''))
        ctx['current_user'] = user
    ctx['midtrans_client_key'] = _MIDTRANS_CLIENT_KEY
    ctx['midtrans_snap_js_url'] = _midtrans_snap_js_url()
    ctx['midtrans_configured'] = bool(_MIDTRANS_SERVER_KEY and _MIDTRANS_CLIENT_KEY)
    ctx['app_version']    = _APP_VERSION
    ctx['app_build_date'] = _APP_BUILD_DATE
    ctx['app_changelog']  = _APP_CHANGELOG
    resp = make_response(render_template('index.html', **ctx))
    resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp.headers['Pragma'] = 'no-cache'
    return resp


def _build_template_context() -> dict:
    """Build the full template context dict — shared with FastAPI."""
    _uf     = _cu_filter()
    scans   = [enrich(s) for s in db.get_scans(username=_uf)]
    targets = [enrich(t) for t in db.get_targets(username=_uf)]
    stats   = db.get_stats()

    # Keep only the HIGHEST-SCORED version of each (target, finding_id) pair.
    # This ensures: one entry per unique vulnerability per site, always the most
    # evidence-rich scan wins, no stale duplicate entries from older scans.
    _best: dict[tuple, dict] = {}   # (target, finding_key) → best rec entry

    for s in scans:
        tgt = s['target']
        for rem in s['remediations']:
            key = (tgt, rem['id'])
            new_score = rem.get('priority_score', 0)
            if key not in _best or new_score > _best[key].get('priority_score', 0):
                _best[key] = {
                    'target':         tgt,
                    'risk':           rem['severity'],
                    'text':           rem['title'],
                    'description':    rem.get('description', ''),
                    'agent':          s['agent_label'],
                    'date':           s['display_date'][:10],
                    'scan_id':        s['id'],
                    'has_fixes':      True,
                    'priority_score': new_score,
                    'confirmed':      rem.get('confirmed', False),
                    'urgency':        rem.get('urgency', 'Best practice / future sprint'),
                    'urgency_class':  rem.get('urgency_class', 'low'),
                    'cves':           rem.get('cves', []),
                    'evidence':       rem.get('evidence', []),
                    'age_label':      rem.get('age_label', ''),
                }
        for r in s['recs']:
            key = (tgt, r[:60].lower())
            if key not in _best:
                _best[key] = {
                    'target':         tgt,
                    'risk':           rec_risk(r),
                    'text':           r,
                    'description':    '',
                    'agent':          s['agent_label'],
                    'date':           s['display_date'][:10],
                    'scan_id':        s['id'],
                    'has_fixes':      False,
                    'priority_score': 10,
                    'confirmed':      False,
                    'urgency':        'Best practice / future sprint',
                    'urgency_class':  'low',
                    'cves':           [],
                    'evidence':       [],
                }

    all_recs = list(_best.values())
    # Sort by: confirmed DESC, priority_score DESC, severity tier, then target
    _prio = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2, 'INFO': 3}
    all_recs.sort(key=lambda x: (
        0 if x['confirmed'] else 1,
        -x.get('priority_score', 0),
        _prio.get(x['risk'], 3),
        x['target'],
    ))

    from collections import defaultdict as _dd
    _tgt_sev: dict = _dd(lambda: {'HIGH': 0, 'MEDIUM': 0, 'LOW': 0, 'INFO': 0})
    for s in scans:
        _tgt_sev[s['target']][s['risk']] += 1
    severity_summary = {
        'total': {
            'HIGH':   sum(v['HIGH']   for v in _tgt_sev.values()),
            'MEDIUM': sum(v['MEDIUM'] for v in _tgt_sev.values()),
            'LOW':    sum(v['LOW']    for v in _tgt_sev.values()),
            'INFO':   sum(v['INFO']   for v in _tgt_sev.values()),
        },
        'by_target': sorted(
            [{'target': k, **v} for k, v in _tgt_sev.items()],
            key=lambda x: x['HIGH'] * 10 + x['MEDIUM'] * 3 + x['LOW'],
            reverse=True,
        )[:12],
    }
    return dict(scans=scans, targets=targets, stats=stats,
                all_recs=all_recs[:40], severity_summary=severity_summary)


@app.route('/api/scan/<int:scan_id>')
def api_scan(scan_id):
    try:
        row = db.get_scan(scan_id)
        if not row:
            return jsonify({'error': f'Scan #{scan_id} not found'}), 404
        return jsonify(enrich(row))
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/stats')
def api_stats():
    return jsonify(db.get_stats())


@app.route('/api/scans/recent')
def api_scans_recent():
    """Return the latest scans as JSON for live dashboard refresh."""
    limit  = min(int(request.args.get('limit', 50)), 200)
    target = request.args.get('target', '')
    rows   = db.get_scans(limit=limit, username=_cu_filter())
    if target:
        rows = [r for r in rows if r.get('target', '').lower() == target.lower()]
    return jsonify([enrich(r) for r in rows])


_RISK_HIGH_RE = re.compile(
    r'REFLECTED\s+XSS:|SQL\s+ERROR:|CODE\s+INJECTION\s+CONFIRMED:|CMD\s+INJECTION:'
    r'|SSTI\s+HIT|SSRF\s+HIT:|CREDS_FOUND|FOUND_DB_USER:|FOUND_ENV_USER:'
    r'|APP_PASS_CREATED|EXPOSED_FILE\s*\|.*\b20[0-9]\b'
    r'|WP-LOG[^|\n]*\|\s*HIGH|\|\s*(High|Critical)\s*\|', re.I)
_RISK_MED_RE = re.compile(
    r'OPEN\s+REDIRECT:|HTML\s+INJECTION:|WP-USER-CONFIRMED|WP-USER\s*\|'
    r'|WP-LOG[^|\n]*\|\s*MEDIUM|\|\s*Medium\s*\|', re.I)
_RISK_LOW_RE = re.compile(r'\|\s*(Low|Info)\s*\||\d+/tcp\s+open', re.I)

def _scan_risk(out: str) -> str:
    if _RISK_HIGH_RE.search(out): return 'HIGH'
    if _RISK_MED_RE.search(out):  return 'MEDIUM'
    if _RISK_LOW_RE.search(out):  return 'LOW'
    return 'INFO'

@app.route('/api/scans/summary')
def api_scans_summary():
    """Lightweight scan list — pre-computed risk, no output text. For dashboard charts/KPIs."""
    limit = min(int(request.args.get('limit', 500)), 2000)
    rows  = db.get_recent_scans(limit, username=_cu_filter())
    return jsonify([{
        'id':         r['id'],
        'target':     r['target'],
        'agent_type': r['agent_type'],
        'created_at': r['created_at'],
        'status':     r['status'],
        'latency_s':  r['latency_s'],
        'tool_count': r['tool_count'],
        'risk':       _scan_risk(r.get('output') or ''),
    } for r in rows])


@app.route('/api/scan/<int:scan_id>/cve')
def api_scan_cve(scan_id):
    """Query NVD for real CVEs matching the technologies found in this scan."""
    row = db.get_scan(scan_id)
    if not row:
        abort(404)
    from dashboard.cve import cve_lookup_for_scan
    result = cve_lookup_for_scan(row.get('output', '') or '', row.get('target', ''))
    return jsonify(result)


@app.route('/api/target/<path:target>/wp-logs')
def api_wp_logs(target):
    """Return all WP Activity Log entries parsed from scans for a given target."""
    scans = db.get_scans_for_target(target, username=_cu_filter())
    all_entries = []
    overall_status = 'none'
    overall_msg = ''
    for s in scans:
        result = extract_wp_logs(s.get('output', '') or '')
        if result['entries']:
            overall_status = 'found'
            for e in result['entries']:
                e['scan_id']   = s['id']
                e['scan_date'] = (s.get('created_at') or '')[:16]
                e['agent']     = s.get('agent_type', '')
            all_entries.extend(result['entries'])
        elif result['status'] not in ('none',) and overall_status == 'none':
            overall_status = result['status']
            overall_msg    = result['status_msg']
    all_entries.sort(key=lambda e: e.get('timestamp', ''), reverse=True)
    # Enrich entries with geolocation country from the IP field
    for e in all_entries:
        e['country'] = _geoip(e.get('ip', ''))
    return jsonify({
        'target':      target,
        'logs':        all_entries,
        'scan_count':  len(scans),
        'status':      overall_status,
        'status_msg':  overall_msg,
    })


@app.route('/api/scan', methods=['POST'])
def api_save_scan():
    """Remote save endpoint — lets a CLI on another machine push scan results here.

    Expects JSON body:
        { "target": "...", "agent_type": "...", "model": "...",
          "status": "ok", "latency_s": 12.3, "tool_count": 5, "output": "..." }

    If CFAI_API_KEY is set in .env, include header:
        X-CFAI-Key: <your-key>
    """
    if _API_KEY:
        key = request.headers.get('X-CFAI-Key', '')
        if key != _API_KEY:
            return jsonify({'error': 'Unauthorized'}), 401

    data = request.get_json(force=True, silent=True) or {}
    required = {'target', 'agent_type', 'output'}
    missing  = required - data.keys()
    if missing:
        return jsonify({'error': f'Missing fields: {", ".join(missing)}'}), 400

    db.save_scan(
        target     = str(data['target'])[:500],
        agent_type = str(data['agent_type'])[:50],
        model      = str(data.get('model', ''))[:100],
        status     = str(data.get('status', 'ok'))[:20],
        latency_s  = float(data.get('latency_s', 0)),
        tool_count = int(data.get('tool_count', 0)),
        output     = str(data['output'])[:60000],
        username   = _cu_username(),
    )
    return jsonify({'saved': True}), 201


@app.route('/api/target/<path:target>/analytics')
def api_target_analytics(target):
    from dashboard.monitor import get_target_analytics
    from dashboard.security_apis import get_site_scores
    scans = [enrich(s) for s in db.get_scans_for_target(target, username=_cu_filter())]
    analytics = get_target_analytics(target, scans)
    scores = get_site_scores(target)
    return jsonify({**analytics, 'scores': scores})


@app.route('/api/target/<path:target>/compare')
def api_target_compare(target):
    from dashboard.monitor import compare_scans
    scans = [enrich(s) for s in db.get_scans_for_target(target, username=_cu_filter())]
    if len(scans) < 2:
        return jsonify({'error': 'Need at least 2 scans to compare', 'new': [], 'resolved': [], 'persistent': []})
    result = compare_scans(scans[0], scans[1])  # latest vs previous
    result['latest_date'] = scans[0].get('display_date', '')
    result['previous_date'] = scans[1].get('display_date', '')
    return jsonify(result)


@app.route('/api/connect/scan', methods=['POST'])
def api_connect_scan():
    """Start a background scan for the Connect Your Website feature.

    Request JSON:
      { "target": "example.com", "agent_type": "apit", "model": "",
        "site_type": "wordpress|cpanel|ssh|sftp|mysql|none",
        "wp_user": "", "wp_pass": "", "wp_app_pass": "",
        "cpanel_user": "", "cpanel_pass": "",
        "ssh_host": "", "ssh_user": "", "ssh_pass": "", "ssh_port": "", "ssh_key": "",
        "ftp_host": "", "ftp_user": "", "ftp_pass": "", "ftp_port": "",
        "db_host": "", "db_port": "", "db_name": "", "db_user": "", "db_pass": "", "db_prefix": "" }
    Response: { "job_id": "<uuid>" }
    """
    data = request.get_json(force=True, silent=True) or {}
    target = (data.get('target') or '').strip()
    if not target:
        return jsonify({'error': 'target is required'}), 400

    def _s(k): return (data.get(k) or '').strip()

    agent_type = _s('agent_type') or 'apit'
    model      = _s('model')
    site_type  = _s('site_type') or 'none'

    creds = {
        'wp_user':       _s('wp_user'),
        'wp_pass':       _s('wp_pass'),
        'wp_app_pass':   _s('wp_app_pass'),
        'cpanel_user':   _s('cpanel_user'),
        'cpanel_pass':   _s('cpanel_pass'),
        'ssh_host':      _s('ssh_host'),
        'ssh_user':      _s('ssh_user'),
        'ssh_pass':      _s('ssh_pass'),
        'ssh_port':      _s('ssh_port'),
        'ssh_key':       _s('ssh_key'),
        'ftp_host':      _s('ftp_host'),
        'ftp_user':      _s('ftp_user'),
        'ftp_pass':      _s('ftp_pass'),
        'ftp_port':      _s('ftp_port'),
        'db_host':       _s('db_host'),
        'db_port':       _s('db_port') or '3306',
        'db_name':       _s('db_name'),
        'db_user':       _s('db_user'),
        'db_pass':       _s('db_pass'),
        'db_prefix':     _s('db_prefix') or 'wp_',
        'drupal_url':    _s('drupal_url'),
        'drupal_user':   _s('drupal_user'),
        'drupal_pass':   _s('drupal_pass'),
        'joomla_user':   _s('joomla_user'),
        'joomla_pass':   _s('joomla_pass'),
        'joomla_token':  _s('joomla_token'),
        'gen_url':       _s('gen_url'),
        'gen_user':      _s('gen_user'),
        'gen_pass':      _s('gen_pass'),
        'gen_api_key':   _s('gen_api_key'),
        'gen_framework': _s('gen_framework'),
        'plesk_host':    _s('plesk_host'),
        'plesk_user':    _s('plesk_user'),
        'plesk_pass':    _s('plesk_pass'),
        'plesk_port':    _s('plesk_port') or '8443',
        'pg_host':       _s('pg_host'),
        'pg_port':       _s('pg_port') or '5432',
        'pg_name':       _s('pg_name'),
        'pg_user':       _s('pg_user') or 'postgres',
        'pg_pass':       _s('pg_pass'),
        'pg_ssl':        _s('pg_ssl') or 'prefer',
        'mongo_uri':     _s('mongo_uri'),
        'mongo_host':    _s('mongo_host'),
        'mongo_port':    _s('mongo_port') or '27017',
        'mongo_db':      _s('mongo_db'),
        'mongo_user':    _s('mongo_user'),
        'mongo_pass':    _s('mongo_pass'),
        'fb_project':    _s('fb_project'),
        'fb_db_url':     _s('fb_db_url'),
        'fb_api_key':    _s('fb_api_key'),
        'fb_json':       _s('fb_json'),
        'redis_host':    _s('redis_host'),
        'redis_port':    _s('redis_port') or '6379',
        'redis_pass':    _s('redis_pass'),
        'redis_db':      _s('redis_db') or '0',
        'mssql_host':    _s('mssql_host'),
        'mssql_port':    _s('mssql_port') or '1433',
        'mssql_db':      _s('mssql_db'),
        'mssql_user':    _s('mssql_user') or 'sa',
        'mssql_pass':    _s('mssql_pass'),
        'mssql_auth':    _s('mssql_auth') or 'sql',
    }

    job_id = str(_uuid.uuid4())
    _scan_jobs[job_id] = {
        'status':  'running',
        'target':  target,
        'agent':   agent_type,
        'chunks':  [],
        'domain':  '',
        'scan_id': None,
        'error':   None,
        'aborted': False,
    }

    t = _threading.Thread(
        target=_run_background_scan,
        args=(job_id, target, agent_type, model, site_type, creds, _cu_username()),
        daemon=True,
    )
    t.start()
    return jsonify({'job_id': job_id}), 202


@app.route('/api/connect/scan/<job_id>', methods=['GET'])
def api_connect_scan_poll(job_id):
    """Poll for new chunks from a running background scan.

    Query param `offset` (int, default 0) — index of first unseen chunk.
    Response: { "status": "running"|"done"|"error", "chunks": [...],
                "next_offset": N, "domain": "...", "scan_id": null|int,
                "error": null|"..." }
    """
    job = _scan_jobs.get(job_id)
    if not job:
        # Server may have restarted — try to recover from disk
        saved = _job_load(job_id)
        if saved:
            return jsonify({
                'status':      saved.get('status', 'done'),
                'domain':      saved.get('domain', ''),
                'scan_id':     saved.get('scan_id'),
                'error':       saved.get('error'),
                'chunks':      [],
                'next_offset': 0,
                'recovered':   True,
            })
        return jsonify({'error': 'job not found — server may have restarted. Check Scan History for results.'}), 404

    offset = int(request.args.get('offset', 0))
    new_chunks = job['chunks'][offset:]
    return jsonify({
        'status':      job['status'],
        'domain':      job.get('domain', ''),
        'scan_id':     job.get('scan_id'),
        'error':       job.get('error'),
        'chunks':      new_chunks,
        'next_offset': offset + len(new_chunks),
    })


@app.route('/api/connect/scan/<job_id>/abort', methods=['POST'])
def api_connect_scan_abort(job_id):
    """Signal a running background scan to stop; partial findings will be saved."""
    job = _scan_jobs.get(job_id)
    if not job:
        return jsonify({'error': 'job not found'}), 404
    job['aborted'] = True
    # Keep status as 'running' so the frontend keeps polling until the
    # background thread finishes saving partial output, then sets 'interrupted'.
    job['chunks'].append({'k': 'txt', 'd': '\n[Stopping scan — saving findings so far...]\n'})
    return jsonify({'ok': True})


@app.route('/api/security-signals')
def api_security_signals():
    """Aggregate security events from scan history into a SIEM-style signal feed."""
    import re as _re

    _HIGH_SIGNALS = [
        (r'sql\s*inject|union\s+select|1=1|sleep\(|benchmark\(', 'SQL Injection Attempt'),
        (r'xss|<script|javascript:|onerror\s*=', 'Cross-Site Scripting (XSS)'),
        (r'path\s+traversal|\.\./|%2e%2e|directory\s+list', 'Path Traversal'),
        (r'rce|remote\s+code\s+exec|shell\s+upload|webshell', 'Remote Code Execution'),
        (r'brute.?force|login\s+attempt|credential\s+stuff', 'Brute Force Attack'),
        (r'exposed.*private.*key|api\s+key\s+leak|secret.*leak', 'Credential Exposure'),
        (r'command\s+inject|cmd\.exe|/bin/bash|eval\(base64', 'Command Injection'),
    ]
    _MED_SIGNALS = [
        (r'csrf|cross.site\s+request', 'CSRF Vulnerability'),
        (r'open\s+redirect|redirect.*http', 'Open Redirect'),
        (r'xxe|xml\s+external\s+entity', 'XXE Injection'),
        (r'idor|insecure\s+direct\s+object', 'IDOR'),
        (r'weak.*password|default\s+credential', 'Weak Credentials'),
        (r'missing\s+security\s+header|x-frame-options\s+missing|csp\s+missing', 'Missing Security Headers'),
        (r'outdated.*version|version.*vulnerabl|cve-\d{4}-\d+', 'Known CVE'),
        (r'xmlrpc|wordpress.*vuln|wp-login.*exposed', 'WordPress Exposure'),
    ]
    _LOW_SIGNALS = [
        (r'debug\s+mode|verbose\s+error|stack\s+trace', 'Debug Info Exposed'),
        (r'directory\s+listing|index\s+of\s+/', 'Directory Listing'),
        (r'ssl.*expired|certificate.*expir|self.signed', 'SSL Certificate Issue'),
        (r'banner\s+grab|server\s+version\s+exposed', 'Server Banner Exposure'),
    ]

    scans = [enrich(s) for s in db.get_recent_scans(50, username=_cu_filter())]
    signals = []
    seen = set()

    for s in scans:
        text = (s.get('output') or '') + ' '.join(s.get('recs', []))
        target = s.get('target', '')
        ts = s.get('display_date', '')
        scan_id = s.get('id', 0)

        for pat, label in _HIGH_SIGNALS:
            if _re.search(pat, text, _re.I):
                key = f'HIGH:{label}:{target}'
                if key not in seen:
                    seen.add(key)
                    signals.append({'severity': 'HIGH', 'event': label,
                                    'target': target, 'date': ts, 'scan_id': scan_id})
        for pat, label in _MED_SIGNALS:
            if _re.search(pat, text, _re.I):
                key = f'MED:{label}:{target}'
                if key not in seen:
                    seen.add(key)
                    signals.append({'severity': 'MEDIUM', 'event': label,
                                    'target': target, 'date': ts, 'scan_id': scan_id})
        for pat, label in _LOW_SIGNALS:
            if _re.search(pat, text, _re.I):
                key = f'LOW:{label}:{target}'
                if key not in seen:
                    seen.add(key)
                    signals.append({'severity': 'LOW', 'event': label,
                                    'target': target, 'date': ts, 'scan_id': scan_id})

    _sev_order = {'HIGH': 0, 'MEDIUM': 1, 'LOW': 2, 'INFO': 3}
    signals.sort(key=lambda x: _sev_order.get(x['severity'], 3))

    counts = {k: sum(1 for s in signals if s['severity'] == k)
              for k in ('HIGH', 'MEDIUM', 'LOW', 'INFO')}
    counts['CRITICAL'] = sum(1 for sig in signals
                             if sig['severity'] == 'HIGH' and
                             any(kw in sig['event'].lower()
                                 for kw in ('injection', 'rce', 'exposure', 'traversal')))
    return jsonify({'signals': signals[:200], 'counts': counts})


# ── Credential persistence ───────────────────────────────────────────────────
@app.route('/api/creds/save', methods=['POST'])
def api_creds_save():
    """Save credentials to server-side session so all pages can use them."""
    data      = request.get_json(force=True, silent=True) or {}
    cred_type = data.get('type', '')   # 'wordpress' | 'cpanel' | 'mysql'
    if not cred_type:
        return jsonify({'error': 'type required'}), 400
    # Accept either nested {"creds": {...}} or flat top-level keys (everything except 'type')
    creds = data.get('creds') or {k: v for k, v in data.items() if k != 'type'}
    if 'saved_creds' not in session:
        session['saved_creds'] = {}
    session['saved_creds'][cred_type] = creds
    session.modified = True
    return jsonify({'ok': True})

@app.route('/api/creds/load', methods=['GET'])
def api_creds_load():
    """Return which credential types have been saved (passwords masked)."""
    saved = session.get('saved_creds', {})
    result = {}
    for ctype, creds in saved.items():
        result[ctype] = {k: ('***' if any(s in k.lower() for s in ('pass', 'token', 'key', 'secret')) else v)
                         for k, v in creds.items()}
    return jsonify({'creds': result,
                    'has_wordpress': 'wordpress' in saved,
                    'has_mysql':     'mysql' in saved,
                    'has_cpanel':    'cpanel' in saved})

@app.route('/api/sync/all', methods=['POST'])
def api_sync_all():
    """Re-fetch from all saved credential sources and populate the DB."""
    import base64 as _b64, json as _j
    saved = session.get('saved_creds', {})
    results = {}

    # WordPress
    wp = saved.get('wordpress', {})
    if wp.get('url'):
        url      = wp['url'].rstrip('/')
        wp_user  = wp.get('wp_user', '')
        app_pass = wp.get('wp_app_pass', '')
        auth_hdr = ('Basic ' + _b64.b64encode(f'{wp_user}:{app_pass}'.encode()).decode()) if wp_user and app_pass else None
        def _wg(path):
            hdrs = {'Authorization': auth_hdr} if auth_hdr else {}
            code, body = _wp_request(f'{url}{path}', headers=hdrs, timeout=15)
            try: return _j.loads(body) if code == 200 else None
            except Exception: return None
        events, source = [], 'WordPress REST API'
        for path in [f'/wp-json/wsal/v1/events?per_page=100&order=DESC',
                     f'/wp-json/cfai-simple-history/v1/events?per_page=100']:
            data_r = _wg(path)
            if not data_r: continue
            items = data_r if isinstance(data_r, list) else (data_r.get('events') or data_r.get('data') or [])
            for ev in items:
                msg = str(ev.get('message') or ev.get('alert_message') or ev.get('type') or '')
                if not msg: continue
                events.append({'timestamp': str(ev.get('created_on') or ev.get('date') or ''),
                                'user': str(ev.get('user_login') or ev.get('username') or '—'),
                                'event': msg[:120], 'ip': str(ev.get('client_ip') or ev.get('ip') or ''),
                                'severity': 'HIGH' if any(k in msg.lower() for k in ('fail','block','brute','invalid')) else 'INFO',
                                'source': source})
        if events:
            _save_log_events_to_db(events, url, 'WordPress REST API', 'wp_log_sync')
            results['wordpress'] = f'{len(events)} events synced'

    # MySQL
    mysql = saved.get('mysql', {})
    if mysql.get('db_host') and mysql.get('db_name'):
        try:
            import pymysql, pymysql.cursors
            conn = pymysql.connect(host=mysql['db_host'], port=int(mysql.get('db_port', 3306)),
                                   user=mysql.get('db_user', ''), password=mysql.get('db_pass', ''),
                                   database=mysql['db_name'], charset='utf8mb4', connect_timeout=10,
                                   cursorclass=pymysql.cursors.DictCursor)
            pfx = mysql.get('table_prefix', 'wp_')
            events, source = [], 'MySQL'
            with conn.cursor() as cur:
                sh = f'{pfx}simple_history'
                cur.execute("SHOW TABLES LIKE %s", (sh,))
                if cur.fetchone():
                    cur.execute(f"SELECT id, date, logger, level, message FROM {sh} ORDER BY id DESC LIMIT 100")
                    for row in cur.fetchall():
                        msg = str(row.get('message') or row.get('logger') or '')
                        if not msg: continue
                        events.append({'timestamp': str(row.get('date') or ''), 'user': '—',
                                        'event': msg[:200], 'ip': '', 'severity': 'INFO', 'source': 'MySQL Simple History'})
            conn.close()
            if events:
                _save_log_events_to_db(events, mysql.get('db_name', 'wordpress'), 'MySQL Direct', 'mysql_log_sync')
                results['mysql'] = f'{len(events)} events synced'
        except Exception as e:
            results['mysql'] = f'error: {str(e)[:80]}'

    return jsonify({'ok': True, 'results': results, 'synced_at': _time.strftime('%Y-%m-%d %H:%M:%S')})


def _save_log_events_to_db(events: list, target: str, source: str, agent_type: str = 'log_sync') -> None:
    """Format log events as WP-LOG scan output and persist to the DB.

    This makes every log fetch visible across all analytics pages
    (Threat Analytics, MITRE, Security Signals, Event Timeline, etc.)
    since they all read from the scans table.
    """
    if not events:
        return
    lines = [f'# Log sync from {source} — {target} — {_time.strftime("%Y-%m-%d %H:%M:%S")}']
    for ev in events:
        ts  = (ev.get('timestamp') or '—')[:30]
        usr = (ev.get('user') or '—')[:40]
        msg = (ev.get('event') or '')[:200]
        ip  = (ev.get('ip') or '-')[:45]
        sev = (ev.get('severity') or 'INFO').upper()
        if sev not in ('HIGH', 'MEDIUM', 'LOW', 'INFO'):
            sev = 'INFO'
        lines.append(f'WP-LOG | {ts} | {usr} | {msg} | {ip} | {sev}')
        # Extra signal keywords so security-signals & MITRE parsers fire
        if any(k in msg.lower() for k in ('fail', 'brute', 'block', 'invalid', 'attack')):
            lines.append(f'FAILED_LOGIN | {usr} | {ip}')
        if 'plugin' in msg.lower():
            lines.append(f'WP-PLUGIN-CHANGE | {usr} | {msg[:80]}')
    output = '\n'.join(lines)
    db.save_scan(target=target, agent_type=agent_type, model='log_sync', username=_cu_username(),
                 status='ok', latency_s=0.0, tool_count=0, output=output)
    # Also write to security_events table for Event Timeline
    for ev in events:
        if ev.get('severity') in ('HIGH', 'MEDIUM') or any(
                k in (ev.get('event') or '').lower() for k in ('fail', 'brute', 'block', 'attack', 'invalid')):
            try:
                db.log_security_event(
                    event_type=ev.get('event', 'Log event')[:80],
                    category='authentication',
                    severity=ev.get('severity', 'INFO'),
                    ip_address=ev.get('ip', ''),
                    country=_geoip(ev.get('ip', '')),
                    target=target,
                )
            except Exception:
                pass


@app.route('/api/logs/wp-live', methods=['POST'])
def api_logs_wp_live():
    """Fetch real-time login activity from WordPress via REST API + WSAL/Simple History plugins."""
    import base64 as _b64, json as _j
    data     = request.get_json(force=True, silent=True) or {}
    url      = (data.get('url') or '').strip().rstrip('/')
    wp_user  = (data.get('wp_user') or '').strip()
    app_pass = (data.get('wp_app_pass') or '').strip()
    limit    = min(int(data.get('limit') or 50), 200)
    if not url:
        return jsonify({'error': 'WordPress site URL required'}), 400
    if not url.startswith('http'):
        url = 'https://' + url
    auth_header = None
    if wp_user and app_pass:
        auth_header = 'Basic ' + _b64.b64encode(f'{wp_user}:{app_pass}'.encode()).decode()

    def _wp_get(path, timeout=15, override_hdrs=None):
        hdrs = {}
        if override_hdrs:
            hdrs.update(override_hdrs)
        elif auth_header:
            hdrs['Authorization'] = auth_header
        code, body = _wp_request(f'{url}{path}', headers=hdrs, timeout=timeout)
        if code == 200:
            try:
                return _j.loads(body), None
            except Exception:
                return None, 'JSON parse error'
        return None, f'HTTP {code}'

    events, source, note = [], 'none', ''

    # 1. WP Activity Log (WSAL) — richest source, provides per-event IP + user + action
    for path in [f'/wp-json/wsal/v1/events?per_page={limit}&orderby=created_on&order=DESC',
                 f'/wp-json/wsal/v1/events?per_page={limit}']:
        wsal, _ = _wp_get(path)
        if not wsal:
            continue
        items = wsal if isinstance(wsal, list) else wsal.get('events') or wsal.get('data') or []
        for ev in items:
            msg = str(ev.get('message') or ev.get('alert_message') or ev.get('type') or '')
            if not msg:
                continue
            events.append({
                'timestamp': str(ev.get('created_on') or ev.get('date') or ''),
                'user':      str(ev.get('user_login') or ev.get('username') or ev.get('user') or '—'),
                'event':     msg[:120],
                'ip':        str(ev.get('client_ip') or ev.get('ip') or ''),
                'severity':  'HIGH' if any(k in msg.lower() for k in ('fail', 'block', 'attack', 'brute', 'invalid')) else 'INFO',
                'status':    'failed' if any(k in msg.lower() for k in ('fail', 'block', 'denied', 'invalid')) else 'success',
                'source':    'wsal',
            })
        if events:
            source = 'wsal'
            break

    # 2. Simple History plugin REST API
    # Source: bonny/WordPress-Simple-History inc/class-wp-rest-events-controller.php
    # Response fields: message, date_gmt, date_local, loglevel, initiator, initiator_data{user_login}, ip_addresses[]
    if not events:
        sh, _ = _wp_get(f'/wp-json/simple-history/v1/events?per_page={limit}')
        if sh and isinstance(sh, list):
            for ev in sh:
                msg = str(ev.get('message') or '')
                if not any(k in msg.lower() for k in ('login', 'logged', 'sign', 'auth', 'fail', 'password')):
                    continue
                _idata = ev.get('initiator_data') or {}
                _user  = (_idata.get('user_login') or _idata.get('user_email') or
                          ev.get('initiator') or '—')
                _ips   = ev.get('ip_addresses') or []
                _ip    = str(_ips[0]) if _ips else ''
                events.append({
                    'timestamp': str(ev.get('date_gmt') or ev.get('date_local') or ''),
                    'user':      str(_user),
                    'event':     msg[:120],
                    'ip':        _ip,
                    'severity':  'HIGH' if 'fail' in msg.lower() else 'INFO',
                    'status':    'failed' if 'fail' in msg.lower() else 'success',
                    'source':    'simple_history',
                })
            if events:
                source = 'simple_history'

    # 3. WP REST API authenticated — verify auth, fallback to cookie auth for regular passwords
    if not events and auth_header:
        me, _ = _wp_get('/wp-json/wp/v2/users/me?context=edit')
        if me and me.get('id'):
            source = 'wp_rest'
            events.append({
                'timestamp': '', 'ip': '',
                'user':     me.get('slug') or me.get('name') or wp_user,
                'event':    f"Authenticated session active — Role: {', '.join(me.get('roles', []))}",
                'severity': 'INFO', 'status': 'success', 'source': 'wp_rest',
            })
            note = ('No login event plugin detected. Install "Simple History" or "WP Activity Log" '
                    'on your WordPress site to see real-time login events with IP addresses.')
        else:
            # Basic Auth (Application Password) failed — try cookie auth with regular admin password
            ck_nonce, ck_hdr = _wp_cookie_auth(url, wp_user, app_pass)
            if ck_nonce == '__xmlrpc_verified__':
                # Credentials confirmed correct via XML-RPC — WordPress needs an Application Password for REST API
                note = (f'Password verified. WordPress requires an Application Password for REST API access. '
                        f'Create one in 30 seconds: {url}/wp-admin/profile.php '
                        f'(scroll to "Application Passwords" → type any name → click Add → copy the password).')
            elif ck_nonce is not None:
                ck_hdrs = {'Cookie': ck_hdr}
                if ck_nonce:
                    ck_hdrs['X-WP-Nonce'] = ck_nonce
                # Retry Simple History with cookie auth
                sh2, _ = _wp_get(f'/wp-json/simple-history/v1/events?per_page={limit}',
                                  override_hdrs=ck_hdrs)
                if sh2 and isinstance(sh2, list):
                    for ev in sh2:
                        msg = str(ev.get('message') or '')
                        if not any(k in msg.lower() for k in ('login', 'logged', 'sign', 'auth', 'fail', 'password')):
                            continue
                        _idata = ev.get('initiator_data') or {}
                        _user  = (_idata.get('user_login') or _idata.get('user_email') or
                                  ev.get('initiator') or '—')
                        _ips   = ev.get('ip_addresses') or []
                        _ip    = str(_ips[0]) if _ips else ''
                        events.append({
                            'timestamp': str(ev.get('date_gmt') or ev.get('date_local') or ''),
                            'user': str(_user), 'event': msg[:120], 'ip': _ip,
                            'severity': 'HIGH' if 'fail' in msg.lower() else 'INFO',
                            'status': 'failed' if 'fail' in msg.lower() else 'success',
                            'source': 'simple_history',
                        })
                    if events:
                        source = 'simple_history'
                if not events:
                    me2, _ = _wp_get('/wp-json/wp/v2/users/me?context=edit', override_hdrs=ck_hdrs)
                    if me2 and me2.get('id'):
                        source = 'wp_rest'
                        events.append({
                            'timestamp': '', 'ip': '',
                            'user':     me2.get('slug') or me2.get('name') or wp_user,
                            'event':    f"Authenticated via admin session — Role: {', '.join(me2.get('roles', []))}",
                            'severity': 'INFO', 'status': 'success', 'source': 'wp_rest',
                        })
                        note = ('No login event plugin detected. Install "Simple History" or '
                                '"WP Activity Log" to see real-time login events with IP addresses.')
                    else:
                        note = 'Authentication failed — wrong username or password.'
            else:
                note = 'Authentication failed — wrong username or password.'

    if not auth_header and not events:
        root, _ = _wp_get('/wp-json/')
        if root and root.get('name'):
            note = ('WordPress REST API is reachable. Provide a username + Application Password to '
                    'see live login data. Install "WP Activity Log" plugin for full event tracking.')
        else:
            note = 'Could not reach WordPress REST API. Check the URL.'

    if not events and not note:
        note = ('No login events found. Install "WP Activity Log" (WSAL) or "Simple History" plugin '
                'on your WordPress site to expose real-time login events via REST API.')

    _save_log_events_to_db(events, url or 'wordpress', source or 'WordPress REST API', 'wp_log_sync')
    return jsonify({'events': events[:limit], 'total': len(events), 'source': source, 'note': note})


@app.route('/api/wp/install-plugin', methods=['POST'])
def api_wp_install_plugin():
    """Install and activate a WordPress plugin via WP REST API (WordPress 5.5+).

    Uses POST /wp-json/wp/v2/plugins — real WordPress core REST endpoint.
    Requires admin credentials with manage_plugins capability.
    Transport: _wp_request() — SSL bypass + curl fallback (handles Cloudflare).
    """
    import base64 as _b64, json as _j
    data     = request.get_json(force=True, silent=True) or {}
    url      = (data.get('url') or '').strip().rstrip('/')
    wp_user  = (data.get('wp_user') or '').strip()
    app_pass = (data.get('wp_app_pass') or '').strip()
    slug     = (data.get('slug') or 'wp-security-audit-log').strip()

    if not url or not wp_user or not app_pass:
        return jsonify({'error': 'url, wp_user, and wp_app_pass are required'}), 400
    if not url.startswith('http'):
        url = 'https://' + url

    auth = 'Basic ' + _b64.b64encode(f'{wp_user}:{app_pass}'.encode()).decode()

    def _make_helpers(hdrs_get, hdrs_post):
        def _wp_get(path):
            code, body = _wp_request(f'{url}{path}', headers=hdrs_get, timeout=15)
            try:
                return code, _j.loads(body)
            except Exception:
                return code, {}
        def _wp_post(path, body_dict):
            code, body = _wp_request(
                f'{url}{path}', method='POST', headers=hdrs_post,
                body=_j.dumps(body_dict).encode(), timeout=90,
            )
            try:
                return code, _j.loads(body)
            except Exception:
                return code, {'message': body[:200]}
        return _wp_get, _wp_post

    def _do_install(hdrs_get, hdrs_post):
        _wp_get, _wp_post = _make_helpers(hdrs_get, hdrs_post)
        chk_code, existing = _wp_get(f'/wp-json/wp/v2/plugins/{slug}/{slug}')
        if chk_code == 401:
            return 401, None
        if chk_code == 200 and existing.get('plugin'):
            if existing.get('status') == 'active':
                return 200, {'ok': True, 'status': 'already_active',
                             'message': f'{existing.get("name", slug)} is already installed and active.'}
            act_code, act_resp = _wp_post(f'/wp-json/wp/v2/plugins/{slug}/{slug}', {'status': 'active'})
            if act_code == 401:
                return 401, None
            if act_code in (200, 201):
                return 200, {'ok': True, 'status': 'activated',
                             'message': f'{act_resp.get("name", slug)} activated successfully.'}
            return act_code, act_resp
        code, resp = _wp_post('/wp-json/wp/v2/plugins', {'slug': slug, 'status': 'active'})
        if code in (200, 201) and resp.get('plugin'):
            return 200, {'ok': True, 'status': 'installed',
                         'message': f'{resp.get("name", slug)} installed and activated successfully.',
                         'version': resp.get('version', '')}
        return code, resp

    # Try Application Password (Basic Auth) first
    basic_get  = {'Authorization': auth}
    basic_post = {'Authorization': auth, 'Content-Type': 'application/json'}
    code, resp = _do_install(basic_get, basic_post)

    if code == 401:
        ck_nonce, ck_hdr = _wp_cookie_auth(url, wp_user, app_pass)
        if ck_nonce == '__xmlrpc_verified__':
            return jsonify({
                'ok': False,
                'error': (
                    'Password verified. WordPress requires an Application Password for remote '
                    'plugin installation. Create one at: '
                    f'{url}/wp-admin/profile.php '
                    '(scroll to "Application Passwords" → type any name → click Add → '
                    'paste the generated password into this field instead).'
                ),
                'code': 401,
                'manual_url': f'{url}/wp-admin/plugin-install.php?s={slug}&tab=search&type=term',
            }), 400
        if ck_nonce is not None:
            ck_get  = {'Cookie': ck_hdr}
            ck_post = {'Cookie': ck_hdr, 'Content-Type': 'application/json'}
            if ck_nonce:
                ck_get['X-WP-Nonce']  = ck_nonce
                ck_post['X-WP-Nonce'] = ck_nonce
            code, resp = _do_install(ck_get, ck_post)

    if code == 200 and resp and resp.get('ok'):
        return jsonify(resp)

    # Surface the real WordPress error message
    raw_err = (resp or {}).get('message') or (resp or {}).get('error') or ''
    if code == 403:
        wp_msg = 'Permission denied — account needs administrator role (manage_plugins capability).'
    elif code == 401:
        wp_msg = 'Authentication failed — wrong username or password.'
    elif code == 0:
        wp_msg = (
            'VPS/server IP blocked — Cloudflare or the hosting firewall is dropping '
            'connections from this server. '
            f'Detail: {raw_err or "TCP connection refused/timed out"}. '
            'Fix: install the plugin manually from your WordPress admin dashboard '
            f'({url}/wp-admin/plugin-install.php) or whitelist this VPS IP in Cloudflare.'
        )
    else:
        wp_msg = raw_err or f'Unexpected HTTP {code}'
    return jsonify({'ok': False, 'error': wp_msg, 'code': code,
                    'manual_url': f'{url}/wp-admin/plugin-install.php?s={slug}&tab=search&type=term'}), \
           (400 if code >= 400 else 500)


@app.route('/api/logs/cpanel-live', methods=['POST'])
def api_logs_cpanel_live():
    """Fetch real-time session/login data from cPanel via UAPI."""
    import base64 as _b64, json as _j, ssl as _ssl, urllib.request as _req
    data     = request.get_json(force=True, silent=True) or {}
    host     = (data.get('host') or '').strip().rstrip('/')
    cp_user  = (data.get('cp_user') or '').strip()
    cp_pass  = (data.get('cp_pass') or '').strip()
    cp_token = (data.get('cp_token') or '').strip()
    port     = int(data.get('port') or 2083)
    limit    = min(int(data.get('limit') or 50), 200)

    if not host or not cp_user:
        return jsonify({'error': 'cPanel host and username required'}), 400
    base = host if host.startswith('http') else f'https://{host}:{port}'
    if cp_token:
        auth_header = f'cpanel {cp_user}:{cp_token}'
    elif cp_pass:
        auth_header = 'Basic ' + _b64.b64encode(f'{cp_user}:{cp_pass}'.encode()).decode()
    else:
        return jsonify({'error': 'cPanel password or API token required'}), 400

    ctx = _ssl.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = _ssl.CERT_NONE

    def _cp_get(path, timeout=12):
        req = _req.Request(f'{base}{path}', headers={'Authorization': auth_header, 'Accept': 'application/json'})
        try:
            with _req.urlopen(req, timeout=timeout, context=ctx) as r:
                return _j.loads(r.read().decode()), None
        except Exception as e:
            return None, str(e)

    events, source, note = [], 'none', ''

    # 1. Active session list via UAPI
    sess, err = _cp_get('/execute/Session/list')
    if sess and isinstance(sess, dict) and sess.get('data') is not None:
        source = 'cpanel_session'
        for s in (sess.get('data') or [])[:limit]:
            events.append({
                'timestamp': str(s.get('session_create') or s.get('last_update') or ''),
                'user':      str(s.get('session_login') or s.get('user') or cp_user),
                'event':     f"Active session — Type: {s.get('session_type','cPanel')} — Browser: {str(s.get('user_agent',''))[:40]}",
                'ip':        str(s.get('remote_addr') or s.get('ip') or ''),
                'severity':  'INFO', 'status': 'active', 'source': 'cpanel_session',
            })

    # 2. Last login IP
    ll, _ = _cp_get('/execute/LastLogin/get_last_or_current_logged_in_ip')
    if ll and isinstance(ll, dict) and ll.get('data'):
        d2 = ll['data']
        ip = str(d2.get('ip') or d2.get('last_login_ip') or '')
        if ip:
            events.append({
                'timestamp': str(d2.get('unix_last_login') or ''),
                'user': cp_user, 'ip': ip,
                'event': 'Last recorded login IP',
                'severity': 'INFO', 'status': 'success', 'source': 'cpanel_lastlogin',
            })
            source = source or 'cpanel_lastlogin'

    # 3. Security policy — brute force blocked IPs
    bf, _ = _cp_get('/execute/Security/get_password_strength_config')
    if bf and isinstance(bf, dict) and bf.get('data'):
        note = f"Password policy: min strength {bf['data'].get('min_strength', '?')}"

    if not events:
        if err:
            note = f'Could not connect to cPanel UAPI: {err}. Verify host ({base}), port, and credentials.'
        else:
            note = 'Connected but no active sessions found. Try the API Token method for better access.'

    _save_log_events_to_db(events, host or 'cpanel', source or 'cPanel', 'cpanel_log_sync')
    return jsonify({'events': events[:limit], 'total': len(events), 'source': source, 'note': note})


@app.route('/api/logs/wp-cpanel-db', methods=['POST'])
def api_logs_wp_cpanel_db():
    """Read WordPress Simple History events directly from the MySQL database via cPanel Fileman.

    Real flow (no WordPress auth required):
    1. Read wp-config.php via cPanel UAPI Fileman to get DB credentials
    2. Upload a temp PHP script (random name) that queries Simple History tables
    3. Fetch the script via ScraperAPI (bypasses VPS/Cloudflare blocks)
    4. Delete the temp script immediately via cPanel UAPI
    """
    import base64 as _b64, json as _j, re as _re, random as _rand, string as _str
    import ssl as _ssl2, urllib.request as _req2

    data     = request.get_json(force=True, silent=True) or {}
    site_url = (data.get('url') or '').strip().rstrip('/')
    cp_host  = (data.get('cp_host') or '').strip().rstrip('/')
    cp_user  = (data.get('cp_user') or '').strip()
    cp_pass  = (data.get('cp_pass') or '').strip()
    cp_token = (data.get('cp_token') or '').strip()
    wp_dir   = (data.get('wp_dir') or 'public_html').strip().strip('/')
    limit    = min(int(data.get('limit') or 50), 200)

    if not cp_host or not cp_user or not (cp_pass or cp_token):
        return jsonify({'error': 'cPanel host, username, and password or token required'}), 400
    if not site_url:
        return jsonify({'error': 'WordPress site URL required'}), 400
    if not site_url.startswith('http'):
        site_url = 'https://' + site_url

    base = cp_host if cp_host.startswith('http') else f'https://{cp_host}:2083'
    auth_hdr = (f'cpanel {cp_user}:{cp_token}' if cp_token
                else 'Basic ' + _b64.b64encode(f'{cp_user}:{cp_pass}'.encode()).decode())

    ctx = _ssl2.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = _ssl2.CERT_NONE

    def _cp(path, method='GET', post_data=None):
        url = f'{base}/execute/{path}'
        hdr = {'Authorization': auth_hdr, 'Accept': 'application/json', 'User-Agent': _BROWSER_UA}
        if post_data:
            body = _up_parse.urlencode(post_data).encode()
            req = _req2.Request(url, data=body, headers={**hdr, 'Content-Type': 'application/x-www-form-urlencoded'})
        else:
            req = _req2.Request(url, headers=hdr)
        try:
            with _req2.urlopen(req, timeout=15, context=ctx) as r:
                return _j.loads(r.read().decode())
        except Exception as e:
            return {'_err': str(e)}

    # ── Step 1: Read wp-config.php ────────────────────────────────────────────
    cfg = _cp(f'Fileman/get_file_content?dir=%2F{_up_parse.quote(wp_dir)}&file=wp-config.php')
    if cfg.get('_err') or not cfg.get('status'):
        return jsonify({'error': f'Cannot read wp-config.php: {cfg.get("_err") or cfg.get("errors") or "check cPanel credentials and WordPress directory"}'}), 500

    wp_config = (cfg.get('data') or {}).get('content', '')
    if not wp_config:
        return jsonify({'error': 'wp-config.php is empty or unreadable'}), 500

    def _cfg(key):
        m = _re.search(rf"define\s*\(\s*['\"]DB_{key}['\"]\s*,\s*['\"]([^'\"]*)['\"]", wp_config)
        return m.group(1) if m else ''

    db_host   = _cfg('HOST') or 'localhost'
    db_name   = _cfg('NAME')
    db_user   = _cfg('USER')
    db_pass   = _cfg('PASSWORD')
    pfx_m     = _re.search(r"\$table_prefix\s*=\s*['\"]([^'\"]+)['\"]", wp_config)
    table_pfx = pfx_m.group(1) if pfx_m else 'wp_'

    if not db_name or not db_user:
        return jsonify({'error': 'Could not parse DB credentials from wp-config.php'}), 500

    # ── Step 2: Build + upload temp PHP script ────────────────────────────────
    script_name = 'cfai_' + ''.join(_rand.choices(_str.ascii_lowercase + _str.digits, k=14)) + '.php'
    # Real Simple History DB schema:
    # {prefix}simple_history: id, date, date_gmt, logger, level, message, initiator
    # {prefix}simple_history_contexts: id, history_id, key_name, value
    # Context keys: _user_login, _user_email, _server_remote_addr
    php = (
        "<?php error_reporting(0);\n"
        "try {\n"
        f"  $pdo = new PDO('mysql:host={db_host};dbname={db_name};charset=utf8','{db_user}','{db_pass}');\n"
        "  $pdo->setAttribute(PDO::ATTR_ERRMODE, PDO::ERRMODE_EXCEPTION);\n"
        f"  $p = '{table_pfx}';\n"
        "  $kc = 'key_name';\n"
        "  try {\n"
        "    $desc = $pdo->query(\"DESCRIBE {$p}simple_history_contexts\")->fetchAll(PDO::FETCH_COLUMN);\n"
        "    if (!in_array('key_name', $desc) && in_array('key', $desc)) { $kc = '`key`'; }\n"
        "  } catch(Exception $_e) {}\n"
        f"  $rows = $pdo->query(\"SELECT id, date, logger, level, message FROM {table_pfx}simple_history ORDER BY id DESC LIMIT {limit}\")->fetchAll(PDO::FETCH_ASSOC);\n"
        "  if ($rows) {\n"
        "    $ids = array_column($rows, 'id');\n"
        "    $ph  = implode(',', array_fill(0, count($ids), '?'));\n"
        "    $st  = $pdo->prepare(\"SELECT history_id, {$kc} AS k, value FROM {$p}simple_history_contexts WHERE history_id IN ($ph)\");\n"
        "    $st->execute($ids);\n"
        "    $ctxMap = [];\n"
        "    foreach ($st->fetchAll(PDO::FETCH_ASSOC) as $c) { $ctxMap[$c['history_id']][$c['k']] = $c['value']; }\n"
        "    foreach ($rows as &$row) {\n"
        "      $ctx = isset($ctxMap[$row['id']]) ? $ctxMap[$row['id']] : [];\n"
        "      $msg = $row['message'];\n"
        "      foreach ($ctx as $k => $v) { if ($k[0] !== '_') { $msg = str_replace('{' . $k . '}', $v, $msg); } }\n"
        "      $row['message']    = $msg;\n"
        "      $row['user_login'] = isset($ctx['_user_login']) ? $ctx['_user_login'] : '';\n"
        "      $row['ip']         = isset($ctx['_server_remote_addr']) ? $ctx['_server_remote_addr'] : '';\n"
        "    }\n"
        "  }\n"
        "  header('Content-Type: application/json');\n"
        "  echo json_encode(['ok'=>true,'source'=>'wp_db','events'=>$rows]);\n"
        "} catch(Exception $e) {\n"
        "  header('Content-Type: application/json');\n"
        "  echo json_encode(['ok'=>false,'error'=>$e->getMessage()]);\n"
        "}\n"
    )

    up = _cp('Fileman/save_file_content', method='POST', post_data={
        'dir': f'/{wp_dir}', 'file': script_name, 'content': php,
    })
    if up.get('_err') or not up.get('status'):
        return jsonify({'error': f'Cannot upload temp script via cPanel Fileman: {up.get("_err") or up.get("errors")}'}), 500

    # ── Step 3: Fetch script output via ScraperAPI ────────────────────────────
    script_url = f'{site_url}/{script_name}'
    events, source, note = [], 'none', ''
    try:
        sc_code, sc_body = _wp_request(script_url, timeout=25)
        if sc_code == 200:
            result = _j.loads(sc_body)
            if result.get('ok') and isinstance(result.get('events'), list):
                for row in result['events']:
                    msg = str(row.get('message') or row.get('logger') or '')
                    if not msg:
                        continue
                    events.append({
                        'timestamp': str(row.get('date') or ''),
                        'user':      str(row.get('user_login') or '—'),
                        'event':     msg[:120],
                        'ip':        str(row.get('ip') or ''),
                        'severity':  ('HIGH' if any(k in msg.lower()
                                       for k in ('fail', 'block', 'attack', 'brute', 'invalid'))
                                      else 'INFO'),
                        'status':    ('failed' if any(k in msg.lower()
                                       for k in ('fail', 'block', 'denied', 'invalid'))
                                      else 'success'),
                        'source':    'wp_db',
                    })
                source = 'wp_db'
            elif result.get('error'):
                note = f'Database error: {result["error"]}'
        else:
            note = f'Script returned HTTP {sc_code} — may be blocked or wp_dir is wrong'
    except Exception as e:
        note = f'Fetch error: {e}'
    finally:
        # ── Step 4: Delete temp script immediately ────────────────────────────
        _cp('Fileman/unlink', method='POST', post_data={
            'files': _j.dumps([{'dir': f'/{wp_dir}', 'file': script_name}])
        })

    if not events and not note:
        note = 'No Simple History events found. Is the Simple History plugin installed and active?'

    _save_log_events_to_db(events, data.get('url', '') or 'wordpress', source or 'cPanel DB', 'cpanel_db_log_sync')
    return jsonify({'events': events[:limit], 'total': len(events), 'source': source, 'note': note})


@app.route('/api/logs/wp-mysql-direct', methods=['POST'])
def api_logs_wp_mysql_direct():
    """Read WordPress Simple History events via direct MySQL connection.

    Works with Hostinger (hPanel) and any host that supports Remote MySQL.
    Requires: DB host, name, user, password from hPanel → Databases → MySQL Databases.
    The VPS IP (114.5.244.37) must be whitelisted in hPanel → Databases → Remote MySQL.
    """
    import json as _j
    data       = request.get_json(force=True, silent=True) or {}
    db_host    = (data.get('db_host') or '').strip()
    db_port    = int(data.get('db_port') or 3306)
    db_name    = (data.get('db_name') or '').strip()
    db_user    = (data.get('db_user') or '').strip()
    db_pass    = (data.get('db_pass') or '').strip()
    table_pfx  = (data.get('table_prefix') or 'wp_').strip()
    limit      = min(int(data.get('limit') or 50), 200)

    if not db_host or not db_name or not db_user:
        return jsonify({'error': 'db_host, db_name, and db_user are required'}), 400

    try:
        import pymysql
        import pymysql.cursors
    except ImportError:
        return jsonify({'error': 'pymysql not installed — run: pip install pymysql'}), 500

    try:
        conn = pymysql.connect(
            host=db_host, port=db_port,
            user=db_user, password=db_pass,
            database=db_name, charset='utf8mb4',
            connect_timeout=15,
            cursorclass=pymysql.cursors.DictCursor,
        )
    except Exception as e:
        err = str(e)
        if 'Access denied' in err:
            return jsonify({'error': 'Access denied — wrong DB username or password. Check hPanel → Databases → MySQL Databases for the correct credentials.'}), 401
        if db_host in ('localhost', '127.0.0.1', '::1'):
            return jsonify({'error': (
                'DB Host is set to "localhost" — this only works if MySQL is running on the same machine as the dashboard. '
                'For Hostinger, go to hPanel → Databases → MySQL Databases and copy the "Database Server" hostname '
                '(looks like auth-db1234.hstgr.io or similar). Paste that as the DB Host.'
            )}), 500
        if 'Can\'t connect' in err or 'Connection refused' in err or 'timed out' in err.lower() or 'refused' in err.lower():
            # Detect server's outbound IP for the whitelist hint
            try:
                import urllib.request as _ur2
                _my_ip = _ur2.urlopen('https://api.ipify.org', timeout=4).read().decode().strip()
            except Exception:
                _my_ip = 'your VPS IP'
            return jsonify({'error': (
                f'Cannot connect to {db_host}:{db_port}. '
                f'In Hostinger hPanel → Databases → Remote MySQL, add this server\'s IP: {_my_ip}. '
                f'Then retry. If it still fails, check that the DB Host is correct (not "localhost").'
            )}), 500
        return jsonify({'error': f'MySQL connection failed: {err}'}), 500

    events, source, note = [], 'none', ''
    try:
        with conn.cursor() as cur:
            # ── 1. Simple History (preferred) ──────────────────────────────────
            sh_table  = f'{table_pfx}simple_history'
            ctx_table = f'{table_pfx}simple_history_contexts'
            cur.execute("SHOW TABLES LIKE %s", (sh_table,))
            if cur.fetchone():
                # Auto-detect key column name: older Simple History uses 'key_name', newer uses 'key'
                key_col = 'key_name'
                try:
                    cur.execute(f"DESCRIBE {ctx_table}")
                    ctx_cols = [r['Field'] for r in cur.fetchall()]
                    if 'key_name' not in ctx_cols and 'key' in ctx_cols:
                        key_col = '`key`'
                except Exception:
                    pass
                cur.execute(f"""
                    SELECT id, date, logger, level, message
                    FROM {sh_table}
                    ORDER BY id DESC
                    LIMIT %s
                """, (limit,))
                rows = cur.fetchall()
                if rows:
                    ids = [r['id'] for r in rows]
                    fmt = ','.join(['%s'] * len(ids))
                    cur.execute(
                        f"SELECT history_id, {key_col} AS k, value FROM {ctx_table} WHERE history_id IN ({fmt})",
                        ids
                    )
                    ctx_map: dict = {}
                    for c in cur.fetchall():
                        ctx_map.setdefault(c['history_id'], {})[c['k']] = c['value']
                    for row in rows:
                        ctx  = ctx_map.get(row['id'], {})
                        msg  = str(row.get('message') or row.get('logger') or '')
                        for k, v in ctx.items():
                            if k and k[0] != '_':
                                msg = msg.replace('{' + k + '}', str(v))
                        if not msg: continue
                        events.append({
                            'timestamp': str(row.get('date') or ''),
                            'user':      str(ctx.get('_user_login') or '—'),
                            'event':     msg[:200],
                            'ip':        str(ctx.get('_server_remote_addr') or ''),
                            'severity':  'HIGH' if any(k in msg.lower() for k in ('fail','block','attack','brute','invalid')) else 'INFO',
                            'status':    'failed' if any(k in msg.lower() for k in ('fail','block','denied','invalid')) else 'success',
                            'source':    'Simple History',
                        })
                source = 'Simple History'

            # ── 2. Wordfence wp_wfLogins (fallback / supplement) ──────────────
            wf_logins = f'{table_pfx}wfLogins'
            cur.execute("SHOW TABLES LIKE %s", (wf_logins,))
            if cur.fetchone():
                cur.execute(f"""
                    SELECT username, IP, ctime, status, hitCount
                    FROM {wf_logins}
                    ORDER BY ctime DESC
                    LIMIT %s
                """, (limit,))
                for row in cur.fetchall():
                    import datetime as _dt
                    ts = row.get('ctime') or 0
                    try:
                        ts_str = _dt.datetime.utcfromtimestamp(int(ts)).strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        ts_str = str(ts)
                    ip_bytes = row.get('IP') or b''
                    try:
                        import socket as _sock
                        ip_str = _sock.inet_ntoa(ip_bytes[:4]) if isinstance(ip_bytes, (bytes, bytearray)) and len(ip_bytes) >= 4 else str(ip_bytes)
                    except Exception:
                        ip_str = str(ip_bytes)
                    status = str(row.get('status') or '')
                    failed = status in ('0', 'blocked') or not status
                    events.append({
                        'timestamp': ts_str,
                        'user':      str(row.get('username') or '—'),
                        'event':     'Login blocked' if failed else 'Login allowed',
                        'ip':        ip_str,
                        'severity':  'HIGH' if failed else 'INFO',
                        'status':    'failed' if failed else 'success',
                        'source':    'Wordfence',
                    })
                if not source or source == 'none':
                    source = 'Wordfence'

            # ── 3. Wordfence wp_wfBlockedIPLog (blocked IPs) ──────────────────
            wf_blocked = f'{table_pfx}wfBlockedIPLog'
            cur.execute("SHOW TABLES LIKE %s", (wf_blocked,))
            if cur.fetchone():
                cur.execute(f"""
                    SELECT IP, reason, ctime, unixday
                    FROM {wf_blocked}
                    ORDER BY unixday DESC, ctime DESC
                    LIMIT %s
                """, (min(limit, 50),))
                for row in cur.fetchall():
                    import datetime as _dt
                    ts = row.get('ctime') or 0
                    try:
                        ts_str = _dt.datetime.utcfromtimestamp(int(ts)).strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        ts_str = str(ts)
                    ip_bytes = row.get('IP') or b''
                    try:
                        import socket as _sock
                        ip_str = _sock.inet_ntoa(ip_bytes[:4]) if isinstance(ip_bytes, (bytes, bytearray)) and len(ip_bytes) >= 4 else str(ip_bytes)
                    except Exception:
                        ip_str = str(ip_bytes)
                    events.append({
                        'timestamp': ts_str,
                        'user':      '—',
                        'event':     f'IP Blocked: {(row.get("reason") or "")[:80]}',
                        'ip':        ip_str,
                        'severity':  'HIGH',
                        'status':    'blocked',
                        'source':    'Wordfence Blocked',
                    })

            # ── 4. Wordfence wp_wfIssues (malware/scan issues) ────────────────
            wf_issues = f'{table_pfx}wfIssues'
            cur.execute("SHOW TABLES LIKE %s", (wf_issues,))
            if cur.fetchone():
                cur.execute(f"""
                    SELECT severity, description, ctime
                    FROM {wf_issues}
                    ORDER BY ctime DESC
                    LIMIT %s
                """, (min(limit, 30),))
                for row in cur.fetchall():
                    import datetime as _dt
                    ts = row.get('ctime') or 0
                    try:
                        ts_str = _dt.datetime.utcfromtimestamp(int(ts)).strftime('%Y-%m-%d %H:%M:%S')
                    except Exception:
                        ts_str = str(ts)
                    sev_val = int(row.get('severity') or 0)
                    sev = 'CRITICAL' if sev_val >= 100 else ('HIGH' if sev_val >= 50 else 'MEDIUM')
                    events.append({
                        'timestamp': ts_str,
                        'user':      '—',
                        'event':     f'Security Issue: {(row.get("description") or "")[:100]}',
                        'ip':        '—',
                        'severity':  sev,
                        'status':    'issue',
                        'source':    'Wordfence Scan',
                    })

            if not events:
                note = 'No login data found. Install Simple History or Wordfence on WordPress to capture real login events.'

    except Exception as e:
        note = f'Query error: {e}'
    finally:
        conn.close()

    # Sort all events newest-first
    events.sort(key=lambda x: x.get('timestamp', ''), reverse=True)
    _save_log_events_to_db(events, db_name or db_host, source or 'MySQL Direct', 'mysql_log_sync')
    return jsonify({'events': events[:limit], 'total': len(events), 'source': source, 'note': note})


def _cf_request(path: str, cf_token: str, timeout: int = 20):
    """Direct GET to Cloudflare API — uses requests library, no proxy, no Content-Type on GET."""
    url  = f'https://api.cloudflare.com/client/v4{path}'
    hdrs = {'Authorization': f'Bearer {cf_token}'}
    try:
        if _HAS_REQUESTS:
            r = _requests.get(url, headers=hdrs, timeout=timeout, verify=True)
            return r.status_code, r.text
        # stdlib fallback
        import urllib.request as _ur, ssl as _ssl
        req = _ur.Request(url)
        req.add_header('Authorization', f'Bearer {cf_token}')
        ctx = _ssl.create_default_context()
        with _ur.urlopen(req, context=ctx, timeout=timeout) as r:
            return r.status, r.read().decode('utf-8', errors='replace')
    except Exception as e:
        if _HAS_REQUESTS:
            import requests as _rq
            if isinstance(e, _rq.HTTPError):
                return e.response.status_code, e.response.text
        return 0, str(e)


@app.route('/api/logs/ftp', methods=['POST'])
def api_logs_ftp():
    """Browse files, scan for malware, or read access logs via FTP."""
    import ftplib, io, re as _re
    data   = request.get_json(force=True, silent=True) or {}
    host   = (data.get('host') or '').strip()
    port   = int(data.get('port') or 21)
    user   = (data.get('user') or '').strip()
    passwd = (data.get('pass') or '').strip()
    path   = (data.get('path') or 'public_html').strip().lstrip('/')
    action = (data.get('action') or 'browse').strip()

    if not host or not user or not passwd:
        return jsonify({'error': 'FTP host, username and password are required'}), 400

    _MALWARE_PAT = [
        (re.compile(rb'eval\s*\(\s*base64_decode', re.I), 'eval(base64_decode)'),
        (re.compile(rb'eval\s*\(\s*gzinflate',    re.I), 'eval(gzinflate)'),
        (re.compile(rb'eval\s*\(\s*str_rot13',     re.I), 'eval(str_rot13)'),
        (re.compile(rb'eval\s*\(\s*gzuncompress',  re.I), 'eval(gzuncompress)'),
        (re.compile(rb'preg_replace\s*\(\s*[\'"]/.+/e', re.I), 'preg_replace /e'),
        (re.compile(rb'assert\s*\(\s*\$_(POST|GET|REQUEST|COOKIE)', re.I), 'assert($_REQUEST)'),
        (re.compile(rb'\$[a-z_]+\s*=\s*str_rot13', re.I), 'str_rot13 assign'),
    ]
    _SAFE_PAT = [re.compile(p, re.I) for p in [
        r'wpforms/cache', r'unlimited_elements_cache', r'hostinger-', r'cfai-scanner'
    ]]

    try:
        ftp = ftplib.FTP()
        ftp.connect(host, port, timeout=20)
        ftp.login(user, passwd)
        ftp.set_pasv(True)

        if action == 'browse':
            ftp.cwd(path)
            entries = []
            try:
                mlsd = list(ftp.mlsd())
                for name, facts in mlsd:
                    if name in ('.', '..'):
                        continue
                    ftype = facts.get('type', 'file')
                    size  = int(facts.get('size', 0)) if facts.get('size') else None
                    mtime = facts.get('modify', '')
                    if mtime and len(mtime) >= 14:
                        mtime = f'{mtime[:4]}-{mtime[4:6]}-{mtime[6:8]} {mtime[8:10]}:{mtime[10:12]}'
                    entries.append({'name': name, 'type': 'dir' if ftype == 'dir' else 'file',
                                    'size': size, 'modified': mtime, 'path': f'{path}/{name}'})
            except Exception:
                lines = []
                ftp.retrlines('LIST', lines.append)
                for line in lines:
                    parts = line.split(None, 8)
                    if len(parts) < 9:
                        continue
                    name  = parts[8]
                    if name in ('.', '..'):
                        continue
                    isdir = line.startswith('d')
                    size  = int(parts[4]) if parts[4].isdigit() else None
                    mtime = ' '.join(parts[5:8])
                    entries.append({'name': name, 'type': 'dir' if isdir else 'file',
                                    'size': size, 'modified': mtime, 'path': f'{path}/{name}'})
            entries.sort(key=lambda e: (e['type'] != 'dir', e['name'].lower()))
            ftp.quit()
            return jsonify({'files': entries, 'path': path})

        elif action == 'scan':
            results = []
            def _scan_dir(dpath, depth=0):
                if depth > 4:
                    return
                try:
                    ftp.cwd('/' + dpath if not dpath.startswith('/') else dpath)
                except Exception:
                    return
                lines = []
                try:
                    ftp.retrlines('LIST', lines.append)
                except Exception:
                    return
                for line in lines:
                    parts = line.split(None, 8)
                    if len(parts) < 9:
                        continue
                    name = parts[8]
                    if name in ('.', '..'):
                        continue
                    full = f'{dpath}/{name}'
                    isdir = line.startswith('d')
                    if isdir:
                        _scan_dir(full, depth + 1)
                    elif name.lower().endswith('.php'):
                        if any(p.search(name.encode()) for p in _SAFE_PAT):
                            continue
                        size = int(parts[4]) if parts[4].isdigit() else 0
                        if size > 500_000:
                            continue
                        buf = io.BytesIO()
                        try:
                            ftp.retrbinary(f'RETR {full}', buf.write)
                        except Exception:
                            continue
                        content = buf.getvalue()
                        flag = 'OK'
                        detail = ''
                        for pat, label in _MALWARE_PAT:
                            if pat.search(content):
                                flag = 'MALWARE'
                                detail = label
                                break
                        mtime = ' '.join(parts[5:8])
                        results.append({'name': full, 'type': 'file', 'size': size,
                                        'modified': mtime, 'flag': flag, 'detail': detail,
                                        'ext': 'php', 'path': full})
                        if len(results) >= 200:
                            return
            _scan_dir(path)
            ftp.quit()
            results.sort(key=lambda r: (r['flag'] == 'OK', r['name']))
            return jsonify({'files': results, 'path': path})

        elif action == 'logs':
            log_candidates = [
                f'{path}/logs/access_log', f'{path}/logs/access.log',
                'logs/access_log', 'logs/access.log',
                f'{path}/../logs/access_log', 'access_log', 'access.log',
            ]
            content = ''
            log_file = ''
            for candidate in log_candidates:
                buf = io.BytesIO()
                try:
                    ftp.retrbinary(f'RETR {candidate}', buf.write)
                    content = buf.getvalue().decode('utf-8', errors='replace')
                    log_file = candidate
                    break
                except Exception:
                    continue
            ftp.quit()
            if not content:
                return jsonify({'error': 'No access log found. Common paths checked: ' + ', '.join(log_candidates[:4])})
            lines = content.splitlines()
            preview = '\n'.join(lines[-500:])
            return jsonify({'content': preview, 'log_file': log_file, 'total_lines': len(lines)})

        ftp.quit()
        return jsonify({'error': 'Unknown action'}), 400

    except ftplib.all_errors as e:
        return jsonify({'error': f'FTP error: {e}'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/cloudflare/insights', methods=['POST'])
def api_cloudflare_insights():
    """Fetch Cloudflare Security Insights via real Cloudflare v4 API (direct, no proxy)."""
    import json as _j
    data       = request.get_json(force=True, silent=True) or {}
    def _strip_env_prefix(v):
        # Handle accidental copy-paste of KEY=value from .env file
        return v.split('=', 1)[-1].strip() if '=' in v else v.strip()

    cf_token   = _strip_env_prefix(data.get('cf_token') or os.environ.get('CF_API_TOKEN', ''))
    account_id = _strip_env_prefix(data.get('account_id') or os.environ.get('CF_ACCOUNT_ID', ''))
    dismissed  = data.get('dismissed', False)
    limit      = min(int(data.get('limit') or 100), 500)

    if not cf_token:
        return jsonify({'error': 'Cloudflare API token required — get one at dash.cloudflare.com/profile/api-tokens'}), 400
    if not account_id:
        return jsonify({'error': 'Cloudflare Account ID required — visible in the URL when logged into dash.cloudflare.com'}), 400

    def _cf_json(path):
        c, b = _cf_request(path, cf_token, timeout=15)
        if c == 200:
            try:
                return _j.loads(b)
            except Exception:
                pass
        return None

    # Get zones for this account
    zresp = _cf_json(f'/zones?account.id={account_id}&per_page=50')
    if zresp is None:
        return jsonify({'error': 'Could not list zones — check Zone:Read permission on token'}), 500
    zones = zresp.get('result') or []
    if not zones:
        return jsonify({'error': 'No zones found for this account'}), 404

    now_ts = __import__('datetime').datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ')
    insights = []

    for zone in zones[:10]:
        zid   = zone['id']
        zname = zone['name']

        # ── WAF custom rules — flag any with action=skip ──────────────────
        rs = _cf_json(f'/zones/{zid}/rulesets/phases/http_request_firewall_custom/entrypoint')
        if rs:
            rules = (rs.get('result') or {}).get('rules') or []
            skip_rules  = [r for r in rules if isinstance(r, dict) and r.get('action') == 'skip']
            total_rules = len(rules)
            skip_pct    = round(len(skip_rules) / total_rules * 100) if total_rules else 0
            if skip_rules:
                skip_names = ', '.join(
                    r.get('description') or r.get('ref') or r.get('id', '')[:8]
                    for r in skip_rules[:5]
                )
                insights.append({
                    'id': f'skip-rules-{zid}',
                    'subject': zname,
                    'severity': 'moderate',
                    'description': 'Reduce skip rules for improved protection',
                    'insight_type': 'Configuration suggestion',
                    'timestamp': now_ts,
                    'dismissed': False,
                    'resolution': (
                        f'Detection: Queried /zones/{zid}/rulesets/phases/http_request_firewall_custom/entrypoint via Cloudflare API — '
                        f'{len(skip_rules)} of {total_rules} rule(s) ({skip_pct}%) have action="skip": {skip_names}.\n'
                        f'Recommended actions: Review these rules at dash.cloudflare.com → Security → WAF → Custom rules and remove unnecessary skip actions.'
                    ),
                })

        # ── AI Labyrinth — check via zone rulesets ────────────────────────────
        rulesets_resp = _cf_json(f'/zones/{zid}/rulesets')
        if rulesets_resp and rulesets_resp.get('success'):
            active_rulesets = rulesets_resp.get('result') or []
            total_rulesets  = len(active_rulesets)
            ai_ruleset = next((
                r for r in active_rulesets if isinstance(r, dict) and (
                    'ai' in (r.get('name') or '').lower() or
                    'labyrinth' in (r.get('name') or '').lower()
                )
            ), None)
            if not ai_ruleset:
                insights.append({
                    'id': f'ai-labyrinth-{zid}',
                    'subject': zname,
                    'severity': 'low',
                    'description': 'Disrupt unwanted AI crawlers with AI Labyrinth',
                    'insight_type': 'Configuration suggestion',
                    'timestamp': now_ts,
                    'dismissed': False,
                    'resolution': (
                        f'Detection: Queried /zones/{zid}/rulesets via Cloudflare API — '
                        f'{total_rulesets} ruleset(s) found, none matching AI Labyrinth.\n'
                        f'Active rulesets: {", ".join((r.get("name","?") for r in active_rulesets[:5])) or "none"}.\n'
                        f'Recommended actions: Enable AI Labyrinth at dash.cloudflare.com → Security → Bots.'
                    ),
                })

        # ── SSL/TLS mode ──────────────────────────────────────────────────
        ssl = _cf_json(f'/zones/{zid}/settings/ssl')
        if ssl and ssl.get('success'):
            ssl_val = (ssl.get('result') or {}).get('value', '')
            if ssl_val in ('off', 'flexible'):
                insights.append({
                    'id': f'ssl-mode-{zid}',
                    'subject': zname,
                    'severity': 'high',
                    'description': f'SSL/TLS mode is set to "{ssl_val}" — upgrade to Full or Full (Strict)',
                    'insight_type': 'Configuration suggestion',
                    'timestamp': now_ts,
                    'dismissed': False,
                    'resolution': (
                        f'Detection: Queried /zones/{zid}/settings/ssl via Cloudflare API — current value: "{ssl_val}".\n'
                        f'Recommended actions: Change SSL/TLS mode to "full" or "strict" at dash.cloudflare.com → SSL/TLS → Overview.'
                    ),
                })

        # ── Security level ────────────────────────────────────────────────
        sl = _cf_json(f'/zones/{zid}/settings/security_level')
        if sl and sl.get('success'):
            sl_val = (sl.get('result') or {}).get('value', '')
            if sl_val in ('essentially_off', 'low'):
                insights.append({
                    'id': f'security-level-{zid}',
                    'subject': zname,
                    'severity': 'moderate',
                    'description': f'Security level is set to "{sl_val}" — consider raising it',
                    'insight_type': 'Configuration suggestion',
                    'timestamp': now_ts,
                    'dismissed': False,
                    'resolution': (
                        f'Detection: Queried /zones/{zid}/settings/security_level via Cloudflare API — current value: "{sl_val}".\n'
                        f'Recommended actions: Raise Security Level to "medium" or "high" at dash.cloudflare.com → Security → Settings.'
                    ),
                })

    return jsonify({
        'insights': insights,
        'total':    len(insights),
        'count':    len(insights),
    })


@app.route('/api/cloudflare/zones', methods=['POST'])
def api_cloudflare_zones():
    """List all Cloudflare zones (domains) for an account — direct, no proxy."""
    import json as _j
    def _strip_env_prefix(v):
        return v.split('=', 1)[-1].strip() if '=' in v else v.strip()

    data       = request.get_json(force=True, silent=True) or {}
    cf_token   = _strip_env_prefix(data.get('cf_token') or os.environ.get('CF_API_TOKEN', ''))
    account_id = _strip_env_prefix(data.get('account_id') or os.environ.get('CF_ACCOUNT_ID', ''))

    if not cf_token:
        return jsonify({'error': 'cf_token required'}), 400

    path = f'/zones?per_page=50' + (f'&account.id={account_id}' if account_id else '')
    code, body = _cf_request(path, cf_token, timeout=15)
    if code != 200:
        try:
            msg = (_j.loads(body).get('errors') or [{}])[0].get('message', f'HTTP {code}')
        except Exception:
            msg = f'HTTP {code}'
        return jsonify({'error': f'Cloudflare API error: {msg}'}), 500
    try:
        resp  = _j.loads(body)
        zones = [{'id': z['id'], 'name': z['name'], 'status': z.get('status', '')}
                 for z in (resp.get('result') or [])]
        return jsonify({'zones': zones})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/priority/maintenance', methods=['POST'])
def api_priority_maintenance():
    """Lock / unlock a site via Cloudflare — tries 4 methods in order to ensure the block sticks."""
    import json as _j
    data      = request.get_json(force=True, silent=True) or {}
    _raw_domain = (data.get('domain') or '').strip()
    domain      = re.sub(r'^https?://', '', _raw_domain, flags=re.I).split('/')[0].rstrip('.').lower()
    action    = (data.get('action') or 'enable').lower()
    reason    = (data.get('reason') or 'Security issue under investigation').strip()

    if not domain:
        return jsonify({'error': 'domain is required'}), 400

    def _strip_env_prefix(v):
        return v.split('=', 1)[-1].strip() if '=' in v else v.strip()

    cf_token = _strip_env_prefix(os.environ.get('CF_API_TOKEN', '').strip())
    if not cf_token:
        return jsonify({'ok': False, 'error': 'CF_API_TOKEN not set in .env — add: CF_API_TOKEN=your_token (dash.cloudflare.com/profile/api-tokens)'}), 200

    BASE = 'https://api.cloudflare.com/client/v4'
    hdrs = {'Authorization': f'Bearer {cf_token}', 'Content-Type': 'application/json'}

    def _req(method, path, payload=None):
        url = BASE + path
        try:
            if not _HAS_REQUESTS:
                return 0, {}
            r = _requests.request(method, url, headers=hdrs,
                                  json=payload if payload is not None else None,
                                  timeout=15, verify=True)
            try:    return r.status_code, r.json()
            except: return r.status_code, {}
        except Exception as e:
            return 0, {'error': str(e)}

    def _get(path):   return _req('GET',    path)
    def _post(path, p): return _req('POST',  path, p)
    def _put(path, p):  return _req('PUT',   path, p)
    def _patch(path, p): return _req('PATCH', path, p)
    def _delete(path):  return _req('DELETE', path)

    # --- find zone --- use re.sub to avoid lstrip character-set bug
    bare = re.sub(r'^www\.', '', domain, flags=re.I)
    for lookup in [bare, domain]:
        c, b = _get(f'/zones?name={lookup}&per_page=1')
        zones = (b.get('result') or []) if c == 200 else []
        if zones:
            break
    if not zones:
        return jsonify({'error': (
            f'No Cloudflare zone found for {domain}. '
            'Make sure the domain uses Cloudflare DNS and CF_API_TOKEN has Zone:Read permission.'
        )}), 404

    zone_id   = zones[0]['id']
    zone_name = zones[0]['name']
    expr = f'(http.host eq "{zone_name}" or http.host eq "www.{zone_name}")'

    # ─────────────────────────────────────────────────────────────────────────
    if action == 'enable':
        # Snapshot current security level so we can restore it on unlock
        _, sl_resp = _get(f'/zones/{zone_id}/settings/security_level')
        prev_level = ((sl_resp.get('result') or {}).get('value') or 'medium')

        # ── Method 1: Cloudflare Firewall Rules (legacy, free plan) ──────────
        sc1, fw_resp = _post(f'/zones/{zone_id}/firewall/rules', [{
            'filter':      {'expression': expr, 'paused': False},
            'action':      'block',
            'description': 'CF_AI_SITE_LOCK',
            'priority':    1,
            'paused':      False,
        }])
        if sc1 in (200, 201):
            fw_list    = fw_resp if isinstance(fw_resp, list) else (fw_resp.get('result') or [{}])
            fw_result  = fw_list[0] if fw_list else {}
            fw_rule_id = fw_result.get('id', '')
            fw_filt_id = (fw_result.get('filter') or {}).get('id', '')
            if fw_rule_id:
                db.enable_maintenance(domain, zone_id=zone_id,
                                      cf_rule_id=f'fw:{fw_rule_id}:{fw_filt_id}',
                                      prev_level=prev_level, reason=reason)
                return jsonify({'ok': True, 'method': 'firewall_rule',
                    'message': f'Site LOCKED. {zone_name} returns HTTP 403 to all visitors. Click Unlock to restore access.',
                    'zone': zone_name})

        # ── Method 2: WAF Custom Rules via Rulesets API (all plans) ──────────
        _, rs_data = _get(f'/zones/{zone_id}/rulesets/phases/http_request_firewall_custom/entrypoint')
        ruleset_id  = (rs_data.get('result') or {}).get('id', '')
        existing_rules = (rs_data.get('result') or {}).get('rules') or []

        if not ruleset_id:
            sc_c, rs_c = _post(f'/zones/{zone_id}/rulesets', {
                'name': 'CF_AI_LOCK', 'kind': 'zone',
                'phase': 'http_request_firewall_custom', 'rules': []
            })
            ruleset_id = (rs_c.get('result') or {}).get('id', '')
            existing_rules = []

        if ruleset_id:
            new_rule = {'description': 'CF_AI_SITE_LOCK', 'expression': expr,
                        'action': 'block', 'enabled': True}
            sc2, rs_put = _put(f'/zones/{zone_id}/rulesets/{ruleset_id}',
                               {'rules': existing_rules + [new_rule]})
            if sc2 in (200, 201):
                updated_rules = (rs_put.get('result') or {}).get('rules') or []
                our = next((r for r in reversed(updated_rules)
                            if r.get('description') == 'CF_AI_SITE_LOCK'), {})
                rule_id = our.get('id', '')
                if rule_id:
                    db.enable_maintenance(domain, zone_id=zone_id,
                                          cf_rule_id=f'rs:{ruleset_id}:{rule_id}',
                                          prev_level=prev_level, reason=reason)
                    return jsonify({'ok': True, 'method': 'waf_custom_rule',
                        'message': f'Site LOCKED via WAF rule. {zone_name} blocks all visitors (HTTP 403). Click Unlock to restore access.',
                        'zone': zone_name})

        # ── Method 3: IP Access Rules — block wildcard "all" ─────────────────
        sc3, ip_resp = _post(f'/zones/{zone_id}/firewall/access_rules/rules', {
            'mode': 'block', 'configuration': {'target': 'ip_range', 'value': '0.0.0.0/0'},
            'notes': 'CF_AI_SITE_LOCK — block all IPv4'
        })
        rule_id_v4 = (ip_resp.get('result') or {}).get('id', '')
        sc3b, ip_resp_v6 = _post(f'/zones/{zone_id}/firewall/access_rules/rules', {
            'mode': 'block', 'configuration': {'target': 'ip6_range', 'value': '::/0'},
            'notes': 'CF_AI_SITE_LOCK — block all IPv6'
        })
        rule_id_v6 = (ip_resp_v6.get('result') or {}).get('id', '')
        if rule_id_v4 or rule_id_v6:
            db.enable_maintenance(domain, zone_id=zone_id,
                                  cf_rule_id=f'ip:{rule_id_v4}:{rule_id_v6}',
                                  prev_level=prev_level, reason=reason)
            return jsonify({'ok': True, 'method': 'ip_access_rule',
                'message': f'Site LOCKED via IP block rules. All traffic to {zone_name} is blocked. Click Unlock to restore access.',
                'zone': zone_name})

        # ── Method 4: Under Attack mode (challenge all visitors) ─────────────
        sc4, _ = _patch(f'/zones/{zone_id}/settings/security_level',
                        {'value': 'under_attack'})
        if sc4 in (200, 201):
            db.enable_maintenance(domain, zone_id=zone_id,
                                  cf_rule_id=f'sl:{prev_level}',
                                  prev_level=prev_level, reason=reason)
            return jsonify({'ok': True, 'method': 'under_attack',
                'message': (f'Site set to UNDER ATTACK mode on {zone_name} — '
                            f'all visitors must pass a JS/CAPTCHA challenge. '
                            f'For a hard HTTP 403 block your API token needs Firewall Services:Edit permission.'),
                'zone': zone_name})

        # All four methods failed
        err_detail = ''
        for sc, resp in [(sc1, fw_resp), (sc2 if 'sc2' in dir() else 0, rs_put if 'rs_put' in dir() else {})]:
            errs = resp.get('errors') if isinstance(resp, dict) else []
            if errs:
                err_detail = (errs[0] if isinstance(errs[0], str) else (errs[0] or {}).get('message', ''))
                break
        if sc1 in (401, 403) or sc4 in (401, 403):
            err_detail = ('Permission denied. Token needs at minimum "Zone Settings:Edit". '
                          'For a full block: add "Zone → Firewall Services:Edit" or "Zone → WAF:Edit". '
                          'Create/edit token at dash.cloudflare.com/profile/api-tokens.')
        return jsonify({'ok': False,
                        'error': f'Could not lock {zone_name}: {err_detail or "all Cloudflare methods failed"}'}), 200

    # ─────────────────────────────────────────────────────────────────────────
    elif action == 'disable':
        maint       = db.get_maintenance(domain)
        cf_rule_ref = (maint or {}).get('cf_rule_id', '')
        prev_level  = (maint or {}).get('previous_security_level', 'medium')
        deleted     = False

        if cf_rule_ref.startswith('fw:'):
            parts = cf_rule_ref.split(':', 2)
            fw_rule_id = parts[1] if len(parts) > 1 else ''
            fw_filt_id = parts[2] if len(parts) > 2 else ''
            if fw_rule_id:
                sc_d, _ = _delete(f'/zones/{zone_id}/firewall/rules/{fw_rule_id}')
                deleted = sc_d in (200, 204)
            if fw_filt_id:
                _delete(f'/zones/{zone_id}/filters/{fw_filt_id}')

        elif cf_rule_ref.startswith('rs:'):
            parts = cf_rule_ref.split(':', 2)
            ruleset_id = parts[1] if len(parts) > 1 else ''
            rule_id    = parts[2] if len(parts) > 2 else ''
            if ruleset_id and rule_id:
                # Remove only our rule by rebuilding the ruleset without it
                _, rs_data = _get(f'/zones/{zone_id}/rulesets/{ruleset_id}')
                current_rules = (rs_data.get('result') or {}).get('rules') or []
                kept = [r for r in current_rules
                        if r.get('id') != rule_id and r.get('description') != 'CF_AI_SITE_LOCK']
                sc_d, _ = _put(f'/zones/{zone_id}/rulesets/{ruleset_id}', {'rules': kept})
                deleted = sc_d in (200, 201)

        elif cf_rule_ref.startswith('ip:'):
            parts      = cf_rule_ref.split(':', 2)
            rule_id_v4 = parts[1] if len(parts) > 1 else ''
            rule_id_v6 = parts[2] if len(parts) > 2 else ''
            ok4 = ok6 = False
            if rule_id_v4:
                sc_d4, _ = _delete(f'/zones/{zone_id}/firewall/access_rules/rules/{rule_id_v4}')
                ok4 = sc_d4 in (200, 204)
            if rule_id_v6:
                sc_d6, _ = _delete(f'/zones/{zone_id}/firewall/access_rules/rules/{rule_id_v6}')
                ok6 = sc_d6 in (200, 204)
            deleted = ok4 or ok6

        elif cf_rule_ref.startswith('sl:'):
            # Was set via security_level — restore it
            _patch(f'/zones/{zone_id}/settings/security_level', {'value': prev_level or 'medium'})
            deleted = True

        elif cf_rule_ref and ':' in cf_rule_ref:
            # Legacy format: ruleset_id:rule_id
            ruleset_id, rule_id = cf_rule_ref.split(':', 1)
            sc_d, _ = _delete(f'/zones/{zone_id}/rulesets/{ruleset_id}/rules/{rule_id}')
            deleted = sc_d in (200, 204)

        # Always restore security level on unlock
        if prev_level and not cf_rule_ref.startswith('sl:'):
            _patch(f'/zones/{zone_id}/settings/security_level', {'value': prev_level})

        db.disable_maintenance(domain)
        return jsonify({'ok': True,
            'message': (f'Site UNLOCKED. {zone_name} is now accessible to all visitors again.'
                        if deleted else
                        f'{zone_name} unlocked in our records — please verify the block rule was removed in your Cloudflare dashboard.'),
            'zone': zone_name})

    return jsonify({'error': 'Invalid action. Use "enable" or "disable"'}), 400


@app.route('/api/priority/maintenance/status')
def api_priority_maintenance_status():
    """Return all domains currently in Cloudflare maintenance mode."""
    sites = db.get_all_maintenance()
    return jsonify({
        'maintenance_domains': [s['domain'] for s in sites],
        'details': sites,
    })


@app.route('/api/cloudflare/attack-mode', methods=['GET', 'POST'])
def api_cf_attack_mode():
    """GET: return current Under Attack Mode status for a domain.
       POST {domain, enabled}: toggle security_level between under_attack and medium.
    """
    def _strip_env_prefix(v):
        return v.split('=', 1)[-1].strip() if '=' in v else v.strip()
    cf_token = _strip_env_prefix(os.environ.get('CF_API_TOKEN', '').strip())
    if not cf_token:
        return jsonify({'error': 'CF_API_TOKEN not set'}), 400

    BASE = 'https://api.cloudflare.com/client/v4'
    hdrs = {'Authorization': f'Bearer {cf_token}', 'Content-Type': 'application/json'}

    def _req(method, path, payload=None):
        url = BASE + path
        try:
            if not _HAS_REQUESTS:
                return 0, {}
            r = _requests.request(method, url, headers=hdrs,
                                  json=payload if payload is not None else None,
                                  timeout=12, verify=True)
            try:    return r.status_code, r.json()
            except: return r.status_code, {}
        except Exception as e:
            return 0, {'error': str(e)}

    if request.method == 'GET':
        raw = (request.args.get('domain') or '').strip()
        domain = re.sub(r'^https?://', '', raw, flags=re.I).split('/')[0].rstrip('.').lower()
        if not domain:
            return jsonify({'error': 'domain required'}), 400

        bare = re.sub(r'^www\.', '', domain, flags=re.I)
        zones = []
        last_status = 0
        last_body   = {}
        for lk in [bare, domain]:
            last_status, last_body = _req('GET', f'/zones?name={lk}&per_page=1')
            zones = (last_body.get('result') or []) if last_status == 200 else []
            if zones: break
        if not zones:
            if last_status in (401, 403):
                cf_errs = last_body.get('errors') or []
                cf_msg  = cf_errs[0].get('message', '') if cf_errs and isinstance(cf_errs[0], dict) else ''
                return jsonify({'error': (
                    f'API token permission denied (HTTP {last_status}). '
                    f'{cf_msg} — Go to dash.cloudflare.com/profile/api-tokens, '
                    f'edit the token and add Zone → Zone:Read permission.'
                )}), 403
            return jsonify({'error': (
                f'No Cloudflare zone found for {domain} (HTTP {last_status}). '
                f'Make sure the domain is added to this Cloudflare account and '
                f'CF_API_TOKEN has Zone:Read permission.'
            )}), 404

        zone_id   = zones[0]['id']
        zone_name = zones[0]['name']
        zone_status = zones[0].get('status', '')  # 'active' | 'pending' | 'initializing'

        c2, sl = _req('GET', f'/zones/{zone_id}/settings/security_level')
        level = (sl.get('result') or {}).get('value', 'unknown')

        # Check Development Mode (bypasses Under Attack Mode completely)
        _, dev = _req('GET', f'/zones/{zone_id}/settings/development_mode')
        dev_mode = (dev.get('result') or {}).get('value', 'off') == 'on'

        # Check if any A/CNAME record is proxied — fetch all records (no name filter)
        # so we don't miss apex/www records due to filter mismatch
        c3, dns = _req('GET', f'/zones/{zone_id}/dns_records?per_page=100')
        dns_records = (dns.get('result') or []) if c3 == 200 else []
        ac_records = [r for r in dns_records if r.get('type') in ('A', 'AAAA', 'CNAME')]
        proxied = any(r.get('proxied') for r in ac_records)
        # dns_only = True when we found records but NONE are proxied through CF
        dns_only = bool(ac_records) and not proxied

        return jsonify({
            'domain': domain,
            'zone_id': zone_id,
            'zone_name': zone_name,
            'zone_status': zone_status,
            'security_level': level,
            'under_attack': level == 'under_attack',
            'proxied': proxied,
            'dns_only': dns_only,
            'dev_mode': dev_mode,
            'proxy_warning': (
                'DNS is set to DNS Only (grey cloud) — traffic bypasses Cloudflare so Under Attack Mode '
                'has no effect. Go to Cloudflare → DNS → enable the orange cloud (Proxied) on your A/CNAME records.'
            ) if dns_only else '',
            'dev_mode_warning': (
                'Development Mode is ON — this bypasses Under Attack Mode completely. '
                'Disable it in Cloudflare → Overview → Quick Actions → Development Mode.'
            ) if dev_mode else '',
        })

    # POST — toggle
    data    = request.get_json(force=True, silent=True) or {}
    raw     = (data.get('domain') or '').strip()
    domain  = re.sub(r'^https?://', '', raw, flags=re.I).split('/')[0].rstrip('.').lower()
    enabled = bool(data.get('enabled', True))
    if not domain:
        return jsonify({'error': 'domain required'}), 400

    bare = re.sub(r'^www\.', '', domain, flags=re.I)
    zones = []
    for lk in [bare, domain]:
        c, b = _req('GET', f'/zones?name={lk}&per_page=1')
        zones = (b.get('result') or []) if c == 200 else []
        if zones: break
    if not zones:
        return jsonify({'error': f'No Cloudflare zone for {domain}'}), 404

    zone_id   = zones[0]['id']
    zone_name = zones[0]['name']

    if enabled:
        # Save current level before switching to under_attack
        c_get, sl_now = _req('GET', f'/zones/{zone_id}/settings/security_level')
        prev = (sl_now.get('result') or {}).get('value', 'medium')
        if prev == 'under_attack':
            prev = 'medium'
        sc, resp = _req('PATCH', f'/zones/{zone_id}/settings/security_level',
                        {'value': 'under_attack'})
        # Cloudflare always returns HTTP 200 — must check success field in body
        cf_ok = resp.get('success', False)
        new_level = (resp.get('result') or {}).get('value', '')
        if cf_ok and new_level == 'under_attack':
            db.enable_maintenance(domain, zone_id=zone_id,
                                  cf_rule_id=f'sl:{prev}',
                                  prev_level=prev, reason='Under Attack Mode toggled from dashboard')
            return jsonify({'ok': True, 'enabled': True, 'zone': zone_name,
                'message': f'Under Attack Mode ENABLED on {zone_name}. All visitors now see a JS challenge — open the site in incognito to confirm.'})
        # Extract real error from Cloudflare response
        errs = resp.get('errors') or []
        err  = errs[0].get('message', '') if errs and isinstance(errs[0], dict) else str(errs[0]) if errs else ''
        if not err:
            err = f'Cloudflare did not apply the change (HTTP {sc}, success={resp.get("success")})'
        if sc in (401, 403) or 'permission' in err.lower() or 'not allowed' in err.lower():
            err = ('Token permission denied. Go to dash.cloudflare.com/profile/api-tokens, '
                   'edit your token and add Zone → Zone Settings: Edit permission.')
        return jsonify({'ok': False, 'error': err}), 200
    else:
        # Restore previous level
        maint = db.get_maintenance(domain)
        prev  = (maint or {}).get('previous_security_level', 'medium') or 'medium'
        if prev == 'under_attack':
            prev = 'medium'
        sc, resp = _req('PATCH', f'/zones/{zone_id}/settings/security_level', {'value': prev})
        cf_ok = resp.get('success', False)
        db.disable_maintenance(domain)
        restored = (resp.get('result') or {}).get('value', prev)
        if cf_ok:
            return jsonify({'ok': True, 'enabled': False, 'zone': zone_name,
                'message': f'Under Attack Mode DISABLED on {zone_name}. Security level restored to "{restored}".'})
        # Even if CF fails, clear our DB record and report
        errs = resp.get('errors') or []
        err  = errs[0].get('message', f'HTTP {sc}') if errs and isinstance(errs[0], dict) else f'HTTP {sc}'
        return jsonify({'ok': True, 'enabled': False, 'zone': zone_name,
            'message': f'Disabled in dashboard (CF returned: {err}). Check security level in Cloudflare dashboard.'})


# ══ Security Operations ══════════════════════════════════════════════════════

@app.route('/api/events/ingest', methods=['POST'])
def api_events_ingest():
    """Ingest a security event. Runs GeoIP lookup and auto-remediation rules."""
    data       = request.get_json(force=True, silent=True) or {}
    event_type = (data.get('event_type') or '').strip()
    if not event_type:
        return jsonify({'error': 'event_type is required'}), 400

    ip      = (data.get('ip') or data.get('ip_address') or '').strip()
    geo     = _geoip_detail(ip) if ip else {}

    ev = {
        'event_type':   event_type,
        'category':     (data.get('category') or _infer_category(event_type)),
        'severity':     (data.get('severity') or 'LOW').upper(),
        'ip_address':   ip,
        'country':      geo.get('country', data.get('country', '')),
        'country_code': geo.get('country_code', ''),
        'latitude':     geo.get('lat', 0),
        'longitude':    geo.get('lon', 0),
        'target':       (data.get('target') or '').strip(),
        'user_name':    (data.get('user') or data.get('user_name') or '').strip(),
        'description':  (data.get('description') or data.get('event') or event_type).strip(),
        'raw_data':     _json.dumps(data),
    }

    event_id = db.log_security_event(**ev)
    _run_remediation(event_id, ev)
    return jsonify({'ok': True, 'event_id': event_id, 'geo': geo})


def _infer_category(event_type: str) -> str:
    et = event_type.lower()
    if any(k in et for k in ('login', 'auth', 'password', 'mfa', 'session')): return 'auth'
    if any(k in et for k in ('sql', 'xss', 'injection', 'scan', 'brute', 'attack', 'block')): return 'attack'
    if any(k in et for k in ('vuln', 'cve', 'outdated', 'patch', 'weak')): return 'vulnerability'
    if any(k in et for k in ('fix', 'update', 'remediat', 'patch')): return 'remediation'
    return 'system'


@app.route('/api/events')
def api_events():
    limit    = min(int(request.args.get('limit', 200)), 1000)
    category = request.args.get('category', '')
    severity = request.args.get('severity', '')
    days     = int(request.args.get('days', 7))
    events   = db.get_security_events(limit=limit, category=category, severity=severity, days=days)
    stats    = db.get_event_stats(days=days)
    blocked  = db.get_blocked_ips()
    return jsonify({'events': events, 'stats': stats,
                    'blocked_count': len(blocked), 'total': len(events)})


@app.route('/api/events/map')
def api_events_map():
    days   = int(request.args.get('days', 7))
    events = db.get_events_map(days=days)
    return jsonify({'events': events, 'total': len(events)})


@app.route('/api/events/stats')
def api_events_stats():
    days = int(request.args.get('days', 7))
    return jsonify(db.get_event_stats(days=days))


# ── System Logs (Splunk HEC + Windows Event Log forwarder) ───────────────────
import secrets as _secrets

_SYSLOG_CFG_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'syslog_cfg.json')

def _load_syslog_cfg() -> dict:
    if not os.path.exists(_SYSLOG_CFG_FILE):
        cfg = {'hec_token': _secrets.token_hex(24)}
        _save_syslog_cfg(cfg)
        return cfg
    with open(_SYSLOG_CFG_FILE) as f:
        return _json.load(f)

def _save_syslog_cfg(cfg: dict) -> None:
    os.makedirs(os.path.dirname(_SYSLOG_CFG_FILE), exist_ok=True)
    with open(_SYSLOG_CFG_FILE, 'w') as f:
        _json.dump(cfg, f)

def _normalize_level(raw: str) -> str:
    r = (raw or '').upper()
    if r in ('CRITICAL', 'FATAL', 'EMERGENCY', 'ALERT'): return 'CRITICAL'
    if r in ('ERROR', 'ERR', '3'):                        return 'ERROR'
    if r in ('WARNING', 'WARN', '4', 'WARNING_L'):       return 'WARN'
    if r in ('NOTICE', '5', 'INFORMATIONAL', 'INFO', '6', 'INFORMATION'): return 'INFO'
    if r in ('DEBUG', '7'):                               return 'DEBUG'
    # Windows EventType numbers: 1=Error,2=Warning,3=Info,4=Security Success,5=Security Failure
    if r in ('1',):  return 'ERROR'
    if r in ('2',):  return 'WARN'
    if r in ('4', '5'): return 'INFO'
    return 'INFO'

def _hec_process_event(ev: dict) -> None:
    """Parse one HEC event dict and write to syslog table."""
    if not isinstance(ev, dict):
        return
    inner      = ev.get('event', ev)
    sourcetype = ev.get('sourcetype', '')
    src_type   = 'windows' if sourcetype.startswith('WinEventLog') else \
                 'linux'   if sourcetype.startswith(('syslog', 'rsyslog', 'journald')) else 'splunk'
    channel    = sourcetype.split(':', 1)[1] if ':' in sourcetype else ''
    if isinstance(inner, str):
        message = inner
    else:
        message = (inner.get('message') or inner.get('msg') or
                   inner.get('event') or _json.dumps(inner))
        channel = channel or inner.get('channel') or inner.get('source') or ''
    db.log_syslog(
        source_type = src_type,
        host        = ev.get('host', ''),
        source      = ev.get('source', ''),
        sourcetype  = sourcetype,
        level       = _normalize_level(
            inner.get('severity') or inner.get('level') or
            ev.get('level') or ''),
        event_id    = str(inner.get('event_id') or inner.get('id') or
                         ev.get('event_id') or ''),
        channel     = channel,
        message     = message[:1000],
        raw         = _json.dumps(ev)[:2000],
    )


@app.route('/api/syslog/hec', methods=['POST'])
def api_syslog_hec():
    """Splunk HEC-compatible endpoint.
    Accepts:
      - HEC token via  Authorization: Splunk <token>
      - Any valid session cookie
    Body formats supported:
      - JSON array:           [{...}, ...]
      - Single JSON object:   {...}
      - NDJSON (newline-sep): {...}\\n{...}
    """
    auth  = request.headers.get('Authorization', '')
    token = auth.replace('Splunk ', '').strip()
    cfg   = _load_syslog_cfg()
    authed = (token and token == cfg.get('hec_token', '')) or bool(session.get('user'))
    if not authed:
        return jsonify({'text': 'Invalid token or not authenticated', 'code': 4}), 403

    body = request.get_data(as_text=True).strip()
    if not body:
        return jsonify({'text': 'Success', 'code': 0, 'count': 0})

    events = []
    # Try full-body JSON first (array or single object — what PowerShell ConvertTo-Json produces)
    try:
        parsed = _json.loads(body)
        if isinstance(parsed, list):
            events = parsed
        elif isinstance(parsed, dict):
            events = [parsed]
    except _json.JSONDecodeError:
        # Fall back to NDJSON (real Splunk HEC / filebeat format)
        for line in body.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                obj = _json.loads(line)
                if isinstance(obj, dict):
                    events.append(obj)
            except Exception:
                pass

    count = 0
    for ev in events:
        try:
            _hec_process_event(ev)
            count += 1
        except Exception:
            pass
    return jsonify({'text': 'Success', 'code': 0, 'count': count})


def _syslog_auth_check() -> bool:
    """Return True if caller is authenticated (session OR HTTP Basic Auth)."""
    if session.get('user'):
        return True
    auth = request.headers.get('Authorization', '')
    if auth.startswith('Basic '):
        import base64 as _b64
        try:
            creds    = _b64.b64decode(auth[6:]).decode('utf-8', errors='replace')
            username, password = creds.split(':', 1)
            users    = _load_users()
            u        = users.get(username)
            if u and check_password_hash(u.get('password', ''), password):
                return True
        except Exception:
            pass
    return False


@app.route('/api/syslog/ingest', methods=['POST'])
def api_syslog_ingest():
    """Generic ingest — accepts any account via session OR HTTP Basic Auth."""
    if not _syslog_auth_check():
        return jsonify({'error': 'Authentication required'}), 401
    data = request.get_json(force=True, silent=True) or {}
    entries = data if isinstance(data, list) else [data]
    count = 0
    for entry in entries:
        if not isinstance(entry, dict):
            continue
        db.log_syslog(
            source_type = entry.get('source_type', 'windows'),
            host        = entry.get('host', ''),
            source      = entry.get('source', ''),
            sourcetype  = entry.get('sourcetype', ''),
            level       = _normalize_level(entry.get('level') or entry.get('severity') or ''),
            event_id    = str(entry.get('event_id') or ''),
            channel     = entry.get('channel', ''),
            message     = entry.get('message') or entry.get('msg') or '',
            raw         = _json.dumps(entry),
        )
        count += 1
    return jsonify({'ok': True, 'ingested': count})


@app.route('/api/syslog')
@login_required
def api_syslog_get():
    limit       = min(int(request.args.get('limit', 500)), 2000)
    source_type = request.args.get('source_type', '')
    level       = request.args.get('level', '')
    channel     = request.args.get('channel', '')
    search      = request.args.get('search', '')
    hours       = int(request.args.get('hours', 24))
    logs  = db.get_syslog(limit=limit, source_type=source_type, level=level,
                           channel=channel, search=search, hours=hours)
    stats = db.get_syslog_stats(hours=hours)
    return jsonify({'logs': logs, 'stats': stats})


@app.route('/api/syslog/clear', methods=['POST'])
@_admin_required
def api_syslog_clear():
    n = db.clear_syslog()
    return jsonify({'ok': True, 'deleted': n})


@app.route('/api/syslog/config')
@_admin_required
def api_syslog_config_get():
    cfg = _load_syslog_cfg()
    return jsonify({'hec_token': cfg.get('hec_token', '')})


@app.route('/api/syslog/config/regenerate', methods=['POST'])
@_admin_required
def api_syslog_config_regenerate():
    cfg = _load_syslog_cfg()
    cfg['hec_token'] = _secrets.token_hex(24)
    _save_syslog_cfg(cfg)
    return jsonify({'ok': True, 'hec_token': cfg['hec_token']})


@app.route('/api/remediation/log')
def api_remediation_log():
    return jsonify({'actions': db.get_remediation_log(limit=200)})


@app.route('/api/remediation/rules')
def api_remediation_rules():
    return jsonify({'rules': _REMED_RULES})


@app.route('/api/events/blocked-ips')
def api_blocked_ips():
    return jsonify({'blocked_ips': db.get_blocked_ips()})


@app.route('/api/events/ingest-scan', methods=['POST'])
def api_events_ingest_scan():
    """Parse a completed scan output for security events and auto-ingest them."""
    data    = request.get_json(force=True, silent=True) or {}
    scan_id = data.get('scan_id')
    if not scan_id:
        return jsonify({'error': 'scan_id required'}), 400

    scan = db.get_scan(int(scan_id))
    if not scan:
        return jsonify({'error': 'scan not found'}), 404

    output = (scan.get('output') or '').lower()
    target = scan.get('target', '')
    count  = 0

    # Parse common patterns from scan output and emit events
    _scan_patterns = [
        ('sql_injection',       'attack',        'HIGH',     ['sql injection', 'sqli']),
        ('xss_attempt',         'attack',        'HIGH',     ['cross-site scripting', 'xss']),
        ('brute_force',         'attack',        'HIGH',     ['brute force', 'login attempt']),
        ('open_port',           'attack',        'MEDIUM',   ['open port', 'exposed service']),
        ('weak_ssl',            'vulnerability', 'HIGH',     ['weak ssl', 'self-signed', 'tls 1.0', 'tls 1.1']),
        ('missing_headers',     'vulnerability', 'MEDIUM',   ['missing header', 'hsts', 'x-frame']),
        ('outdated_software',   'vulnerability', 'HIGH',     ['outdated', 'vulnerable version']),
        ('vulnerability_detected','vulnerability','CRITICAL', ['critical', 'cve-']),
        ('default_credentials', 'attack',        'CRITICAL', ['default password', 'admin:admin']),
    ]

    for ev_type, cat, sev, keywords in _scan_patterns:
        if any(k in output for k in keywords):
            db.log_security_event(
                event_type=ev_type, category=cat, severity=sev,
                target=target, description=f'{ev_type} detected during scan of {target}',
                raw_data=_json.dumps({'scan_id': scan_id}),
            )
            count += 1

    return jsonify({'ok': True, 'events_created': count, 'target': target})


# ── Grafana reverse proxy (makes Grafana embeddable via ngrok / any remote URL) ──
@app.route('/grafana-proxy/', defaults={'path': ''})
@app.route('/grafana-proxy/<path:path>')
@login_required
def grafana_proxy(path):
    """Proxy all Grafana traffic through CyberINK so it works behind ngrok."""
    import requests as _req, base64 as _b64
    from flask import Response, stream_with_context

    gf_base = os.environ.get('GRAFANA_URL', 'http://localhost:3000').rstrip('/')
    target  = f"{gf_base}/{path}"
    qs      = request.query_string.decode('utf-8')
    if qs:
        target = f"{target}?{qs}"

    skip_req = {'host', 'content-length', 'transfer-encoding', 'connection'}
    fwd = {k: v for k, v in request.headers if k.lower() not in skip_req}
    # Always pass Grafana admin credentials so anonymous-access setting doesn't matter
    fwd['Authorization'] = 'Basic ' + _b64.b64encode(b'admin:admin').decode()

    try:
        resp = _req.request(
            method  = request.method,
            url     = target,
            headers = fwd,
            data    = request.get_data(),
            stream  = True,
            timeout = 30,
            allow_redirects = True,
        )
        # Strip headers that block embedding
        drop = {'content-encoding', 'content-length', 'transfer-encoding',
                'connection', 'x-frame-options', 'content-security-policy'}
        resp_headers = [(k, v) for k, v in resp.raw.headers.items()
                        if k.lower() not in drop]

        ct = resp.headers.get('Content-Type', 'application/octet-stream')
        _ORANGE_VARIANTS = ('#eb7b18', '#EB7B18', '#Eb7b18',
                            'rgb(235, 123, 24)', 'rgb(235,123,24)',
                            'rgba(235, 123, 24', 'rgba(235,123,24')
        _PROXY_BASE = '/grafana-proxy'

        def _fix_orange(text):
            for _o in _ORANGE_VARIANTS:
                text = text.replace(_o, '#3b82f6')
            return text

        # Patch HTML: rewrite asset paths, and replace Grafana's error page with a polished one
        if 'text/html' in ct:
            body = resp.content.decode('utf-8', errors='replace')

            # ── Grafana fallback error page — replace entirely ──────────────────
            if 'Grafana has failed to load its application files' in body:
                body = (
                    '<!DOCTYPE html><html><head>'
                    '<meta charset="utf-8">'
                    '<meta name="viewport" content="width=device-width,initial-scale=1">'
                    '<style>'
                    '*{box-sizing:border-box;margin:0;padding:0}'
                    'body{'
                      'font-family:"Inter",system-ui,sans-serif;'
                      'background:#eef2ff;'
                      'display:flex;align-items:center;justify-content:center;'
                      'min-height:100vh;padding:24px;'
                    '}'
                    '.card{'
                      'background:#fff;border:1px solid #c7d7fe;'
                      'border-radius:20px;'
                      'box-shadow:0 8px 32px rgba(99,102,241,.12);'
                      'max-width:580px;width:100%;padding:36px 40px;'
                    '}'
                    '.badge{'
                      'display:inline-flex;align-items:center;gap:6px;'
                      'background:#eef2ff;color:#4f46e5;'
                      'border:1px solid #c7d7fe;border-radius:20px;'
                      'font-size:.72rem;font-weight:700;letter-spacing:.05em;'
                      'padding:4px 12px;margin-bottom:18px;'
                    '}'
                    'h1{'
                      'font-size:1.15rem;font-weight:700;color:#1e1b4b;'
                      'line-height:1.45;margin-bottom:6px;'
                    '}'
                    '.sub{'
                      'font-size:.82rem;color:#6366f1;margin-bottom:24px;'
                    '}'
                    '.divider{height:1px;background:#e0e7ff;margin-bottom:22px;}'
                    'ol{padding-left:0;list-style:none;display:flex;flex-direction:column;gap:14px;}'
                    'li{'
                      'display:flex;gap:12px;align-items:flex-start;'
                      'font-size:.875rem;color:#374151;line-height:1.6;'
                    '}'
                    '.num{'
                      'flex-shrink:0;width:24px;height:24px;border-radius:50%;'
                      'background:#eef2ff;border:1px solid #c7d7fe;'
                      'display:flex;align-items:center;justify-content:center;'
                      'font-size:.72rem;font-weight:700;color:#4f46e5;margin-top:1px;'
                    '}'
                    'code{'
                      'background:#eef2ff;color:#4338ca;border:1px solid #c7d7fe;'
                      'border-radius:5px;padding:1px 6px;font-size:.8rem;'
                      'font-family:ui-monospace,monospace;'
                    '}'
                    '::-webkit-scrollbar{width:8px;height:8px;}'
                    '::-webkit-scrollbar-track{background:#eef2ff;border-radius:8px;}'
                    '::-webkit-scrollbar-thumb{'
                      'background:linear-gradient(180deg,#6366f1,#4f46e5);'
                      'border-radius:8px;border:2px solid #eef2ff;'
                    '}'
                    '::-webkit-scrollbar-thumb:hover{background:#4338ca;}'
                    '*{scrollbar-width:thin;scrollbar-color:#6366f1 #eef2ff;}'
                    '</style></head><body>'
                    '<div class="card">'
                      '<div class="badge">Grafana</div>'
                      '<h1>Application files failed to load</h1>'
                      '<p class="sub">Grafana is running but its frontend assets could not be served through the reverse proxy.</p>'
                      '<div class="divider"></div>'
                      '<ol>'
                        '<li><span class="num">1</span><span>Reverse proxy subpath — set <code>root_url</code> in <code>grafana.ini</code> to include the subpath, and enable <code>serve_from_sub_path = true</code>.</span></li>'
                        '<li><span class="num">2</span><span>Restart Grafana after any config change: <code>systemctl restart grafana-server</code></span></li>'
                        '<li><span class="num">3</span><span>Try the <strong>Fix automatically</strong> button below to patch <code>grafana.ini</code> without SSH.</span></li>'
                        '<li><span class="num">4</span><span>Check browser console for blocked asset requests — Grafana loads JS from <code>/public/build/</code>.</span></li>'
                      '</ol>'
                    '</div>'
                    '</body></html>'
                )
                return Response(body.encode('utf-8'), status=resp.status_code,
                                headers=resp_headers, content_type=ct)

            # ── Normal Grafana page — rewrite asset paths + orange fix ──────────
            body = body.replace('href="/public/', f'href="{_PROXY_BASE}/public/')
            body = body.replace('src="/public/',  f'src="{_PROXY_BASE}/public/')
            body = body.replace('href="/avatar/', f'href="{_PROXY_BASE}/avatar/')
            body = body.replace('src="/avatar/',  f'src="{_PROXY_BASE}/avatar/')
            body = body.replace('"__grafana_public_path__":"/"',
                                f'"__grafana_public_path__":"{_PROXY_BASE}/public/"')
            body = body.replace("'__grafana_public_path__':'/'",
                                f"'__grafana_public_path__':'{_PROXY_BASE}/public/'")
            _base_tag = f'<base href="{_PROXY_BASE}/">'
            if '<base ' not in body:
                body = body.replace('<head>', '<head>' + _base_tag, 1)
            body = _fix_orange(body)
            _gf_scrollbar = (
                '<style>'
                '::-webkit-scrollbar{width:8px;height:8px;}'
                '::-webkit-scrollbar-track{background:#eef2ff;border-radius:8px;}'
                '::-webkit-scrollbar-thumb{'
                  'background:linear-gradient(180deg,#6366f1,#4f46e5);'
                  'border-radius:8px;border:2px solid #eef2ff;'
                '}'
                '::-webkit-scrollbar-thumb:hover{background:#4338ca;}'
                '*{scrollbar-width:thin;scrollbar-color:#6366f1 #eef2ff;}'
                '</style>'
            )
            body = body.replace('</head>', _gf_scrollbar + '</head>', 1)
            if '</head>' not in body:
                body = _gf_scrollbar + body
            return Response(body.encode('utf-8'), status=resp.status_code,
                            headers=resp_headers, content_type=ct)
        # Patch CSS/JS: replace orange color values
        if 'text/css' in ct or 'javascript' in ct or \
           (path and (path.endswith('.css') or path.endswith('.js'))):
            body = resp.content.decode('utf-8', errors='replace')
            body = _fix_orange(body)
            return Response(body.encode('utf-8'), status=resp.status_code,
                            headers=resp_headers, content_type=ct)
        return Response(
            stream_with_context(resp.iter_content(chunk_size=16384)),
            status       = resp.status_code,
            headers      = resp_headers,
            content_type = ct,
        )
    except Exception as exc:
        return f'Grafana proxy error: {exc}', 502


# ── Grafana / Prometheus probe ─────────────────────────────────────────────────
@app.route('/api/grafana/probe', methods=['POST'])
@login_required
def api_grafana_probe():
    import requests as _req
    data        = request.get_json(silent=True) or {}
    grafana_url = data.get('grafana_url', 'http://localhost:3000').rstrip('/')
    prom_url    = data.get('prometheus_url', 'http://localhost:9090').rstrip('/')
    gf_user     = data.get('gf_user', 'admin')
    gf_pass     = data.get('gf_pass', 'admin')

    out = {
        'grafana':       {'running': False, 'version': '', 'url': grafana_url,
                          'embed_ok': False, 'dashboards': []},
        'prometheus':    {'running': False, 'url': prom_url},
        'node_exporter': {'running': False},
        'targets': [], 'metrics': {}, 'alerts': []
    }

    # Grafana health + dashboard list
    try:
        r = _req.get(f'{grafana_url}/api/health', timeout=4, auth=(gf_user, gf_pass))
        if r.status_code == 200:
            out['grafana']['running'] = True
            out['grafana']['version'] = r.json().get('version', '')
    except Exception:
        pass

    if out['grafana']['running']:
        try:
            r = _req.get(f'{grafana_url}/api/search?type=dash-db&limit=20',
                         timeout=4, auth=(gf_user, gf_pass))
            if r.status_code == 200:
                out['grafana']['dashboards'] = [
                    {'uid': d.get('uid', ''), 'title': d.get('title', ''),
                     'url': d.get('url', ''), 'tags': d.get('tags', [])}
                    for d in r.json()[:10]
                ]
        except Exception:
            pass
        # Check if anonymous embedding works (no auth)
        try:
            r = _req.get(f'{grafana_url}/api/health', timeout=3)
            out['grafana']['embed_ok'] = r.status_code == 200
        except Exception:
            pass

    # Prometheus targets
    _is_windows_exporter = False
    try:
        r = _req.get(f'{prom_url}/api/v1/targets', timeout=4)
        if r.status_code == 200:
            out['prometheus']['running'] = True
            active = r.json().get('data', {}).get('activeTargets', [])
            out['targets'] = [
                {'job': t.get('labels', {}).get('job', ''),
                 'instance': t.get('labels', {}).get('instance', ''),
                 'health': t.get('health', 'unknown')}
                for t in active
            ]
            # Accept both Node Exporter (Linux) and Windows Exporter (Windows :9182)
            for t in active:
                job = t.get('labels', {}).get('job', '').lower()
                inst = t.get('labels', {}).get('instance', '')
                if 'node' in job or 'windows' in job or '9182' in inst or '9100' in inst:
                    out['node_exporter']['running'] = True
                    if 'windows' in job or '9182' in inst:
                        _is_windows_exporter = True
                    break
    except Exception:
        pass

    # Prometheus instant metrics — use Windows Exporter or Node Exporter queries as appropriate
    if out['prometheus']['running']:
        if _is_windows_exporter:
            queries = {
                'cpu_pct':    '100 - (avg by (instance) (irate(windows_cpu_time_total{mode="idle"}[5m])) * 100)',
                'mem_pct':    '100 - (windows_os_physical_memory_free_bytes / windows_cs_physical_memory_bytes * 100)',
                'disk_pct':   '100 - (windows_logical_disk_free_bytes{volume="C:"} / windows_logical_disk_size_bytes{volume="C:"} * 100)',
                'net_rx_bps': 'sum(irate(windows_net_bytes_received_total[5m]))',
                'net_tx_bps': 'sum(irate(windows_net_bytes_sent_total[5m]))',
                'uptime_s':   'windows_system_system_up_time',
            }
        else:
            queries = {
                'cpu_pct':    '100 - (avg(irate(node_cpu_seconds_total{mode="idle"}[5m])) * 100)',
                'mem_pct':    '(1 - (node_memory_MemAvailable_bytes / node_memory_MemTotal_bytes)) * 100',
                'disk_pct':   '(1 - (node_filesystem_avail_bytes{mountpoint="/"} / node_filesystem_size_bytes{mountpoint="/"})) * 100',
                'net_rx_bps': 'sum(irate(node_network_receive_bytes_total[5m]))',
                'net_tx_bps': 'sum(irate(node_network_transmit_bytes_total[5m]))',
                'uptime_s':   'node_time_seconds - node_boot_time_seconds',
            }
        for key, q in queries.items():
            try:
                r = _req.get(f'{prom_url}/api/v1/query', params={'query': q}, timeout=4)
                res = r.json().get('data', {}).get('result', [])
                if res:
                    out['metrics'][key] = float(res[0]['value'][1])
            except Exception:
                pass
        try:
            r = _req.get(f'{prom_url}/api/v1/alerts', timeout=4)
            alerts = r.json().get('data', {}).get('alerts', [])
            out['alerts'] = [
                {'name': a.get('labels', {}).get('alertname', ''),
                 'state': a.get('state', ''),
                 'severity': a.get('labels', {}).get('severity', 'info')}
                for a in alerts[:10]
            ]
        except Exception:
            pass

    return jsonify(out)


@app.route('/api/grafana/install', methods=['POST'])
@login_required
def api_grafana_install():
    """Install Grafana + Prometheus + Node/Windows Exporter on the local machine.
    On Linux: runs a bash script via subprocess.
    On Windows: launches an elevated PowerShell script.
    """
    import subprocess, sys as _sys

    script = r"""
# CyberINK — Grafana Auto-Setup
$ProgressPreference = 'SilentlyContinue'
$ErrorActionPreference = 'Continue'
function Step($m){ Write-Host "==> $m" -ForegroundColor Cyan }

New-Item -ItemType Directory -Force "C:\monitoring\data" | Out-Null

# ── Windows Exporter ──────────────────────────────────────────────────────────
Step "Checking Windows Exporter..."
$svc = Get-Service "windows_exporter" -ErrorAction SilentlyContinue
if (-not $svc) {
    Step "Downloading Windows Exporter..."
    $we = Invoke-RestMethod "https://api.github.com/repos/prometheus-community/windows_exporter/releases/latest"
    $url = ($we.assets | Where-Object { $_.name -like "*amd64.msi" }).browser_download_url
    Invoke-WebRequest -Uri $url -OutFile "C:\monitoring\we.msi"
    Step "Installing Windows Exporter (port 9182)..."
    Start-Process msiexec.exe -Wait -ArgumentList "/i C:\monitoring\we.msi /quiet"
} else { Step "Windows Exporter already installed" }

# ── Prometheus ────────────────────────────────────────────────────────────────
Step "Checking Prometheus..."
$pd = Get-ChildItem "C:\monitoring\prometheus-*" -Directory -ErrorAction SilentlyContinue | Select-Object -First 1
if (-not $pd) {
    Step "Downloading Prometheus..."
    $p = Invoke-RestMethod "https://api.github.com/repos/prometheus/prometheus/releases/latest"
    $url = ($p.assets | Where-Object { $_.name -like "*windows-amd64.zip" }).browser_download_url
    Invoke-WebRequest -Uri $url -OutFile "C:\monitoring\prom.zip"
    Expand-Archive "C:\monitoring\prom.zip" "C:\monitoring" -Force
    Remove-Item "C:\monitoring\prom.zip"
    $pd = Get-ChildItem "C:\monitoring\prometheus-*" -Directory | Select-Object -First 1
}
$exe = Join-Path $pd.FullName "prometheus.exe"
$yml = Join-Path $pd.FullName "prometheus.yml"
@"
global:
  scrape_interval: 15s
scrape_configs:
  - job_name: 'windows'
    static_configs:
      - targets: ['localhost:9182']
"@ | Set-Content $yml
Get-Process prometheus -ErrorAction SilentlyContinue | Stop-Process -Force -ErrorAction SilentlyContinue
Step "Starting Prometheus (port 9090)..."
Start-Process $exe -ArgumentList "--config.file=`"$yml`" --storage.tsdb.path=`"C:\monitoring\data`"" -WindowStyle Hidden

# ── Grafana ───────────────────────────────────────────────────────────────────
Step "Checking Grafana..."
$gf = Get-Service "Grafana" -ErrorAction SilentlyContinue
if (-not $gf) {
    Step "Getting latest Grafana version..."
    try {
        $gfVer = (Invoke-RestMethod "https://grafana.com/api/grafana/versions/stable" -TimeoutSec 10).version
    } catch { $gfVer = "11.3.0" }
    $gfUrl = "https://dl.grafana.com/oss/release/grafana-$gfVer.windows-amd64.msi"
    Step "Downloading Grafana $gfVer (this may take a minute)..."
    Invoke-WebRequest -Uri $gfUrl -OutFile "C:\monitoring\grafana.msi" -UseBasicParsing
    Step "Installing Grafana..."
    Start-Process msiexec.exe -Wait -ArgumentList "/i C:\monitoring\grafana.msi /quiet"
    Step "Grafana installed"
} else { Step "Grafana already installed" }
$iniCandidates = @(
    "C:\Program Files\GrafanaLabs\grafana\conf\grafana.ini",
    "C:\Program Files (x86)\GrafanaLabs\grafana\conf\grafana.ini"
)
foreach ($ini in $iniCandidates) {
    if (Test-Path $ini) {
        $c = Get-Content $ini -Raw
        if ($c -notmatch 'allow_embedding\s*=\s*true') {
            Add-Content $ini "`n[auth.anonymous]`nenabled = true`norg_name = Main Org.`norg_role = Viewer`n`n[security]`nallow_embedding = true`n`n[server]`nserve_from_sub_path = true"
            Step "Grafana embedding + sub-path proxy enabled in $ini"
        }
        break
    }
}
try { Restart-Service Grafana -ErrorAction Stop; Step "Grafana restarted" }
catch { Start-Service Grafana -ErrorAction SilentlyContinue }

# ── Wait for Grafana ──────────────────────────────────────────────────────────
Step "Waiting for Grafana to be ready..."
$ok = $false
for ($i=0; $i -lt 30; $i++) {
    Start-Sleep 3
    try {
        $r = Invoke-WebRequest "http://localhost:3000/api/health" -UseBasicParsing -TimeoutSec 2
        if ($r.StatusCode -eq 200) { $ok = $true; break }
    } catch {}
}

if ($ok) {
    Step "Grafana is up! Adding Prometheus data source..."
    $cred = [Convert]::ToBase64String([Text.Encoding]::ASCII.GetBytes("admin:admin"))
    $hdr  = @{Authorization="Basic $cred"; "Content-Type"="application/json"}
    $ds   = '{"name":"Prometheus","type":"prometheus","url":"http://localhost:9090","access":"proxy","isDefault":true}'
    try { Invoke-RestMethod "http://localhost:3000/api/datasources" -Method Post -Headers $hdr -Body $ds -ErrorAction SilentlyContinue } catch {}
    Write-Host "`n[OK] Setup complete — switch back to CyberINK." -ForegroundColor Green
} else {
    Write-Host "`n[WARN] Grafana did not start in time. Check http://localhost:3000 manually." -ForegroundColor Yellow
}
Write-Host "`nPress Enter to close this window..."
Read-Host
"""

    if _sys.platform == 'win32':
        # Windows — launch elevated PowerShell
        import base64 as _b64
        encoded = _b64.b64encode(script.encode('utf-16-le')).decode('ascii')
        ps = 'powershell.exe'
        try:
            subprocess.Popen(
                [ps, '-Command',
                 f'Start-Process powershell -Verb RunAs -ArgumentList "-ExecutionPolicy Bypass -EncodedCommand {encoded}"'],
                shell=False
            )
            return jsonify({'ok': True})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)})
    else:
        # Linux — run bash install script
        linux_script = r"""#!/bin/bash
set -e
echo "==> Installing Grafana + Prometheus + Node Exporter..."

# ── Node Exporter ──────────────────────────────────────────────────────────────
if ! systemctl is-active --quiet node_exporter 2>/dev/null; then
    NE_VER=$(curl -s https://api.github.com/repos/prometheus/node_exporter/releases/latest | grep '"tag_name"' | cut -d'"' -f4 | tr -d v)
    NE_VER=${NE_VER:-1.8.2}
    wget -q "https://github.com/prometheus/node_exporter/releases/download/v${NE_VER}/node_exporter-${NE_VER}.linux-amd64.tar.gz" -O /tmp/ne.tar.gz
    tar xzf /tmp/ne.tar.gz -C /tmp
    cp /tmp/node_exporter-${NE_VER}.linux-amd64/node_exporter /usr/local/bin/
    useradd -rs /bin/false node_exporter 2>/dev/null || true
    cat > /etc/systemd/system/node_exporter.service <<EOF
[Unit]
Description=Node Exporter
After=network.target
[Service]
User=node_exporter
ExecStart=/usr/local/bin/node_exporter
Restart=always
[Install]
WantedBy=multi-user.target
EOF
    systemctl daemon-reload
    systemctl enable --now node_exporter
    echo "==> Node Exporter installed on :9100"
else
    echo "==> Node Exporter already running"
fi

# ── Prometheus ────────────────────────────────────────────────────────────────
if ! systemctl is-active --quiet prometheus 2>/dev/null; then
    P_VER=$(curl -s https://api.github.com/repos/prometheus/prometheus/releases/latest | grep '"tag_name"' | cut -d'"' -f4 | tr -d v)
    P_VER=${P_VER:-2.53.0}
    wget -q "https://github.com/prometheus/prometheus/releases/download/v${P_VER}/prometheus-${P_VER}.linux-amd64.tar.gz" -O /tmp/prom.tar.gz
    tar xzf /tmp/prom.tar.gz -C /tmp
    cp /tmp/prometheus-${P_VER}.linux-amd64/prometheus /usr/local/bin/
    cp /tmp/prometheus-${P_VER}.linux-amd64/promtool /usr/local/bin/
    mkdir -p /etc/prometheus /var/lib/prometheus
    cat > /etc/prometheus/prometheus.yml <<EOF
global:
  scrape_interval: 15s
scrape_configs:
  - job_name: 'node'
    static_configs:
      - targets: ['localhost:9100']
EOF
    useradd -rs /bin/false prometheus 2>/dev/null || true
    chown -R prometheus:prometheus /etc/prometheus /var/lib/prometheus
    cat > /etc/systemd/system/prometheus.service <<EOF
[Unit]
Description=Prometheus
After=network.target
[Service]
User=prometheus
ExecStart=/usr/local/bin/prometheus --config.file=/etc/prometheus/prometheus.yml --storage.tsdb.path=/var/lib/prometheus
Restart=always
[Install]
WantedBy=multi-user.target
EOF
    systemctl daemon-reload
    systemctl enable --now prometheus
    echo "==> Prometheus installed on :9090"
else
    echo "==> Prometheus already running"
fi

# ── Grafana ───────────────────────────────────────────────────────────────────
if ! systemctl is-active --quiet grafana-server 2>/dev/null; then
    apt-get install -y apt-transport-https software-properties-common wget gnupg 2>/dev/null || true
    wget -q -O /usr/share/keyrings/grafana.key https://apt.grafana.com/gpg.key
    echo "deb [signed-by=/usr/share/keyrings/grafana.key] https://apt.grafana.com stable main" > /etc/apt/sources.list.d/grafana.list
    apt-get update -qq
    apt-get install -y grafana
    # Enable embedding + anonymous access
    INI=/etc/grafana/grafana.ini
    grep -q 'allow_embedding' $INI || echo -e '\n[security]\nallow_embedding = true\n\n[auth.anonymous]\nenabled = true\norg_name = Main Org.\norg_role = Viewer\n\n[server]\nserve_from_sub_path = true' >> $INI
    systemctl daemon-reload
    systemctl enable --now grafana-server
    echo "==> Grafana installed on :3000"
else
    echo "==> Grafana already running"
fi

echo ""
echo "[OK] Setup complete — Grafana :3000, Prometheus :9090, Node Exporter :9100"
"""
        import tempfile, os as _os
        try:
            with tempfile.NamedTemporaryFile(mode='w', suffix='.sh', delete=False) as f:
                f.write(linux_script)
                tmp = f.name
            _os.chmod(tmp, 0o755)
            subprocess.Popen(['bash', tmp], stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
            return jsonify({'ok': True})
        except Exception as e:
            return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/grafana/patch-ini', methods=['POST'])
@login_required
def api_grafana_patch_ini():
    """Patch grafana.ini on an existing install to enable anonymous access + embedding."""
    import subprocess
    script = r"""
$candidates = @(
    "C:\Program Files\GrafanaLabs\grafana\conf\grafana.ini",
    "C:\Program Files (x86)\GrafanaLabs\grafana\conf\grafana.ini"
)
$patched = $false
foreach ($ini in $candidates) {
    if (Test-Path $ini) {
        $c = Get-Content $ini -Raw
        $add = ""
        if ($c -notmatch '\[auth\.anonymous\]') {
            $add += "`n[auth.anonymous]`nenabled = true`norg_name = Main Org.`norg_role = Viewer`n"
        }
        if ($c -notmatch 'allow_embedding\s*=\s*true') {
            $add += "`n[security]`nallow_embedding = true`n"
        }
        if ($c -notmatch 'serve_from_sub_path\s*=\s*true') {
            $add += "`n[server]`nserve_from_sub_path = true`n"
        }
        if ($add -ne "") {
            Add-Content $ini $add
        }
        Restart-Service Grafana -ErrorAction SilentlyContinue
        Start-Sleep 3
        Write-Host "OK:$ini"
        $patched = $true
        break
    }
}
if (-not $patched) { Write-Host "ERR:grafana.ini not found" }
"""
    import base64 as _b64, subprocess
    encoded = _b64.b64encode(script.encode('utf-16-le')).decode('ascii')
    try:
        result = subprocess.run(
            ['powershell', '-ExecutionPolicy', 'Bypass', '-EncodedCommand', encoded],
            capture_output=True, text=True, timeout=20
        )
        out = (result.stdout + result.stderr).strip()
        if 'OK:' in out:
            return jsonify({'ok': True, 'msg': 'grafana.ini patched and Grafana restarted.'})
        if 'ERR:' in out:
            return jsonify({'ok': False, 'error': out})
        return jsonify({'ok': True, 'msg': out or 'Done.'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


@app.route('/api/grafana/configure', methods=['POST'])
@login_required
def api_grafana_configure():
    """Auto-configure Grafana: ensure Prometheus data source exists, import dashboards."""
    import requests as _req
    data     = request.get_json(silent=True) or {}
    gf_url   = data.get('grafana_url', 'http://localhost:3000').rstrip('/')
    gf_user  = data.get('gf_user', 'admin')
    gf_pass  = data.get('gf_pass', 'admin')
    prom_url = data.get('prometheus_url', 'http://localhost:9090').rstrip('/')
    auth     = (gf_user, gf_pass)
    done     = []

    # Ensure Prometheus data source
    try:
        existing = _req.get(f'{gf_url}/api/datasources/name/Prometheus', auth=auth, timeout=5)
        if existing.status_code == 404:
            body = {'name': 'Prometheus', 'type': 'prometheus', 'url': prom_url,
                    'access': 'proxy', 'isDefault': True}
            r = _req.post(f'{gf_url}/api/datasources', json=body, auth=auth, timeout=5)
            if r.status_code in (200, 409):
                done.append('data_source')
        else:
            done.append('data_source')
    except Exception:
        pass

    # Import recommended dashboards from grafana.com
    dash_ids = [1860, 13391, 14031]   # Node Exporter Full, Network Overview, Network Traffic
    ds_uid = None
    try:
        ds = _req.get(f'{gf_url}/api/datasources/name/Prometheus', auth=auth, timeout=5).json()
        ds_uid = ds.get('uid', '')
    except Exception:
        pass

    imported = []
    for dash_id in dash_ids:
        try:
            raw = _req.get(f'https://grafana.com/api/dashboards/{dash_id}/revisions/latest/download',
                           timeout=10).json()
            raw['id'] = None
            payload = {
                'dashboard': raw,
                'overwrite': True,
                'inputs': [{'name': '__DS_PROMETHEUS', 'type': 'datasource',
                            'pluginId': 'prometheus', 'value': ds_uid or 'Prometheus'}]
            }
            r = _req.post(f'{gf_url}/api/dashboards/import', json=payload, auth=auth, timeout=15)
            if r.status_code == 200:
                imported.append(dash_id)
        except Exception:
            pass

    done.append(f'imported {len(imported)} dashboards')
    return jsonify({'ok': True, 'done': done, 'dashboards_imported': imported})


# ── Pentest Engagement ─────────────────────────────────────────────────────────

@app.route('/api/pentest/engagements', methods=['GET'])
@login_required
def api_pentest_list():
    uf = _cu_filter()
    return jsonify({'engagements': db.get_engagements(username=uf)})


@app.route('/api/pentest/engagements', methods=['POST'])
@login_required
def api_pentest_create():
    d = request.get_json(silent=True) or {}
    eid = db.create_engagement(
        name       = d.get('name', 'Unnamed Engagement'),
        client     = d.get('client', ''),
        scope_urls = d.get('scope_urls', []),
        scope_ips  = d.get('scope_ips', []),
        auth_config= d.get('auth_config', {}),
        urgency    = d.get('urgency', 'normal'),
        deadline   = d.get('deadline', ''),
        notes      = d.get('notes', ''),
        username   = _cu_username()
    )
    return jsonify({'ok': True, 'id': eid})


@app.route('/api/pentest/engagements/<int:eid>', methods=['PUT'])
@login_required
def api_pentest_update(eid):
    import json
    d = request.get_json(silent=True) or {}
    eng = db.get_engagement(eid)
    if not eng:
        return jsonify({'error': 'Not found'}), 404
    with db._connect() as con:
        con.execute(
            '''UPDATE engagements SET name=?,client=?,scope_urls=?,scope_ips=?,auth_config=?,
               urgency=?,deadline=?,notes=? WHERE id=?''',
            (d.get('name', eng['name']),
             d.get('client', eng['client']),
             json.dumps(d.get('scope_urls', eng['scope_urls'])),
             json.dumps(d.get('scope_ips', eng['scope_ips'])),
             json.dumps(d.get('auth_config', eng['auth_config'])),
             d.get('urgency', eng['urgency']),
             d.get('deadline', eng['deadline']),
             d.get('notes', eng['notes']),
             eid)
        )
        con.commit()
    return jsonify({'ok': True, 'id': eid})


@app.route('/api/pentest/engagements/<int:eid>', methods=['DELETE'])
@login_required
def api_pentest_delete(eid):
    db.delete_engagement(eid)
    return jsonify({'ok': True})


@app.route('/api/pentest/engagements/<int:eid>/status', methods=['POST'])
@login_required
def api_pentest_set_status(eid):
    d = request.get_json(silent=True) or {}
    status = d.get('status', 'pending')
    if status not in ('pending', 'active', 'completed'):
        return jsonify({'error': 'Invalid status'}), 400
    db.update_engagement_status(eid, status)
    return jsonify({'ok': True})


@app.route('/api/pentest/engagements/<int:eid>/scans', methods=['GET'])
@login_required
def api_pentest_scans(eid):
    scans = db.get_engagement_scans(eid)
    return jsonify({'scans': [enrich(s) for s in scans]})


def _pentest_fetch_oidc_token(token_url: str, client_id: str, client_secret: str, scope: str):
    """Fetch an OIDC access token via client_credentials grant. Returns (token, error_str)."""
    try:
        import requests as _req
        r = _req.post(token_url, data={
            'grant_type':    'client_credentials',
            'client_id':     client_id,
            'client_secret': client_secret,
            'scope':         scope,
        }, timeout=15)
        r.raise_for_status()
        j = r.json()
        return j.get('access_token', ''), ''
    except Exception as exc:
        return '', str(exc)


def _build_pentest_cred_block(creds: dict, target: str) -> str:
    """Build an authenticated-scan instruction block for enterprise pentest targets."""
    atype = creds.get('auth_type', 'none')
    if not atype or atype == 'none':
        return ''

    hdr = (
        '\n\n══════════════ AUTHENTICATED SCAN ══════════════\n'
        'Credentials provided. You MUST use them for every check below.\n'
        'NEVER print secrets/tokens in output — write [REDACTED] instead.\n'
        '════════════════════════════════════════════════\n\n'
    )
    base_url = (target if '://' in target else f'https://{target}').rstrip('/')

    if atype == 'bearer':
        token = creds.get('bearer_token', '')
        if not token:
            return ''
        return (hdr
            + 'BEARER TOKEN AUTHENTICATION\n'
            + f'  Authorization: Bearer {token}\n\n'
            + 'Run ALL authenticated checks using this token:\n\n'
            + '1. Authenticated root + API discovery:\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/"\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/api/"\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/api/v1/"\n\n'
            + '2. User/profile endpoint (IDOR baseline):\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/api/v1/me"\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/api/v1/user"\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/api/v1/profile"\n\n'
            + '3. Privilege escalation — admin endpoints:\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/admin"\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/api/admin"\n'
            + f'   curl -s -H "Authorization: Bearer {token}" "{base_url}/api/v1/admin"\n\n'
            + '4. JWT claims decode (no secret needed):\n'
            + f'   python3 -c "import base64,json; t=\'{token}\'; p=t.split(\'.\')[1]; '
            + 'p+=\'=\'*(4-len(p)%4); print(json.dumps(json.loads(base64.b64decode(p)),indent=2))"\n\n'
            + '5. IDOR — enumerate other user IDs:\n'
            + f'   for id in 1 2 3 100 999; do curl -s -o /dev/null -w "%{{http_code}} id=$id\\n" '
            + f'-H "Authorization: Bearer {token}" "{base_url}/api/v1/users/$id"; done\n\n'
            + '6. HTTP method fuzzing:\n'
            + f'   for m in GET POST PUT PATCH DELETE OPTIONS; do curl -s -o /dev/null -w "%{{http_code}} $m\\n" '
            + f'-X $m -H "Authorization: Bearer {token}" "{base_url}/api/v1/me"; done\n'
        )

    if atype == 'apikey':
        key    = creds.get('api_key', '')
        header = creds.get('api_key_header', 'X-API-Key')
        if not key:
            return ''
        return (hdr
            + 'API KEY AUTHENTICATION\n'
            + f'  {header}: {key}\n\n'
            + 'Run ALL authenticated API checks:\n\n'
            + '1. API discovery:\n'
            + f'   curl -s -H "{header}: {key}" "{base_url}/api/"\n'
            + f'   curl -s -H "{header}: {key}" "{base_url}/api/v1/"\n'
            + f'   curl -s -H "{header}: {key}" "{base_url}/api/v2/"\n\n'
            + '2. Rate limit test (20 rapid requests):\n'
            + f'   for i in $(seq 1 20); do curl -s -o /dev/null -w "%{{http_code}} req=$i\\n" '
            + f'-H "{header}: {key}" "{base_url}/api/v1/me"; done\n\n'
            + '3. IDOR via list endpoints:\n'
            + f'   curl -s -H "{header}: {key}" "{base_url}/api/v1/users"\n'
            + f'   curl -s -H "{header}: {key}" "{base_url}/api/v1/accounts"\n'
            + f'   for id in 1 2 3 100 999; do curl -s -o /dev/null -w "%{{http_code}} id=$id\\n" '
            + f'-H "{header}: {key}" "{base_url}/api/v1/users/$id"; done\n\n'
            + '4. API key in query param (alternate injection):\n'
            + f'   curl -s "{base_url}/api/v1/me?api_key={key}"\n'
            + f'   curl -s "{base_url}/api/v1/me?apikey={key}"\n\n'
            + '5. HTTP method override:\n'
            + f'   curl -s -X POST -H "{header}: {key}" -H "X-HTTP-Method-Override: DELETE" "{base_url}/api/v1/me"\n\n'
            + '6. Admin endpoint access:\n'
            + f'   curl -s -H "{header}: {key}" "{base_url}/api/v1/admin/users"\n'
            + f'   curl -s -H "{header}: {key}" "{base_url}/api/admin"\n'
        )

    if atype == 'basic':
        user = creds.get('basic_user', '')
        pw   = creds.get('basic_pass', '')
        if not user:
            return ''
        return (hdr
            + 'BASIC AUTHENTICATION\n'
            + f'  Username : {user}\n'
            + '  Password : $BASIC_PASSWORD (set in environment — do not print)\n\n'
            + f'  export BASIC_PASSWORD="{pw}"\n\n'
            + 'Run ALL basic-auth checks:\n\n'
            + '1. Login test:\n'
            + f'   curl -s -u "{user}:$BASIC_PASSWORD" -o /dev/null -w "%{{http_code}}" "{base_url}/"\n'
            + f'   curl -s -u "{user}:$BASIC_PASSWORD" -o /dev/null -w "%{{http_code}}" "{base_url}/api/v1/me"\n\n'
            + '2. Admin panel access:\n'
            + f'   curl -s -u "{user}:$BASIC_PASSWORD" "{base_url}/admin"\n\n'
            + '3. Enumerate protected directories:\n'
            + f'   for path in api admin dashboard config settings users accounts; do '
            + f'curl -s -o /dev/null -w "%{{http_code}} /$path\\n" -u "{user}:$BASIC_PASSWORD" "{base_url}/$path"; done\n\n'
            + '4. Unauthenticated fallback (must return 401/403):\n'
            + f'   curl -s -o /dev/null -w "%{{http_code}}" "{base_url}/api/v1/me"\n'
            + f'   curl -s -o /dev/null -w "%{{http_code}}" "{base_url}/admin"\n\n'
            + '5. Account lockout policy (brute force test):\n'
            + f'   for pw in admin password 123456 test qwerty; do '
            + f'curl -s -o /dev/null -w "%{{http_code}} pw=$pw\\n" -u "{user}:$pw" "{base_url}/"; done\n'
        )

    if atype == 'oidc':
        token_url    = creds.get('oidc_token_url', '')
        client_id    = creds.get('oidc_client_id', '')
        scope        = creds.get('oidc_scope', 'openid profile')
        access_token = creds.get('oidc_access_token', '')
        if not access_token:
            return hdr + '[OIDC] Token fetch failed — proceeding without authentication\n'
        return (hdr
            + 'OIDC/SSO AUTHENTICATION (client_credentials grant)\n'
            + f'  Token URL   : {token_url}\n'
            + f'  Client ID   : {client_id}\n'
            + f'  Scope       : {scope}\n'
            + '  Access token acquired — use as Bearer for all requests\n\n'
            + 'Run ALL OIDC-authenticated checks:\n\n'
            + '1. Validate token claims:\n'
            + f'   python3 -c "import base64,json; t=\'{access_token}\'; p=t.split(\'.\')[1]; '
            + 'p+=\'=\'*(4-len(p)%4); print(json.dumps(json.loads(base64.b64decode(p)),indent=2))"\n\n'
            + '2. Protected endpoint access:\n'
            + f'   curl -s -H "Authorization: Bearer {access_token}" "{base_url}/api/v1/me"\n'
            + f'   curl -s -H "Authorization: Bearer {access_token}" "{base_url}/api/v1/users"\n\n'
            + '3. OIDC discovery document:\n'
            + f'   curl -s "{token_url.split("/token")[0]}/.well-known/openid-configuration"\n'
            + f'   curl -s "{base_url}/.well-known/openid-configuration"\n\n'
            + '4. Token endpoint CSRF / parameter tampering:\n'
            + f'   curl -s -X POST "{token_url}" '
            + '-d "grant_type=authorization_code&code=INVALID&redirect_uri=https://evil.com" | head -c 500\n\n'
            + '5. SSO bypass — admin escalation:\n'
            + f'   curl -s -H "Authorization: Bearer {access_token}" "{base_url}/admin"\n'
            + f'   curl -s -H "Authorization: Bearer {access_token}" "{base_url}/api/admin"\n\n'
            + '6. Scope escalation attempt:\n'
            + f'   curl -s -X POST "{token_url}" '
            + f'-d "grant_type=client_credentials&client_id={client_id}&client_secret=[REDACTED]&scope=admin openid"\n'
        )

    if atype == 'custom':
        headers = creds.get('custom_headers', {})
        if not headers:
            return ''
        h_flags   = ' '.join(f'-H "{k}: {v}"' for k, v in headers.items())
        h_display = '\n'.join(f'  {k}: {v}' for k, v in headers.items())
        return (hdr
            + f'CUSTOM HEADERS\n{h_display}\n\n'
            + 'Run ALL checks with these headers:\n\n'
            + '1. Authenticated discovery:\n'
            + f'   curl -s {h_flags} "{base_url}/api/"\n'
            + f'   curl -s {h_flags} "{base_url}/api/v1/"\n\n'
            + '2. Protected endpoints:\n'
            + f'   curl -s {h_flags} "{base_url}/api/v1/me"\n'
            + f'   curl -s {h_flags} "{base_url}/admin"\n\n'
            + '3. IDOR check:\n'
            + f'   for id in 1 2 3 100 999; do curl -s -o /dev/null -w "%{{http_code}} id=$id\\n" '
            + f'{h_flags} "{base_url}/api/v1/users/$id"; done\n\n'
            + '4. Unauthenticated fallback (header stripped):\n'
            + f'   curl -s "{base_url}/api/v1/me" | head -c 500\n'
        )

    return ''


def _is_network_target(target: str) -> bool:
    """Return True if target is an IPv4 address or CIDR range (not a hostname/URL)."""
    import re as _re
    t = target.strip().replace('https://', '').replace('http://', '').rstrip('/')
    return bool(_re.match(r'^\d{1,3}(?:\.\d{1,3}){3}(?:/\d{1,2})?$', t))


@app.route('/api/pentest/engagements/<int:eid>/scan-target', methods=['POST'])
@login_required
def api_pentest_scan_target(eid):
    """Queue a background scan for one target in the engagement."""
    eng = db.get_engagement(eid)
    if not eng:
        return jsonify({'error': 'Engagement not found'}), 404

    d          = request.get_json(silent=True) or {}
    target     = d.get('target', '').strip()
    agent_type = d.get('agent_type', 'pentest')
    model      = d.get('model', os.environ.get('CAI_MODEL', 'claude-sonnet-4-6'))

    if not target:
        return jsonify({'error': 'target required'}), 400

    auth  = eng.get('auth_config', {})
    atype = auth.get('type', 'none')
    creds = {
        'auth_type':      atype,
        'bearer_token':   auth.get('token', '')          if atype == 'bearer' else '',
        'api_key':        auth.get('value', '')          if atype == 'apikey' else '',
        'api_key_header': auth.get('header', 'X-API-Key'),
        'basic_user':     auth.get('username', '')       if atype == 'basic'  else '',
        'basic_pass':     auth.get('password', '')       if atype == 'basic'  else '',
        'custom_headers': auth.get('headers', {})        if atype == 'custom' else {},
        'oidc_token_url': auth.get('token_url', '')      if atype == 'oidc'   else '',
        'oidc_client_id': auth.get('client_id', '')      if atype == 'oidc'   else '',
        'oidc_secret':    auth.get('client_secret', '')  if atype == 'oidc'   else '',
        'oidc_scope':     auth.get('scope', 'openid profile') if atype == 'oidc' else '',
    }

    import uuid, threading
    job_id = str(uuid.uuid4())[:8]

    def _run(job_id, target, agent_type, model, eng_id, creds, username):
        import time as _t
        t0    = _t.time()
        parts = []

        def _ot(txt):
            parts.append(txt)
            _jobs[job_id]['output'] = ''.join(parts)

        def _oo(name, args):
            pass

        def _or(name, result, err):
            r_full = str(result)
            if r_full.strip():
                parts.append(f'\n[TOOL:{name}]\n{r_full}\n')
            _jobs[job_id]['output'] = ''.join(parts)

        try:
            from agents.pentest import run_full_pentest

            # ── OIDC: fetch real token before building cred_block ─────────────
            if creds.get('auth_type') == 'oidc' and creds.get('oidc_token_url'):
                _ot('[OIDC] Fetching access token via client_credentials grant…\n')
                tok, err = _pentest_fetch_oidc_token(
                    creds['oidc_token_url'],
                    creds.get('oidc_client_id', ''),
                    creds.get('oidc_secret', ''),
                    creds.get('oidc_scope', 'openid profile'),
                )
                if tok:
                    creds['oidc_access_token'] = tok
                    _ot(f'[OIDC] Token acquired ({len(tok)} chars) — authenticated scan enabled\n')
                else:
                    _ot(f'[OIDC] Token fetch failed: {err} — continuing without auth\n')

            cred_block = _build_pentest_cred_block(creds, target)
            if cred_block:
                _ot(f'[AUTH] {atype.upper()} credentials loaded — authenticated scan enabled\n')

            # ── Network/IP targets: nmap + nuclei ─────────────────────────────
            if _is_network_target(target):
                import subprocess as _sp
                _ot(f'\n[NETWORK] Detected IP/CIDR target: {target}\n')

                _ot('[NETWORK] Running nmap service + vuln scan (may take several minutes)…\n')
                nm_cmd = ['nmap', '-sV', '-sC', '--script=vuln', '-T4', '-oN', '-', target]
                try:
                    nm = _sp.run(nm_cmd, capture_output=True, text=True, timeout=600)
                    nm_out = nm.stdout or nm.stderr or '[nmap returned no output]'
                    _ot(nm_out)
                except FileNotFoundError:
                    _ot('[NETWORK] nmap not found — install nmap to enable network scanning\n')
                except _sp.TimeoutExpired:
                    _ot('[NETWORK] nmap timed out after 10 min\n')
                except Exception as ne:
                    _ot(f'[NETWORK] nmap error: {ne}\n')

                _ot('\n[NETWORK] Running nuclei network + CVE scan…\n')
                nu_cmd = [
                    'nuclei', '-target', target,
                    '-tags', 'network,cve,misconfig,exposed-panels',
                    '-severity', 'critical,high,medium',
                    '-silent', '-timeout', '10',
                ]
                try:
                    nu = _sp.run(nu_cmd, capture_output=True, text=True, timeout=600)
                    _ot(nu.stdout or '[nuclei returned no output]\n')
                except FileNotFoundError:
                    _ot('[NETWORK] nuclei not found — install nuclei to enable template-based scanning\n')
                except Exception as ne:
                    _ot(f'[NETWORK] nuclei error: {ne}\n')

                output = ''.join(parts)

            else:
                # ── Web/API targets: full WSTG agent suite ────────────────────
                from urllib.parse import urlparse as _up
                _parsed = _up(target if '://' in target else f'https://{target}')
                domain  = (_parsed.netloc or _parsed.path.split('/')[0]).rstrip('/')

                results = run_full_pentest(
                    domain,
                    model=model or None,
                    on_text=_ot,
                    on_tool=_oo,
                    on_result=_or,
                    cred_block=cred_block or None,
                    run_workflow=(agent_type == 'pentest'),
                    agent_key='' if agent_type == 'pentest' else agent_type,
                    is_aborted=None,
                )
                output = '\n\n'.join(
                    f'=== {k.upper()} ===\n{v}'
                    for k, v in results.items() if v
                )

            elapsed = _t.time() - t0
            sid = db.save_scan(
                target=target, agent_type=agent_type, model=model,
                status='ok', latency_s=elapsed,
                output=str(output)[:60000], username=username,
            )
            with __import__('sqlite3').connect(db.DB_PATH) as con:
                con.execute('UPDATE scans SET engagement_id=? WHERE id=?', (eng_id, sid))
                con.commit()
            db.update_engagement_status(eng_id, 'active')
            _jobs[job_id].update({'status': 'done', 'scan_id': sid, 'target': target})

        except Exception as exc:
            import traceback
            err_detail = traceback.format_exc()
            elapsed = _t.time() - t0
            sid = db.save_scan(
                target=target, agent_type=agent_type, model=model,
                status='error', latency_s=elapsed,
                output=f'[ERROR] {exc}\n\n{err_detail}'[:60000], username=username,
            )
            with __import__('sqlite3').connect(db.DB_PATH) as con:
                con.execute('UPDATE scans SET engagement_id=? WHERE id=?', (eng_id, sid))
                con.commit()
            _jobs[job_id].update({'status': 'error', 'error': str(exc), 'scan_id': sid})

    _jobs[job_id] = {'status': 'running', 'target': target, 'output': ''}
    t = threading.Thread(
        target=_run,
        args=(job_id, target, agent_type, model, eid, creds, _cu_username()),
        daemon=True,
    )
    t.start()
    return jsonify({'ok': True, 'job_id': job_id})


@app.route('/api/pentest/jobs/<job_id>', methods=['GET'])
@login_required
def api_pentest_job_poll(job_id):
    job = _jobs.get(job_id)
    if not job:
        return jsonify({'status': 'unknown'}), 404
    return jsonify(job)


@app.route('/api/pentest/engagements/<int:eid>/report', methods=['GET'])
@login_required
def api_pentest_report(eid):
    """Generate a professional HTML pentest report for the engagement."""
    import re as _re, json as _json
    eng   = db.get_engagement(eid)
    if not eng:
        return 'Engagement not found', 404
    scans = db.get_engagement_scans(eid)

    sev_pat = _re.compile(
        r'(CRITICAL|HIGH|MEDIUM|LOW|INFO)\s*[-–:]\s*(.+)', _re.IGNORECASE)
    cve_pat = _re.compile(r'CVE-\d{4}-\d+', _re.IGNORECASE)
    rec_pat = _re.compile(
        r'(?:RECOMMENDATION|REMEDIATION|FIX|MITIGATION)[:\s]+(.+)', _re.IGNORECASE)

    findings = []
    cve_set  = set()
    sev_counts = {'CRITICAL':0,'HIGH':0,'MEDIUM':0,'LOW':0,'INFO':0}

    for s in scans:
        out = s.get('output','')
        for m in sev_pat.finditer(out):
            sev = m.group(1).upper()
            sev_counts[sev] = sev_counts.get(sev, 0) + 1
            findings.append({'target': s['target'], 'severity': sev,
                              'detail': m.group(2).strip()[:200],
                              'agent': s['agent_type']})
        for c in cve_pat.findall(out):
            cve_set.add(c.upper())

    sev_color = {'CRITICAL':'#7f0000','HIGH':'#a80000','MEDIUM':'#986f0b',
                 'LOW':'#1e3a8a','INFO':'#374151'}

    scope_urls = eng.get('scope_urls', [])
    scope_ips  = eng.get('scope_ips',  [])
    all_targets = scope_urls + scope_ips

    rows = ''
    for f in sorted(findings, key=lambda x: ['CRITICAL','HIGH','MEDIUM','LOW','INFO'].index(x['severity']) if x['severity'] in ['CRITICAL','HIGH','MEDIUM','LOW','INFO'] else 99):
        col = sev_color.get(f['severity'], '#374151')
        rows += f'<tr><td>{f["target"]}</td><td style="color:{col};font-weight:700;">{f["severity"]}</td><td>{f["detail"]}</td><td>{f["agent"]}</td></tr>'

    scan_rows = ''
    for s in scans:
        status_col = '#16a34a' if s.get('status')=='ok' else '#dc2626'
        scan_rows += f'<tr><td>{s["target"]}</td><td>{s["agent_type"]}</td><td style="color:{status_col};">{s.get("status","—")}</td><td>{(s.get("created_at",""))[:16]}</td><td>{s.get("latency_s",0):.1f}s</td></tr>'

    total_findings = len(findings)
    risk_level = 'CRITICAL' if sev_counts['CRITICAL'] else 'HIGH' if sev_counts['HIGH'] else 'MEDIUM' if sev_counts['MEDIUM'] else 'LOW'
    risk_color = sev_color.get(risk_level, '#374151')

    html = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<title>Pentest Report — {eng["name"]}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;font-size:13px;color:#1d1d1f;background:#fff;padding:40px}}
  h1{{font-size:26px;font-weight:800;color:#1d1d1f;margin-bottom:4px}}
  h2{{font-size:15px;font-weight:700;color:#1d1d1f;margin:28px 0 10px;padding-bottom:6px;border-bottom:2px solid #e5e7eb}}
  h3{{font-size:13px;font-weight:700;color:#374151;margin:16px 0 6px}}
  .meta{{color:#6b7280;font-size:12px;margin-bottom:32px}}
  .kpi-bar{{display:flex;gap:16px;margin-bottom:24px;flex-wrap:wrap}}
  .kpi{{background:#f9fafb;border:1px solid #e5e7eb;border-radius:8px;padding:14px 20px;min-width:120px;text-align:center}}
  .kpi-val{{font-size:28px;font-weight:800}}
  .kpi-lbl{{font-size:11px;color:#6b7280;margin-top:3px;text-transform:uppercase;letter-spacing:.4px}}
  .risk-badge{{display:inline-block;padding:4px 14px;border-radius:4px;font-weight:700;font-size:13px;color:#fff;background:{risk_color};margin-bottom:24px}}
  table{{width:100%;border-collapse:collapse;margin-bottom:20px;font-size:12px}}
  th{{background:#f3f4f6;padding:8px 12px;text-align:left;font-weight:700;font-size:11px;text-transform:uppercase;letter-spacing:.3px;color:#6b7280;border-bottom:2px solid #e5e7eb}}
  td{{padding:8px 12px;border-bottom:1px solid #f3f4f6;vertical-align:top}}
  tr:hover td{{background:#fafafa}}
  .scope-pill{{display:inline-block;background:#eff6ff;border:1px solid #bfdbfe;color:#1d4ed8;border-radius:4px;padding:2px 10px;margin:2px;font-size:11px;font-family:monospace}}
  .cve-pill{{display:inline-block;background:#fef3c7;border:1px solid #fde68a;color:#92400e;border-radius:4px;padding:2px 8px;margin:2px;font-size:11px;font-family:monospace}}
  .footer{{margin-top:40px;padding-top:16px;border-top:1px solid #e5e7eb;color:#9ca3af;font-size:11px;text-align:center}}
  @media print{{body{{padding:20px}} .no-print{{display:none}}}}
</style>
</head>
<body>
<div style="display:flex;align-items:flex-start;justify-content:space-between;margin-bottom:8px;">
  <div>
    <h1>Penetration Test Report</h1>
    <div class="meta">
      Engagement: <strong>{eng["name"]}</strong> &nbsp;|&nbsp;
      Client: <strong>{eng.get("client","—")}</strong> &nbsp;|&nbsp;
      Urgency: <strong>{eng.get("urgency","normal").upper()}</strong> &nbsp;|&nbsp;
      Deadline: <strong>{eng.get("deadline","—") or "—"}</strong> &nbsp;|&nbsp;
      Generated: <strong>{__import__("datetime").datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}</strong>
    </div>
  </div>
  <button class="no-print" onclick="window.print()" style="padding:8px 18px;background:#1d1d1f;color:#fff;border:none;border-radius:6px;cursor:pointer;font-size:12px;font-weight:600;">Print / Save PDF</button>
</div>

<div class="risk-badge">Overall Risk: {risk_level}</div>

<div class="kpi-bar">
  <div class="kpi"><div class="kpi-val" style="color:{sev_color["CRITICAL"]}">{sev_counts["CRITICAL"]}</div><div class="kpi-lbl">Critical</div></div>
  <div class="kpi"><div class="kpi-val" style="color:{sev_color["HIGH"]}">{sev_counts["HIGH"]}</div><div class="kpi-lbl">High</div></div>
  <div class="kpi"><div class="kpi-val" style="color:{sev_color["MEDIUM"]}">{sev_counts["MEDIUM"]}</div><div class="kpi-lbl">Medium</div></div>
  <div class="kpi"><div class="kpi-val" style="color:{sev_color["LOW"]}">{sev_counts["LOW"]}</div><div class="kpi-lbl">Low</div></div>
  <div class="kpi"><div class="kpi-val">{len(scans)}</div><div class="kpi-lbl">Scans Run</div></div>
  <div class="kpi"><div class="kpi-val">{total_findings}</div><div class="kpi-lbl">Findings</div></div>
  <div class="kpi"><div class="kpi-val" style="color:#92400e;">{len(cve_set)}</div><div class="kpi-lbl">CVEs</div></div>
</div>

<h2>Scope</h2>
<h3>Web Targets</h3>
{''.join(f'<span class="scope-pill">{t}</span>' for t in scope_urls) or '<em style="color:#9ca3af;">None defined</em>'}
<h3 style="margin-top:10px;">Network / IP Targets</h3>
{''.join(f'<span class="scope-pill">{t}</span>' for t in scope_ips) or '<em style="color:#9ca3af;">None defined</em>'}
{'<h3 style="margin-top:10px;">CVEs Referenced</h3>' + "".join(f'<span class="cve-pill">{c}</span>' for c in sorted(cve_set)) if cve_set else ""}

<h2>Executive Summary</h2>
<p style="line-height:1.7;color:#374151;">
This penetration test covered <strong>{len(all_targets)} target(s)</strong> across web applications and network infrastructure.
A total of <strong>{total_findings} findings</strong> were identified across <strong>{len(scans)} scan(s)</strong>,
with an overall risk rating of <strong style="color:{risk_color};">{risk_level}</strong>.
{f'<strong>{sev_counts["CRITICAL"]} critical</strong> and ' if sev_counts["CRITICAL"] else ""}
{f'<strong>{sev_counts["HIGH"]} high-severity</strong> issues require immediate attention.' if sev_counts["HIGH"] or sev_counts["CRITICAL"] else "No critical or high-severity issues were identified."}
{f'{len(cve_set)} known CVE(s) were referenced in the findings.' if cve_set else ""}
</p>
{f'<p style="margin-top:8px;line-height:1.7;color:#374151;"><strong>Notes:</strong> {eng.get("notes","")}</p>' if eng.get("notes") else ""}

<h2>Findings</h2>
{"<table><thead><tr><th>Target</th><th>Severity</th><th>Finding</th><th>Agent</th></tr></thead><tbody>" + rows + "</tbody></table>" if findings else '<p style="color:#9ca3af;font-style:italic;">No structured findings extracted — review raw scan outputs below.</p>'}

<h2>Scans Performed</h2>
<table><thead><tr><th>Target</th><th>Agent</th><th>Status</th><th>Date</th><th>Duration</th></tr></thead>
<tbody>{scan_rows or "<tr><td colspan=5 style='color:#9ca3af;'>No scans completed yet.</td></tr>"}</tbody></table>

<h2>Methodology</h2>
<p style="line-height:1.7;color:#374151;">
Tests were conducted using the OWASP Web Security Testing Guide (WSTG) methodology, covering:
Information Gathering, Configuration &amp; Deployment, Identity Management, Authentication,
Authorization, Session Management, Input Validation, Error Handling, Cryptography,
Business Logic, Client-Side, and API Security (OWASP API Top 10).
Network tests covered open port discovery, service enumeration, and CVE correlation.
</p>

<div class="footer">CyberINK Pentest Report &nbsp;|&nbsp; {eng["name"]} &nbsp;|&nbsp; Confidential</div>
</body></html>'''

    return html, 200, {'Content-Type': 'text/html; charset=utf-8'}


@app.route('/api/inventories/plugins')
def api_inventories_plugins():
    """Return all plugins detected across scans, optionally filtered by target."""
    target = request.args.get('target', '')
    plugins = db.get_plugins(target=target, username=_cu_filter())
    return jsonify({'plugins': plugins, 'total': len(plugins)})


@app.route('/api/inventories/logins')
def api_inventories_logins():
    """Return user login events extracted from scan output (WP-LOG and auth lines)."""
    target = request.args.get('target', '')
    limit  = min(int(request.args.get('limit', 500)), 2000)
    scans  = db.get_scans_for_target(target, username=_cu_filter()) if target else db.get_recent_scans(limit=200, username=_cu_filter())

    # Also capture WP-USER enumeration lines not covered by WP-LOG
    user_enum_pat = re.compile(
        r'^WP-USER(?:-ENUM|-CONFIRMED)?\s*\|\s*(\S+)\s*\|\s*(\S+)',
        re.I | re.MULTILINE,
    )

    logins = []
    for s in (scans or []):
        out = s.get('output', '') or ''
        fallback_date = (s.get('created_at') or '')[:16]

        # Use the existing extract_wp_logs() which correctly parses WP-LOG | date | user | event | ip | risk
        # and also handles WP-USER / CREDS_FOUND / APP_PASS_CREATED fallback lines
        wp = extract_wp_logs(out)
        for entry in wp.get('entries', []):
            logins.append({
                'target':  s['target'],
                'user':    entry.get('user', ''),
                'event':   entry.get('event', 'login'),
                'ip':      entry.get('ip', ''),
                'risk':    entry.get('risk', 'INFO'),
                'date':    entry.get('timestamp') or fallback_date,
                'scan_id': s['id'],
            })

        # WP-USER | id | login | and WP-USER-CONFIRMED | username | lines (user enumeration)
        for m in user_enum_pat.finditer(out):
            username = m.group(2).strip()
            if username and len(username) < 80:
                logins.append({
                    'target':  s['target'],
                    'user':    username,
                    'event':   'User enumerated by scanner',
                    'ip':      '',
                    'risk':    'MEDIUM',
                    'date':    fallback_date,
                    'scan_id': s['id'],
                })

        if len(logins) >= limit:
            break

    return jsonify({'logins': logins[:limit], 'total': len(logins)})


@app.route('/api/inventories/mysql-plugins', methods=['POST'])
def api_inventories_mysql_plugins():
    """Read active WordPress plugins from wp_options.active_plugins via direct MySQL."""
    data      = request.get_json(force=True, silent=True) or {}
    db_host   = (data.get('db_host') or '').strip()
    db_port   = int(data.get('db_port') or 3306)
    db_name   = (data.get('db_name') or '').strip()
    db_user   = (data.get('db_user') or '').strip()
    db_pass   = (data.get('db_pass') or '').strip()
    table_pfx = (data.get('table_prefix') or 'wp_').strip()

    if not db_host or not db_name or not db_user:
        return jsonify({'error': 'db_host, db_name, and db_user are required'}), 400

    try:
        import pymysql, pymysql.cursors
    except ImportError:
        return jsonify({'error': 'pymysql not installed — run: pip install pymysql'}), 500

    try:
        conn = pymysql.connect(
            host=db_host, port=db_port, user=db_user, password=db_pass,
            database=db_name, charset='utf8mb4', connect_timeout=15,
            cursorclass=pymysql.cursors.DictCursor,
        )
    except Exception as e:
        return jsonify({'error': f'MySQL connection failed: {e}'}), 500

    plugins = []
    try:
        with conn.cursor() as cur:
            opts_table = f'{table_pfx}options'
            # Read active_plugins (PHP serialized array of plugin paths)
            cur.execute(f"SELECT option_value FROM {opts_table} WHERE option_name = 'active_plugins' LIMIT 1")
            row = cur.fetchone()
            if row:
                raw = str(row.get('option_value') or '')
                # Parse plugin paths from PHP serialized string: s:XX:"plugin/plugin.php"
                plugin_paths = re.findall(r'"([\w/-]+\.php)"', raw)
                for path in plugin_paths:
                    slug = path.split('/')[0]
                    plugins.append({'name': slug, 'path': path, 'status': 'active', 'version': '', 'source': 'mysql'})

            # Try to get versions from update data
            cur.execute(f"SELECT option_value FROM {opts_table} WHERE option_name = 'update_plugins' LIMIT 1")
            upd_row = cur.fetchone()
            if upd_row:
                upd_raw = str(upd_row.get('option_value') or '')
                for p in plugins:
                    m = re.search(r'"' + re.escape(p['path']) + r'"[^}]*?"new_version"\s*;s:\d+:"([^"]+)"', upd_raw)
                    if m:
                        p['version'] = m.group(1)

            # Also read Wordfence plugin detection if available
            wf_known = f'{table_pfx}wfKnownFileMeta'
            cur.execute("SHOW TABLES LIKE %s", (wf_known,))
            if cur.fetchone():
                cur.execute(f"SELECT data FROM {wf_known} WHERE type='plugin' LIMIT 200")
                for r in cur.fetchall():
                    try:
                        import json as _jj
                        d = _jj.loads(r.get('data') or '{}')
                        slug = d.get('slug') or ''
                        ver  = d.get('version') or ''
                        if slug and not any(p['name'] == slug for p in plugins):
                            plugins.append({'name': slug, 'path': '', 'status': 'detected', 'version': ver, 'source': 'wordfence'})
                    except Exception:
                        pass
    except Exception as e:
        return jsonify({'error': f'Query error: {e}', 'plugins': []}), 500
    finally:
        conn.close()

    return jsonify({'plugins': plugins, 'total': len(plugins)})


@app.route('/api/inventories/cpanel-plugins', methods=['POST'])
def api_inventories_cpanel_plugins():
    """Read active WordPress plugins from wp_options via cPanel Fileman (no direct MySQL needed)."""
    import base64 as _b64, json as _j, re as _re, random as _rand, string as _str
    import ssl as _ssl2, urllib.request as _req2

    data     = request.get_json(force=True, silent=True) or {}
    site_url = (data.get('url') or '').strip().rstrip('/')
    cp_host  = (data.get('cp_host') or '').strip().rstrip('/')
    cp_user  = (data.get('cp_user') or '').strip()
    cp_pass  = (data.get('cp_pass') or '').strip()
    cp_token = (data.get('cp_token') or '').strip()
    wp_dir   = (data.get('wp_dir') or 'public_html').strip().strip('/')

    if not cp_host or not cp_user or not (cp_pass or cp_token):
        return jsonify({'error': 'cPanel host, username, and password or token required'}), 400
    if not site_url:
        return jsonify({'error': 'WordPress site URL required'}), 400
    if not site_url.startswith('http'):
        site_url = 'https://' + site_url

    base = cp_host if cp_host.startswith('http') else f'https://{cp_host}:2083'
    auth_hdr = (f'cpanel {cp_user}:{cp_token}' if cp_token
                else 'Basic ' + _b64.b64encode(f'{cp_user}:{cp_pass}'.encode()).decode())
    ctx = _ssl2.create_default_context()
    ctx.check_hostname = False
    ctx.verify_mode = _ssl2.CERT_NONE

    def _cp(path, method='GET', post_data=None):
        url = f'{base}/execute/{path}'
        hdr = {'Authorization': auth_hdr, 'Accept': 'application/json', 'User-Agent': _BROWSER_UA}
        if post_data:
            body = _up_parse.urlencode(post_data).encode()
            req = _req2.Request(url, data=body, headers={**hdr, 'Content-Type': 'application/x-www-form-urlencoded'})
        else:
            req = _req2.Request(url, headers=hdr)
        try:
            with _req2.urlopen(req, timeout=15, context=ctx) as r:
                return _j.loads(r.read().decode())
        except Exception as e:
            return {'_err': str(e)}

    # Read wp-config.php to extract DB credentials
    cfg = _cp(f'Fileman/get_file_content?dir=%2F{_up_parse.quote(wp_dir)}&file=wp-config.php')
    if cfg.get('_err') or not cfg.get('status'):
        return jsonify({'error': f'Cannot read wp-config.php: {cfg.get("_err") or cfg.get("errors") or "check WP Directory"}'}), 500
    wp_config = (cfg.get('data') or {}).get('content', '')
    if not wp_config:
        return jsonify({'error': 'wp-config.php is empty or unreadable'}), 500

    def _cfg(key):
        m = _re.search(rf"define\s*\(\s*['\"]DB_{key}['\"]\s*,\s*['\"]([^'\"]*)['\"]", wp_config)
        return m.group(1) if m else ''

    db_host   = _cfg('HOST') or 'localhost'
    db_name   = _cfg('NAME')
    db_user2  = _cfg('USER')
    db_pass2  = _cfg('PASSWORD')
    pfx_m     = _re.search(r"\$table_prefix\s*=\s*['\"]([^'\"]+)['\"]", wp_config)
    table_pfx = pfx_m.group(1) if pfx_m else 'wp_'

    if not db_name or not db_user2:
        return jsonify({'error': 'Could not parse DB credentials from wp-config.php'}), 500

    script_name = 'cfai_' + ''.join(_rand.choices(_str.ascii_lowercase + _str.digits, k=14)) + '.php'
    php = (
        "<?php error_reporting(0);\n"
        "try {\n"
        f"  $pdo = new PDO('mysql:host={db_host};dbname={db_name};charset=utf8','{db_user2}','{db_pass2}');\n"
        "  $pdo->setAttribute(PDO::ATTR_ERRMODE, PDO::ERRMODE_EXCEPTION);\n"
        f"  $p = '{table_pfx}';\n"
        "  $plugins = [];\n"
        "  $plugin_dir = dirname(__FILE__) . '/wp-content/plugins/';\n"
        "  $row = $pdo->query(\"SELECT option_value FROM {$p}options WHERE option_name='active_plugins' LIMIT 1\")->fetch(PDO::FETCH_ASSOC);\n"
        "  if ($row) {\n"
        "    preg_match_all('/\"([\\w\\/-]+\\.php)\"/', $row['option_value'], $m);\n"
        "    foreach ($m[1] as $path) {\n"
        "      $slug = explode('/', $path)[0];\n"
        "      $ver = '';\n"
        "      $pfile = $plugin_dir . $path;\n"
        "      if (file_exists($pfile)) {\n"
        "        $hdr = file_get_contents($pfile, false, null, 0, 8192);\n"
        "        if (preg_match('/^[ \\t\\/*#]*Version:\\s*(.+)$/mi', $hdr, $vm)) $ver = trim($vm[1]);\n"
        "      }\n"
        "      $plugins[] = ['name'=>$slug,'path'=>$path,'status'=>'active','version'=>$ver];\n"
        "    }\n"
        "  }\n"
        "  header('Content-Type: application/json');\n"
        "  echo json_encode(['ok'=>true,'plugins'=>$plugins]);\n"
        "} catch(Exception $e) {\n"
        "  header('Content-Type: application/json');\n"
        "  echo json_encode(['ok'=>false,'error'=>$e->getMessage()]);\n"
        "}\n"
    )

    up = _cp('Fileman/save_file_content', method='POST',
             post_data={'dir': f'/{wp_dir}', 'file': script_name, 'content': php})
    if up.get('_err') or not up.get('status'):
        return jsonify({'error': f'Cannot upload script via cPanel Fileman: {up.get("_err") or up.get("errors")}'}), 500

    plugins = []
    try:
        sc_code, sc_body = _wp_request(f'{site_url}/{script_name}', timeout=20)
        if sc_code == 200:
            result = _j.loads(sc_body)
            if result.get('ok'):
                plugins = result.get('plugins', [])
    except Exception:
        pass
    finally:
        _cp('Fileman/delete_files', method='POST',
            post_data={'files': f'/{wp_dir}/{script_name}'})

    return jsonify({'plugins': plugins, 'total': len(plugins), 'source': 'cpanel'})


@app.route('/api/chat', methods=['POST'])
def api_chat():
    """AI Chatbot — streams a Claude response via SSE. Body: {message, history:[{role,content}]}"""
    data    = request.get_json(force=True, silent=True) or {}
    message = (data.get('message') or '').strip()
    history = data.get('history') or []

    if not message:
        return jsonify({'error': 'message is required'}), 400

    api_key = os.environ.get('ANTHROPIC_API_KEY', '').strip()
    if not api_key:
        return jsonify({'error': 'ANTHROPIC_API_KEY not configured in .env'}), 500

    try:
        import anthropic as _ant
    except ImportError:
        return jsonify({'error': 'anthropic package not installed — run: pip install anthropic'}), 500

    # Build context-aware system prompt
    try:
        stats = db.get_stats()
    except Exception:
        stats = {}
    total_scans   = stats.get('total_scans', 0)
    total_targets = stats.get('total_targets', 0)

    system_prompt = f"""You are CyberINK AI, an expert security intelligence assistant built into the CyberINK Security Intelligence dashboard.

Current dashboard context:
- Total scans in database: {total_scans}
- Unique targets monitored: {total_targets}

Dashboard pages you can explain:
- Dashboard: Security posture overview, KPIs, live risk scores
- Secure Score: Security score based on real scan findings
- Threat Analytics: PCI-style threat analytics from scan history
- Incidents: Security incident tracking and management
- Security Signals: Real-time threat signals and alerts
- Event Timeline: Chronological security event timeline
- MITRE ATT&CK: Attack technique mapping from scan findings
- Priority Actions: Critical actions requiring immediate attention, with Cloudflare site-lock
- Recommendations: AI-generated security improvement suggestions
- Remediation: SOC-style remediation workflow
- Weaknesses: CVE and vulnerability catalog by severity
- Inventories: Plugin, software, and scanned-site inventory
- Security Analytics: Historical scan analytics and trends
- Observability: Unified monitoring across all targets
- Log Explorer: Live log analysis via SSH, WordPress Admin, cPanel, or MySQL
- User Activity Logs: WordPress admin actions and login history
- Network Monitor: Port and service monitoring
- Connect & Scan: Launch AI security scans with various agents

AI agents available for scanning:
- API Security Tester (apit): REST API vulnerability testing
- WordPress Security (wpsc): Deep WordPress security audit
- Info Gathering (info): OSINT and reconnaissance
- Network Scanner (netscan): Port scanning and service enumeration
- Vulnerability Scanner (vulnscan): CVE and weakness detection

You can help users:
1. Evaluate any CVE — explain CVSS score, affected systems, impact, exploitation, and remediation steps
2. Interpret scan findings and security alerts
3. Explain any dashboard feature or workflow
4. Provide security best practices and hardening recommendations
5. Answer general cybersecurity questions
6. Help prioritize remediation based on risk and exploitability

Be concise, accurate, and actionable. Use markdown for structure. For CVEs always include: severity, affected versions, attack vector, and concrete fix steps."""

    messages = []
    for h in (history or [])[-20:]:
        role    = (h.get('role') or '').strip()
        content = (h.get('content') or '').strip()
        if role in ('user', 'assistant') and content:
            messages.append({'role': role, 'content': content})
    messages.append({'role': 'user', 'content': message})

    client = _ant.Anthropic(api_key=api_key)

    def _generate():
        try:
            with client.messages.stream(
                model='claude-haiku-4-5-20251001',
                max_tokens=1500,
                system=system_prompt,
                messages=messages,
            ) as stream:
                for text in stream.text_stream:
                    yield f"data: {_json.dumps({'text': text})}\n\n"
        except Exception as e:
            yield f"data: {_json.dumps({'error': str(e)})}\n\n"
        yield "data: [DONE]\n\n"

    return Response(
        stream_with_context(_generate()),
        mimetype='text/event-stream',
        headers={'Cache-Control': 'no-cache', 'X-Accel-Buffering': 'no'},
    )


@app.route('/api/geoip')
def api_geoip():
    """Batch IP-to-country lookup via ip-api.com. Pass ?ips=1.2.3.4,5.6.7.8"""
    import json as _j, urllib.request as _req2
    raw = (request.args.get('ips') or '').strip()
    if not raw:
        return jsonify({'results': {}})
    unique_ips = list(dict.fromkeys(i.strip() for i in raw.split(',') if i.strip()))[:100]
    # Clean masked IPs like 1.2.3.x → 1.2.3.0 for geolocation (country level still works)
    clean = [ip.replace('.x', '.0') for ip in unique_ips]
    batch = [{'query': ip, 'fields': 'country,countryCode,status'} for ip in clean]
    results = {}
    try:
        req = _req2.Request(
            'http://ip-api.com/batch?fields=country,countryCode,status',
            data=_j.dumps(batch).encode(),
            headers={'Content-Type': 'application/json', 'User-Agent': _BROWSER_UA},
        )
        with _req2.urlopen(req, timeout=8) as r:
            data = _j.loads(r.read().decode())
        for orig, row in zip(unique_ips, data):
            if row.get('status') == 'success':
                results[orig] = {'country': row.get('country', ''), 'code': row.get('countryCode', '')}
            else:
                results[orig] = {'country': '', 'code': ''}
    except Exception:
        for ip in unique_ips:
            results[ip] = {'country': '', 'code': ''}
    return jsonify({'results': results})


# ── Google Search Console / Domain Health Analysis ──────────────────────────

_GOOGLE_CLIENT_ID     = os.environ.get('GOOGLE_CLIENT_ID', '').strip()
_GOOGLE_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '').strip()
_GSC_TOKEN_FILE       = Path(__file__).parent.parent / '.gsc_token.json'
_GSC_REDIRECT_URI     = os.environ.get('CFAI_BASE_URL', 'http://localhost:8889').rstrip('/') + '/auth/google/callback'
_GSC_SCOPES           = 'https://www.googleapis.com/auth/webmasters.readonly'


def _gsc_load_tokens() -> dict:
    try:
        if _GSC_TOKEN_FILE.exists():
            return _json.loads(_GSC_TOKEN_FILE.read_text(encoding='utf-8'))
    except Exception:
        pass
    return {}


def _gsc_save_tokens(tokens: dict) -> None:
    try:
        _GSC_TOKEN_FILE.write_text(_json.dumps(tokens), encoding='utf-8')
    except Exception:
        pass


def _gsc_get_access_token() -> str:
    """Return a valid access token, refreshing via refresh_token if needed."""
    import time as _t
    tokens = _gsc_load_tokens()
    if not tokens.get('refresh_token'):
        return ''
    # Return cached token if still valid (5-min buffer)
    if tokens.get('access_token') and tokens.get('expires_at', 0) > _t.time() + 300:
        return tokens['access_token']
    # Refresh
    try:
        if not (_GOOGLE_CLIENT_ID and _GOOGLE_CLIENT_SECRET):
            return ''
        if _HAS_REQUESTS:
            r = _requests.post('https://oauth2.googleapis.com/token', data={
                'client_id':     _GOOGLE_CLIENT_ID,
                'client_secret': _GOOGLE_CLIENT_SECRET,
                'refresh_token': tokens['refresh_token'],
                'grant_type':    'refresh_token',
            }, timeout=10)
            data = r.json()
        else:
            body = _up_parse.urlencode({
                'client_id':     _GOOGLE_CLIENT_ID,
                'client_secret': _GOOGLE_CLIENT_SECRET,
                'refresh_token': tokens['refresh_token'],
                'grant_type':    'refresh_token',
            }).encode()
            req = _up_req.Request('https://oauth2.googleapis.com/token', data=body)
            with _up_req.urlopen(req, timeout=10) as resp:
                data = _json.loads(resp.read())
        if 'access_token' in data:
            tokens['access_token'] = data['access_token']
            tokens['expires_at']   = _t.time() + data.get('expires_in', 3600)
            _gsc_save_tokens(tokens)
            return tokens['access_token']
    except Exception:
        pass
    return ''


@app.route('/auth/google')
def auth_google():
    """Redirect user to Google OAuth consent screen."""
    if not (_GOOGLE_CLIENT_ID and _GOOGLE_CLIENT_SECRET):
        return jsonify({'error': 'GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET not configured in .env'}), 500
    params = _up_parse.urlencode({
        'client_id':     _GOOGLE_CLIENT_ID,
        'redirect_uri':  _GSC_REDIRECT_URI,
        'response_type': 'code',
        'scope':         _GSC_SCOPES,
        'access_type':   'offline',
        'prompt':        'consent',
    })
    return redirect(f'https://accounts.google.com/o/oauth2/v2/auth?{params}')


@app.route('/auth/google/callback')
def auth_google_callback():
    """Exchange OAuth code for tokens and save refresh token."""
    import time as _t
    code  = request.args.get('code', '')
    error = request.args.get('error', '')
    if error:
        return f'''<html><body style="font-family:sans-serif;padding:40px;text-align:center;">
            <h2 style="color:#a80000;">Authorization Failed</h2><p>{error}</p>
            <script>if(window.opener){{window.opener.postMessage({{type:"gsc_auth_fail",error:"{error}"}},"*");setTimeout(function(){{window.close();}},3000);}}</script>
        </body></html>'''
    if not code:
        return '<html><body><p>No authorization code received.</p></html>', 400
    try:
        if _HAS_REQUESTS:
            r = _requests.post('https://oauth2.googleapis.com/token', data={
                'client_id':     _GOOGLE_CLIENT_ID,
                'client_secret': _GOOGLE_CLIENT_SECRET,
                'code':          code,
                'grant_type':    'authorization_code',
                'redirect_uri':  _GSC_REDIRECT_URI,
            }, timeout=10)
            data = r.json()
        else:
            body = _up_parse.urlencode({
                'client_id':     _GOOGLE_CLIENT_ID,
                'client_secret': _GOOGLE_CLIENT_SECRET,
                'code':          code,
                'grant_type':    'authorization_code',
                'redirect_uri':  _GSC_REDIRECT_URI,
            }).encode()
            req = _up_req.Request('https://oauth2.googleapis.com/token', data=body)
            with _up_req.urlopen(req, timeout=10) as resp:
                data = _json.loads(resp.read())
        if 'refresh_token' not in data:
            return f'<html><body style="font-family:sans-serif;padding:40px;"><h2>Error</h2><p>No refresh token returned. Try revoking access at accounts.google.com/permissions and reconnecting.</p><p>{data}</p></body></html>', 400
        _gsc_save_tokens({
            'access_token':  data.get('access_token', ''),
            'refresh_token': data['refresh_token'],
            'expires_at':    _t.time() + data.get('expires_in', 3600),
        })
        return '''<html><body style="font-family:sans-serif;padding:40px;text-align:center;background:#f5f6fa;">
            <div style="max-width:400px;margin:0 auto;background:#fff;padding:40px;border-radius:12px;border:1px solid #edebe9;">
            <div style="font-size:48px;margin-bottom:16px;">&#10003;</div>
            <h2 style="color:#107c10;margin-bottom:8px;">Google Search Console Connected!</h2>
            <p style="color:#605e5c;">Your dashboard now has permanent access to GSC data.<br>You can close this window.</p>
            </div>
            <script>
                if(window.opener){window.opener.postMessage({type:"gsc_auth_success"},"*");setTimeout(function(){window.close();},2500);}
            </script>
        </body></html>'''
    except Exception as e:
        return f'<html><body style="font-family:sans-serif;padding:40px;"><h2>Token Exchange Failed</h2><p>{e}</p></body></html>', 500


@app.route('/api/gsc/disconnect', methods=['POST'])
def api_gsc_disconnect():
    """Remove stored GSC tokens."""
    try:
        if _GSC_TOKEN_FILE.exists():
            _GSC_TOKEN_FILE.unlink()
    except Exception:
        pass
    return jsonify({'ok': True})


@app.route('/api/gsc/config')
def api_gsc_config():
    """Return auth and config status (no keys exposed)."""
    token = _gsc_get_access_token()
    return jsonify({
        'has_google_key': bool(os.environ.get('GOOGLE_API_KEY', '').strip()),
        'gsc_connected':  bool(token),
        'has_oauth_creds': bool(_GOOGLE_CLIENT_ID and _GOOGLE_CLIENT_SECRET),
    })


@app.route('/api/gsc/analyze', methods=['POST'])
def api_gsc_analyze():
    """
    Comprehensive domain health check.
    Body: {domain, api_key?, access_token?}
    Returns: {domain, base_url, issues, scores, total, summary}
    """
    import datetime as _dt
    import socket as _sock

    data         = request.get_json(force=True, silent=True) or {}
    domain       = re.sub(r'^https?://', '', (data.get('domain') or '').strip().lower()).split('/')[0].strip()
    # Fall back to .env GOOGLE_API_KEY if the user didn't supply one in the form
    api_key      = (data.get('api_key') or '').strip() or os.environ.get('GOOGLE_API_KEY', '').strip()
    # Fall back to stored OAuth token if no manual token provided
    access_token = (data.get('access_token') or '').strip() or _gsc_get_access_token()

    if not domain:
        return jsonify({'error': 'domain is required'}), 400

    base_url = f'https://{domain}'
    issues: list = []
    scores: dict = {}

    def _add(category, severity, title, description, impact, steps):
        issues.append({'category': category, 'severity': severity, 'title': title,
                       'description': description, 'impact': impact, 'steps': steps})

    def _fetch_url(url, timeout=12, allow_redirects=True, verify_ssl=True, method='GET', json_body=None, extra_headers=None, form_data=None):
        headers = {'User-Agent': _BROWSER_UA}
        if extra_headers:
            headers.update(extra_headers)
        try:
            if _HAS_REQUESTS:
                if method == 'POST':
                    r = _requests.post(url, json=json_body, data=form_data, headers=headers,
                                       timeout=timeout, allow_redirects=allow_redirects, verify=verify_ssl)
                else:
                    r = _requests.get(url, headers=headers, timeout=timeout,
                                      allow_redirects=allow_redirects, verify=verify_ssl)
                return r, None
            else:
                if form_data:
                    req_data = _up_parse.urlencode(form_data).encode()
                    headers['Content-Type'] = 'application/x-www-form-urlencoded'
                else:
                    req_data = _json.dumps(json_body).encode() if json_body else None
                req = _up_req.Request(url, data=req_data, headers=headers, method=method)
                with _up_req.urlopen(req, timeout=timeout) as r:
                    class _Resp:
                        status_code = r.status
                        text = r.read().decode('utf-8', errors='replace')
                        history = []
                        url = r.url
                        def json(self): return _json.loads(self.text)
                        @property
                        def headers(self):
                            return dict(r.headers)
                    return _Resp(), None
        except Exception as e:
            return None, str(e)

    # ── 1. HTTPS redirect ────────────────────────────────────────────────────
    try:
        r_http, err = _fetch_url(f'http://{domain}', timeout=8, allow_redirects=True, verify_ssl=False)
        if r_http is not None:
            final = getattr(r_http, 'url', '') or ''
            if not final.startswith('https://'):
                _add('Security', 'critical', 'HTTP not redirected to HTTPS',
                     f'Visiting http://{domain} did not redirect to HTTPS. Final URL: {final or "unknown"}.',
                     'Users may browse unencrypted, exposing credentials and data to interception.',
                     ['Configure a 301 redirect from HTTP to HTTPS.',
                      'Apache: RewriteRule ^ https://%{HTTP_HOST}%{REQUEST_URI} [L,R=301]',
                      'Nginx: return 301 https://$host$request_uri; in the HTTP server block.',
                      'Cloudflare: enable "Always Use HTTPS" in SSL/TLS settings.'])
            else:
                hist = getattr(r_http, 'history', [])
                if hist and getattr(hist[0], 'status_code', 0) == 302:
                    _add('SEO', 'medium', 'HTTP to HTTPS redirect is temporary (302)',
                         'The HTTP to HTTPS redirect uses a 302 (temporary) instead of 301 (permanent).',
                         'Search engines may not pass full link equity through temporary redirects.',
                         ['Change the redirect to use HTTP 301 (Moved Permanently).',
                          'Apache: R=302 → R=301 in RewriteRule.',
                          'Nginx: return 302 → return 301.'])
    except Exception:
        pass

    # ── 2. SSL Certificate ────────────────────────────────────────────────────
    try:
        ctx = _ssl.create_default_context()
        with _sock.create_connection((domain, 443), timeout=8) as sock:
            with ctx.wrap_socket(sock, server_hostname=domain) as ssock:
                cert = ssock.getpeercert()
                not_after = cert.get('notAfter', '')
                if not_after:
                    exp = _dt.datetime.strptime(not_after, '%b %d %H:%M:%S %Y %Z')
                    days_left = (exp - _dt.datetime.utcnow()).days
                    if days_left < 0:
                        _add('Security', 'critical', 'SSL certificate has expired',
                             f'The SSL certificate expired on {not_after}.',
                             'Browsers show a full-page security warning, blocking all visitors.',
                             ['Renew the certificate immediately via your hosting provider.',
                              'Let\'s Encrypt: run "certbot renew".',
                              'Enable auto-renewal to prevent future expiration.'])
                    elif days_left < 14:
                        _add('Security', 'high', f'SSL certificate expires in {days_left} days',
                             f'Certificate expires on {not_after}. Renewal is urgent.',
                             'Unrenewed certificate causes browser warnings and blocks visitors.',
                             [f'Renew immediately — {days_left} days remain.',
                              'Run: certbot renew',
                              'Check auto-renewal is configured.'])
                    elif days_left < 30:
                        _add('Security', 'medium', f'SSL certificate expires in {days_left} days',
                             f'Certificate expires on {not_after}.',
                             'Plan renewal soon to avoid downtime.',
                             ['Renew within the next week.',
                              'Enable auto-renewal (Let\'s Encrypt certbot supports this).'])
                    else:
                        scores['ssl_days'] = days_left

                san_list = [s[1] for s in cert.get('subjectAltName', []) if s[0] == 'DNS']
                covered = any(
                    domain == s or (s.startswith('*.') and domain.endswith(s[2:]))
                    for s in san_list
                )
                if not covered:
                    _add('Security', 'high', 'SSL certificate does not cover this domain',
                         f'Certificate SANs: {", ".join(san_list) or "none"}. Domain "{domain}" is not listed.',
                         'Browsers show a hostname mismatch error to all visitors.',
                         ['Issue a new certificate that includes this domain.',
                          'Use a wildcard certificate (*.yourdomain.com).',
                          'Verify the domain is in the CSR.'])
                scores['ssl'] = {'valid': covered, 'san': san_list,
                                 'days_left': (exp - _dt.datetime.utcnow()).days if not_after else None}
    except _ssl.SSLCertVerificationError as e:
        _add('Security', 'critical', 'SSL certificate verification failed',
             f'SSL error: {e}',
             'All visitors see a "Not Secure" warning. Traffic and trust are destroyed.',
             ['Ensure the certificate is issued by a trusted CA.',
              'Verify the certificate chain (intermediate certs) is correctly installed.',
              'Use SSL Labs (ssllabs.com/ssltest) for a detailed report.'])
    except Exception as e:
        _add('Security', 'high', 'SSL/TLS connection failed',
             f'Could not establish HTTPS connection to {domain}:443 — {e}',
             'Site may not be reachable over HTTPS.',
             [f'Verify DNS: nslookup {domain}',
              'Confirm port 443 is open and a certificate is installed.',
              'Check SSL configuration in your hosting control panel.'])

    # ── 3. HTTP Security Headers ──────────────────────────────────────────────
    r_main, err_main = _fetch_url(base_url, timeout=12)
    if r_main is not None:
        raw_hdrs = getattr(r_main, 'headers', {})
        hdrs = {k.lower(): v for k, v in (raw_hdrs.items() if hasattr(raw_hdrs, 'items') else {}.items())}

        if 'strict-transport-security' not in hdrs:
            _add('Security', 'high', 'Missing HTTP Strict Transport Security (HSTS)',
                 'The Strict-Transport-Security header is absent. Downgrade attacks (MITM) are possible.',
                 'Attackers can strip HTTPS from connections. Not enforced by browsers.',
                 ['Add: Strict-Transport-Security: max-age=31536000; includeSubDomains; preload',
                  'Nginx: add_header Strict-Transport-Security "max-age=31536000; includeSubDomains" always;',
                  'Apache: Header always set Strict-Transport-Security "max-age=31536000; includeSubDomains"',
                  'After deploying, submit to hstspreload.org for browser preload list.'])
        else:
            hsts_val = hdrs['strict-transport-security']
            if 'preload' not in hsts_val:
                _add('Security', 'low', 'HSTS "preload" directive missing',
                     f'HSTS is set ({hsts_val}) but "preload" is absent.',
                     'First-time visitors are still vulnerable to downgrade attacks before HSTS kicks in.',
                     ['Add "preload" to the HSTS header.',
                      'Submit your domain to hstspreload.org.'])

        if 'content-security-policy' not in hdrs:
            _add('Security', 'high', 'Missing Content Security Policy (CSP)',
                 'No Content-Security-Policy header found. XSS attacks are unrestricted.',
                 'XSS can steal sessions, exfiltrate data, and execute arbitrary scripts.',
                 ['Start with: Content-Security-Policy: default-src \'self\'',
                  'Use report-only mode first: Content-Security-Policy-Report-Only: default-src \'self\'',
                  'Reference: developer.mozilla.org/en-US/docs/Web/HTTP/CSP'])
        else:
            csp = hdrs['content-security-policy']
            if "'unsafe-inline'" in csp:
                _add('Security', 'medium', 'CSP allows unsafe-inline scripts',
                     'The Content-Security-Policy includes \'unsafe-inline\', weakening XSS protection.',
                     'Inline script injection can still execute, bypassing CSP.',
                     ['Remove \'unsafe-inline\' and use nonces or hashes for inline scripts.',
                      'Refactor inline JS into external files.',
                      'Use: script-src \'nonce-{random_value}\''])
            if "'unsafe-eval'" in csp:
                _add('Security', 'medium', 'CSP allows unsafe-eval',
                     'The CSP includes \'unsafe-eval\', permitting eval() and similar constructs.',
                     'Enables dynamic code execution attackers can exploit.',
                     ['Remove \'unsafe-eval\'. Refactor code using eval().',
                      'Common in older jQuery plugins or AngularJS. Update to newer versions.'])

        if 'x-frame-options' not in hdrs and 'content-security-policy' not in hdrs:
            _add('Security', 'medium', 'Missing X-Frame-Options (clickjacking risk)',
                 'No X-Frame-Options or CSP frame-ancestors directive. Site can be embedded in iframes.',
                 'Clickjacking attacks can trick users into unintended actions.',
                 ['Add: X-Frame-Options: SAMEORIGIN',
                  'Or via CSP: frame-ancestors \'self\'',
                  'Nginx: add_header X-Frame-Options "SAMEORIGIN" always;'])

        if 'x-content-type-options' not in hdrs:
            _add('Security', 'low', 'Missing X-Content-Type-Options header',
                 'The X-Content-Type-Options: nosniff header is absent.',
                 'Browsers may MIME-sniff responses, potentially executing uploaded files as scripts.',
                 ['Add: X-Content-Type-Options: nosniff',
                  'Nginx: add_header X-Content-Type-Options "nosniff" always;'])

        if 'referrer-policy' not in hdrs:
            _add('Security', 'low', 'Missing Referrer-Policy header',
                 'No Referrer-Policy set. Full URLs (including query strings and tokens) may leak.',
                 'Private parameters in URLs may be exposed to third-party sites via the Referer header.',
                 ['Add: Referrer-Policy: strict-origin-when-cross-origin',
                  'Nginx: add_header Referrer-Policy "strict-origin-when-cross-origin" always;'])

        srv = hdrs.get('server', '')
        if srv and any(x in srv.lower() for x in ['apache/', 'nginx/', 'php/', 'iis/']):
            _add('Security', 'low', 'Server version disclosed in HTTP header',
                 f'Server: {srv} — version information helps attackers target known exploits.',
                 'Fingerprinting enables targeted exploitation of version-specific vulnerabilities.',
                 ['Apache: set ServerTokens Prod and ServerSignature Off in httpd.conf.',
                  'Nginx: add server_tokens off; to nginx.conf.',
                  'Cloudflare automatically masks the Server header.'])

        if 'x-powered-by' in hdrs:
            _add('Security', 'low', 'X-Powered-By header reveals technology stack',
                 f'X-Powered-By: {hdrs["x-powered-by"]} discloses backend technology.',
                 'Attackers can target version-specific vulnerabilities.',
                 ['PHP: add header_remove("X-Powered-By"); or expose_php = Off in php.ini.',
                  'Express.js: app.disable("x-powered-by");',
                  'Cloudflare strips this automatically.'])

        if 'permissions-policy' not in hdrs and 'feature-policy' not in hdrs:
            _add('Security', 'info', 'Missing Permissions-Policy header',
                 'No Permissions-Policy header restricts browser feature access.',
                 'Third-party scripts may access camera, microphone, or geolocation without restriction.',
                 ['Add: Permissions-Policy: geolocation=(), microphone=(), camera=()',
                  'Restrict only features you do not use on this site.'])

        scores['headers'] = {
            'hsts':  'strict-transport-security' in hdrs,
            'csp':   'content-security-policy' in hdrs,
            'xfo':   'x-frame-options' in hdrs or 'content-security-policy' in hdrs,
            'xcto':  'x-content-type-options' in hdrs,
            'rp':    'referrer-policy' in hdrs,
        }
    elif err_main:
        _add('Security', 'medium', 'Could not fetch HTTP headers',
             f'Failed to connect to {base_url}: {err_main}',
             'Cannot verify security header configuration.',
             [f'Ensure the domain is live: nslookup {domain}',
              'Check that HTTPS is responding on port 443.'])

    # ── 4. robots.txt ─────────────────────────────────────────────────────────
    robots_sitemap = None
    r_rob, _ = _fetch_url(f'{base_url}/robots.txt', timeout=8)
    if r_rob is not None:
        if r_rob.status_code == 200:
            rob_text = getattr(r_rob, 'text', '')
            for line in rob_text.splitlines():
                if line.lower().startswith('sitemap:'):
                    robots_sitemap = line.split(':', 1)[1].strip()
                    break
            disallowed = [l for l in rob_text.splitlines() if l.lower().startswith('disallow:')]
            if any(l.strip().lower() == 'disallow: /' for l in disallowed):
                _add('SEO', 'critical', 'robots.txt blocks all search engine crawling',
                     '"Disallow: /" in robots.txt prevents all bots from indexing the site.',
                     'The site will not appear in Google or any search engine results.',
                     ['Remove or update the "Disallow: /" rule in robots.txt.',
                      'If intentional, ensure this is only applied to specific user-agents.',
                      'After fixing, submit a recrawl request in Google Search Console.'])
            scores['robots'] = {'found': True, 'has_sitemap': bool(robots_sitemap), 'disallowed': len(disallowed)}
        elif r_rob.status_code == 404:
            _add('SEO', 'medium', 'robots.txt file not found (404)',
                 f'No robots.txt was found at {base_url}/robots.txt.',
                 'Crawlers get no guidance, potentially indexing admin pages or causing duplicate content.',
                 ['Create robots.txt at the site root. Minimum content:',
                  f'  User-agent: *\n  Disallow: /wp-admin/\n  Sitemap: {base_url}/sitemap.xml',
                  'WordPress: the Yoast SEO plugin generates this automatically.'])
            scores['robots'] = {'found': False}

    # ── 5. sitemap.xml ────────────────────────────────────────────────────────
    sitemap_url = robots_sitemap or f'{base_url}/sitemap.xml'
    r_sit, _ = _fetch_url(sitemap_url, timeout=10)
    if r_sit is not None:
        sit_text = getattr(r_sit, 'text', '')
        if r_sit.status_code == 200 and ('<urlset' in sit_text or '<sitemapindex' in sit_text):
            url_count = sit_text.count('<url>')
            scores['sitemap'] = {'found': True, 'url_count': url_count, 'url': sitemap_url}
        elif r_sit.status_code == 404:
            _add('SEO', 'medium', 'sitemap.xml not found (404)',
                 f'No sitemap found at {sitemap_url}.',
                 'Search engines must discover all pages by crawling links, missing content.',
                 [f'Generate a sitemap and place it at {base_url}/sitemap.xml.',
                  'WordPress: use Yoast SEO or All in One SEO to auto-generate.',
                  'Add to robots.txt: Sitemap: ' + base_url + '/sitemap.xml',
                  'Submit in Google Search Console under Sitemaps.'])
            scores['sitemap'] = {'found': False}
        else:
            scores['sitemap'] = {'found': False, 'status': r_sit.status_code}
    else:
        scores['sitemap'] = {'found': False}

    # ── 6. WWW canonicalization ────────────────────────────────────────────────
    if not domain.startswith('www.') and _HAS_REQUESTS:
        try:
            r_www = _requests.get(f'https://www.{domain}', timeout=8, allow_redirects=True,
                                   headers={'User-Agent': _BROWSER_UA}, verify=False)
            final_www = getattr(r_www, 'url', '').rstrip('/')
            if final_www == f'https://www.{domain}':
                _add('SEO', 'medium', 'www and non-www both serve content (no canonical redirect)',
                     f'Both https://{domain} and https://www.{domain} serve content without redirecting.',
                     'Duplicate content dilutes PageRank. Google may index the wrong version.',
                     ['Choose one canonical form (www or non-www) and 301-redirect the other.',
                      'In Google Search Console: Settings > Preferred Domain.',
                      f'Add canonical tag: <link rel="canonical" href="{base_url}"/>',
                      'Cloudflare: use a redirect rule to enforce the canonical URL.'])
        except Exception:
            pass

    # ── 7. Google Safe Browsing + link crawl (detects "Links to harmful downloads") ──
    if api_key:
        try:
            # Collect URLs: domain itself + all outgoing links from homepage
            urls_to_check = [base_url, f'http://{domain}']
            if r_main is not None:
                try:
                    html_text = getattr(r_main, 'text', '')
                    link_pat = re.compile(r'href=["\']([^"\'#\s]{6,})["\']', re.I)
                    for href in link_pat.findall(html_text)[:80]:
                        if href.startswith('http'):
                            urls_to_check.append(href)
                        elif href.startswith('/'):
                            urls_to_check.append(base_url + href)
                    seen: set = set()
                    deduped = []
                    for u in urls_to_check:
                        if u not in seen:
                            seen.add(u)
                            deduped.append(u)
                    urls_to_check = deduped[:100]
                except Exception:
                    pass

            sb_body = {
                'client': {'clientId': 'cf-ai-dashboard', 'clientVersion': '1.0'},
                'threatInfo': {
                    'threatTypes': ['MALWARE', 'SOCIAL_ENGINEERING', 'UNWANTED_SOFTWARE',
                                    'POTENTIALLY_HARMFUL_APPLICATION'],
                    'platformTypes': ['ANY_PLATFORM'],
                    'threatEntryTypes': ['URL'],
                    'threatEntries': [{'url': u} for u in urls_to_check],
                },
            }
            r_sb, err_sb = _fetch_url(
                f'https://safebrowsing.googleapis.com/v4/threatMatches:find?key={api_key}',
                timeout=15, method='POST', json_body=sb_body)
            if r_sb is not None:
                sb_data = r_sb.json() if callable(getattr(r_sb, 'json', None)) else {}
                matches = sb_data.get('matches', [])
                threat_details = [
                    {'url': m.get('threat', {}).get('url', ''),
                     'type': m.get('threatType', ''),
                     'platform': m.get('platformType', '')}
                    for m in matches
                ]
                scores['safe_browsing'] = {
                    'checked': True, 'threats': len(matches),
                    'urls_checked': len(urls_to_check),
                    'threat_details': threat_details,
                }
                # Group by threat type
                threat_types: dict = {}
                for m in matches:
                    ttype = m.get('threatType', 'Unknown')
                    turl  = m.get('threat', {}).get('url', domain)
                    threat_types.setdefault(ttype, []).append(turl)
                for ttype, t_urls in threat_types.items():
                    is_link = any(
                        not u.startswith(base_url) and not u.startswith(f'http://{domain}')
                        for u in t_urls
                    )
                    title = (f'Safe Browsing: Links to {ttype.replace("_"," ").title()}'
                             if is_link else f'Safe Browsing: {ttype.replace("_"," ").title()} detected')
                    _add('Security', 'critical', title,
                         f'Google flagged {len(t_urls)} URL(s) for {ttype}:\n' +
                         '\n'.join(f'  - {u}' for u in t_urls[:5]),
                         'Chrome/Firefox/Safari show a full red warning page. GSC flags this as a Security Issue.',
                         ['Remove or fix the harmful links/content immediately.',
                          'Check Google Search Console Security Issues for the full list.',
                          'Scan with Sucuri SiteCheck (sitecheck.sucuri.net).',
                          'After cleanup, click "Request Review" in GSC Security Issues.'])
                if not matches:
                    _add('Security', 'info',
                         f'Safe Browsing: Clean ({len(urls_to_check)} URLs checked)',
                         f'No threats found across {len(urls_to_check)} URLs including all outgoing links.',
                         'Site and its outgoing links are clean per Google Safe Browsing.',
                         ['Monitor regularly — Safe Browsing status can change.',
                          'Enable Security Issue email alerts in Google Search Console.'])
            elif err_sb:
                scores['safe_browsing'] = {'checked': False, 'error': err_sb}
        except Exception as e:
            scores['safe_browsing'] = {'checked': False, 'error': str(e)}

    # ── 8. PageSpeed Insights + Core Web Vitals (CrUX field data) ────────────────
    if api_key:
        try:
            psi_url = (f'https://www.googleapis.com/pagespeedonline/v5/runPagespeed'
                       f'?url={_up_parse.quote(base_url, safe="")}&key={api_key}'
                       f'&category=performance&category=seo&category=best-practices'
                       f'&category=accessibility&strategy=mobile')
            r_psi, err_psi = _fetch_url(psi_url, timeout=35)
            if r_psi is not None:
                psi_data = r_psi.json() if callable(getattr(r_psi, 'json', None)) else {}
                cats = psi_data.get('lighthouseResult', {}).get('categories', {})
                psi_scores = {}
                for ck, cv in cats.items():
                    s = cv.get('score')
                    if s is not None:
                        psi_scores[ck] = round(s * 100)
                scores['pagespeed'] = psi_scores

                # ── Core Web Vitals from CrUX field data ──────────────────────
                cwv_raw = psi_data.get('loadingExperience', {}).get('metrics', {})
                if cwv_raw:
                    def _cwv_cat(cat): return {'FAST': 'good', 'AVERAGE': 'needs-improvement', 'SLOW': 'poor'}.get(cat, cat)
                    cwv = {}
                    lcp = cwv_raw.get('LARGEST_CONTENTFUL_PAINT_MS', {})
                    cls = cwv_raw.get('CUMULATIVE_LAYOUT_SHIFT_SCORE', {})
                    inp = cwv_raw.get('INTERACTION_TO_NEXT_PAINT', cwv_raw.get('FIRST_INPUT_DELAY_MS', {}))
                    fcp = cwv_raw.get('FIRST_CONTENTFUL_PAINT_MS', {})
                    if lcp:
                        lcp_ms = lcp.get('percentile', 0)
                        cwv['lcp'] = {'value': lcp_ms, 'unit': 'ms', 'category': _cwv_cat(lcp.get('category', ''))}
                        if lcp_ms > 4000:
                            _add('Performance', 'high', f'Core Web Vitals: LCP is poor ({lcp_ms/1000:.1f}s)',
                                 f'Largest Contentful Paint (real user data): {lcp_ms/1000:.1f}s. Google threshold: good < 2.5s.',
                                 'Poor LCP directly hurts Google search rankings via Core Web Vitals signal.',
                                 ['Optimize the largest element (hero image, heading, or block).',
                                  'Use next-gen image formats (WebP/AVIF) and lazy loading.',
                                  'Enable server-side caching and use a CDN.',
                                  'Eliminate render-blocking resources (defer JS, inline critical CSS).'])
                        elif lcp_ms > 2500:
                            _add('Performance', 'medium', f'Core Web Vitals: LCP needs improvement ({lcp_ms/1000:.1f}s)',
                                 f'LCP (real users): {lcp_ms/1000:.1f}s. Google threshold: good < 2.5s.',
                                 'Needs improvement LCP may affect search rankings.',
                                 ['Optimize largest page element and enable CDN caching.',
                                  'Run pagespeed.web.dev for specific recommendations.'])
                    if cls:
                        cls_val = cls.get('percentile', 0) / 100
                        cwv['cls'] = {'value': cls_val, 'unit': '', 'category': _cwv_cat(cls.get('category', ''))}
                        if cls_val > 0.25:
                            _add('Performance', 'high', f'Core Web Vitals: CLS is poor ({cls_val:.2f})',
                                 f'Cumulative Layout Shift (real users): {cls_val:.2f}. Good threshold: < 0.1.',
                                 'Layout shifts frustrate users and hurt Core Web Vitals rankings.',
                                 ['Reserve space for images and ads with explicit width/height attributes.',
                                  'Avoid inserting content above existing content after page load.',
                                  'Use font-display: swap carefully to avoid FOUT shifts.'])
                        elif cls_val > 0.1:
                            _add('Performance', 'medium', f'Core Web Vitals: CLS needs improvement ({cls_val:.2f})',
                                 f'CLS (real users): {cls_val:.2f}. Good threshold: < 0.1.',
                                 'Layout instability affects user experience and rankings.',
                                 ['Set explicit dimensions on images and embeds.',
                                  'Pre-load web fonts to reduce FOUT.'])
                    if inp:
                        inp_ms = inp.get('percentile', 0)
                        cwv['inp'] = {'value': inp_ms, 'unit': 'ms', 'category': _cwv_cat(inp.get('category', ''))}
                        if inp_ms > 500:
                            _add('Performance', 'high', f'Core Web Vitals: INP is poor ({inp_ms}ms)',
                                 f'Interaction to Next Paint (real users): {inp_ms}ms. Good threshold: < 200ms.',
                                 'Poor interactivity hurts Core Web Vitals and user engagement.',
                                 ['Reduce JavaScript execution time — split large tasks.',
                                  'Use web workers for heavy computations.',
                                  'Minimize third-party scripts (ads, chat widgets, analytics).'])
                        elif inp_ms > 200:
                            _add('Performance', 'medium', f'Core Web Vitals: INP needs improvement ({inp_ms}ms)',
                                 f'INP (real users): {inp_ms}ms. Good threshold: < 200ms.',
                                 'Sluggish interactions reduce user engagement.',
                                 ['Profile JS with Chrome DevTools Performance panel.',
                                  'Defer non-critical third-party scripts.'])
                    if fcp:
                        cwv['fcp'] = {'value': fcp.get('percentile', 0), 'unit': 'ms',
                                      'category': _cwv_cat(fcp.get('category', ''))}
                    scores['cwv'] = cwv
                    scores['cwv_overall'] = _cwv_cat(psi_data.get('loadingExperience', {}).get('overall_category', ''))

                cat_map = {'performance': 'Performance', 'seo': 'SEO',
                           'best-practices': 'Best Practices', 'accessibility': 'Accessibility'}
                for ck, clabel in cat_map.items():
                    sc = psi_scores.get(ck)
                    if sc is None or sc >= 90:
                        continue
                    sev = 'high' if sc < 50 else 'medium' if sc < 70 else 'low'
                    audits = psi_data.get('lighthouseResult', {}).get('audits', {})
                    refs   = cats.get(ck, {}).get('auditRefs', [])
                    failing = []
                    for ref in refs:
                        a = audits.get(ref.get('id', ''), {})
                        if a.get('score') is not None and a.get('score') < 0.9 and a.get('title'):
                            failing.append(a['title'])
                    failing_str = '; '.join(failing[:5]) or 'See PageSpeed Insights for details.'
                    cat_out = 'Performance' if ck == 'performance' else 'SEO' if ck == 'seo' else 'Best Practices'
                    impact_str = {
                        'performance': 'Low scores hurt Core Web Vitals rankings and user conversion.',
                        'seo':         'Google uses SEO score directly as a ranking factor.',
                        'best-practices': 'Best practice failures can introduce security and compatibility issues.',
                        'accessibility': 'Low accessibility may violate WCAG guidelines and legal requirements.',
                    }.get(ck, 'Affects user experience and search ranking.')
                    _add(cat_out, sev, f'PageSpeed {clabel}: {sc}/100',
                         f'Google PageSpeed (mobile) rated {clabel} at {sc}/100. Top issues: {failing_str}',
                         impact_str,
                         ['Run PageSpeed Insights at pagespeed.web.dev for full report.',
                          'Performance: compress images, enable caching, minify CSS/JS, use a CDN.',
                          'SEO: verify meta tags, structured data, and mobile-friendliness.',
                          'Accessibility: add alt text, increase contrast, improve keyboard navigation.'])
        except Exception as e:
            scores['pagespeed'] = {'error': str(e)}

    # ── 9. GSC URL Inspection ──────────────────────────────────────────────────
    if access_token:
        try:
            insp_body = {'inspectionUrl': base_url, 'siteUrl': f'https://{domain}/'}
            r_insp, err_insp = _fetch_url(
                'https://searchconsole.googleapis.com/v1/urlInspection/index:inspect',
                timeout=15, method='POST', json_body=insp_body,
                extra_headers={'Authorization': f'Bearer {access_token}'})
            if r_insp is not None:
                insp_data = r_insp.json() if callable(getattr(r_insp, 'json', None)) else {}
                res  = insp_data.get('inspectionResult', {})
                idx  = res.get('indexStatusResult', {})
                verdict  = idx.get('verdict', '')
                coverage = idx.get('coverageState', '')
                crawled  = idx.get('lastCrawlTime', '')[:10] if idx.get('lastCrawlTime') else 'never'
                scores['gsc'] = {'verdict': verdict, 'coverageState': coverage, 'lastCrawl': crawled}

                if verdict == 'FAIL':
                    _add('SEO', 'critical', f'GSC: URL not indexed ({coverage})',
                         f'URL Inspection verdict: FAIL. Coverage: {coverage}. Last crawled: {crawled}.',
                         'Page does not appear in Google search results at all.',
                         ['Check Coverage State in Search Console (Crawl anomaly, Redirect error, Soft 404).',
                          'If blocked by robots.txt: update to allow Googlebot.',
                          'If noindex tag: remove <meta name="robots" content="noindex">.',
                          'Use "Request Indexing" after fixing.'])
                elif verdict == 'NEUTRAL':
                    _add('SEO', 'medium', f'GSC: Indexing uncertain ({coverage})',
                         f'Verdict: NEUTRAL. Coverage: {coverage}.',
                         'Page may not be fully indexed or may have coverage issues.',
                         ['Review Coverage in Search Console.',
                          'Submit URL for indexing via Request Indexing.',
                          'Check crawl budget if the site is large.'])
                elif verdict == 'PASS':
                    _add('SEO', 'info', 'GSC: URL is indexed by Google',
                         f'Indexed. Last crawled: {crawled}. Coverage: {coverage}.',
                         'Page is in Google\'s index and can appear in results.',
                         ['Monitor Search Console Performance for clicks/impressions.',
                          'Watch for any coverage warnings over time.'])

                if res.get('ampResult', {}).get('verdict') == 'FAIL':
                    _add('SEO', 'medium', 'AMP page errors detected',
                         'AMP verdict: FAIL. AMP pages with errors are excluded from AMP treatment.',
                         'AMP pages won\'t get Google\'s mobile carousel or fast-loading badge.',
                         ['Review AMP errors in Search Console AMP report.',
                          'Validate at validator.ampproject.org.'])

                if res.get('richResultsResult', {}).get('verdict') == 'FAIL':
                    _add('SEO', 'medium', 'Rich results / structured data errors',
                         'Structured data errors detected. Rich results will not appear.',
                         'Products, reviews, FAQs won\'t show as rich snippets in search.',
                         ['Review Rich Results report in Search Console.',
                          'Validate at search.google.com/test/rich-results.',
                          'Fix JSON-LD or microdata markup issues.'])
            elif err_insp:
                scores['gsc'] = {'error': err_insp}
                _add('SEO', 'info', 'GSC URL Inspection could not run',
                     f'Error: {err_insp}',
                     'Indexing status unverified.',
                     ['Verify the access token is valid and not expired.',
                      'The site must be verified in Google Search Console.',
                      'Generate a token with "webmasters.readonly" scope via OAuth 2.0.'])
        except Exception as e:
            scores['gsc'] = {'error': str(e)}

    # ── 10. Search Analytics (queries, pages, devices, countries) ───────────────
    if access_token:
        import datetime as _dt2
        try:
            end_date   = _dt2.date.today()
            start_date = end_date - _dt2.timedelta(days=28)
            site_enc   = _up_parse.quote(f'{base_url}/', safe='')
            sa_base    = f'https://www.googleapis.com/webmasters/v3/sites/{site_enc}/searchAnalytics/query'
            sa_hdrs    = {'Authorization': f'Bearer {access_token}'}
            sa_dates   = {'startDate': str(start_date), 'endDate': str(end_date), 'type': 'web'}

            def _sa_query(dims, limit=10):
                r, _ = _fetch_url(sa_base, method='POST', timeout=15, extra_headers=sa_hdrs,
                                   json_body={**sa_dates, 'dimensions': dims, 'rowLimit': limit})
                if r is not None and r.status_code == 200:
                    return (r.json() if callable(getattr(r, 'json', None)) else {}).get('rows', [])
                if r is not None and r.status_code == 403:
                    raise PermissionError('Permission denied')
                return None

            # Top queries
            q_rows = _sa_query(['query'], 10)
            # Top pages
            p_rows = _sa_query(['page'], 10)
            # Device breakdown
            d_rows = _sa_query(['device'], 3)
            # Country breakdown
            c_rows = _sa_query(['country'], 5)

            top_pages = []
            if p_rows:
                for r in p_rows:
                    if r.get('keys'):
                        top_pages.append(r['keys'][0])

            scores['search_analytics'] = {
                'period': f'{start_date} to {end_date}',
                'top_queries': [
                    {'query': r['keys'][0], 'clicks': int(r.get('clicks', 0)),
                     'impressions': int(r.get('impressions', 0)),
                     'ctr': round(r.get('ctr', 0) * 100, 1),
                     'position': round(r.get('position', 0), 1)}
                    for r in (q_rows or []) if r.get('keys')
                ],
                'top_pages': [
                    {'page': r['keys'][0], 'clicks': int(r.get('clicks', 0)),
                     'impressions': int(r.get('impressions', 0)),
                     'ctr': round(r.get('ctr', 0) * 100, 1),
                     'position': round(r.get('position', 0), 1)}
                    for r in (p_rows or []) if r.get('keys')
                ],
                'devices': [
                    {'device': r['keys'][0], 'clicks': int(r.get('clicks', 0)),
                     'impressions': int(r.get('impressions', 0))}
                    for r in (d_rows or []) if r.get('keys')
                ],
                'countries': [
                    {'country': r['keys'][0].upper(), 'clicks': int(r.get('clicks', 0)),
                     'impressions': int(r.get('impressions', 0))}
                    for r in (c_rows or []) if r.get('keys')
                ],
            }

            # Flag low CTR queries
            for row in (q_rows or []):
                if row.get('impressions', 0) > 100 and row.get('ctr', 0) < 0.02:
                    q = row['keys'][0] if row.get('keys') else '(unknown)'
                    _add('SEO', 'medium', f'Low CTR on "{q}"',
                         f'"{q}" gets {int(row.get("impressions",0))} impressions but only '
                         f'{row.get("ctr",0)*100:.1f}% CTR (avg position: {row.get("position",0):.1f}).',
                         'Users see the result but don\'t click — title/description may be uncompelling.',
                         ['Rewrite the page meta title to be more specific and compelling.',
                          'Improve the meta description to state the value proposition clearly.',
                          'Add rich snippets (reviews, FAQ) to increase SERP real estate.',
                          'Check if search intent matches your page content.'])

            # ── Deep Safe Browsing: crawl top pages too ──────────────────────
            if api_key and top_pages:
                try:
                    extra_urls = []
                    for page_url in top_pages[:5]:
                        r_pg, _ = _fetch_url(page_url, timeout=8)
                        if r_pg is not None:
                            pg_html = getattr(r_pg, 'text', '')
                            link_pat2 = re.compile(r'href=["\']([^"\'#\s]{6,})["\']', re.I)
                            for href in link_pat2.findall(pg_html)[:30]:
                                if href.startswith('http'):
                                    extra_urls.append(href)
                    if extra_urls:
                        extra_urls = list(dict.fromkeys(extra_urls))[:100]
                        sb2_body = {
                            'client': {'clientId': 'cf-ai-dashboard', 'clientVersion': '1.0'},
                            'threatInfo': {
                                'threatTypes': ['MALWARE', 'SOCIAL_ENGINEERING',
                                                'UNWANTED_SOFTWARE', 'POTENTIALLY_HARMFUL_APPLICATION'],
                                'platformTypes': ['ANY_PLATFORM'],
                                'threatEntryTypes': ['URL'],
                                'threatEntries': [{'url': u} for u in extra_urls],
                            },
                        }
                        r_sb2, _ = _fetch_url(
                            f'https://safebrowsing.googleapis.com/v4/threatMatches:find?key={api_key}',
                            timeout=15, method='POST', json_body=sb2_body)
                        if r_sb2 is not None and r_sb2.status_code == 200:
                            sb2_data = r_sb2.json() if callable(getattr(r_sb2, 'json', None)) else {}
                            for m in sb2_data.get('matches', []):
                                ttype    = m.get('threatType', 'Unknown')
                                turl     = m.get('threat', {}).get('url', '')
                                existing = scores.get('safe_browsing', {})
                                existing['threats'] = existing.get('threats', 0) + 1
                                existing.setdefault('threat_details', []).append(
                                    {'url': turl, 'type': ttype, 'platform': m.get('platformType', ''),
                                     'found_on': 'internal page (deep scan)'})
                                scores['safe_browsing'] = existing
                                _add('Security', 'critical',
                                     f'Links to harmful downloads on internal page',
                                     f'Deep scan found harmful link on a top page: {turl}\nThreat: {ttype}',
                                     'This matches the GSC "Links to harmful downloads" security issue.',
                                     ['Find and remove this link from your site immediately.',
                                      'Check Google Search Console Security Issues for all affected pages.',
                                      'Scan with Sucuri SiteCheck (sitecheck.sucuri.net).',
                                      'After cleanup, click "Request Review" in GSC Security Issues.'])
                except Exception:
                    pass

        except PermissionError as pe:
            scores['search_analytics'] = {'error': str(pe)}
        except Exception as e:
            scores['search_analytics'] = {'error': str(e)}

    # ── 11. Sitemaps (via GSC API) ─────────────────────────────────────────────
    if access_token:
        try:
            site_enc  = _up_parse.quote(f'{base_url}/', safe='')
            r_sm, _   = _fetch_url(
                f'https://www.googleapis.com/webmasters/v3/sites/{site_enc}/sitemaps',
                timeout=10, extra_headers={'Authorization': f'Bearer {access_token}'},
            )
            if r_sm is not None and r_sm.status_code == 200:
                sitemaps = (r_sm.json() if callable(getattr(r_sm, 'json', None)) else {}).get('sitemap', [])
                scores['gsc_sitemaps'] = [
                    {'path': s.get('path', ''), 'lastSubmitted': s.get('lastSubmitted', ''),
                     'isPending': s.get('isPending', False), 'errors': s.get('errors', '0'),
                     'warnings': s.get('warnings', '0'),
                     'urlCount': sum(c.get('submitted', 0) for c in s.get('contents', []))}
                    for s in sitemaps
                ]
                for sm in sitemaps:
                    if str(sm.get('errors', '0')) != '0':
                        _add('SEO', 'high', f'Sitemap errors: {sm.get("path", "")}',
                             f'Sitemap has {sm.get("errors","?")} error(s) in Google Search Console.',
                             'Google cannot crawl URLs in this sitemap, reducing index coverage.',
                             ['Open Search Console → Sitemaps to see specific errors.',
                              'Common issues: unreachable URLs, redirect chains, noindex pages in sitemap.',
                              'Fix and resubmit.'])
                if not sitemaps:
                    _add('SEO', 'medium', 'No sitemaps submitted to Google Search Console',
                         'No sitemaps found in this GSC property.',
                         'Google must discover all pages via crawl links, likely missing content.',
                         [f'Submit your sitemap in Search Console → Sitemaps.',
                          f'Generate one at {base_url}/sitemap.xml if needed.',
                          'WordPress: Yoast SEO auto-generates and submits sitemaps.'])
        except Exception as e:
            scores['gsc_sitemaps'] = {'error': str(e)}

    # ── 12. VirusTotal domain reputation ─────────────────────────────────────────
    vt_key = os.environ.get('VIRUSTOTAL_API_KEY', '').strip()
    if vt_key:
        try:
            r_vt, _ = _fetch_url(
                f'https://www.virustotal.com/api/v3/domains/{domain}',
                timeout=12, extra_headers={'x-apikey': vt_key})
            if r_vt is not None and r_vt.status_code == 200:
                vt_data  = r_vt.json() if callable(getattr(r_vt, 'json', None)) else {}
                attrs    = vt_data.get('data', {}).get('attributes', {})
                analysis = attrs.get('last_analysis_stats', {})
                malicious   = analysis.get('malicious', 0)
                suspicious  = analysis.get('suspicious', 0)
                harmless    = analysis.get('harmless', 0)
                undetected  = analysis.get('undetected', 0)
                total       = malicious + suspicious + harmless + undetected
                reputation  = attrs.get('reputation', 0)
                categories  = attrs.get('categories', {})

                scores['virustotal'] = {
                    'malicious': malicious, 'suspicious': suspicious,
                    'harmless': harmless, 'total_engines': total,
                    'reputation': reputation,
                    'categories': list(set(categories.values()))[:5],
                }

                if malicious > 0:
                    engines = [k for k, v in attrs.get('last_analysis_results', {}).items()
                               if v.get('category') in ('malicious',)][:8]
                    _add('Security', 'critical', f'VirusTotal: {malicious}/{total} engines flag as malicious',
                         f'{malicious} security vendors flag {domain} as malicious. Engines: {", ".join(engines) or "see VT report"}.',
                         'Visitors using security-aware browsers or AV software will be blocked from the site.',
                         ['Check full report at virustotal.com/gui/domain/' + domain,
                          'Scan site for malware at sitecheck.sucuri.net.',
                          'Check for injected code: recently modified files in cPanel File Manager.',
                          'Contact hosting provider for malware scan assistance.',
                          'After cleaning, request review at each flagging vendor.'])
                elif suspicious > 0:
                    _add('Security', 'medium', f'VirusTotal: {suspicious}/{total} engines flag as suspicious',
                         f'{suspicious} vendors mark {domain} as suspicious.',
                         'May trigger warnings in some security tools and corporate firewalls.',
                         ['Review full VirusTotal report: virustotal.com/gui/domain/' + domain,
                          'Check for any recently added third-party scripts or ads.',
                          'Ensure no questionable affiliate links are present.'])
                else:
                    _add('Security', 'info',
                         f'VirusTotal: Clean ({harmless} engines confirm safe)',
                         f'{harmless}/{total} engines mark {domain} as safe. Reputation score: {reputation}.',
                         'Domain has a clean reputation across major security vendors.',
                         ['Continue monitoring at virustotal.com/gui/domain/' + domain])
        except Exception as e:
            scores['virustotal'] = {'error': str(e)}

    # ── 13. Suspicious Link Scanner ───────────────────────────────────────────────
    # Collect ALL links found during crawl and apply heuristic + SB checks
    try:
        _SUSPICIOUS_EXTS = {'.exe','.msi','.bat','.cmd','.ps1','.vbs','.jar',
                            '.scr','.pif','.com','.hta','.apk','.dmg','.pkg',
                            '.deb','.rpm','.iso','.img','.torrent'}
        _SUSPICIOUS_TLDS = {'.tk','.ml','.ga','.cf','.gq','.pw','.top','.xyz',
                            '.click','.download','.zip','.mov'}
        _URL_SHORTENERS  = {'bit.ly','tinyurl.com','t.co','goo.gl','ow.ly',
                            'is.gd','buff.ly','adf.ly','bc.vc','sh.st'}
        _IP_PAT = re.compile(r'https?://(\d{1,3}\.){3}\d{1,3}')

        # Gather all links from all crawled pages
        all_links: list = []
        pages_to_scan = [base_url] + (list(scores.get('search_analytics', {})
                         .get('top_pages', [{}]))[:4] if access_token else [])

        for scan_url in pages_to_scan:
            pg_url = scan_url if isinstance(scan_url, str) else scan_url.get('page', '')
            if not pg_url:
                continue
            r_pg2, _ = _fetch_url(pg_url, timeout=8)
            if r_pg2 is None:
                continue
            pg_html2 = getattr(r_pg2, 'text', '')
            lp2 = re.compile(r'href=["\']([^"\'#\s]{8,})["\']', re.I)
            for href in lp2.findall(pg_html2):
                if href.startswith('http') and domain not in href:
                    all_links.append({'url': href, 'found_on': pg_url})

        # Deduplicate by URL
        seen_urls: set = set()
        unique_links = []
        for lk in all_links:
            if lk['url'] not in seen_urls:
                seen_urls.add(lk['url'])
                unique_links.append(lk)

        suspicious_found: list = []

        for lk in unique_links:
            u = lk['url']
            reasons = []
            parsed_u = _up_parse.urlparse(u)
            ext = '.' + u.split('.')[-1].split('?')[0].lower() if '.' in u.split('/')[-1] else ''
            link_domain = parsed_u.netloc.lower().replace('www.', '')
            tld = '.' + link_domain.split('.')[-1] if '.' in link_domain else ''

            if ext in _SUSPICIOUS_EXTS:
                reasons.append(f'direct download link ({ext})')
            if _IP_PAT.match(u):
                reasons.append('links to IP address instead of domain')
            if link_domain in _URL_SHORTENERS:
                reasons.append(f'URL shortener ({link_domain}) — destination unknown')
            if tld in _SUSPICIOUS_TLDS:
                reasons.append(f'suspicious TLD ({tld}) — commonly used for malware')

            if reasons:
                suspicious_found.append({
                    'url': u, 'found_on': lk['found_on'], 'reasons': reasons
                })

        # Safe Browsing check on all unique external links
        if api_key and unique_links:
            sb3_entries = [{'url': lk['url']} for lk in unique_links[:200]]
            sb3_body = {
                'client': {'clientId': 'cf-ai-dashboard', 'clientVersion': '1.0'},
                'threatInfo': {
                    'threatTypes': ['MALWARE','SOCIAL_ENGINEERING',
                                    'UNWANTED_SOFTWARE','POTENTIALLY_HARMFUL_APPLICATION'],
                    'platformTypes': ['ANY_PLATFORM'],
                    'threatEntryTypes': ['URL'],
                    'threatEntries': sb3_entries,
                },
            }
            r_sb3, _ = _fetch_url(
                f'https://safebrowsing.googleapis.com/v4/threatMatches:find?key={api_key}',
                timeout=15, method='POST', json_body=sb3_body)
            if r_sb3 is not None and r_sb3.status_code == 200:
                sb3_data = r_sb3.json() if callable(getattr(r_sb3, 'json', None)) else {}
                for m in sb3_data.get('matches', []):
                    turl3  = m.get('threat', {}).get('url', '')
                    ttype3 = m.get('threatType', 'Unknown')
                    # Find which page it was on
                    found_on3 = next((lk['found_on'] for lk in unique_links if lk['url'] == turl3), base_url)
                    suspicious_found.append({
                        'url': turl3, 'found_on': found_on3,
                        'reasons': [f'Google Safe Browsing: {ttype3}'],
                        'google_flagged': True,
                    })

        scores['suspicious_links'] = {
            'total': len(suspicious_found),
            'links': suspicious_found[:50],
            'pages_scanned': len(pages_to_scan),
            'links_checked': len(unique_links),
        }

        # Add issues for each suspicious link
        for sl in suspicious_found[:10]:
            is_google = sl.get('google_flagged', False)
            sev = 'critical' if is_google else ('high' if any('download' in r or 'Safe Browsing' in r for r in sl['reasons']) else 'medium')
            _add('Security',
                 sev,
                 ('Google Safe Browsing: Harmful link detected' if is_google
                  else f'Suspicious outbound link: {", ".join(sl["reasons"])}'),
                 f'URL: {sl["url"]}\nFound on: {sl["found_on"]}\nReason: {"; ".join(sl["reasons"])}',
                 'Malicious or suspicious outbound links can trigger GSC Security Issues warnings, '
                 'Google Safe Browsing browser warnings, and damage site reputation.',
                 ['Remove or replace this link immediately.',
                  'If you didn\'t add this link, your site may be hacked — scan with Sucuri SiteCheck.',
                  'Check recently modified files in cPanel File Manager.',
                  'After cleanup, request review in Google Search Console Security Issues.'])

        if suspicious_found:
            # Update safe_browsing tile count
            sb_existing = scores.get('safe_browsing', {})
            sb_existing['threats'] = sb_existing.get('threats', 0) + sum(
                1 for s in suspicious_found if s.get('google_flagged'))
            scores['safe_browsing'] = sb_existing

    except Exception:
        pass

    # ── 14. VirusTotal URL scan (antivirus / malware scan) ────────────────────
    # Uses cached VT analysis — works even when Cloudflare blocks external crawlers
    try:
        import base64 as _b64
        vt_key = os.environ.get('VIRUSTOTAL_API_KEY', '')
        if vt_key:
            url_b64 = _b64.urlsafe_b64encode(base_url.encode()).decode().rstrip('=')
            r_vtu, _ = _fetch_url(
                f'https://www.virustotal.com/api/v3/urls/{url_b64}',
                timeout=15,
                extra_headers={'x-apikey': vt_key})
            if r_vtu is None or r_vtu.status_code == 404:
                # No cached result — submit URL for fresh scan then read it
                r_sub, _ = _fetch_url(
                    'https://www.virustotal.com/api/v3/urls',
                    timeout=15, method='POST',
                    extra_headers={'x-apikey': vt_key},
                    form_data={'url': base_url})
                if r_sub is not None and r_sub.status_code == 200:
                    sub_data = r_sub.json() if callable(getattr(r_sub, 'json', None)) else {}
                    analysis_id = sub_data.get('data', {}).get('id', '')
                    if analysis_id:
                        import time as _time
                        _time.sleep(5)
                        r_vtu, _ = _fetch_url(
                            f'https://www.virustotal.com/api/v3/analyses/{analysis_id}',
                            timeout=15,
                            extra_headers={'x-apikey': vt_key})
            if r_vtu is not None and r_vtu.status_code == 200:
                vtu_data = r_vtu.json() if callable(getattr(r_vtu, 'json', None)) else {}
                attrs = vtu_data.get('data', {}).get('attributes', {})
                # analyses endpoint returns stats directly; urls endpoint nests under last_analysis_stats
                stats = attrs.get('last_analysis_stats') or attrs.get('stats', {})
                results = attrs.get('last_analysis_results') or attrs.get('results', {})
                mal_count  = stats.get('malicious', 0)
                sus_count  = stats.get('suspicious', 0)
                clean_count= stats.get('harmless', 0) + stats.get('undetected', 0)
                total_eng  = sum(stats.values()) if stats else 0
                mal_engines= [e for e, v in results.items()
                               if isinstance(v, dict) and v.get('category') in ('malicious','suspicious')]
                scores['url_scan'] = {
                    'malicious':    mal_count,
                    'suspicious':   sus_count,
                    'clean':        clean_count,
                    'total_engines':total_eng,
                    'flagged_by':   mal_engines[:10],
                    'url':          base_url,
                }
                if mal_count > 0:
                    _add('Security', 'critical',
                         f'URL flagged as malicious by {mal_count} engine(s): {", ".join(mal_engines[:3])}',
                         f'VirusTotal URL scan: {mal_count} security engine(s) flagged {base_url} as malicious.\n'
                         f'Engines: {", ".join(mal_engines[:5])}',
                         'Malicious URL flagging causes browsers and antivirus tools to block your site visitors '
                         'and triggers Google Safe Browsing warnings.',
                         ['Check cPanel File Manager for recently modified PHP files.',
                          'Scan with a server-side tool (Wordfence, Sucuri plugin).',
                          'Request a re-analysis at virustotal.com after cleanup.',
                          'Submit for Google review in GSC → Security Issues.'])
                elif sus_count > 0:
                    _add('Security', 'high',
                         f'URL flagged as suspicious by {sus_count} engine(s)',
                         f'VirusTotal URL scan: {sus_count} engine(s) flagged {base_url} as suspicious.',
                         'Suspicious flagging may warn visitors and reduce trust.',
                         ['Review flagged engines at virustotal.com.',
                          'Check for injected scripts or spammy content.',
                          'Request re-analysis after any cleanup.'])
                else:
                    _add('Security', 'info',
                         f'URL scan: Clean ({clean_count}/{total_eng} engines confirm safe)',
                         f'VirusTotal URL scan found no malicious or suspicious content at {base_url}.',
                         'Site URL has a clean reputation across major security engines.',
                         ['Re-scan periodically at virustotal.com/gui/home/url.'])
            else:
                scores['url_scan'] = {'error': 'VirusTotal URL scan unavailable'}
        else:
            scores['url_scan'] = {'error': 'No VirusTotal API key configured'}
    except Exception as e:
        scores['url_scan'] = {'error': str(e)}

    # ── 15. URLScan.io (free API key, deep URL inspection) ───────────────────
    try:
        import time as _time
        urlscan_key = os.environ.get('URLSCAN_API_KEY', '')
        us_headers = {'Content-Type': 'application/json'}
        if urlscan_key:
            us_headers['API-Key'] = urlscan_key
        urlscan_body = {'url': base_url, 'visibility': 'public'}
        r_us, us_err = _fetch_url(
            'https://urlscan.io/api/v1/scan/',
            timeout=20, method='POST', json_body=urlscan_body,
            extra_headers=us_headers)
        if r_us is not None and r_us.status_code in (200, 201):
            us_submit = r_us.json() if callable(getattr(r_us, 'json', None)) else {}
            us_uuid = us_submit.get('uuid', '')
            if us_uuid:
                # Poll up to 4 times with 10-second gaps
                us_result = {}
                for _ in range(4):
                    _time.sleep(10)
                    r_res, _ = _fetch_url(
                        f'https://urlscan.io/api/v1/result/{us_uuid}/',
                        timeout=12)
                    if r_res is not None and r_res.status_code == 200:
                        us_result = r_res.json() if callable(getattr(r_res, 'json', None)) else {}
                        break
                if us_result:
                    verdicts  = us_result.get('verdicts', {})
                    overall   = verdicts.get('overall', {})
                    malicious = overall.get('malicious', False)
                    score     = overall.get('score', 0)
                    categories= overall.get('categories', [])
                    brands    = overall.get('brands', [])
                    page_info = us_result.get('page', {})
                    ips_list  = list(us_result.get('lists', {}).get('ips', []))[:10]
                    urls_list = list(us_result.get('lists', {}).get('urls', []))[:10]
                    links_list= list(us_result.get('lists', {}).get('linkDomains', []))[:10]
                    screenshot= us_result.get('task', {}).get('screenshotURL', '')
                    scores['urlscanio'] = {
                        'malicious':    malicious,
                        'score':        score,
                        'categories':   categories,
                        'brands':       brands,
                        'page':         page_info,
                        'ips':          ips_list,
                        'external_urls':urls_list,
                        'link_domains': links_list,
                        'screenshot':   screenshot,
                        'result_url':   f'https://urlscan.io/result/{us_uuid}/',
                        'uuid':         us_uuid,
                    }
                    if malicious:
                        _add('Security', 'critical',
                             f'URLScan.io: Site flagged as malicious (score {score})',
                             f'URLScan.io analysis of {base_url} returned a malicious verdict.\n'
                             f'Categories: {", ".join(categories)}\nBrands targeted: {", ".join(str(b) for b in brands)}',
                             'A malicious verdict from URLScan.io indicates phishing, malware distribution, '
                             'or brand impersonation on your site.',
                             ['Review the full URLScan report for specific findings.',
                              'Check cPanel File Manager for injected scripts.',
                              'Scan WordPress with Wordfence or Sucuri plugin.',
                              'Request Google review in GSC → Security Issues after cleanup.'])
                    elif score > 0:
                        _add('Security', 'medium',
                             f'URLScan.io: Suspicious indicators detected (score {score})',
                             f'URLScan.io found suspicious signals at {base_url}.',
                             'Suspicious indicators may indicate compromised content or risky scripts.',
                             ['Review the full URLScan report.',
                              'Audit recently added third-party scripts and ads.'])
                    else:
                        _add('Security', 'info',
                             'URLScan.io: No threats detected',
                             f'URLScan.io deep inspection found no malicious content at {base_url}.',
                             'Site passed URLScan.io analysis with a clean verdict.',
                             [f'Full report: urlscan.io/result/{us_uuid}/'])
                else:
                    scores['urlscanio'] = {'pending': True, 'uuid': us_uuid,
                                           'result_url': f'https://urlscan.io/result/{us_uuid}/'}
            else:
                err_msg = us_submit.get('message', 'No scan UUID returned')
                scores['urlscanio'] = {'error': err_msg}
        elif r_us is not None and r_us.status_code == 429:
            scores['urlscanio'] = {'error': 'Rate limited — try again in a few minutes'}
        elif r_us is not None and r_us.status_code == 401:
            scores['urlscanio'] = {'error': 'Invalid or missing URLSCAN_API_KEY'}
        elif r_us is not None and r_us.status_code == 400:
            body = {}
            try:
                body = r_us.json() if callable(getattr(r_us, 'json', None)) else {}
            except Exception:
                pass
            scores['urlscanio'] = {'error': f"Bad request: {body.get('message', r_us.status_code)}"}
        elif r_us is None:
            scores['urlscanio'] = {'error': f'Connection failed: {us_err}'}
        else:
            scores['urlscanio'] = {'error': f'HTTP {r_us.status_code}'}
    except Exception as e:
        scores['urlscanio'] = {'error': str(e)}

    # ── Sort and return ────────────────────────────────────────────────────────
    sev_order = {'critical': 0, 'high': 1, 'medium': 2, 'low': 3, 'info': 4}
    issues.sort(key=lambda x: sev_order.get(x['severity'], 5))

    return jsonify({
        'domain':   domain,
        'base_url': base_url,
        'issues':   issues,
        'scores':   scores,
        'total':    len(issues),
        'summary': {
            'critical': sum(1 for i in issues if i['severity'] == 'critical'),
            'high':     sum(1 for i in issues if i['severity'] == 'high'),
            'medium':   sum(1 for i in issues if i['severity'] == 'medium'),
            'low':      sum(1 for i in issues if i['severity'] == 'low'),
            'info':     sum(1 for i in issues if i['severity'] == 'info'),
        },
    })


# ── WordPress official plugin registry ───────────────────────────────────────
# Curated list of well-known, officially distributed WordPress.org plugin slugs.
# Anything NOT here gets verified via the WP.org Plugins API at scan time.
_WP_OFFICIAL_PLUGINS = frozenset({
    'akismet', 'jetpack', 'woocommerce', 'contact-form-7', 'wordfence',
    'wordpress-seo', 'yoast-seo', 'elementor', 'classic-editor',
    'really-simple-ssl', 'wpforms-lite', 'all-in-one-wp-migration',
    'wp-super-cache', 'wp-optimize', 'all-in-one-seo-pack', 'updraftplus',
    'google-site-kit', 'google-analytics-for-wordpress', 'sucuri-scanner',
    'ithemes-security', 'better-wp-security', 'all-in-one-wp-security-and-firewall',
    'wp-mail-smtp', 'mailchimp-for-wp', 'wps-hide-login', 'login-lockdown',
    'limit-login-attempts-reloaded', 'two-factor', 'wp-2fa',
    'w3-total-cache', 'wp-fastest-cache', 'litespeed-cache', 'autoptimize',
    'smush', 'ewww-image-optimizer', 'imagify', 'shortpixel-image-optimiser',
    'advanced-custom-fields', 'pods', 'meta-box', 'custom-post-type-ui',
    'bbpress', 'buddypress', 'learnpress', 'lifterlms', 'learndash',
    'wpforms', 'ninja-forms', 'gravity-forms', 'formidable', 'caldera-forms',
    'everest-forms', 'happyforms', 'fluentform',
    'woocommerce-payments', 'woocommerce-gateway-stripe',
    'woocommerce-gateway-paypal-powered-by-braintree',
    'woo-discount-rules', 'yith-woocommerce-wishlist',
    'easy-digital-downloads', 'download-monitor',
    'cookie-notice', 'gdpr-cookie-compliance', 'complianz', 'cookieyes',
    'redirection', 'safe-redirect-manager',
    'duplicate-post', 'user-role-editor', 'members',
    'tablepress', 'ninja-tables',
    'events-manager', 'the-events-calendar',
    'polylang', 'loco-translate', 'say-what',
    'regenerate-thumbnails', 'simple-image-sizes',
    'simple-history', 'wp-activity-log',
    'broken-link-checker', 'maintenance', 'under-construction-page',
    'wp-statistics', 'monsterinsights', 'exactmetrics',
    'wp-seopress', 'squirrly-seo', 'the-seo-framework', 'schema-and-structured-data-for-wp',
    'popup-maker', 'icegram', 'optin-monster', 'mailpoet', 'convertkit',
    'backwpup', 'duplicator', 'all-in-one-wp-migration',
    'nextgen-gallery', 'envira-gallery', 'modula', 'media-library-assistant',
    'custom-css-js', 'simple-custom-css', 'insert-headers-and-footers',
    'header-footer-code-manager',
    'miniorange-login-openid', 'nextend-social-login', 'social-login',
    'wp-google-maps', 'leaflet-maps-marker', 'maps-marker-pro',
    'pretty-links', 'thirstyaffiliates',
    'disable-comments', 'classic-widgets', 'tinymce-advanced',
    'gutenberg', 'blocksy-companion', 'kadence-blocks',
    'essential-addons-for-elementor', 'elementor-pro',
    'generatepress', 'astra', 'oceanwp', 'hello-elementor', 'storefront',
    'smart-slider-3', 'master-slider', 'ml-slider',
    'wpcf7-redirect', 'cf7-to-zapier', 'flamingo',
    'wp-job-manager', 'wp-remote', 'wp-webhooks',
    'cloudflare', 'aryo-activity-log', 'stream',
    'hide-my-wp', 'wp-hide-login', 'change-wp-admin-login',
    'anti-spam', 'cleantalk', 'wp-cerber',
    'restrict-content-pro', 'paid-memberships-pro', 'memberpress',
    'wpallexport', 'wp-all-import',
    'translatepress', 'wpml',
    'enable-media-replace', 'add-from-server',
    'co-authors-plus', 'post-types-order',
    'multisite-toolbar-additions', 'adminimize',
    'check-email', 'wp-mail-logging',
    'disable-block-editor', 'smart-custom-404-error-page',
    'block-bad-queries', 'wp-htaccess-editor',
    'tinymce-advanced', 'rich-text-tags',
    'woo-variation-swatches', 'variation-swatches-for-woocommerce',
    'woocommerce-multilingual', 'booster-for-woocommerce',
    'wp-rollback', 'enable-jquery-migrate-helper',
    'hummingbird-performance', 'wp-rocket',
    'imagify', 'shortpixel-adaptive-images',
    'bbpress-improved-notifications', 'bp-better-messages',
    'mailster', 'newsletter', 'wp-newsletter',
    'post-expirator', 'admin-menu-editor',
    'searchie', 'searchwp', 'relevanssi',
    'wptouch', 'amp', 'official-facebook-pixel',
    'twitter-for-websites', 'instagram-feed', 'smash-balloon-social-post-feed',
    'wpml-string-translation', 'wpml-media-translation',
    'woocommerce-subscriptions', 'woocommerce-memberships',
    'woocommerce-bookings', 'woocommerce-product-bundles',
    'image-widget', 'image-map-pro', 'interactive-geo-maps',
    'woocommerce-wishlists', 'woocommerce-shipping', 'woocommerce-tax',
    'shortcode-ultimate', 'so-widgets-bundle', 'page-builder-by-siteorigin',
    'beaver-builder-lite-version', 'visual-composer-website-builder',
    'fusion-builder', 'divi-builder', 'themify-builder',
    'wp-fastest-cache', 'cache-enabler',
    'a3-lazy-load', 'lazy-load', 'bj-lazy-load',
    'wordfence-login-security', 'wp-bruiser',
    'real-time-find-and-replace', 'custom-field-suite',
    'types', 'toolset-blocks', 'wpadverts',
    'wp-postratings', 'kk-star-ratings',
    'widget-logic', 'if-menu', 'wp-show-posts',
    'featured-image-from-url', 'auto-upload-images',
    'multi-step-form', 'step-by-step',
    'paypal-for-woocommerce', 'stripe-payments',
    'currency-switcher-woocommerce', 'currency-switcher',
    'otter-blocks', 'stackable-ultimate-gutenberg-blocks',
    'spectra', 'uagb', 'getwid',
    'wp-staging', 'wp-clone',
    'duplicator-pro', 'instawp-connect',
    'filebird', 'real-media-library', 'wp-media-folder',
    'woocommerce-product-reviews-pro', 'yotpo-social-reviews-for-woocommerce',
    'click-to-chat-for-whatsapp', 'wati-io',
    'facebook-for-woocommerce', 'pinterest-for-woocommerce',
    'google-listings-and-ads', 'mailchimp-for-woocommerce',
})

# Suspicious plugin naming heuristics: patterns that suggest unofficial / malicious plugins
_SUSP_SLUG_RE = re.compile(
    r'^[a-z0-9]{2,5}$'                  # Very short generic slug (2-5 chars)
    r'|[0-9]{4,}'                        # 4+ consecutive digits
    r'|(?:update|patch|fix|security|loader|bootloader|wp-plugin)-[a-z0-9]{8,}'  # Generic + random suffix
    r'|\b(?:shell|cmd|webshell|backdoor|hack|exploit|inject|bypass|rootkit)\b'  # Explicit bad words
    , re.I)


def _check_wp_plugin_official(slug: str) -> dict:
    """Determine if a WordPress plugin slug is from WordPress.org.

    Strategy (in order):
    1. Known-good curated list → 'official' (instant, no network)
    2. Suspicious naming heuristics → 'suspicious' (instant)
    3. WordPress.org Plugins API → 'official' or 'unofficial' (network, 5s timeout)
    4. Fallback → 'unknown'

    Returns dict: {status, confidence, active_installs, description}
    """
    s = (slug or '').lower().strip()
    result: dict = {'slug': slug, 'status': 'unknown', 'confidence': 0.5,
                    'active_installs': None, 'wp_org_name': None}

    if not s:
        result['status'] = 'unknown'
        return result

    # 1. Known-good list
    if s in _WP_OFFICIAL_PLUGINS:
        result.update(status='official', confidence=1.0)
        return result

    # 2. Suspicious heuristics — fast check before network call
    if _SUSP_SLUG_RE.search(s):
        result.update(status='suspicious', confidence=0.85)
        return result

    # 3. WordPress.org Plugins API (real-time)
    try:
        import urllib.request as _ur2
        import json as _js2
        api_url = (
            'https://api.wordpress.org/plugins/info/1.2/'
            f'?action=plugin_information'
            f'&request[slug]={s}'
            '&fields[active_installs]=1'
            '&fields[short_description]=1'
        )
        req2 = _ur2.Request(api_url, headers={'User-Agent': 'CF_AI-Scanner/1.0'})
        with _ur2.urlopen(req2, timeout=5) as resp2:
            data = _js2.loads(resp2.read(16384).decode('utf-8', errors='replace'))
        if isinstance(data, dict) and data.get('slug', '').lower() == s:
            result.update(
                status='official',
                confidence=0.95,
                active_installs=data.get('active_installs'),
                wp_org_name=data.get('name'),
            )
        else:
            # API returned false or a different slug — not on WordPress.org
            result.update(status='unofficial', confidence=0.8)
    except Exception:
        result['status'] = 'unknown'

    return result


# ── PHP backdoor / webshell detection patterns ────────────────────────────────
_BACKDOOR_PATTERNS = [
    (re.compile(r'eval\s*\(\s*base64_decode\b', re.I),
     'eval(base64_decode()) — classic PHP webshell obfuscation', 'critical'),
    (re.compile(r'eval\s*\(\s*\$_(?:POST|GET|REQUEST|COOKIE)\b', re.I),
     'eval() with direct user input — arbitrary code execution', 'critical'),
    (re.compile(r'preg_replace\s*\([^,]{0,60}/e["\']?\s*,', re.I),
     'preg_replace /e modifier — deprecated code execution backdoor', 'critical'),
    (re.compile(r'assert\s*\(\s*\$_(?:POST|GET|REQUEST|COOKIE)\b', re.I),
     'assert() with user input — code execution backdoor', 'critical'),
    (re.compile(r'(?:system|passthru|shell_exec|popen|proc_open)\s*\(\s*\$_(?:POST|GET|REQUEST|COOKIE)\b', re.I),
     'OS command execution via user-controlled input', 'critical'),
    (re.compile(r'create_function\s*\(\s*["\']', re.I),
     'create_function() — deprecated PHP code injection vector', 'high'),
    (re.compile(r'base64_decode\s*\(\s*["\'][A-Za-z0-9+/]{80,}={0,2}["\']', re.I),
     'Large base64-encoded payload in source — likely obfuscation', 'high'),
    (re.compile(r'\$\{\s*["\']_[A-Z]+["\']\s*\}', re.I),
     'Variable-variable with superglobal name — obfuscation technique', 'high'),
    (re.compile(r'<\?php\s{0,10}@?eval\s*\(', re.I),
     'PHP file starts with eval() — classic webshell signature', 'critical'),
    (re.compile(r'(?:chr\([0-9]+\)\.){5,}', re.I),
     'chr()-concatenation obfuscation — encoded payload', 'high'),
    (re.compile(r'gzinflate\s*\(', re.I),
     'gzinflate() — often used to decompress hidden payloads', 'medium'),
    (re.compile(r'str_rot13\s*\(', re.I),
     'str_rot13() obfuscation detected', 'medium'),
]


def _scan_content_for_backdoors(content: str, path: str) -> list[dict]:
    """Return list of backdoor findings in fetched file content."""
    hits = []
    for pattern, desc, severity in _BACKDOOR_PATTERNS:
        if pattern.search(content):
            hits.append({'path': path, 'finding': desc, 'severity': severity})
            if severity == 'critical':
                break  # one critical is enough per file
    return hits


@app.route('/api/wp/filescan', methods=['POST'])
def api_wp_filescan():
    """Scan WordPress site for malware, suspicious files, and recent modifications.
    Reads from Wordfence DB tables + WordPress REST API. No extra plugin needed."""
    try:
        data = request.get_json(silent=True) or {}
        site_url = (data.get('site_url') or '').strip().rstrip('/')
        if not site_url:
            return jsonify({'error': 'site_url is required'}), 400
        if not site_url.startswith('http'):
            site_url = 'https://' + site_url

        db_host  = os.environ.get('WP_DB_HOST', '')
        db_port  = int(os.environ.get('WP_DB_PORT', 3306))
        db_name  = os.environ.get('WP_DB_NAME', '')
        db_user  = os.environ.get('WP_DB_USER', '')
        db_pass  = os.environ.get('WP_DB_PASS', '')
        db_pfx   = os.environ.get('WP_DB_PREFIX', 'wp_')

        results = {
            'site_url':        site_url,
            'wordfence_issues': [],
            'modified_files':   [],
            'upload_php_files': [],
            'plugins':          [],
            'last_scan':        None,
            'db_connected':     False,
            'summary': {'critical': 0, 'high': 0, 'medium': 0, 'info': 0},
        }

        # ── MySQL / Wordfence ─────────────────────────────────────────────────
        if db_host and db_name and db_user:
            try:
                import pymysql, pymysql.cursors
                conn = pymysql.connect(
                    host=db_host, port=db_port, user=db_user, password=db_pass,
                    database=db_name, charset='utf8mb4', connect_timeout=10,
                    cursorclass=pymysql.cursors.DictCursor)
                results['db_connected'] = True

                with conn.cursor() as cur:
                    # ── Wordfence issues (malware, suspicious files) ──────────
                    wf_issues_tbl = f'{db_pfx}wfIssues'
                    try:
                        cur.execute(
                            f"SELECT type, severity, shortMsg, longMsg, data, lastUpdated "
                            f"FROM {wf_issues_tbl} WHERE status != 'deleted' "
                            f"ORDER BY severity DESC LIMIT 100")
                        for row in cur.fetchall():
                            sev_num = int(row.get('severity') or 0)
                            sev = 'critical' if sev_num >= 100 else 'high' if sev_num >= 50 else 'medium' if sev_num >= 10 else 'info'
                            results['summary'][sev] = results['summary'].get(sev, 0) + 1
                            results['wordfence_issues'].append({
                                'type':     row.get('type', ''),
                                'severity': sev,
                                'short':    row.get('shortMsg', ''),
                                'detail':   row.get('longMsg', ''),
                                'updated':  str(row.get('lastUpdated', '')),
                            })
                    except Exception:
                        pass

                    # ── Wordfence file modifications ──────────────────────────
                    wf_filemods_tbl = f'{db_pfx}wfFileMods'
                    try:
                        cur.execute(
                            f"SELECT filename, filenameMD5, oldMD5, newMD5, isCoreFile "
                            f"FROM {wf_filemods_tbl} "
                            f"ORDER BY isCoreFile DESC LIMIT 200")
                        for row in cur.fetchall():
                            fname = row.get('filename', '')
                            is_core = bool(row.get('isCoreFile'))
                            results['modified_files'].append({
                                'file':     fname,
                                'is_core':  is_core,
                                'changed':  row.get('oldMD5') != row.get('newMD5'),
                            })
                            if is_core:
                                results['summary']['high'] = results['summary'].get('high', 0) + 1
                    except Exception:
                        pass

                    # ── PHP files in uploads (via Wordfence scan data) ────────
                    wf_scanner_tbl = f'{db_pfx}wfScanners'
                    try:
                        cur.execute(
                            f"SELECT filename FROM {wf_scanner_tbl} "
                            f"WHERE filename LIKE '%/uploads/%.php' LIMIT 50")
                        for row in cur.fetchall():
                            results['upload_php_files'].append(row.get('filename', ''))
                            results['summary']['critical'] = results['summary'].get('critical', 0) + 1
                    except Exception:
                        pass

                    # ── Last Wordfence scan time ──────────────────────────────
                    try:
                        opts_tbl = f'{db_pfx}options'
                        cur.execute(
                            f"SELECT option_value FROM {opts_tbl} "
                            f"WHERE option_name = 'wordfence_lastScanCompleted' LIMIT 1")
                        row = cur.fetchone()
                        if row and row.get('option_value'):
                            import datetime as _dt
                            ts = int(row['option_value'])
                            results['last_scan'] = _dt.datetime.utcfromtimestamp(ts).strftime('%Y-%m-%d %H:%M UTC')
                    except Exception:
                        pass

                    # ── Active plugins from DB — with official/unofficial check ──
                    try:
                        opts_tbl = f'{db_pfx}options'
                        cur.execute(
                            f"SELECT option_value FROM {opts_tbl} "
                            f"WHERE option_name = 'active_plugins' LIMIT 1")
                        row = cur.fetchone()
                        if row:
                            raw = str(row.get('option_value') or '')
                            plugin_paths = re.findall(r'"([\w/.-]+\.php)"', raw)
                            for path in plugin_paths:
                                slug = path.split('/')[0]
                                check = _check_wp_plugin_official(slug)
                                entry = {
                                    'slug':           slug,
                                    'path':           path,
                                    'status':         check['status'],
                                    'confidence':     check['confidence'],
                                    'active_installs':check.get('active_installs'),
                                    'wp_org_name':    check.get('wp_org_name'),
                                }
                                results['plugins'].append(entry)
                                if check['status'] in ('unofficial', 'suspicious'):
                                    results['summary']['high'] = results['summary'].get('high', 0) + 1
                    except Exception:
                        pass

                conn.close()
            except Exception as e:
                results['db_error'] = str(e)

        # ── CyberINK mu-plugin deep scan (if installed) ──────────────────────
        cfai_token = os.environ.get('CFAI_FILESCAN_TOKEN', '')
        plugin_url = f'{site_url}/wp-json/cfai/v1/filescan'
        try:
            import urllib.request as _ur2, json as _js2
            req2 = _ur2.Request(plugin_url, headers={
                'User-Agent': 'Mozilla/5.0',
                'X-CFAI-Token': cfai_token,
            })
            with _ur2.urlopen(req2, timeout=20) as resp2:
                plugin_data = _js2.loads(resp2.read().decode('utf-8', errors='replace'))
                results['plugin_installed'] = True
                results['wp_version']  = plugin_data.get('wp_version', '')
                results['php_version'] = plugin_data.get('php_version', '')
                results['scan_time']   = plugin_data.get('scan_time', '')
                for f in plugin_data.get('php_in_uploads', []):
                    if f not in results['upload_php_files']:
                        results['upload_php_files'].append(f)
                        results['summary']['critical'] = results['summary'].get('critical', 0) + 1
                results['recent_modified'] = plugin_data.get('recent_modified', [])
                # Executables / scripts in WP dirs are always suspicious; JS/PY files only
                # if they appear outside wp-admin/wp-includes (i.e. user-writable areas).
                _EXEC_SUSP = {'.exe', '.bat', '.sh', '.cmd', '.ps1', '.vbs', '.jar'}
                _SCRIPT_SUSP = {'.py', '.rb', '.pl', '.go'}
                for fm in results['recent_modified']:
                    ext  = (fm.get('ext') or '').lower()
                    fpath = (fm.get('path') or '').lower()
                    in_core = any(seg in fpath for seg in ('/wp-admin/', '/wp-includes/'))
                    if ext in _EXEC_SUSP:
                        fm['suspicious'] = True
                        fm['reason'] = 'Executable file type in WordPress directory'
                        results['summary']['high'] = results['summary'].get('high', 0) + 1
                    elif ext in _SCRIPT_SUSP and not in_core:
                        fm['suspicious'] = True
                        fm['reason'] = f'Script file ({ext}) outside core directories'
                        results['summary']['medium'] = results['summary'].get('medium', 0) + 1
                    elif ext == '.js' and '/uploads/' in fpath:
                        fm['suspicious'] = True
                        fm['reason'] = 'JavaScript file in uploads directory'
                        results['summary']['medium'] = results['summary'].get('medium', 0) + 1

                # Also scan any plugin-provided file content for backdoor patterns
                if plugin_data.get('file_samples'):
                    backdoor_hits = []
                    for sample in plugin_data['file_samples']:
                        content = sample.get('content', '')
                        fpath2  = sample.get('path', '')
                        backdoor_hits.extend(_scan_content_for_backdoors(content, fpath2))
                    if backdoor_hits:
                        results['backdoor_findings'] = backdoor_hits
                        results['summary']['critical'] = results['summary'].get('critical', 0) + len(backdoor_hits)
        except Exception:
            results['plugin_installed'] = False

        # ── WordPress REST API — detect PHP in uploads via media endpoint ─────
        if not results.get('plugin_installed'):
            try:
                import urllib.request as _ur, json as _js3
                media_url = f"{site_url}/wp-json/wp/v2/media?mime_type=application/x-php&per_page=20"
                req = _ur.Request(media_url, headers={'User-Agent': 'Mozilla/5.0'})
                with _ur.urlopen(req, timeout=8) as resp:
                    media = _js3.loads(resp.read().decode('utf-8', errors='replace'))
                    for m in (media or []):
                        src = m.get('source_url', '')
                        if src and '.php' in src.lower():
                            if src not in results['upload_php_files']:
                                results['upload_php_files'].append(src)
                                results['summary']['critical'] = results['summary'].get('critical', 0) + 1
            except Exception:
                pass

        return jsonify(results)

    except Exception as e:
        import traceback as _tb
        return jsonify({'error': str(e), 'detail': _tb.format_exc()[-400:]}), 500


@app.route('/api/filescan/generic', methods=['POST'])
@login_required
def api_filescan_generic():
    """Generic website file scanner — HTTP-based checks for any platform."""
    import urllib.request as _req
    import urllib.error as _uerr
    data = request.get_json(silent=True) or {}
    site_url = (data.get('site_url') or '').strip().rstrip('/')
    platform = (data.get('platform') or 'generic').lower()
    if not site_url:
        return jsonify({'error': 'site_url is required'}), 400
    if not site_url.startswith('http'):
        site_url = 'https://' + site_url

    def _get(path, timeout=8):
        try:
            req = _req.Request(site_url + path, headers={'User-Agent': _BROWSER_UA})
            with _req.urlopen(req, timeout=timeout) as r:
                return r.status, dict(r.headers), r.read(4096).decode('utf-8', errors='ignore')
        except _uerr.HTTPError as e:
            return e.code, {}, ''
        except Exception:
            return 0, {}, ''

    results = {'platform': platform, 'site_url': site_url,
                'exposed_files': [], 'missing_headers': [],
                'security_headers': {}, 'platform_findings': [],
                'server_info': {}, 'backdoor_findings': [],
                'summary': {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'info': 0}}

    # ── 1. Security headers + server fingerprinting ───────────────────────────
    code, hdrs, body_root = _get('/')
    hdrs_lower = {k.lower(): v for k, v in hdrs.items()}

    # Fingerprint the server from response headers
    server_banner = hdrs_lower.get('server', '')
    powered_by    = hdrs_lower.get('x-powered-by', '')
    generator_m   = re.search(r'<meta[^>]+name=["\']generator["\'][^>]+content=["\']([^"\']+)', body_root, re.I)
    generator_tag = generator_m.group(1) if generator_m else ''
    results['server_info'] = {
        'server':       server_banner,
        'x_powered_by': powered_by,
        'generator':    generator_tag,
        'http_status':  code,
    }
    # If server exposes version → flag it
    if re.search(r'[0-9]+\.[0-9]+', server_banner or powered_by or ''):
        results['exposed_files'].append({
            'path': '/ (response header)',
            'severity': 'low',
            'description': f'Server version disclosed in HTTP headers: {server_banner or powered_by}',
            'status': str(code),
        })
        results['summary']['low'] += 1

    hdr_checks = [
        ('strict-transport-security', 'Strict-Transport-Security', 'Enforces HTTPS — prevents protocol downgrade attacks'),
        ('content-security-policy',   'Content-Security-Policy',   'Mitigates XSS and data injection attacks'),
        ('x-frame-options',           'X-Frame-Options',           'Prevents clickjacking via iframes'),
        ('x-content-type-options',    'X-Content-Type-Options',    'Prevents MIME-type sniffing'),
        ('referrer-policy',           'Referrer-Policy',           'Controls how much referrer info is sent'),
        ('permissions-policy',        'Permissions-Policy',        'Restricts access to browser features'),
    ]
    for key, name, desc in hdr_checks:
        val = hdrs_lower.get(key, '')
        results['security_headers'][name] = val
        if not val:
            results['missing_headers'].append({'header': name, 'description': desc})
            results['summary']['medium'] += 1

    # ── 2. Exposed sensitive files — fetch content for PHP files ─────────────
    common_paths = [
        ('/.env',            'critical', 'Environment file — may contain DB passwords, API keys, secrets'),
        ('/.env.local',      'critical', 'Local env file — may contain secrets'),
        ('/.env.production', 'critical', 'Production env file — may contain secrets'),
        ('/.git/HEAD',       'critical', 'Git repo exposed — full source code downloadable'),
        ('/.git/config',     'high',     'Git config exposed — reveals remote URLs and branches'),
        ('/backup.sql',      'critical', 'SQL database dump exposed publicly'),
        ('/backup.zip',      'high',     'Site backup archive exposed publicly'),
        ('/dump.sql',        'critical', 'SQL dump exposed publicly'),
        ('/.htpasswd',       'high',     'Password file exposed'),
        ('/phpinfo.php',     'high',     'PHP info page exposes server configuration'),
        ('/info.php',        'high',     'PHP info page exposes server configuration'),
        ('/test.php',        'medium',   'Test PHP file left on server'),
        ('/shell.php',       'critical', 'Possible PHP webshell on server'),
        ('/cmd.php',         'critical', 'Possible PHP webshell on server'),
        ('/wp-content/uploads/shell.php', 'critical', 'PHP webshell in WordPress uploads'),
        ('/composer.json',   'medium',   'Composer config — reveals dependencies and versions'),
        ('/package.json',    'medium',   'NPM config — reveals dependencies and versions'),
        ('/config.php',      'high',     'Config file potentially exposed'),
        ('/db.php',          'high',     'Database config file potentially exposed'),
    ]

    # Platform-specific sensitive paths
    platform_paths = {
        'drupal':  [('/sites/default/settings.php', 'critical', 'Drupal config — may contain DB credentials'),
                    ('/sites/default/default.settings.php', 'medium', 'Drupal default settings exposed'),
                    ('/CHANGELOG.txt', 'medium', 'Drupal version fingerprint'),
                    ('/core/CHANGELOG.txt', 'medium', 'Drupal core version exposed'),
                    ('/modules/', 'info', 'Drupal modules directory browsable')],
        'joomla':  [('/configuration.php', 'critical', 'Joomla config — may contain DB credentials'),
                    ('/administrator/', 'medium', 'Joomla admin panel exposed'),
                    ('/logs/', 'medium', 'Joomla logs directory potentially browsable'),
                    ('/cache/', 'info', 'Joomla cache directory potentially browsable'),
                    ('/tmp/', 'medium', 'Joomla temp directory potentially browsable')],
        'laravel': [('/storage/logs/laravel.log', 'high', 'Laravel log file — may contain stack traces with credentials'),
                    ('/.env', 'critical', 'Laravel .env with APP_KEY and DB credentials'),
                    ('/public/.htaccess', 'info', 'Laravel htaccess exposed'),
                    ('/artisan', 'medium', 'Artisan CLI script exposed'),
                    ('/telescope/api/requests', 'high', 'Laravel Telescope exposed — reveals all HTTP requests')],
        'django':  [('/admin/', 'medium', 'Django admin panel accessible'),
                    ('/__debug__/', 'high', 'Django Debug Toolbar exposed — reveals queries and environment'),
                    ('/static/admin/', 'info', 'Django static admin files'),
                    ('/media/', 'info', 'Django media directory potentially browsable'),
                    ('/api/schema/', 'medium', 'DRF schema exposed — reveals API endpoints')],
        'nodejs':  [('/package.json', 'medium', 'npm package.json — reveals dependencies'),
                    ('/.env', 'critical', 'Node.js env file with secrets'),
                    ('/node_modules/', 'info', 'node_modules directory potentially browsable'),
                    ('/graphql', 'medium', 'GraphQL endpoint — may be introspectable'),
                    ('/metrics', 'medium', 'Prometheus metrics endpoint potentially exposed')],
        'wordpress': [('/wp-config.php', 'critical', 'WordPress config — contains DB credentials'),
                      ('/wp-content/debug.log', 'high', 'WordPress debug log exposed — may contain paths and errors'),
                      ('/readme.html', 'low', 'WordPress readme — reveals version'),
                      ('/license.txt', 'low', 'WordPress license — reveals version'),
                      ('/wp-json/wp/v2/users', 'medium', 'WordPress REST API exposes user list')],
    }

    # Auto-detect WordPress from generator tag or server info
    wp_detected = (
        'wordpress' in generator_tag.lower()
        or 'wp-content' in body_root.lower()
        or (platform == 'wordpress')
    )
    if wp_detected and platform not in ('drupal', 'joomla', 'laravel', 'django', 'nodejs'):
        platform = 'wordpress'
        results['platform'] = 'wordpress'

    all_paths = common_paths + platform_paths.get(platform, [])
    for fpath, severity, description in all_paths:
        fcode, fhdrs, fbody = _get(fpath)
        if fcode == 200:
            entry = {'path': fpath, 'severity': severity,
                     'description': description, 'status': str(fcode)}
            results['exposed_files'].append(entry)
            results['summary'][severity] = results['summary'].get(severity, 0) + 1
            # Scan PHP file content for backdoor patterns
            if fpath.endswith('.php') and fbody:
                hits = _scan_content_for_backdoors(fbody, fpath)
                if hits:
                    results['backdoor_findings'].extend(hits)
                    for h in hits:
                        sev = h['severity']
                        results['summary'][sev] = results['summary'].get(sev, 0) + 1

    # ── 3. Platform-specific checks ────────────────────────────────────────────
    def _check(label, path, expect_not=None, expect=None, note=''):
        chk_code, _, chk_body = _get(path)
        if expect_not is not None:
            passed = chk_code not in expect_not
        elif expect is not None:
            passed = chk_code in expect
        else:
            passed = chk_code == 200
        detail = note or ('HTTP ' + str(chk_code))
        if not passed:
            results['summary']['medium'] += 1
        return {'check': label, 'pass': passed, 'detail': detail, 'http_status': chk_code}

    if platform == 'drupal':
        results['platform_findings'] = [
            _check('CHANGELOG.txt not public',      '/CHANGELOG.txt',            expect_not=[200], note='Version info should not be publicly readable'),
            _check('install.php removed',           '/install.php',              expect_not=[200], note='install.php must not exist on live sites'),
            _check('update.php protected',          '/update.php',               expect_not=[200], note='update.php must require admin auth'),
            _check('sites/default not browsable',   '/sites/default/',           expect_not=[200], note='Directory listing should be disabled'),
            _check('xmlrpc.php blocked',            '/xmlrpc.php',               expect_not=[200], note='XML-RPC should be disabled if unused'),
        ]
    elif platform == 'joomla':
        results['platform_findings'] = [
            _check('Admin panel requires auth',     '/administrator/',            expect_not=[200], note='Should redirect to login, not expose admin panel'),
            _check('Logs not public',               '/logs/',                     expect_not=[200], note='Log directory must not be browsable'),
            _check('configuration.php not readable','/configuration.php',        expect_not=[200], note='Config must not be readable via HTTP'),
            _check('Cache not browsable',           '/cache/',                    expect_not=[200], note='Cache directory must not be browsable'),
            _check('Tmp not browsable',             '/tmp/',                      expect_not=[200], note='Temp directory must not be browsable'),
        ]
    elif platform == 'laravel':
        results['platform_findings'] = [
            _check('Debug mode off',                '/_debugbar/open',            expect_not=[200], note='Laravel Debugbar should not be publicly accessible'),
            _check('Telescope not public',          '/telescope/api/requests',    expect_not=[200], note='Laravel Telescope must require authentication'),
            _check('Storage logs not public',       '/storage/logs/laravel.log',  expect_not=[200], note='Log files must not be publicly readable'),
            _check('.env not exposed',              '/.env',                      expect_not=[200], note='.env must never be publicly accessible'),
            _check('Horizon not public',            '/horizon',                   expect_not=[200], note='Laravel Horizon must require authentication'),
        ]
    elif platform == 'django':
        results['platform_findings'] = [
            _check('Debug Toolbar not public',      '/__debug__/',                expect_not=[200], note='Django Debug Toolbar must not be publicly accessible'),
            _check('Admin requires auth',           '/admin/',                    expect_not=[200], note='Admin should redirect to login — not expose Django admin'),
            _check('Media not browsable',           '/media/',                    expect_not=[200], note='Media directory listing should be disabled'),
            _check('API schema not public',         '/api/schema/',               expect_not=[200], note='DRF API schema reveals all endpoints and methods'),
        ]
    elif platform == 'nodejs':
        results['platform_findings'] = [
            _check('node_modules not browsable',    '/node_modules/',             expect_not=[200], note='node_modules must never be web-accessible'),
            _check('.env not exposed',              '/.env',                      expect_not=[200], note='.env must never be publicly accessible'),
            _check('package.json not exposed',      '/package.json',              expect_not=[200], note='Dependency list should not be public'),
            _check('GraphQL introspection',         '/graphql?query={__schema{types{name}}}', expect_not=[200], note='GraphQL introspection in production leaks API schema'),
        ]
    elif platform == 'wordpress':
        results['platform_findings'] = [
            _check('wp-config.php blocked',         '/wp-config.php',             expect_not=[200], note='wp-config.php must never be publicly readable'),
            _check('xmlrpc.php blocked',            '/xmlrpc.php',                expect_not=[200], note='XML-RPC should be disabled if not explicitly needed'),
            _check('readme.html removed',           '/readme.html',               expect_not=[200], note='readme.html reveals WordPress version'),
            _check('wp-login rate limited',         '/wp-login.php',              expect=[302, 200], note='wp-login.php should require auth (200=exposed, 302=redirecting)'),
            _check('debug.log not public',          '/wp-content/debug.log',      expect_not=[200], note='Debug log may contain sensitive paths and credentials'),
            _check('wp-json users hidden',          '/wp-json/wp/v2/users',       expect_not=[200], note='User enumeration via REST API should be disabled'),
        ]
    else:
        results['platform_findings'] = [
            _check('.env not exposed',              '/.env',                      expect_not=[200], note='.env must never be publicly accessible'),
            _check('.git not exposed',              '/.git/HEAD',                 expect_not=[200], note='Git repo must not be publicly accessible'),
            _check('Backup files not exposed',      '/backup.sql',                expect_not=[200], note='Database backups must not be publicly accessible'),
            _check('Admin panel protected',         '/admin',                     expect_not=[200], note='Admin panel must require authentication'),
            _check('Config file protected',         '/config.php',                expect_not=[200], note='Configuration files must not be web-accessible'),
        ]

    return jsonify(results)


@app.route('/api/logs/drupal-live', methods=['POST'])
@login_required
def api_logs_drupal_live():
    """Fetch Drupal watchdog events via JSON:API (requires admin credentials)."""
    import urllib.request as _req
    import urllib.parse as _up
    import base64 as _b64
    data = request.get_json(silent=True) or {}
    site_url  = (data.get('site_url') or '').strip().rstrip('/')
    username  = (data.get('username') or '').strip()
    password  = (data.get('password') or '').strip()
    limit     = min(int(data.get('limit') or 50), 200)
    if not site_url:
        return jsonify({'error': 'site_url is required'}), 400
    if not site_url.startswith('http'):
        site_url = 'https://' + site_url

    auth = _b64.b64encode(f'{username}:{password}'.encode()).decode() if username else None
    headers = {'Accept': 'application/vnd.api+json', 'Content-Type': 'application/vnd.api+json'}
    if auth:
        headers['Authorization'] = 'Basic ' + auth

    def _get_json(url):
        req = _req.Request(url, headers=headers)
        with _req.urlopen(req, timeout=15) as r:
            return _json.loads(r.read().decode('utf-8', errors='ignore'))

    try:
        # Drupal JSON:API dblog endpoint (requires dblog module enabled)
        api_url = (f'{site_url}/jsonapi/dblog_message/message'
                   f'?sort=-timestamp&page[limit]={limit}')
        resp = _get_json(api_url)
        events = []
        severity_map = {'0':'emergency','1':'alert','2':'critical','3':'error',
                        '4':'warning','5':'notice','6':'info','7':'debug'}
        for item in (resp.get('data') or []):
            attrs = item.get('attributes', {})
            sev_num = str(attrs.get('severity', 6))
            events.append({
                'timestamp': (attrs.get('timestamp') or '')[:19].replace('T', ' '),
                'type':      attrs.get('type', ''),
                'severity':  severity_map.get(sev_num, 'info'),
                'message':   attrs.get('message', ''),
                'user':      (item.get('relationships', {}).get('uid', {}).get('data') or {}).get('id', ''),
                'ip':        attrs.get('hostname', ''),
            })
        return jsonify({'events': events, 'count': len(events)})
    except Exception as exc:
        return jsonify({'error': f'Failed to fetch Drupal logs: {exc}. '
                                 'Ensure JSON:API and dblog modules are enabled and credentials are correct.'}), 500


@app.route('/api/logs/joomla-live', methods=['POST'])
@login_required
def api_logs_joomla_live():
    """Fetch Joomla action logs via Web Services API (Joomla 4+)."""
    import urllib.request as _req
    import base64 as _b64
    data = request.get_json(silent=True) or {}
    site_url  = (data.get('site_url') or '').strip().rstrip('/')
    username  = (data.get('username') or '').strip()
    password  = (data.get('password') or '').strip()
    token     = (data.get('token') or '').strip()
    limit     = min(int(data.get('limit') or 50), 200)
    if not site_url:
        return jsonify({'error': 'site_url is required'}), 400
    if not site_url.startswith('http'):
        site_url = 'https://' + site_url

    headers = {'Accept': 'application/vnd.api+json'}
    if token:
        headers['X-Joomla-Token'] = token
    elif username:
        auth = _b64.b64encode(f'{username}:{password}'.encode()).decode()
        headers['Authorization'] = 'Basic ' + auth

    def _get_json(url):
        req = _req.Request(url, headers=headers)
        with _req.urlopen(req, timeout=15) as r:
            return _json.loads(r.read().decode('utf-8', errors='ignore'))

    try:
        api_url = f'{site_url}/api/index.php/v1/privacy/actionlogs?page[limit]={limit}&sort=-id'
        resp = _get_json(api_url)
        events = []
        for item in (resp.get('data') or []):
            attrs = item.get('attributes', {})
            events.append({
                'timestamp': (attrs.get('log_date') or '')[:19].replace('T', ' '),
                'user':      attrs.get('username', '') or attrs.get('name', ''),
                'action':    attrs.get('message_language_key', '').replace('PLG_ACTIONLOG_', '').replace('_', ' ').title(),
                'item':      attrs.get('item_id', ''),
                'ip':        attrs.get('ip_address', ''),
                'extension': attrs.get('extension', ''),
            })
        return jsonify({'events': events, 'count': len(events)})
    except Exception as exc:
        return jsonify({'error': f'Failed to fetch Joomla logs: {exc}. '
                                 'Ensure Joomla 4+ Web Services API is enabled and provide an API token or admin credentials.'}), 500


@app.route('/api/logs/app-ssh', methods=['POST'])
@login_required
def api_logs_app_ssh():
    """Read any log file from a server via SSH."""
    data = request.get_json(silent=True) or {}
    ssh_host = (data.get('ssh_host') or '').strip()
    ssh_user = (data.get('ssh_user') or 'root').strip()
    ssh_pass = (data.get('ssh_pass') or '').strip()
    ssh_port = int(data.get('ssh_port') or 22)
    log_path = (data.get('log_path') or '').strip()
    lines    = min(int(data.get('lines') or 200), 2000)
    grep_filter = (data.get('filter') or '').strip()

    if not ssh_host or not log_path:
        return jsonify({'error': 'ssh_host and log_path are required'}), 400

    try:
        import paramiko
    except ImportError:
        return jsonify({'error': 'paramiko not installed on server. Run: pip install paramiko --break-system-packages'}), 500

    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(ssh_host, port=ssh_port, username=ssh_user, password=ssh_pass, timeout=15)

        if grep_filter:
            cmd = f'tail -n {lines} {log_path} | grep -i {_subprocess.list2cmdline([grep_filter])} | tail -n {lines}'
        else:
            cmd = f'tail -n {lines} {log_path}'

        _, stdout, stderr = client.exec_command(cmd, timeout=20)
        out = stdout.read().decode('utf-8', errors='ignore')
        err = stderr.read().decode('utf-8', errors='ignore')
        client.close()

        if err and not out:
            return jsonify({'error': err.strip()}), 500

        log_lines = [l for l in out.splitlines() if l.strip()]
        return jsonify({'lines': log_lines, 'count': len(log_lines), 'path': log_path})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/api/plugins/drupal', methods=['POST'])
@login_required
def api_plugins_drupal():
    """Fetch installed Drupal modules via JSON:API."""
    import urllib.request as _req
    import base64 as _b64
    data = request.get_json(silent=True) or {}
    site_url = (data.get('site_url') or '').strip().rstrip('/')
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    token    = (data.get('token') or '').strip()
    if not site_url:
        return jsonify({'error': 'site_url is required'}), 400
    if not site_url.startswith('http'):
        site_url = 'https://' + site_url

    headers = {'Accept': 'application/vnd.api+json'}
    if token:
        headers['Authorization'] = 'Bearer ' + token
    elif username:
        auth = _b64.b64encode(f'{username}:{password}'.encode()).decode()
        headers['Authorization'] = 'Basic ' + auth

    def _get_json(url):
        req = _req.Request(url, headers=headers)
        with _req.urlopen(req, timeout=15) as r:
            return _json.loads(r.read().decode('utf-8', errors='ignore'))

    try:
        resp = _get_json(f'{site_url}/jsonapi/node_type/node_type')
        # Try update status endpoint for modules
        items = []
        try:
            mod_resp = _get_json(f'{site_url}/admin/modules/json')
            for m in (mod_resp if isinstance(mod_resp, list) else []):
                items.append({'name': m.get('name',''), 'type': 'Module',
                              'version': m.get('version',''), 'status': 'enabled' if m.get('status') else 'disabled',
                              'vulnerable': False, 'updated': ''})
        except Exception:
            # Fallback: parse from update status page
            upd = _get_json(f'{site_url}/jsonapi/update_status/release?filter[status]=1&page[limit]=100')
            for item in (upd.get('data') or []):
                attrs = item.get('attributes', {})
                items.append({'name': attrs.get('name',''), 'type': 'Module',
                              'version': attrs.get('existing_version',''),
                              'status': 'enabled', 'vulnerable': bool(attrs.get('security_update')),
                              'updated': ''})
        if not items:
            return jsonify({'error': 'No module data returned. Ensure JSON:API and admin credentials are correct.'}), 404
        return jsonify({'items': items, 'count': len(items)})
    except Exception as exc:
        return jsonify({'error': f'Failed to fetch Drupal modules: {exc}'}), 500


@app.route('/api/plugins/joomla', methods=['POST'])
@login_required
def api_plugins_joomla():
    """Fetch installed Joomla extensions via Web Services API (Joomla 4+)."""
    import urllib.request as _req
    import base64 as _b64
    data = request.get_json(silent=True) or {}
    site_url = (data.get('site_url') or '').strip().rstrip('/')
    username = (data.get('username') or '').strip()
    password = (data.get('password') or '').strip()
    token    = (data.get('token') or '').strip()
    if not site_url:
        return jsonify({'error': 'site_url is required'}), 400
    if not site_url.startswith('http'):
        site_url = 'https://' + site_url

    headers = {'Accept': 'application/vnd.api+json'}
    if token:
        headers['X-Joomla-Token'] = token
    elif username:
        auth = _b64.b64encode(f'{username}:{password}'.encode()).decode()
        headers['Authorization'] = 'Basic ' + auth

    def _get_json(url):
        req = _req.Request(url, headers=headers)
        with _req.urlopen(req, timeout=15) as r:
            return _json.loads(r.read().decode('utf-8', errors='ignore'))

    try:
        resp = _get_json(f'{site_url}/api/index.php/v1/extensions?page[limit]=200&sort=name')
        items = []
        for item in (resp.get('data') or []):
            attrs = item.get('attributes', {})
            items.append({
                'name':       attrs.get('name', ''),
                'type':       attrs.get('type', 'extension').capitalize(),
                'version':    attrs.get('version', ''),
                'status':     'enabled' if attrs.get('enabled') else 'disabled',
                'vulnerable': False,
                'updated':    (attrs.get('manifest_cache') or {}).get('creation_date', '') if isinstance(attrs.get('manifest_cache'), dict) else '',
            })
        return jsonify({'items': items, 'count': len(items)})
    except Exception as exc:
        return jsonify({'error': f'Failed to fetch Joomla extensions: {exc}. Ensure Joomla 4+ and Web Services API plugin enabled.'}), 500


@app.route('/api/plugins/packages', methods=['POST'])
@login_required
def api_plugins_packages():
    """Fetch package list (composer/npm/pip) from a server via SSH."""
    data = request.get_json(silent=True) or {}
    manager  = (data.get('manager') or 'composer').lower()
    ssh_host = (data.get('ssh_host') or '').strip()
    ssh_user = (data.get('ssh_user') or 'root').strip()
    ssh_pass = (data.get('ssh_pass') or '').strip()
    ssh_port = int(data.get('ssh_port') or 22)
    path     = (data.get('path') or '/var/www/html').strip()

    if not ssh_host:
        return jsonify({'error': 'ssh_host is required'}), 400

    try:
        import paramiko
    except ImportError:
        return jsonify({'error': 'paramiko not installed. Run: pip install paramiko --break-system-packages'}), 500

    cmds = {
        'composer': f'cat {path}/composer.lock 2>/dev/null || cat {path}/composer.json 2>/dev/null',
        'npm':      f'cat {path}/package-lock.json 2>/dev/null || cat {path}/package.json 2>/dev/null',
        'pip':      f'pip freeze 2>/dev/null || pip3 freeze 2>/dev/null || cat {path}/requirements.txt 2>/dev/null',
    }

    try:
        client = paramiko.SSHClient()
        client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
        client.connect(ssh_host, port=ssh_port, username=ssh_user, password=ssh_pass, timeout=15)
        _, stdout, _ = client.exec_command(cmds.get(manager, cmds['pip']), timeout=20)
        out = stdout.read().decode('utf-8', errors='ignore')
        client.close()

        items = []
        if manager == 'pip':
            for line in out.splitlines():
                if '==' in line:
                    parts = line.split('==')
                    items.append({'name': parts[0].strip(), 'version': parts[1].strip() if len(parts)>1 else ''})
        else:
            try:
                pkg_data = _json.loads(out)
                if manager == 'composer':
                    for pkg in pkg_data.get('packages', []):
                        items.append({'name': pkg.get('name',''), 'version': pkg.get('version','')})
                elif manager == 'npm':
                    deps = {**pkg_data.get('dependencies',{}), **pkg_data.get('devDependencies',{})}
                    for name, val in deps.items():
                        ver = val.get('version','') if isinstance(val, dict) else str(val)
                        items.append({'name': name, 'version': ver})
            except Exception:
                for line in out.splitlines():
                    if line.strip():
                        items.append({'name': line.strip(), 'version': ''})

        return jsonify({'items': items, 'count': len(items)})
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500


@app.route('/api/wp/filescan-plugin')
def api_wp_filescan_plugin():
    """Return a ready-to-upload WordPress mu-plugin that exposes a secure file-scan REST endpoint."""
    token = os.environ.get('CFAI_FILESCAN_TOKEN', '')
    if not token:
        import secrets
        token = secrets.token_hex(24)
    plugin_code = f'''<?php
/**
 * Plugin Name: CyberINK File Scanner
 * Description: Secure REST endpoint for CyberINK dashboard file scanning.
 * Version: 1.0
 * — Drop this file into wp-content/mu-plugins/ (no activation needed)
 */
if ( ! defined( 'ABSPATH' ) ) exit;

add_action( 'rest_api_init', function () {{
    register_rest_route( 'cfai/v1', '/filescan', array(
        'methods'             => 'GET',
        'callback'            => 'cfai_filescan_handler',
        'permission_callback' => '__return_true',
    ));
}});

function cfai_filescan_handler( $request ) {{
    $token = $request->get_header( 'X-CFAI-Token' );
    if ( $token !== '{token}' ) {{
        return new WP_Error( 'forbidden', 'Invalid token', array( 'status' => 403 ) );
    }}

    $upload_dir  = wp_upload_dir();
    $upload_base = $upload_dir['basedir'];
    $results     = array(
        'php_in_uploads'   => array(),
        'recent_modified'  => array(),
        'wp_version'       => get_bloginfo('version'),
        'php_version'      => PHP_VERSION,
        'scan_time'        => date('Y-m-d H:i:s'),
    );

    // Scan uploads recursively for PHP files
    $iter = new RecursiveIteratorIterator(
        new RecursiveDirectoryIterator( $upload_base, RecursiveDirectoryIterator::SKIP_DOTS )
    );
    $cutoff = time() - (90 * 86400); // files modified in last 90 days
    foreach ( $iter as $file ) {{
        if ( ! $file->isFile() ) continue;
        $path = $file->getPathname();
        $ext  = strtolower( pathinfo( $path, PATHINFO_EXTENSION ) );
        // Flag PHP files in uploads (always suspicious)
        if ( $ext === 'php' ) {{
            $results['php_in_uploads'][] = str_replace( $upload_base, '/uploads', $path );
        }}
        // Flag recently modified non-image files in uploads
        if ( $file->getMTime() > $cutoff && ! in_array($ext, ['jpg','jpeg','png','gif','webp','svg','pdf','mp4','mp3','zip']) ) {{
            $results['recent_modified'][] = array(
                'file'     => str_replace( $upload_base, '/uploads', $path ),
                'modified' => date('Y-m-d H:i', $file->getMTime()),
                'size'     => $file->getSize(),
                'ext'      => $ext,
            );
        }}
    }}

    // Check wp-content for recently modified PHP (outside uploads)
    $wc_dir = WP_CONTENT_DIR;
    $wc_iter = new RecursiveIteratorIterator(
        new RecursiveDirectoryIterator( $wc_dir, RecursiveDirectoryIterator::SKIP_DOTS )
    );
    $recent_cutoff = time() - (7 * 86400); // last 7 days
    foreach ( $wc_iter as $file ) {{
        if ( ! $file->isFile() ) continue;
        $path = $file->getPathname();
        if ( strpos($path, '/uploads/') !== false ) continue; // already covered
        if ( strpos($path, '/wflogs/') !== false ) continue;
        $ext = strtolower( pathinfo( $path, PATHINFO_EXTENSION ) );
        if ( $ext === 'php' && $file->getMTime() > $recent_cutoff ) {{
            $results['recent_modified'][] = array(
                'file'     => str_replace( $wc_dir, '/wp-content', $path ),
                'modified' => date('Y-m-d H:i', $file->getMTime()),
                'size'     => $file->getSize(),
                'ext'      => $ext,
            );
        }}
    }}

    // Sort recent_modified by date desc, limit 50
    usort( $results['recent_modified'], function($a,$b){{ return strcmp($b['modified'], $a['modified']); }} );
    $results['recent_modified'] = array_slice( $results['recent_modified'], 0, 50 );

    return rest_ensure_response( $results );
}}
'''
    from flask import Response
    resp = Response(plugin_code, mimetype='application/octet-stream')
    resp.headers['Content-Disposition'] = 'attachment; filename="cfai-scanner.php"'
    return resp


@app.route('/api/analytics/pci')
def api_analytics_pci():
    """PCI-style threat analytics derived entirely from real scan history in the DB."""
    try:
        return _api_analytics_pci_inner()
    except Exception as _e:
        import traceback as _tb
        return jsonify({'error': str(_e), 'detail': _tb.format_exc()[-600:],
                        'mitigation_severity': [], 'compliance_keyword': [],
                        'most_vulnerable': [], 'vuln_summary': [], 'top_failures': [],
                        'trends': {'labels': [], 'vulnerabilities': [], 'compliance': []},
                        'config_summary': {}, 'meta': {'total_scans': 0, 'total_targets': 0, 'last_updated': ''}})

def _api_analytics_pci_inner():
    import json as _j
    from datetime import datetime, timedelta

    scans = db.get_scans(limit=3000, username=_cu_filter())
    now   = datetime.utcnow()

    # ── 1. Mitigation by severity ──────────────────────────────────────────────
    sev_bkts = {s: {'lt10': 0, 'lt30': 0, 'gt30': 0}
                for s in ('CRITICAL', 'HIGH', 'MEDIUM', 'LOW')}
    for sc in scans:
        out = sc.get('output', '').upper()
        try:
            dt  = datetime.fromisoformat(sc.get('created_at', '').replace('Z', '')[:19])
            age = (now - dt).days
        except Exception:
            age = 31
        bkt = 'lt10' if age < 10 else ('lt30' if age < 30 else 'gt30')
        for sev in sev_bkts:
            sev_bkts[sev][bkt] += min(out.count(sev), 15)

    t10 = t30 = tgt = 0
    panel1 = []
    for sev in ('CRITICAL', 'HIGH', 'MEDIUM', 'LOW'):
        b = sev_bkts[sev]
        tot = b['lt10'] + b['lt30'] + b['gt30']
        t10 += b['lt10']; t30 += b['lt30']; tgt += b['gt30']
        panel1.append({'severity': sev,
                       'lt10': b['lt10'], 'lt10_pct': round(b['lt10']/tot*100) if tot else 0,
                       'lt30': b['lt30'], 'lt30_pct': round(b['lt30']/tot*100) if tot else 0,
                       'gt30': b['gt30'], 'gt30_pct': round(b['gt30']/tot*100) if tot else 0})
    ttot = t10 + t30 + tgt
    panel1 = [{'severity': 'Total Vulnerabilities',
               'lt10': t10, 'lt10_pct': round(t10/ttot*100) if ttot else 0,
               'lt30': t30, 'lt30_pct': round(t30/ttot*100) if ttot else 0,
               'gt30': tgt, 'gt30_pct': round(tgt/ttot*100) if ttot else 0}] + panel1

    # ── 2. Weekly trends (8 weeks) ─────────────────────────────────────────────
    labels, v_series, c_series = [], [], []
    for w in range(8, 0, -1):
        ws = now - timedelta(weeks=w)
        we = now - timedelta(weeks=w - 1)
        labels.append(ws.strftime('%b %d'))
        ws_s, we_s = ws.isoformat()[:10], we.isoformat()[:10]
        wk = [s for s in scans if ws_s <= (s.get('created_at', '') or '')[:10] < we_s]
        v = sum(min(s.get('output', '').upper().count('HIGH') + s.get('output', '').upper().count('CRITICAL') * 2, 20) for s in wk)
        c = sum(min(len([ln for ln in s.get('output', '').split('\n')
                         if any(k in ln.lower() for k in ('ssl', 'tls', 'config', 'header', 'csrf', 'cors'))]), 5) for s in wk)
        v_series.append(v); c_series.append(c)

    # ── 3. Compliance by keyword ───────────────────────────────────────────────
    week_ago = (now - timedelta(days=7)).isoformat()[:10]
    kws = ('Auth', 'Account', 'Audit', 'Disable', 'Enable', 'Log', 'Password', 'Permission', 'User')
    panel3 = []
    for kw in kws:
        kl = kw.lower()
        matched  = [s for s in scans if kl in (s.get('output', '') or '').lower()]
        systems  = len(set(s['target'] for s in matched))
        last7    = sum(1 for s in matched if (s.get('created_at', '') or '')[:10] >= week_ago)
        ok_cnt   = sum(1 for s in matched if s.get('status') == 'ok')
        total    = len(matched)
        failed   = max(0, total - ok_cnt)
        manual   = max(0, failed // 4)
        panel3.append({'keyword': kw, 'systems': systems, 'last7d': last7,
                       'passed': ok_cnt, 'passed_pct': round(ok_cnt / total * 100) if total else 0,
                       'manual': manual, 'failed': max(0, failed - manual)})

    # ── 4. Most vulnerable hosts ───────────────────────────────────────────────
    hosts: dict = {}
    for s in scans:
        t   = s['target']
        out = (s.get('output', '') or '').upper()
        v   = out.count('HIGH') + out.count('CRITICAL') * 2 + out.count('MEDIUM') // 2
        if t not in hosts:
            hosts[t] = {'target': t, 'total': 0, 'critical': 0, 'high': 0, 'medium': 0}
        hosts[t]['total']    += v
        hosts[t]['critical'] += min(out.count('CRITICAL'), 15)
        hosts[t]['high']     += min(out.count('HIGH'), 15)
        hosts[t]['medium']   += min(out.count('MEDIUM'), 15)
    panel4 = sorted(hosts.values(), key=lambda x: x['total'], reverse=True)[:25]
    max_v  = max((h['total'] for h in panel4), default=1) or 1
    for h in panel4:
        h['score'] = min(10.0, round(h['total'] / max_v * 10, 1))
        h['bar_pct'] = round(h['total'] / max_v * 100)

    # ── 5. Vulnerability summary by period ────────────────────────────────────
    def _vsum(slist):
        tot  = sum(min((s.get('output','') or '').upper().count('HIGH') +
                       (s.get('output','') or '').upper().count('CRITICAL') * 2, 20) for s in slist)
        ok   = sum(1 for s in slist if s.get('status') == 'ok')
        mit  = min(ok * 3, tot)
        crit = sum(min((s.get('output','') or '').upper().count('CRITICAL'), 5) for s in slist)
        high = sum(min((s.get('output','') or '').upper().count('HIGH'), 10) for s in slist)
        med  = sum(min((s.get('output','') or '').upper().count('MEDIUM'), 10) for s in slist)
        return {'total': tot, 'mitigated': mit, 'unmitigated': max(0, tot - mit),
                'crit_pct':  round(crit / max(1, tot) * 100),
                'high_pct':  round(high / max(1, tot) * 100),
                'med_pct':   round(med  / max(1, tot) * 100)}

    mo  = now.replace(day=1).isoformat()[:10]
    lm  = (now.replace(day=1) - timedelta(days=1)).replace(day=1).isoformat()[:10]
    q   = now.replace(month=((now.month - 1) // 3) * 3 + 1, day=1).isoformat()[:10]
    d180 = (now - timedelta(days=180)).isoformat()[:10]
    panel5 = [
        {'period': 'Total',           **_vsum(scans)},
        {'period': 'Current Month',   **_vsum([s for s in scans if (s.get('created_at','') or '')[:10] >= mo])},
        {'period': 'Last Month',      **_vsum([s for s in scans if lm <= (s.get('created_at','') or '')[:10] < mo])},
        {'period': 'Current Quarter', **_vsum([s for s in scans if (s.get('created_at','') or '')[:10] >= q])},
        {'period': '>180 Days',       **_vsum([s for s in scans if (s.get('created_at','') or '')[:10] < d180])},
    ]

    # ── 6. Top failures ────────────────────────────────────────────────────────
    patterns = [
        ('Rate limiting missing',    ['rate limit','brute force','login attempt']),
        ('Outdated TLS/SSL',         ['tls 1.0','tls 1.1','sslv3','weak cipher']),
        ('SQL Injection risk',       ['sql injection','sqli','union select']),
        ('XSS vulnerability',        ['cross-site scripting','xss','script injection']),
        ('Exposed admin panel',      ['wp-admin','phpmyadmin','admin interface']),
        ('Missing security headers', ['hsts','x-frame','content-security','missing header']),
        ('Weak SSL/TLS cert',        ['self-signed','expired cert','certificate error']),
        ('Default credentials',      ['default password','default credential','admin:admin']),
        ('Open ports exposed',       ['open port','exposed service','unnecessary port']),
        ('CSRF vulnerability',       ['csrf','cross-site request','forgery']),
    ]
    panel6 = []
    for name, keys in patterns:
        cnt = sum(1 for s in scans if any(k in (s.get('output','') or '').lower() for k in keys))
        if cnt > 0:
            panel6.append({'name': name, 'severity': 'HIGH' if cnt > 3 else 'MEDIUM', 'total': cnt})
    panel6.sort(key=lambda x: x['total'], reverse=True)

    # ── 7. Config summary ──────────────────────────────────────────────────────
    total_chk = max(1, len(scans))
    passed    = sum(1 for s in scans if s.get('status') == 'ok')
    failed    = total_chk - passed
    manual    = max(0, failed // 5)
    uniq      = len(hosts)
    panel7 = {
        'check_count':      total_chk,
        'check_passed':     passed,
        'check_manual':     manual,
        'check_failed':     failed - manual,
        'check_pass_pct':   round(passed   / total_chk * 100),
        'check_manual_pct': round(manual   / total_chk * 100),
        'check_fail_pct':   round((failed - manual) / total_chk * 100),
        'system_count':     uniq,
        'system_pass':      min(uniq, passed),
        'system_manual':    min(uniq, manual),
        'system_fail':      max(0, uniq - passed - manual),
        'system_pass_pct':  round(passed / max(1, uniq) * 100),
        'system_manual_pct':round(manual / max(1, uniq) * 100),
        'system_fail_pct':  round(max(0, uniq-passed-manual) / max(1, uniq) * 100),
    }

    return jsonify({
        'mitigation_severity': panel1,
        'trends':              {'labels': labels, 'vulnerabilities': v_series, 'compliance': c_series},
        'compliance_keyword':  panel3,
        'most_vulnerable':     panel4,
        'vuln_summary':        panel5,
        'top_failures':        panel6,
        'config_summary':      panel7,
        'meta': {
            'total_scans':   len(scans),
            'total_targets': uniq,
            'last_updated':  now.strftime('%Y-%m-%d %H:%M UTC'),
        },
    })


@app.route('/api/logs/analyze', methods=['POST'])
def api_logs_analyze():
    """Fetch + analyze real server access logs via SSH or HTTP probe."""
    from tools.log_analyzer import analyze_from_ssh, analyze_from_probe, check_latency, check_error_rate

    data    = request.get_json(force=True, silent=True) or {}
    domain  = (data.get('domain') or '').strip().replace('https://', '').replace('http://', '').rstrip('/')
    if not domain:
        return jsonify({'error': 'domain is required'}), 400

    ssh_host = (data.get('ssh_host') or domain).strip()
    ssh_user = (data.get('ssh_user') or 'root').strip()
    ssh_pass = (data.get('ssh_pass') or '').strip()
    ssh_port = int(data.get('ssh_port') or 22)

    if ssh_user and ssh_pass:
        result = analyze_from_ssh(ssh_host, ssh_user, ssh_pass, ssh_port)
    else:
        result = analyze_from_probe(domain)

    result['latency'] = check_latency(domain)
    return jsonify(result)


@app.route('/api/monitor/network', methods=['POST'])
def api_monitor_network():
    """Discover services + network topology via Nmap + SSH netstat."""
    import subprocess, json as _j, re as _re

    data   = request.get_json(force=True, silent=True) or {}
    domain = (data.get('domain') or '').strip().replace('https://', '').replace('http://', '').rstrip('/')
    if not domain:
        return jsonify({'error': 'domain is required'}), 400

    ssh_host = (data.get('ssh_host') or domain).strip()
    ssh_user = (data.get('ssh_user') or '').strip()
    ssh_pass = (data.get('ssh_pass') or '').strip()
    ssh_port = int(data.get('ssh_port') or 22)

    # ── Nmap service scan (external) ──────────────────────────────────────────
    nodes = [{'id': domain, 'label': domain, 'type': 'target', 'group': 'target'}]
    edges = []
    services = []

    try:
        nmap_out = subprocess.run(
            ['nmap', '-Pn', '-sV', '--top-ports', '20', '--host-timeout', '30s',
             '--open', '-oG', '-', domain],
            capture_output=True, text=True, timeout=45,
        ).stdout
        for line in nmap_out.splitlines():
            m_ports = _re.findall(r'(\d+)/open/tcp//([^/]+)//([^/]*)', line)
            for port, proto, ver in m_ports:
                svc_id  = f'{proto.strip()}:{port}'
                svc_label = f'{proto.strip()} ({port})'
                services.append({'port': int(port), 'service': proto.strip(),
                                  'version': ver.strip()[:40], 'state': 'open'})
                nodes.append({'id': svc_id, 'label': svc_label, 'type': 'service', 'group': proto.strip()})
                edges.append({'from': domain, 'to': svc_id,
                               'label': f':{port}', 'arrows': 'to'})
    except Exception as e:
        services.append({'error': str(e)[:60]})

    # ── SSH netstat — active connections ─────────────────────────────────────
    connections = []
    traffic     = {}
    if ssh_user and ssh_pass:
        import os as _os
        ssh_base = ['sshpass', '-e', 'ssh', '-o', 'StrictHostKeyChecking=no',
                    '-o', 'ConnectTimeout=10', '-p', str(ssh_port), f'{ssh_user}@{ssh_host}']
        run_env = _os.environ.copy()
        run_env['SSHPASS'] = ssh_pass

        try:
            ss_out = subprocess.run(
                ssh_base + ["ss -tuanp 2>/dev/null | head -40 || netstat -tunap 2>/dev/null | head -40"],
                capture_output=True, text=True, timeout=15, env=run_env,
            ).stdout
            for line in ss_out.splitlines()[1:]:
                parts = line.split()
                if len(parts) >= 5 and parts[0] in ('tcp', 'udp', 'ESTAB', 'LISTEN', 'TIME-WAIT'):
                    connections.append({'proto': parts[0], 'local': parts[3] if len(parts) > 3 else '',
                                        'remote': parts[4] if len(parts) > 4 else '', 'state': parts[1] if len(parts) > 1 else ''})
        except Exception:
            pass

        try:
            dev_out = subprocess.run(
                ssh_base + ["cat /proc/net/dev 2>/dev/null | tail -n +3"],
                capture_output=True, text=True, timeout=10, env=run_env,
            ).stdout
            for line in dev_out.splitlines():
                if ':' not in line:
                    continue
                iface, rest = line.split(':', 1)
                nums = rest.split()
                if len(nums) >= 9:
                    traffic[iface.strip()] = {
                        'rx_bytes': int(nums[0]),
                        'tx_bytes': int(nums[8]),
                        'rx_mb': round(int(nums[0]) / 1048576, 2),
                        'tx_mb': round(int(nums[8]) / 1048576, 2),
                    }
        except Exception:
            pass

    # Add remote connection nodes
    remote_ips = set()
    for c in connections:
        remote = c.get('remote', '')
        if remote and remote not in ('*:*', '0.0.0.0:*', ':::*'):
            ip = remote.rsplit(':', 1)[0].strip('[]')
            if ip and ip not in ('0.0.0.0', '::', '127.0.0.1', '::1') and ip not in remote_ips:
                remote_ips.add(ip)
                geo = _geoip(ip)
                label = f'{ip}\n{geo}' if geo else ip
                nodes.append({'id': ip, 'label': label, 'type': 'remote', 'group': 'remote'})
                edges.append({'from': domain, 'to': ip, 'arrows': 'to'})

    return jsonify({
        'nodes': nodes, 'edges': edges,
        'services': services, 'connections': connections[:30],
        'traffic': traffic,
    })


@app.route('/api/monitor/latency')
def api_monitor_latency():
    """Quick HTTP latency probe for a domain."""
    from tools.log_analyzer import check_latency
    domain = (request.args.get('domain') or '').strip().replace('https://', '').replace('http://', '').rstrip('/')
    if not domain:
        return jsonify({'error': 'domain required'}), 400
    return jsonify({'results': check_latency(domain)})


@app.route('/api/mitre/coverage')
def api_mitre_coverage():
    """MITRE ATT&CK coverage from all scan history."""
    from dashboard.mitre_rules import get_coverage, TACTICS
    scans   = db.get_recent_scans(100, username=_cu_filter())
    result  = get_coverage(scans)
    result['tactic_order'] = [name for _, name in TACTICS]
    return jsonify(result)


@app.route('/api/incidents', methods=['GET'])
def api_incidents_get():
    status = request.args.get('status')
    uf = _cu_filter()
    return jsonify({'incidents': db.get_incidents(status=status, username=uf),
                    'stats': db.get_incident_stats(username=uf)})


@app.route('/api/incidents', methods=['POST'])
def api_incidents_create():
    data = request.get_json(force=True, silent=True) or {}
    title = (data.get('title') or '').strip()
    if not title:
        return jsonify({'error': 'title required'}), 400
    iid = db.create_incident(
        title=title,
        description=data.get('description', ''),
        severity=data.get('severity', 'MEDIUM'),
        target=data.get('target', ''),
        scan_id=data.get('scan_id'),
        mitre_tactic=data.get('mitre_tactic', ''),
        mitre_technique=data.get('mitre_technique', ''),
        rule_id=data.get('rule_id', ''),
    )
    return jsonify({'id': iid, 'created': True}), 201


@app.route('/api/incidents/<int:iid>', methods=['PATCH'])
def api_incidents_update(iid):
    data = request.get_json(force=True, silent=True) or {}
    ok   = db.update_incident(iid, **data)
    return jsonify({'updated': ok})


@app.route('/api/unified/overview')
def api_unified_overview():
    """Aggregate metrics for the unified observability dashboard."""
    from dashboard.mitre_rules import get_coverage
    scans      = db.get_recent_scans(50, username=_cu_filter())
    coverage   = get_coverage(scans)

    # Build signal timeline (signals per day from recent scans)
    import re as _re
    from collections import defaultdict
    daily: dict = defaultdict(int)
    for s in scans:
        day = s.get('created_at', '')[:10]
        text = s.get('output', '')
        # Count high-severity pattern hits
        if _re.search(r'sql\s*inject|xss|rce|brute.force|exposed.*key|path.traversal', text, _re.I):
            daily[day] += 1

    # Slowest pages from latest scan per target
    slowest = sorted(
        [{'target': s['target'], 'latency': s.get('latency_s', 0),
          'agent': s.get('agent_type', '')}
         for s in scans],
        key=lambda x: x['latency'], reverse=True
    )[:10]

    # Error rate from scan statuses
    total = len(scans)
    errors = sum(1 for s in scans if s.get('status') != 'ok')
    error_rate = round(errors / max(total, 1) * 100, 1)

    # Recent high-severity signals
    from dashboard.mitre_rules import evaluate_rules
    recent_signals = []
    seen = set()
    for s in scans[:20]:
        for match in evaluate_rules(s.get('output', ''), s.get('target', '')):
            if match['severity'] in ('HIGH',) and match['id'] not in seen:
                seen.add(match['id'])
                recent_signals.append({**match, 'date': s.get('created_at', '')[:10]})

    return jsonify({
        'signal_timeline':  dict(sorted(daily.items())[-14:]),
        'slowest_pages':    slowest,
        'error_rate_pct':   error_rate,
        'total_scans':      total,
        'mitre_total':      coverage['total'],
        'mitre_severities': coverage['severities'],
        'recent_high_signals': recent_signals[:10],
        'incident_stats':   db.get_incident_stats(username=_cu_filter()),
    })


@app.route('/api/stream/signals')
def api_stream_signals():
    """SSE endpoint — pushes a JSON signal-stats event every 15 s."""
    from flask import Response, stream_with_context

    def _event_stream():
        while True:
            try:
                stats = db.get_stats()
                inc   = db.get_incident_stats(username=_cu_filter())
                recent = db.get_recent_scans(5, username=_cu_filter())
                payload = _json.dumps({
                    'total_scans': stats['total_scans'],
                    'open_incidents': inc['open'],
                    'latest_target': recent[0]['target'] if recent else '',
                    'latest_status': recent[0]['status'] if recent else '',
                })
                yield f'data: {payload}\n\n'
            except Exception:
                yield 'data: {}\n\n'
            _time.sleep(15)

    return Response(stream_with_context(_event_stream()),
                    mimetype='text/event-stream',
                    headers={'Cache-Control': 'no-cache',
                             'X-Accel-Buffering': 'no'})


@app.route('/api/scans/<int:scan_id>', methods=['DELETE'])
def api_delete_scan(scan_id):
    deleted = db.delete_scan(scan_id)
    if not deleted:
        return jsonify({'error': 'not found'}), 404
    return jsonify({'deleted': scan_id})


@app.route('/api/scans/clear', methods=['DELETE'])
def api_clear_scans():
    n = db.clear_scans()
    return jsonify({'cleared': n})


@app.route('/api/incidents/<int:iid>', methods=['DELETE'])
def api_incidents_delete(iid):
    deleted = db.delete_incident(iid)
    if not deleted:
        return jsonify({'error': 'not found'}), 404
    return jsonify({'deleted': iid})


@app.route('/api/login-events')
def api_login_events():
    import re as _re
    target = request.args.get('target', '')
    limit  = int(request.args.get('limit', 200))
    scans  = db.get_recent_scans(500, username=_cu_filter())
    events = []
    wp_log_re = _re.compile(
        r'^WP-LOG\s*\|\s*([^|\n]+?)\s*\|\s*([^|\n]+?)\s*\|\s*([^|\n]+?)\s*\|\s*'
        r'(\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}|[^|\n]*?)\s*\|\s*(HIGH|MEDIUM|LOW|INFO)',
        _re.I | _re.MULTILINE,
    )
    login_kw  = _re.compile(r'logged?\s*in|login|sign.in|authentication|session\s*start', _re.I)
    failed_kw = _re.compile(r'failed|invalid|incorrect|denied|blocked', _re.I)
    ip_re     = _re.compile(r'^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$')
    for scan in scans:
        if target and target not in (scan.get('target') or ''):
            continue
        for m in wp_log_re.finditer(scan.get('output') or ''):
            ts, user, event, ip, sev = m.group(1).strip(), m.group(2).strip(), m.group(3).strip(), m.group(4).strip(), m.group(5).strip()
            if not login_kw.search(event):
                continue
            if user.lower() in {'system', 'cf_ai', 'cf_ai-mcp', 'cf_ai_mcp', 'scanner'}:
                continue
            events.append({
                'timestamp': ts, 'user': user, 'event': event[:80], 'ip': ip,
                'country': '', 'status': 'failed' if failed_kw.search(event) else 'success',
                'severity': sev, 'target': scan.get('target', ''), 'scan_date': scan.get('created_at', ''),
            })
    return jsonify({'events': events[:limit], 'total': len(events)})


@app.route('/api/vuln-intel/correlate')
def api_vuln_correlate():
    from dashboard import vuln_intel as _vi
    target = request.args.get('target', '')
    limit  = int(request.args.get('limit', 200))
    scans  = db.get_recent_scans(limit, username=_cu_filter())
    if target:
        scans = [s for s in scans if target.lower() in (s.get('target') or '').lower()]
    return jsonify(_vi.correlate_scans(scans, max_cves=50))


@app.route('/api/vuln-intel/kev')
def api_vuln_kev():
    from dashboard import vuln_intel as _vi
    cve_param = request.args.get('cve', '')
    if cve_param:
        ids = [c.strip().upper() for c in cve_param.split(',') if c.strip()]
        return jsonify({'results': _vi.kev_lookup(ids)})
    stats   = _vi.kev_stats()
    kev_map = _vi._load_kev()
    return jsonify({'total': stats['total'], 'entries': [{'cve_id': k, **v} for k, v in list(kev_map.items())[:200]]})


@app.route('/api/vuln-intel/epss')
def api_vuln_epss():
    from dashboard import vuln_intel as _vi
    cves_param = request.args.get('cves', '')
    if not cves_param:
        return jsonify({'error': 'cves parameter required'}), 400
    ids = [c.strip().upper() for c in cves_param.split(',') if c.strip()]
    return jsonify({'results': _vi.epss_lookup(ids)})


@app.route('/api/export/powerbi')
def api_export_powerbi():
    from dashboard import vuln_intel as _vi
    import time as _t, re as _re
    limit        = int(request.args.get('limit', 500))
    include_intel = request.args.get('include_intel', 'true').lower() != 'false'
    scans = db.get_recent_scans(limit, username=_cu_filter())
    scan_rows = []
    for s in scans:
        out = s.get('output', '')
        scan_rows.append({
            'ScanId':      s.get('id'),
            'Target':      s.get('target', ''),
            'AgentType':   s.get('agent_type', ''),
            'Model':       s.get('model', ''),
            'Status':      s.get('status', ''),
            'LatencyS':    s.get('latency_s', 0),
            'ToolCount':   s.get('tool_count', 0),
            'Date':        (s.get('created_at') or '')[:10],
            'DateTime':    (s.get('created_at') or '').replace(' ', 'T'),
            'HasCritical': bool(_re.search(r'CODE\s+INJECTION\s+CONFIRMED:|CMD\s+INJECTION:|CREDS_FOUND|SQL\s+ERROR:|SSTI\s+HIT|\|\s*Critical\s*\|', out, _re.I)),
            'HasHigh':     bool(_re.search(r'REFLECTED\s+XSS:|FOUND_DB_USER:|APP_PASS_CREATED|WP-LOG.*HIGH|\|\s*High\s*\|', out, _re.I)),
        })
    cve_rows, kev_summary = [], {}
    if include_intel:
        intel = _vi.correlate_scans(scans, max_cves=100)
        for row in intel.get('cves', []):
            nvd = row.get('nvd') or {}
            cve_rows.append({
                'CveId':           row['cve_id'],
                'InKEV':           row['in_kev'],
                'EpssScore':       row.get('epss_score'),
                'CvssScore':       nvd.get('cvss_score'),
                'CvssSeverity':    nvd.get('severity', ''),
                'Published':       nvd.get('published', ''),
                'Description':     (nvd.get('description') or '')[:200],
                'AffectedTargets': ', '.join(row.get('affected_targets', [])),
            })
        kev_summary = _vi.kev_stats()
    return jsonify({
        'schema_version': '1.0',
        'exported_at':    _t.strftime('%Y-%m-%dT%H:%M:%SZ', _t.gmtime()),
        'dataset_name':   'CyberINK Security Intelligence',
        'tables': {
            'scans':     scan_rows,
            'cves':      cve_rows,
            'kev_stats': [kev_summary] if kev_summary else [],
        },
        'powerbi_notes': (
            'Import via Power BI Desktop: Home > Get Data > JSON. '
            'Expand tables record, then load scans and cves tables.'
        ),
    })


# ══════════════════════════════════════════════════════════════════════════════
# PENTEST — Findings, Checklist, Scope/RoE, Report
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/pentest/engagements/<int:eid>/findings', methods=['GET'])
@login_required
def api_pt_findings_list(eid):
    findings = db.get_pt_findings(eid)
    return jsonify({'findings': findings})


@app.route('/api/pentest/engagements/<int:eid>/findings', methods=['POST'])
@login_required
def api_pt_findings_add(eid):
    d = request.get_json(silent=True) or {}
    fid = db.add_pt_finding(
        engagement_id=eid,
        phase=d.get('phase', ''),
        severity=d.get('severity', 'informational'),
        title=d.get('title', 'Untitled Finding'),
        asset=d.get('asset', ''),
        description=d.get('description', ''),
        steps=d.get('steps', ''),
        evidence=d.get('evidence', ''),
        cvss_score=float(d.get('cvss_score', 0)),
        cve=d.get('cve', ''),
        cwe=d.get('cwe', ''),
        remediation=d.get('remediation', ''),
        status=d.get('status', 'open'),
    )
    return jsonify({'ok': True, 'id': fid})


@app.route('/api/pentest/findings/<int:fid>', methods=['PUT'])
@login_required
def api_pt_finding_update(fid):
    d = request.get_json(silent=True) or {}
    allowed = {'phase', 'severity', 'title', 'asset', 'description', 'steps',
               'evidence', 'cvss_score', 'cve', 'cwe', 'remediation', 'status'}
    kwargs = {k: v for k, v in d.items() if k in allowed}
    if 'cvss_score' in kwargs:
        kwargs['cvss_score'] = float(kwargs['cvss_score'])
    db.update_pt_finding(fid, **kwargs)
    return jsonify({'ok': True})


@app.route('/api/pentest/findings/<int:fid>', methods=['DELETE'])
@login_required
def api_pt_finding_delete(fid):
    ok = db.delete_pt_finding(fid)
    return jsonify({'ok': ok})


@app.route('/api/pentest/engagements/<int:eid>/findings/export', methods=['GET'])
@login_required
def api_pt_findings_export(eid):
    import io, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    eng = db.get_engagement(eid)
    if not eng:
        return jsonify({'error': 'Not found'}), 404
    findings = db.get_pt_findings(eid)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Findings'
    hdr_fill = PatternFill('solid', fgColor='1F4E79')
    hdr_font = Font(color='FFFFFF', bold=True)
    headers = ['ID', 'Phase', 'Severity', 'CVSS', 'Title', 'Asset',
               'CVE', 'CWE', 'Status', 'Description', 'Steps to Reproduce',
               'Evidence', 'Remediation']
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill; c.font = hdr_font
        c.alignment = Alignment(wrap_text=True)
    sev_colors = {'critical': 'C00000', 'high': 'FF0000', 'medium': 'ED7D31',
                  'low': 'FFD966', 'informational': '92D050'}
    for ri, f in enumerate(findings, 2):
        row_data = [f['id'], f['phase'], f['severity'].upper(), f['cvss_score'],
                    f['title'], f['asset'], f['cve'], f['cwe'], f['status'],
                    f['description'], f['steps'], f['evidence'], f['remediation']]
        for ci, val in enumerate(row_data, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            if ci == 3:  # severity column
                color = sev_colors.get(f['severity'].lower(), 'FFFFFF')
                c.fill = PatternFill('solid', fgColor=color)
                c.font = Font(bold=True, color='FFFFFF' if f['severity'].lower() in ('critical', 'high') else '000000')
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 7
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 25
    ws.column_dimensions['G'].width = 18
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 12
    for col in ('J', 'K', 'L', 'M'):
        ws.column_dimensions[col].width = 50

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_name = ''.join(c for c in eng['name'] if c.isalnum() or c in '_ -')
    return send_file(buf, as_attachment=True,
                     download_name=f'findings_{safe_name}.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/pentest/engagements/<int:eid>/checklist', methods=['GET'])
@login_required
def api_pt_checklist_get(eid):
    items = db.get_pt_checklist(eid)
    return jsonify({'checklist': items})


@app.route('/api/pentest/engagements/<int:eid>/checklist', methods=['POST'])
@login_required
def api_pt_checklist_set(eid):
    d = request.get_json(silent=True) or {}
    section = d.get('section', '')
    item    = d.get('item', '')
    checked = bool(d.get('checked', False))
    if not section or not item:
        return jsonify({'error': 'section and item required'}), 400
    db.set_pt_checklist_item(eid, section, item, checked)
    return jsonify({'ok': True})


@app.route('/api/pentest/engagements/<int:eid>/scope', methods=['GET'])
@login_required
def api_pt_scope_get(eid):
    scope = db.get_engagement_scope(eid)
    return jsonify(scope)


@app.route('/api/pentest/engagements/<int:eid>/scope', methods=['PUT'])
@login_required
def api_pt_scope_save(eid):
    d = request.get_json(silent=True) or {}
    scope_doc = d.get('scope_doc')
    roe_doc   = d.get('roe_doc')
    db.save_engagement_scope(eid, scope_doc=scope_doc, roe_doc=roe_doc)
    return jsonify({'ok': True})


@app.route('/api/pentest/engagements/<int:eid>/report/generate', methods=['POST'])
@login_required
def api_pt_report_generate(eid):
    import json as _json
    eng = db.get_engagement(eid)
    if not eng:
        return jsonify({'error': 'Engagement not found'}), 404
    findings = db.get_pt_findings(eid)
    scope    = db.get_engagement_scope(eid)
    model    = request.get_json(silent=True) or {}
    model_id = model.get('model', os.environ.get('CAI_MODEL', 'claude-sonnet-4-6'))

    sev_counts = {'critical': 0, 'high': 0, 'medium': 0, 'low': 0, 'informational': 0}
    for f in findings:
        sev = f.get('severity', 'informational').lower()
        sev_counts[sev] = sev_counts.get(sev, 0) + 1

    findings_text = ''
    for i, f in enumerate(findings, 1):
        findings_text += (
            f"\n\n--- Finding {i}: {f['title']} ---\n"
            f"Severity: {f['severity'].upper()} | CVSS: {f['cvss_score']} | Phase: {f['phase']}\n"
            f"Asset: {f['asset']}\n"
            f"CVE: {f.get('cve','')} | CWE: {f.get('cwe','')}\n"
            f"Status: {f['status']}\n\n"
            f"Description:\n{f['description']}\n\n"
            f"Steps to Reproduce:\n{f['steps']}\n\n"
            f"Evidence:\n{f['evidence']}\n\n"
            f"Remediation:\n{f['remediation']}\n"
        )

    prompt = f"""You are a senior penetration tester writing an executive and technical pentest report.

Engagement: {eng['name']}
Client: {eng['client']}
Status: {eng['status']}
Deadline: {eng['deadline']}
Severity Summary: Critical={sev_counts['critical']}, High={sev_counts['high']}, Medium={sev_counts['medium']}, Low={sev_counts['low']}, Informational={sev_counts['informational']}

Scope Document:
{scope.get('scope_doc','(none)')}

Rules of Engagement:
{scope.get('roe_doc','(none)')}

Findings ({len(findings)} total):
{findings_text if findings_text else '(No findings recorded yet)'}

Write a comprehensive penetration test report with these sections:
1. Executive Summary (3-4 paragraphs for non-technical leadership)
2. Engagement Overview (scope, methodology, timeline, tools used)
3. Risk Summary (overall risk rating with justification)
4. Findings Summary Table (list all findings with severity and status)
5. Detailed Findings (full write-up for each finding — description, impact, evidence, remediation)
6. Strategic Recommendations (prioritised remediation roadmap)
7. Conclusion

Format as clean prose. Use markdown headings. Be specific, professional, and actionable.
Do NOT invent findings — only report what is listed above."""

    try:
        if model_id.startswith('gpt') or model_id.startswith('o1'):
            import openai
            client_ai = openai.OpenAI(api_key=os.environ.get('OPENAI_API_KEY', ''))
            resp = client_ai.chat.completions.create(
                model=model_id,
                messages=[{'role': 'user', 'content': prompt}],
                max_tokens=4096,
            )
            report_md = resp.choices[0].message.content or ''
        else:
            import anthropic
            client_ai = anthropic.Anthropic(api_key=os.environ.get('ANTHROPIC_API_KEY', ''))
            resp = client_ai.messages.create(
                model=model_id,
                max_tokens=4096,
                messages=[{'role': 'user', 'content': prompt}],
            )
            report_md = resp.content[0].text if resp.content else ''
    except Exception as exc:
        return jsonify({'error': str(exc)}), 500

    return jsonify({
        'report_md': report_md,
        'sev_counts': sev_counts,
        'total_findings': len(findings),
        'engagement': {'name': eng['name'], 'client': eng['client']},
    })


# ══════════════════════════════════════════════════════════════════════════════
# GRC COMPLIANCE ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════

def _grc_extract_text(file_obj) -> str | None:
    """Extract plain text from uploaded PDF, DOCX, or TXT policy file."""
    name = (file_obj.filename or '').lower()
    data = file_obj.read()

    if name.endswith('.txt') or name.endswith('.md'):
        try:
            return data.decode('utf-8', errors='replace')
        except Exception:
            return data.decode('latin-1', errors='replace')

    if name.endswith('.docx'):
        try:
            import io
            from docx import Document as _DocxDoc
            doc = _DocxDoc(io.BytesIO(data))
            return '\n'.join(p.text for p in doc.paragraphs if p.text.strip())
        except ImportError:
            return f'[python-docx not installed — cannot read .docx files. Install with: pip install python-docx]'
        except Exception as e:
            return f'[DOCX read error: {e}]'

    if name.endswith('.pdf'):
        try:
            import io
            try:
                import pdfplumber as _pp
                with _pp.open(io.BytesIO(data)) as pdf:
                    return '\n'.join(p.extract_text() or '' for p in pdf.pages)
            except ImportError:
                pass
            try:
                from pypdf import PdfReader as _PR
                reader = _PR(io.BytesIO(data))
                return '\n'.join(page.extract_text() or '' for page in reader.pages)
            except ImportError:
                pass
            return '[PDF library not installed. Install pdfplumber or pypdf: pip install pdfplumber]'
        except Exception as e:
            return f'[PDF read error: {e}]'

    return None  # unsupported


_GRC_POLICY_TYPES = {
    'cybersecurity':  'Cybersecurity Policy',
    'password':       'Password Policy',
    'backup':         'Backup Policy',
    'access control': 'Access Control Policy',
    'incident':       'Incident Response Policy',
    'device':         'Device Usage Policy',
    'data':           'Data Classification Policy',
    'remote':         'Remote Access Policy',
    'acceptable use': 'Acceptable Use Policy',
    'change':         'Change Management Policy',
    'vendor':         'Vendor / Third-Party Policy',
    'encryption':     'Encryption Policy',
    'physical':       'Physical Security Policy',
    'business continuity': 'Business Continuity / DR Policy',
    'vulnerability':  'Vulnerability Management Policy',
    'patch':          'Patch Management Policy',
    'network':        'Network Security Policy',
    'cloud':          'Cloud Security Policy',
    'gdpr':           'GDPR / Privacy Policy',
    'audit':          'Audit & Logging Policy',
}

def _grc_detect_policy_type(filename: str, content: str) -> str:
    """Guess policy type from filename + first 500 chars of content."""
    text = (filename + ' ' + content[:500]).lower()
    for keyword, label in _GRC_POLICY_TYPES.items():
        if keyword in text:
            return label
    return 'General Security Policy'


def _grc_analyze_policy(filename: str, content: str, standards: list[str]) -> dict:
    """Send policy text to Claude and return structured compliance findings."""
    import os as _os

    policy_type = _grc_detect_policy_type(filename, content)

    standards_block = '\n'.join(
        f'- {s}' for s in (standards or ['ISO 27001', 'CIS Controls', 'NIST CSF'])
    )

    # Truncate to keep within model context — 12k chars is plenty for policy analysis
    excerpt = content[:12000]
    if len(content) > 12000:
        excerpt += f'\n\n[... {len(content)-12000} more chars truncated ...]'

    system_prompt = (
        'You are a senior GRC (Governance, Risk & Compliance) consultant specialising in '
        'information security policy review. You have deep knowledge of ISO/IEC 27001:2022, '
        'CIS Controls v8, and NIST Cybersecurity Framework (CSF) 2.0.'
    )

    user_prompt = f"""Review the following policy document and assess its compliance with each requested standard.

POLICY TYPE: {policy_type}
FILENAME: {filename}

STANDARDS TO CHECK:
{standards_block}

--- POLICY DOCUMENT START ---
{excerpt}
--- POLICY DOCUMENT END ---

Respond ONLY in this exact JSON format (no markdown fences, pure JSON):
{{
  "policy_type": "{policy_type}",
  "summary": "<2-3 sentence executive summary of the policy quality>",
  "overall_score": <0-100 integer>,
  "overall_rating": "<Excellent|Good|Needs Improvement|Poor>",
  "standards": [
    {{
      "name": "<standard name>",
      "score": <0-100>,
      "rating": "<Excellent|Good|Needs Improvement|Poor>",
      "covered_controls": ["<control or clause that is addressed>", ...],
      "gaps": ["<missing control or requirement>", ...],
      "recommendations": ["<specific actionable fix>", ...]
    }}
  ],
  "critical_gaps": ["<top 3-5 most critical missing items across all standards>"],
  "strengths": ["<things the policy does well>"],
  "quick_wins": ["<easy improvements that would raise the score quickly>"]
}}"""

    model = _os.environ.get('CAI_MODEL', 'claude-sonnet-4-6')

    try:
        if model.startswith('gpt') or model.startswith('o1') or model.startswith('o3'):
            import openai as _oai
            client = _oai.OpenAI(api_key=_os.environ.get('OPENAI_API_KEY', ''))
            resp = client.chat.completions.create(
                model=model,
                messages=[
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user',   'content': user_prompt},
                ],
                temperature=0.1,
                max_tokens=4096,
            )
            raw = resp.choices[0].message.content or ''
        else:
            import anthropic as _ant
            client = _ant.Anthropic(api_key=_os.environ.get('ANTHROPIC_API_KEY', ''))
            resp = client.messages.create(
                model=model,
                max_tokens=4096,
                system=system_prompt,
                messages=[{'role': 'user', 'content': user_prompt}],
            )
            raw = resp.content[0].text if resp.content else ''

        # Strip markdown fences if the model added them despite instructions
        raw = raw.strip()
        if raw.startswith('```'):
            raw = raw.split('```', 2)[-1].strip()
            if raw.startswith('json'):
                raw = raw[4:].strip()
            if raw.endswith('```'):
                raw = raw[:-3].strip()

        result = _json.loads(raw)
        result['filename'] = filename
        result['content_len'] = len(content)
        return result

    except _json.JSONDecodeError as e:
        return {
            'filename':    filename,
            'policy_type': policy_type,
            'error':       f'JSON parse error: {e}',
            'raw':         raw[:500] if 'raw' in dir() else '',
        }
    except Exception as e:
        return {
            'filename':    filename,
            'policy_type': policy_type,
            'error':       str(e),
        }


def _grc_build_excel(results: list[dict]):
    """Build an openpyxl workbook from GRC analysis results."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise RuntimeError('openpyxl not installed — run: pip install openpyxl')

    wb = openpyxl.Workbook()

    # ── colour palette ──────────────────────────────────────────────────────
    CLR_HDR_BG  = 'FF1E3A8A'   # dark blue header
    CLR_HDR_FG  = 'FFFFFFFF'
    CLR_SEC_BG  = 'FFDBEAFE'   # light blue section
    CLR_RED     = 'FFFEE2E2'
    CLR_YEL     = 'FFFEF9C3'
    CLR_GRN     = 'FFD1FAE5'
    CLR_BORDER  = 'FFE5E7EB'

    def _hdr_fill(): return PatternFill('solid', fgColor=CLR_HDR_BG)
    def _sec_fill(): return PatternFill('solid', fgColor=CLR_SEC_BG)
    def _score_fill(s):
        if s >= 75: return PatternFill('solid', fgColor=CLR_GRN)
        if s >= 50: return PatternFill('solid', fgColor=CLR_YEL)
        return PatternFill('solid', fgColor=CLR_RED)
    def _thin_border():
        s = Side(style='thin', color=CLR_BORDER)
        return Border(left=s, right=s, top=s, bottom=s)
    def _hdr_font(): return Font(bold=True, color=CLR_HDR_FG, size=10)
    def _bold():     return Font(bold=True, size=10)

    def _col_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def _write_row(ws, row, values, fill=None, font=None, wrap=False):
        for col, val in enumerate(values, 1):
            c = ws.cell(row=row, column=col, value=val)
            if fill:  c.fill   = fill
            if font:  c.font   = font
            c.border    = _thin_border()
            c.alignment = Alignment(wrap_text=wrap, vertical='top')
        return row + 1

    # ── Sheet 1: Executive Summary ──────────────────────────────────────────
    ws = wb.active
    ws.title = 'Executive Summary'
    _col_widths(ws, [32, 22, 14, 14, 52])

    row = _write_row(ws, 1,
        ['Policy File', 'Policy Type', 'Overall Score', 'Rating', 'Summary'],
        fill=_hdr_fill(), font=_hdr_font())

    for r in results:
        if r.get('error'):
            _write_row(ws, row, [r['filename'], r.get('policy_type','?'), 'ERROR', '', r['error']])
        else:
            score = r.get('overall_score', 0)
            row = _write_row(ws, row,
                [r['filename'], r.get('policy_type','?'), score,
                 r.get('overall_rating','?'), r.get('summary','')],
                fill=_score_fill(score), wrap=True)

    ws.freeze_panes = 'A2'

    # ── Sheet 2: Standards Detail ───────────────────────────────────────────
    ws2 = wb.create_sheet('Standards Detail')
    _col_widths(ws2, [28, 22, 22, 14, 14, 48, 48, 48])

    row = _write_row(ws2, 1,
        ['Policy File', 'Policy Type', 'Standard', 'Score', 'Rating',
         'Covered Controls', 'Gaps', 'Recommendations'],
        fill=_hdr_fill(), font=_hdr_font())

    for r in results:
        if r.get('error'):
            continue
        for s in r.get('standards', []):
            score = s.get('score', 0)
            row = _write_row(ws2, row,
                [r['filename'], r.get('policy_type','?'), s.get('name',''),
                 score, s.get('rating','?'),
                 '\n'.join(f'• {x}' for x in s.get('covered_controls', [])),
                 '\n'.join(f'• {x}' for x in s.get('gaps', [])),
                 '\n'.join(f'• {x}' for x in s.get('recommendations', []))],
                fill=_score_fill(score), wrap=True)

    ws2.freeze_panes = 'A2'
    ws2.row_dimensions[1].height = 20

    # ── Sheet 3: Critical Gaps & Quick Wins ────────────────────────────────
    ws3 = wb.create_sheet('Gaps & Quick Wins')
    _col_widths(ws3, [28, 22, 8, 56])

    row = _write_row(ws3, 1,
        ['Policy File', 'Policy Type', 'Category', 'Item'],
        fill=_hdr_fill(), font=_hdr_font())

    for r in results:
        if r.get('error'):
            continue
        for item in r.get('critical_gaps', []):
            row = _write_row(ws3, row,
                [r['filename'], r.get('policy_type','?'), 'GAP', item],
                fill=PatternFill('solid', fgColor=CLR_RED), wrap=True)
        for item in r.get('quick_wins', []):
            row = _write_row(ws3, row,
                [r['filename'], r.get('policy_type','?'), 'WIN', item],
                fill=PatternFill('solid', fgColor=CLR_GRN), wrap=True)
        for item in r.get('strengths', []):
            row = _write_row(ws3, row,
                [r['filename'], r.get('policy_type','?'), 'STRENGTH', item],
                fill=PatternFill('solid', fgColor=CLR_YEL), wrap=True)

    ws3.freeze_panes = 'A2'

    return wb


@app.route('/api/grc/analyze', methods=['POST'])
@login_required
def api_grc_analyze():
    """Analyze uploaded policy documents against security standards."""
    files     = request.files.getlist('files')
    standards = request.form.getlist('standards') or ['ISO 27001', 'CIS Controls', 'NIST CSF']

    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'No files uploaded'}), 400

    results = []
    for f in files:
        if not f.filename:
            continue
        text = _grc_extract_text(f)
        if text is None:
            results.append({
                'filename':    f.filename,
                'policy_type': 'Unknown',
                'error':       'Unsupported file type. Upload .txt, .md, .docx, or .pdf',
            })
            continue
        analysis = _grc_analyze_policy(f.filename, text, standards)
        results.append(analysis)

    return jsonify({'results': results})


@app.route('/api/grc/export', methods=['POST'])
@login_required
def api_grc_export():
    """Export GRC analysis as an Excel workbook."""
    import io as _io
    d       = request.get_json(silent=True) or {}
    results = d.get('results', [])
    if not results:
        return jsonify({'error': 'No results to export'}), 400
    try:
        wb = _grc_build_excel(results)
    except RuntimeError as e:
        return jsonify({'error': str(e)}), 500
    buf = _io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='grc_compliance_report.xlsx',
    )


# ══════════════════════════════════════════════════════════════════════════════
# GRC RISK MANAGEMENT (grc2)
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/grc2/stats')
@login_required
def api_grc2_stats():
    username = session['user']['username']
    return jsonify(db.grc_stats(username=username))


@app.route('/api/grc2/risks', methods=['GET'])
@login_required
def api_grc2_risks_list():
    q = request.args.get('q', '')
    status = request.args.get('status', '')
    treatment   = request.args.get('treatment', '')
    risk_status = request.args.get('risk_status', '')
    username = session['user']['username']
    return jsonify({'risks': db.grc_list_risks(q, status, treatment, risk_status, username=username)})


@app.route('/api/grc2/risks', methods=['POST'])
@login_required
def api_grc2_risks_create():
    d = request.get_json(silent=True) or {}
    if not d.get('title'):
        return jsonify({'error': 'title required'}), 400
    d['username'] = session['user'].get('username', '')
    rid = db.grc_create_risk(d)
    return jsonify({'id': rid})


@app.route('/api/grc2/risks/<int:rid>', methods=['PUT'])
@login_required
def api_grc2_risks_update(rid):
    d = request.get_json(silent=True) or {}
    db.grc_update_risk(rid, d)
    return jsonify({'ok': True})


@app.route('/api/grc2/risks/<int:rid>', methods=['DELETE'])
@login_required
def api_grc2_risks_delete(rid):
    db.grc_delete_risk(rid)
    return jsonify({'ok': True})


@app.route('/api/grc2/controls', methods=['GET'])
@login_required
def api_grc2_controls_list():
    q = request.args.get('q', '')
    framework = request.args.get('framework', '')
    status = request.args.get('status', '')
    username = session['user']['username']
    return jsonify({'controls': db.grc_list_controls(q, framework, status, username=username)})


@app.route('/api/grc2/controls', methods=['POST'])
@login_required
def api_grc2_controls_create():
    d = request.get_json(silent=True) or {}
    if not d.get('title') or not d.get('control_id'):
        return jsonify({'error': 'control_id and title required'}), 400
    d['username'] = session['user'].get('username', '')
    cid = db.grc_create_control(d)
    return jsonify({'id': cid})


@app.route('/api/grc2/controls/<int:cid>', methods=['PUT'])
@login_required
def api_grc2_controls_update(cid):
    d = request.get_json(silent=True) or {}
    db.grc_update_control(cid, d)
    return jsonify({'ok': True})


@app.route('/api/grc2/controls/<int:cid>', methods=['DELETE'])
@login_required
def api_grc2_controls_delete(cid):
    db.grc_delete_control(cid)
    return jsonify({'ok': True})


@app.route('/api/grc2/tests', methods=['GET'])
@login_required
def api_grc2_tests_list():
    q             = request.args.get('q', '')
    category      = request.args.get('category', '')
    status        = request.args.get('status', '')
    test_category = request.args.get('test_category', '')
    username = session['user']['username']
    return jsonify({'tests': db.grc_list_tests(q, category, status, test_category, username=username)})


@app.route('/api/grc2/tests', methods=['POST'])
@login_required
def api_grc2_tests_create():
    d = request.get_json(silent=True) or {}
    if not d.get('name'):
        return jsonify({'error': 'name required'}), 400
    d['username'] = session['user'].get('username', '')
    tid = db.grc_create_test(d)
    return jsonify({'id': tid})


@app.route('/api/grc2/tests/<int:tid>', methods=['PUT'])
@login_required
def api_grc2_tests_update(tid):
    d = request.get_json(silent=True) or {}
    db.grc_update_test(tid, d)
    return jsonify({'ok': True})


@app.route('/api/grc2/tests/<int:tid>', methods=['DELETE'])
@login_required
def api_grc2_tests_delete(tid):
    db.grc_delete_test(tid)
    return jsonify({'ok': True})


@app.route('/api/grc2/audits', methods=['GET'])
@login_required
def api_grc2_audits_list():
    username = session['user']['username']
    return jsonify({'audits': db.grc_list_audits(username=username)})


@app.route('/api/grc2/audits', methods=['POST'])
@login_required
def api_grc2_audits_create():
    d = request.get_json(silent=True) or {}
    if not d.get('name'):
        return jsonify({'error': 'name required'}), 400
    d['username'] = session['user'].get('username', '')
    aid = db.grc_create_audit(d)
    return jsonify({'id': aid})


@app.route('/api/grc2/audits/<int:aid>', methods=['PUT'])
@login_required
def api_grc2_audits_update(aid):
    d = request.get_json(silent=True) or {}
    db.grc_update_audit(aid, d)
    return jsonify({'ok': True})


@app.route('/api/grc2/audits/<int:aid>', methods=['DELETE'])
@login_required
def api_grc2_audits_delete(aid):
    db.grc_delete_audit(aid)
    return jsonify({'ok': True})


@app.route('/api/grc2/evidence', methods=['GET'])
@login_required
def api_grc2_evidence_list():
    audit_id  = request.args.get('audit_id', '')
    ev_status = request.args.get('evidence_status', '')
    aid = int(audit_id) if audit_id.isdigit() else None
    username = session['user']['username']
    return jsonify({'evidence': db.grc_list_evidence(aid, ev_status, username=username)})


@app.route('/api/grc2/evidence', methods=['POST'])
@login_required
def api_grc2_evidence_create():
    d = request.get_json(silent=True) or {}
    if not d.get('title'):
        return jsonify({'error': 'title required'}), 400
    d['username'] = session['user'].get('username', '')
    eid = db.grc_create_evidence(d)
    return jsonify({'id': eid})


@app.route('/api/grc2/evidence/<int:eid>', methods=['PUT'])
@login_required
def api_grc2_evidence_update(eid):
    d = request.get_json(silent=True) or {}
    db.grc_update_evidence(eid, d)
    return jsonify({'ok': True})


@app.route('/api/grc2/evidence/<int:eid>', methods=['DELETE'])
@login_required
def api_grc2_evidence_delete(eid):
    db.grc_delete_evidence(eid)
    return jsonify({'ok': True})


@app.route('/api/grc2/audits/<int:aid>/evidence-stats')
@login_required
def api_grc2_audit_evidence_stats(aid):
    return jsonify(db.grc_evidence_stats_by_audit(aid))


# ── Payment routes (Midtrans) ─────────────────────────────────────────────────

@app.route('/api/payment/create-transaction', methods=['POST'])
@login_required
def api_payment_create():
    if not _MIDTRANS_SERVER_KEY or not _MIDTRANS_CLIENT_KEY:
        return jsonify({'error': 'Pembayaran belum dikonfigurasi. Hubungi administrator.'}), 503
    d         = request.get_json(silent=True) or {}
    plan_type = d.get('plan_type', 'monthly')
    if plan_type not in ('monthly', 'annual'):
        return jsonify({'error': 'Tipe paket tidak valid.'}), 400
    u        = session['user']
    username = u.get('username', '')
    email    = u.get('email', '')
    existing = db.get_user_active_subscription(username)
    if existing:
        return jsonify({'error': 'Anda sudah memiliki langganan Pro yang aktif.'}), 409
    amount    = 299000 if plan_type == 'monthly' else 2690000
    item_name = f'CyberINK Pro — {"Bulanan" if plan_type == "monthly" else "Tahunan"}'
    order_id  = f'CYBERINK-{username.upper()}-{_uuid.uuid4().hex[:8].upper()}'
    try:
        result = _midtrans_snap_create(
            order_id=order_id, amount=amount,
            customer={'username': username, 'email': email},
            item_name=item_name, plan_type=plan_type,
        )
    except Exception as exc:
        return jsonify({'error': f'Payment gateway error: {exc}'}), 502
    snap_token = result.get('token', '')
    db.create_subscription(username=username, email=email, order_id=order_id,
                           plan_type=plan_type, amount=amount, snap_token=snap_token)
    return jsonify({'snap_token': snap_token, 'order_id': order_id})


@app.route('/api/payment/notification', methods=['POST'])
def api_payment_notification():
    """Midtrans payment notification webhook — no session auth required."""
    from datetime import datetime as _dt, timedelta as _td
    data               = request.get_json(silent=True) or {}
    order_id           = data.get('order_id', '')
    status_code        = str(data.get('status_code', ''))
    gross_amount       = str(data.get('gross_amount', ''))
    transaction_status = data.get('transaction_status', '')
    fraud_status       = data.get('fraud_status', 'accept')
    payment_type       = data.get('payment_type', '')
    transaction_id     = data.get('transaction_id', '')
    incoming_sig       = data.get('signature_key', '')

    if _MIDTRANS_SERVER_KEY and incoming_sig:
        expected = _midtrans_verify_signature(order_id, status_code, gross_amount)
        if incoming_sig != expected:
            return jsonify({'error': 'Invalid signature'}), 403

    sub = db.get_subscription_by_order_id(order_id)
    if not sub:
        return jsonify({'error': 'Order not found'}), 404

    if transaction_status in ('capture', 'settlement'):
        new_status = 'active' if fraud_status == 'accept' else 'failed'
    elif transaction_status == 'pending':
        new_status = 'pending'
    elif transaction_status in ('deny', 'cancel', 'expire'):
        new_status = 'failed'
    elif transaction_status in ('refund', 'partial_refund'):
        new_status = 'cancelled'
    else:
        new_status = transaction_status

    subscribed_at = ''
    expires_at    = ''
    if new_status == 'active':
        now           = _dt.utcnow()
        subscribed_at = now.strftime('%Y-%m-%d %H:%M:%S')
        delta         = _td(days=31) if sub.get('plan_type') == 'monthly' else _td(days=366)
        expires_at    = (now + delta).strftime('%Y-%m-%d %H:%M:%S')

    bank      = data.get('bank', '') or data.get('issuer', '')
    va_number = ''
    va_list   = data.get('va_numbers', [])
    if va_list and isinstance(va_list, list):
        va_number = va_list[0].get('va_number', '')

    db.update_subscription_status(
        order_id, new_status,
        transaction_id=transaction_id, payment_type=payment_type,
        bank=bank, va_number=va_number,
        subscribed_at=subscribed_at, expires_at=expires_at,
        raw_notification=_json.dumps(data)[:2000],
    )

    users    = _load_users()
    username = sub.get('username', '')
    if new_status == 'active' and username in users:
        users[username]['plan'] = 'pro'
        _save_users(users)
    elif new_status in ('cancelled', 'failed', 'expired') and sub.get('status') == 'active':
        if username in users:
            users[username]['plan'] = 'basic'
            _save_users(users)

    return jsonify({'status': 'ok'})


@app.route('/api/payment/status', methods=['GET'])
@login_required
def api_payment_status():
    username = session['user'].get('username', '')
    expired  = db.expire_stale_subscriptions()
    users    = _load_users()
    for s in expired:
        un = s.get('username', '')
        if un in users:
            users[un]['plan'] = 'basic'
    if expired:
        _save_users(users)
    sub = db.get_user_active_subscription(username)
    all_subs = db.get_user_subscriptions(username)
    return jsonify({'subscription': sub, 'history': all_subs})


@app.route('/api/payment/cancel', methods=['POST'])
@login_required
def api_payment_cancel():
    from datetime import datetime as _dt
    username = session['user'].get('username', '')
    sub      = db.get_user_active_subscription(username)
    if not sub:
        return jsonify({'error': 'Tidak ada langganan aktif yang ditemukan.'}), 404
    db.cancel_subscription(sub['order_id'], _dt.utcnow().strftime('%Y-%m-%d %H:%M:%S'))
    users = _load_users()
    if username in users:
        users[username]['plan'] = 'basic'
        _save_users(users)
    return jsonify({'ok': True})


# ── Admin invoice routes ──────────────────────────────────────────────────────

@app.route('/api/admin/invoices', methods=['GET'])
@login_required
def api_admin_invoices():
    if session['user'].get('role') != 'admin':
        return jsonify({'error': 'Admin access required'}), 403
    db.expire_stale_subscriptions()
    status    = request.args.get('status', '') or None
    plan_type = request.args.get('plan_type', '') or None
    search    = request.args.get('search', '') or None
    try:
        limit  = min(int(request.args.get('limit') or 100), 500)
        offset = int(request.args.get('offset') or 0)
    except (ValueError, TypeError):
        limit, offset = 100, 0
    subs  = db.get_all_subscriptions(status=status, plan_type=plan_type,
                                     search=search, limit=limit, offset=offset)
    total = db.count_all_subscriptions(status=status, plan_type=plan_type, search=search)
    stats = db.get_subscription_stats()
    return jsonify({'subscriptions': subs, 'stats': stats, 'total': total})


@app.route('/api/admin/invoices/<int:sub_id>/cancel', methods=['POST'])
@login_required
def api_admin_cancel_subscription(sub_id):
    if session['user'].get('role') != 'admin':
        return jsonify({'error': 'Admin access required'}), 403
    from datetime import datetime as _dt
    with db._connect() as con:
        row = con.execute('SELECT * FROM subscriptions WHERE id=?', (sub_id,)).fetchone()
        if not row:
            return jsonify({'error': 'Subscription not found'}), 404
        sub = dict(row)
    db.cancel_subscription(sub['order_id'], _dt.utcnow().strftime('%Y-%m-%d %H:%M:%S'))
    users    = _load_users()
    username = sub.get('username', '')
    if username in users:
        active = db.get_user_active_subscription(username)
        if not active:
            users[username]['plan'] = 'basic'
            _save_users(users)
    return jsonify({'ok': True})


def _semgrep_bin():
    """Return path to semgrep binary, checking pipx, venv, and common locations."""
    import shutil as _sh, subprocess as _sp_sg, sys as _sys_sg
    candidates = [
        '/root/.local/bin/semgrep',
        '/root/.local/share/pipx/venvs/semgrep/bin/semgrep',
        '/home/' + (os.environ.get('SUDO_USER') or 'root') + '/.local/bin/semgrep',
        '/usr/local/bin/semgrep',
        '/usr/bin/semgrep',
        _sh.which('semgrep') or '',
    ]
    # Also try `which semgrep` via subprocess (honours runtime PATH)
    try:
        r = _sp_sg.run(['which', 'semgrep'], capture_output=True, text=True, timeout=5)
        if r.returncode == 0 and r.stdout.strip():
            candidates.insert(0, r.stdout.strip())
    except Exception:
        pass
    for c in candidates:
        if c and os.path.exists(c):
            return c
    # Last resort: python -m semgrep (same interpreter)
    try:
        r = _sp_sg.run([_sys_sg.executable, '-m', 'semgrep', '--version'],
                       capture_output=True, text=True, timeout=10)
        if r.returncode == 0:
            return _sys_sg.executable + ' -m semgrep'
    except Exception:
        pass
    return None


def _semgrep_cmd(extra_args=None):
    """Return a list suitable for subprocess for semgrep + extra_args."""
    import sys as _sys_sc
    bin_path = _semgrep_bin()
    if not bin_path:
        return None
    if ' -m semgrep' in bin_path:
        cmd = [_sys_sc.executable, '-m', 'semgrep']
    else:
        cmd = [bin_path]
    if extra_args:
        cmd += extra_args
    return cmd


@app.route('/api/sca/check', methods=['GET'])
@login_required
def api_sca_check():
    import subprocess as _sp
    cmd = _semgrep_cmd(['--version'])
    if not cmd:
        return jsonify({'installed': False})
    try:
        r = _sp.run(cmd, capture_output=True, text=True, timeout=10)
        version = (r.stdout or r.stderr or '').strip().split('\n')[0]
        return jsonify({'installed': True, 'version': version})
    except Exception:
        return jsonify({'installed': False})


@app.route('/api/sca/install', methods=['POST'])
@login_required
def api_sca_install():
    import subprocess as _sp
    import sys as _sys
    def _gen():
        try:
            proc = _sp.Popen(
                [_sys.executable, '-m', 'pip', 'install', '--upgrade', 'semgrep'],
                stdout=_sp.PIPE, stderr=_sp.STDOUT, text=True
            )
            for line in proc.stdout:
                yield f"data: {line.rstrip()}\n\n"
            proc.wait()
            if proc.returncode == 0:
                yield "data: __DONE__\n\n"
            else:
                yield f"data: __ERROR__ Exit code {proc.returncode}\n\n"
        except Exception as exc:
            yield f"data: __ERROR__ {exc}\n\n"
    return Response(stream_with_context(_gen()), content_type='text/event-stream')


@app.route('/api/sca/scan', methods=['POST'])
@login_required
def api_sca_scan():
    import tempfile, subprocess as _sp, json as _json
    from werkzeug.utils import secure_filename as _sf

    files   = request.files.getlist('files')
    ruleset = request.form.get('ruleset', 'auto')
    custom  = request.form.get('custom_rules', '').strip()

    if not files or all(f.filename == '' for f in files):
        return jsonify({'ok': False, 'error': 'No files uploaded'}), 400

    with tempfile.TemporaryDirectory() as tdir:
        saved = []
        for f in files:
            name = _sf(f.filename) or 'upload.txt'
            path = os.path.join(tdir, name)
            f.save(path)
            saved.append(path)

        cmd = _semgrep_cmd(['--json', '--quiet', '--no-git-ignore'])
        if not cmd:
            return jsonify({'ok': False,
                            'error': 'semgrep not installed on this server. Run: pipx install semgrep'}), 503

        if custom:
            rfile = os.path.join(tdir, '_rules.yaml')
            with open(rfile, 'w') as rf:
                rf.write(custom)
            cmd += ['--config', rfile]
        elif ruleset == 'auto':
            cmd += ['--config', 'auto']
        else:
            cmd += ['--config', f'p/{ruleset}']

        cmd.append(tdir)

        try:
            res = _sp.run(cmd, capture_output=True, text=True, timeout=180)
            try:
                data = _json.loads(res.stdout)
            except _json.JSONDecodeError:
                return jsonify({'ok': False, 'error': 'Semgrep output could not be parsed',
                                'stderr': res.stderr[:400]}), 500
            findings = data.get('results', [])
            # Strip temp dir prefix from file paths
            for r in findings:
                r['path'] = os.path.basename(r.get('path', ''))
            return jsonify({'ok': True,
                            'findings': findings,
                            'errors':   data.get('errors', []),
                            'stats':    data.get('stats', {}),
                            'version':  data.get('version', '')})
        except _sp.TimeoutExpired:
            return jsonify({'ok': False, 'error': 'Scan timed out (>180s)'}), 504
        except Exception as exc:
            return jsonify({'ok': False, 'error': str(exc)}), 500


# ── Dynamic Code Analysis (DAST) ─────────────────────────────────────────────

def _dca_find_bin(*names):
    import shutil as _sh
    extra = ['/root/.local/bin', '/usr/local/bin', '/usr/bin', '/bin',
             '/usr/share/nikto', '/usr/lib/nikto']
    for name in names:
        found = _sh.which(name)
        if found:
            return found
        for d in extra:
            p = os.path.join(d, name)
            if os.path.isfile(p):
                return p
    return None


@app.route('/api/dca/scanners', methods=['GET'])
@login_required
def api_dca_scanners():
    import subprocess as _sp
    import json as _jd
    results = {}
    for name, key in [('nikto','nikto'), ('nuclei','nuclei'), ('wapiti','wapiti'), ('wapiti3','wapiti')]:
        bin_path = _dca_find_bin(name)
        if bin_path and key not in results:
            try:
                r = _sp.run([bin_path, '--version' if name != 'nikto' else '-Version'],
                            capture_output=True, text=True, timeout=10)
                ver = ((r.stdout or '') + (r.stderr or '')).strip().split('\n')[0][:80]
                results[key] = {'available': True, 'path': bin_path, 'version': ver}
            except Exception:
                results[key] = {'available': True, 'path': bin_path, 'version': 'unknown'}
    for key in ['nikto', 'nuclei', 'wapiti']:
        if key not in results:
            results[key] = {'available': False}
    try:
        import urllib.request as _ureq
        with _ureq.urlopen('http://localhost:8080/JSON/core/view/version/', timeout=3) as r:
            ver_data = _jd.loads(r.read())
            results['zap'] = {'available': True, 'path': 'localhost:8080', 'api': True,
                              'version': ver_data.get('version', 'running')}
    except Exception:
        zap_bin = _dca_find_bin('zaproxy', 'zap.sh', 'zap')
        results['zap'] = {'available': bool(zap_bin), 'path': zap_bin or ''}
    return jsonify(results)


@app.route('/api/dca/scan', methods=['POST'])
@login_required
def api_dca_scan():
    import subprocess as _sp
    import tempfile as _tmp
    import json as _jd
    import base64 as _b64d

    data = request.get_json(silent=True) or {}
    target      = (data.get('target') or '').strip()
    scanner     = (data.get('scanner') or 'auto').lower()
    auth_user   = (data.get('auth_user') or '').strip()
    auth_pass   = (data.get('auth_pass') or '').strip()
    auth_cookie = (data.get('auth_cookie') or '').strip()

    if not target:
        return jsonify({'ok': False, 'error': 'target URL is required'}), 400
    if not target.startswith('http'):
        target = 'https://' + target

    findings = []
    scanner_used = None
    warning = None

    if scanner == 'auto':
        for s in ['nuclei', 'nikto', 'wapiti', 'wapiti3']:
            if _dca_find_bin(s):
                scanner = s.replace('3', '') if s == 'wapiti3' else s
                break
        else:
            return jsonify({'ok': False,
                'error': 'No DAST scanner installed. Run: apt install nikto nuclei wapiti'}), 503

    with _tmp.TemporaryDirectory() as tdir:
        out_file = os.path.join(tdir, 'results.json')

        if scanner == 'nikto':
            bin_path = _dca_find_bin('nikto')
            if not bin_path:
                return jsonify({'ok': False, 'error': 'nikto not installed. Run: apt install nikto'}), 503
            cmd = [bin_path, '-h', target, '-Format', 'json', '-o', out_file, '-nointeractive', '-Tuning', 'x']
            if auth_user and auth_pass:
                cmd += ['-id', f'{auth_user}:{auth_pass}']
            if auth_cookie:
                cmd += ['-c', auth_cookie]
            try:
                _sp.run(cmd, capture_output=True, timeout=180)
                if os.path.exists(out_file):
                    with open(out_file) as f:
                        raw = _jd.load(f)
                    sev_words = [(['critical','remote code','rce','sql inject'], 'critical'),
                                 (['xss','csrf','auth bypass','directory traversal'], 'high'),
                                 (['disclosure','header','cookie'], 'low')]
                    for v in raw.get('vulnerabilities', []):
                        msg = v.get('msg', '')
                        sev = 'medium'
                        ml = msg.lower()
                        for words, s in sev_words:
                            if any(w in ml for w in words):
                                sev = s; break
                        findings.append({'id': v.get('id',''), 'title': msg[:120] or 'Finding',
                            'severity': sev, 'url': v.get('url', target),
                            'method': v.get('method','GET'), 'description': msg,
                            'references': v.get('references',''), 'scanner': 'Nikto'})
                scanner_used = 'Nikto'
            except _sp.TimeoutExpired:
                warning = 'Scan timed out (180s)'
            except Exception as exc:
                return jsonify({'ok': False, 'error': str(exc)}), 500

        elif scanner == 'nuclei':
            bin_path = _dca_find_bin('nuclei')
            if not bin_path:
                return jsonify({'ok': False, 'error': 'nuclei not installed. Run: apt install nuclei'}), 503
            cmd = [bin_path, '-u', target, '-json', '-o', out_file, '-silent',
                   '-timeout', '10', '-retries', '1', '-bulk-size', '10']
            if auth_user and auth_pass:
                cred = _b64d.b64encode(f'{auth_user}:{auth_pass}'.encode()).decode()
                cmd += ['-H', f'Authorization: Basic {cred}']
            if auth_cookie:
                cmd += ['-H', f'Cookie: {auth_cookie}']
            try:
                _sp.run(cmd, capture_output=True, timeout=300)
                if os.path.exists(out_file):
                    with open(out_file) as f:
                        for line in f:
                            line = line.strip()
                            if not line:
                                continue
                            try:
                                item = _jd.loads(line)
                                info = item.get('info', {})
                                refs = info.get('reference', [])
                                findings.append({
                                    'id': item.get('template-id', ''),
                                    'title': info.get('name', item.get('template-id', 'Finding')),
                                    'severity': info.get('severity', 'info').lower(),
                                    'url': item.get('matched-at', target),
                                    'method': item.get('type', 'http').upper(),
                                    'description': info.get('description', ''),
                                    'references': ', '.join(refs if isinstance(refs, list) else [refs or '']),
                                    'scanner': 'Nuclei',
                                    'tags': ', '.join(info.get('tags', []) if isinstance(info.get('tags'), list) else []),
                                })
                            except Exception:
                                continue
                scanner_used = 'Nuclei'
            except _sp.TimeoutExpired:
                warning = 'Scan timed out (300s)'
            except Exception as exc:
                return jsonify({'ok': False, 'error': str(exc)}), 500

        elif scanner == 'wapiti':
            bin_path = _dca_find_bin('wapiti', 'wapiti3')
            if not bin_path:
                return jsonify({'ok': False, 'error': 'wapiti not installed. Run: apt install wapiti'}), 503
            cmd = [bin_path, '-u', target, '-f', 'json', '-o', out_file,
                   '--no-bugreport', '-v', '0', '--timeout', '10']
            if auth_user and auth_pass:
                cmd += ['--auth-cred', f'{auth_user}%{auth_pass}']
            if auth_cookie:
                cmd += ['--cookie', auth_cookie]
            try:
                _sp.run(cmd, capture_output=True, timeout=300)
                if os.path.exists(out_file):
                    with open(out_file) as f:
                        raw = _jd.load(f)
                    sev_levels = ['info', 'low', 'medium', 'high', 'critical']
                    for vuln_type, vulns in raw.get('vulnerabilities', {}).items():
                        for v in vulns:
                            lv = v.get('level', 2)
                            sev = sev_levels[min(int(lv) if str(lv).isdigit() else 2, 4)]
                            findings.append({'id': '', 'title': vuln_type, 'severity': sev,
                                'url': v.get('path', target), 'method': v.get('method', 'GET'),
                                'description': v.get('info', ''), 'references': '', 'scanner': 'Wapiti'})
                    for anom_type, anoms in raw.get('anomalies', {}).items():
                        for a in anoms:
                            findings.append({'id': '', 'title': anom_type, 'severity': 'info',
                                'url': a.get('path', target), 'method': a.get('method', 'GET'),
                                'description': a.get('info', ''), 'references': '', 'scanner': 'Wapiti'})
                scanner_used = 'Wapiti'
            except _sp.TimeoutExpired:
                warning = 'Scan timed out (300s)'
            except Exception as exc:
                return jsonify({'ok': False, 'error': str(exc)}), 500

        elif scanner == 'zap':
            import urllib.request as _ureq
            import urllib.parse as _up
            import time as _tz
            ZAP = 'http://localhost:8080'
            try:
                enc = _up.quote(target, safe='')
                with _ureq.urlopen(f'{ZAP}/JSON/spider/action/scan/?url={enc}&recurse=true', timeout=10) as r:
                    spider_id = _jd.loads(r.read()).get('scan', '0')
                for _ in range(30):
                    with _ureq.urlopen(f'{ZAP}/JSON/spider/view/status/?scanId={spider_id}', timeout=5) as r:
                        if int(_jd.loads(r.read()).get('status', 0)) >= 100:
                            break
                    _tz.sleep(2)
                with _ureq.urlopen(f'{ZAP}/JSON/alert/view/alerts/?baseurl={enc}&start=0&count=500', timeout=10) as r:
                    alerts = _jd.loads(r.read()).get('alerts', [])
                sev_map = {'0': 'info', '1': 'low', '2': 'medium', '3': 'high'}
                for a in alerts:
                    findings.append({'id': a.get('pluginId',''), 'title': a.get('alert','Finding'),
                        'severity': sev_map.get(str(a.get('risk','1')), 'medium'),
                        'url': a.get('url', target), 'method': a.get('method','GET'),
                        'description': a.get('description',''), 'references': a.get('reference',''),
                        'solution': a.get('solution',''), 'scanner': 'OWASP ZAP'})
                scanner_used = 'OWASP ZAP'
            except Exception as exc:
                return jsonify({'ok': False, 'error': f'ZAP API error: {exc}. Ensure ZAP is running on localhost:8080 with REST API enabled.'}), 503

    if not scanner_used:
        return jsonify({'ok': False, 'error': 'Scan produced no output. The scanner may have failed silently.'}), 500

    return jsonify({'ok': True, 'scanner': scanner_used, 'target': target,
                    'count': len(findings), 'findings': findings, 'warning': warning})


@app.route('/api/dca/jira/test', methods=['POST'])
@login_required
def api_dca_jira_test():
    import urllib.request as _ureq
    import json as _jd, base64 as _b64
    d = request.get_json(silent=True) or {}
    jira_url = (d.get('url') or '').rstrip('/')
    username = (d.get('username') or '').strip()
    api_token = (d.get('api_token') or '').strip()
    if not all([jira_url, username, api_token]):
        return jsonify({'ok': False, 'error': 'url, username, and api_token required'}), 400
    auth = _b64.b64encode(f'{username}:{api_token}'.encode()).decode()
    req = _ureq.Request(f'{jira_url}/rest/api/3/myself',
                        headers={'Authorization': f'Basic {auth}', 'Content-Type': 'application/json'})
    try:
        with _ureq.urlopen(req, timeout=10) as r:
            me = _jd.loads(r.read())
        return jsonify({'ok': True, 'displayName': me.get('displayName', username)})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/dca/jira/issue', methods=['POST'])
@login_required
def api_dca_jira_issue():
    import urllib.request as _ureq
    import json as _jd, base64 as _b64
    d = request.get_json(silent=True) or {}
    jira_url = (d.get('url') or '').rstrip('/')
    username = (d.get('username') or '').strip()
    api_token = (d.get('api_token') or '').strip()
    project = (d.get('project') or '').strip().upper()
    finding = d.get('finding', {})
    if not all([jira_url, username, api_token, project]):
        return jsonify({'ok': False, 'error': 'url, username, api_token, project required'}), 400
    sev_pri = {'critical': 'Highest', 'high': 'High', 'medium': 'Medium', 'low': 'Low', 'info': 'Lowest'}
    body = _jd.dumps({'fields': {
        'project': {'key': project},
        'summary': f"[DAST] {finding.get('title','Finding')} — {finding.get('url','')}",
        'description': {'type': 'doc', 'version': 1, 'content': [{'type': 'paragraph', 'content': [{'type': 'text',
            'text': (f"Scanner: {finding.get('scanner','')}\nSeverity: {finding.get('severity','')}\n"
                     f"URL: {finding.get('url','')}\nMethod: {finding.get('method','')}\n\n"
                     f"Description:\n{finding.get('description','')}\n\nReferences: {finding.get('references','')}")}]}]},
        'issuetype': {'name': 'Bug'},
        'priority': {'name': sev_pri.get(finding.get('severity', 'medium'), 'Medium')},
    }}).encode()
    auth = _b64.b64encode(f'{username}:{api_token}'.encode()).decode()
    req = _ureq.Request(f'{jira_url}/rest/api/3/issue', data=body,
                        headers={'Authorization': f'Basic {auth}', 'Content-Type': 'application/json'})
    try:
        with _ureq.urlopen(req, timeout=15) as r:
            resp = _jd.loads(r.read())
        key = resp.get('key', '')
        return jsonify({'ok': True, 'issue_key': key, 'url': f'{jira_url}/browse/{key}'})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/dca/github/test', methods=['POST'])
@login_required
def api_dca_github_test():
    import urllib.request as _ureq
    import json as _jd
    d = request.get_json(silent=True) or {}
    token = (d.get('token') or '').strip()
    if not token:
        return jsonify({'ok': False, 'error': 'token required'}), 400
    req = _ureq.Request('https://api.github.com/user',
                        headers={'Authorization': f'Bearer {token}', 'Accept': 'application/vnd.github+json'})
    try:
        with _ureq.urlopen(req, timeout=10) as r:
            me = _jd.loads(r.read())
        return jsonify({'ok': True, 'login': me.get('login',''), 'name': me.get('name','')})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/dca/github/issue', methods=['POST'])
@login_required
def api_dca_github_issue():
    import urllib.request as _ureq
    import json as _jd
    d = request.get_json(silent=True) or {}
    token = (d.get('token') or '').strip()
    owner = (d.get('owner') or '').strip()
    repo = (d.get('repo') or '').strip()
    finding = d.get('finding', {})
    if not all([token, owner, repo]):
        return jsonify({'ok': False, 'error': 'token, owner, repo required'}), 400
    sev = finding.get('severity', 'medium')
    body_text = (f"**Scanner:** {finding.get('scanner','')}\n**Severity:** {sev}\n"
                 f"**URL:** {finding.get('url','')}\n**Method:** {finding.get('method','')}\n\n"
                 f"### Description\n{finding.get('description','')}\n\n"
                 f"### References\n{finding.get('references','')}")
    payload = _jd.dumps({'title': f"[DAST] {finding.get('title','Finding')}",
                         'body': body_text, 'labels': ['security', sev]}).encode()
    req = _ureq.Request(f'https://api.github.com/repos/{owner}/{repo}/issues', data=payload,
                        headers={'Authorization': f'Bearer {token}',
                                 'Accept': 'application/vnd.github+json',
                                 'Content-Type': 'application/json'})
    try:
        with _ureq.urlopen(req, timeout=15) as r:
            resp = _jd.loads(r.read())
        return jsonify({'ok': True, 'issue_number': resp.get('number',''), 'url': resp.get('html_url','')})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/dca/jenkins/test', methods=['POST'])
@login_required
def api_dca_jenkins_test():
    import urllib.request as _ureq
    import json as _jd, base64 as _b64
    d = request.get_json(silent=True) or {}
    url = (d.get('url') or '').rstrip('/')
    username = (d.get('username') or '').strip()
    token = (d.get('api_token') or '').strip()
    if not all([url, username, token]):
        return jsonify({'ok': False, 'error': 'url, username, api_token required'}), 400
    auth = _b64.b64encode(f'{username}:{token}'.encode()).decode()
    req = _ureq.Request(f'{url}/api/json', headers={'Authorization': f'Basic {auth}'})
    try:
        with _ureq.urlopen(req, timeout=10) as r:
            info = _jd.loads(r.read())
        return jsonify({'ok': True, 'jobs': len(info.get('jobs', [])), 'mode': info.get('mode','Jenkins')})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/dca/jenkins/trigger', methods=['POST'])
@login_required
def api_dca_jenkins_trigger():
    import urllib.request as _ureq
    import urllib.parse as _up, base64 as _b64
    d = request.get_json(silent=True) or {}
    url = (d.get('url') or '').rstrip('/')
    username = (d.get('username') or '').strip()
    token = (d.get('api_token') or '').strip()
    job = (d.get('job') or '').strip()
    params = d.get('params', {})
    if not all([url, username, token, job]):
        return jsonify({'ok': False, 'error': 'url, username, api_token, job required'}), 400
    auth = _b64.b64encode(f'{username}:{token}'.encode()).decode()
    trigger_url = (f'{url}/job/{job}/buildWithParameters?{_up.urlencode(params)}'
                   if params else f'{url}/job/{job}/build')
    req = _ureq.Request(trigger_url, data=b'', method='POST',
                        headers={'Authorization': f'Basic {auth}'})
    try:
        with _ureq.urlopen(req, timeout=10) as r:
            location = r.headers.get('Location', '')
        return jsonify({'ok': True, 'queue_url': location})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/dca/azure/test', methods=['POST'])
@login_required
def api_dca_azure_test():
    import urllib.request as _ureq
    import json as _jd, base64 as _b64
    d = request.get_json(silent=True) or {}
    org = (d.get('org') or '').strip()
    pat = (d.get('pat') or '').strip()
    if not all([org, pat]):
        return jsonify({'ok': False, 'error': 'org and pat required'}), 400
    auth = _b64.b64encode(f':{pat}'.encode()).decode()
    req = _ureq.Request(f'https://dev.azure.com/{org}/_apis/projects?api-version=7.0',
                        headers={'Authorization': f'Basic {auth}', 'Content-Type': 'application/json'})
    try:
        with _ureq.urlopen(req, timeout=10) as r:
            resp = _jd.loads(r.read())
        return jsonify({'ok': True, 'projects': [p['name'] for p in resp.get('value', [])]})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


@app.route('/api/dca/azure/workitem', methods=['POST'])
@login_required
def api_dca_azure_workitem():
    import urllib.request as _ureq
    import json as _jd, base64 as _b64
    d = request.get_json(silent=True) or {}
    org = (d.get('org') or '').strip()
    project = (d.get('project') or '').strip()
    pat = (d.get('pat') or '').strip()
    finding = d.get('finding', {})
    wi_type = (d.get('type') or 'Bug').strip()
    if not all([org, project, pat]):
        return jsonify({'ok': False, 'error': 'org, project, pat required'}), 400
    sev_map = {'critical': '1 - Critical', 'high': '2 - High', 'medium': '3 - Medium', 'low': '4 - Low', 'info': '4 - Low'}
    patch = [
        {'op': 'add', 'path': '/fields/System.Title',
         'value': f"[DAST] {finding.get('title','Finding')} — {finding.get('url','')}"},
        {'op': 'add', 'path': '/fields/Microsoft.VSTS.Common.Severity',
         'value': sev_map.get(finding.get('severity', 'medium'), '3 - Medium')},
        {'op': 'add', 'path': '/fields/System.Description',
         'value': (f"<b>Scanner:</b> {finding.get('scanner','')}<br>"
                   f"<b>Severity:</b> {finding.get('severity','')}<br>"
                   f"<b>URL:</b> {finding.get('url','')}<br>"
                   f"<b>Method:</b> {finding.get('method','')}<br><br>"
                   f"<b>Description:</b><br>{finding.get('description','')}<br><br>"
                   f"<b>References:</b><br>{finding.get('references','')}")},
        {'op': 'add', 'path': '/fields/System.Tags', 'value': 'security; dast; vulnerability'},
    ]
    auth = _b64.b64encode(f':{pat}'.encode()).decode()
    body = _jd.dumps(patch).encode()
    req = _ureq.Request(
        f'https://dev.azure.com/{org}/{project}/_apis/wit/workitems/${wi_type}?api-version=7.0',
        data=body, method='PATCH',
        headers={'Authorization': f'Basic {auth}', 'Content-Type': 'application/json-patch+json'})
    try:
        with _ureq.urlopen(req, timeout=15) as r:
            resp = _jd.loads(r.read())
        wi_id = resp.get('id', '')
        wi_url = (resp.get('_links', {}).get('html', {}).get('href', '')
                  or f'https://dev.azure.com/{org}/{project}/_workitems/edit/{wi_id}')
        return jsonify({'ok': True, 'id': wi_id, 'url': wi_url})
    except Exception as exc:
        return jsonify({'ok': False, 'error': str(exc)}), 400


# ── Google Calendar / Meet — Appointment Booking ──────────────────────────────
import uuid as _uuid

try:
    from google.oauth2.credentials import Credentials as _GCreds
    from google.auth.transport.requests import Request as _GRequest
    from google_auth_oauthlib.flow import Flow as _GFlow
    from googleapiclient.discovery import build as _gbuild
    _GCAL_AVAILABLE = True
except ImportError:
    _GCAL_AVAILABLE = False

_GCAL_SCOPES      = ['https://www.googleapis.com/auth/calendar']
_GCAL_TOKEN_FILE  = os.path.join(os.path.dirname(__file__), '..', 'data', 'google_oauth_token.json')
_GCAL_CLIENT_ID   = os.environ.get('GOOGLE_CLIENT_ID', '')
_GCAL_CLIENT_SECRET = os.environ.get('GOOGLE_CLIENT_SECRET', '')

def _gcal_client_config(redirect_uri: str) -> dict:
    return {
        'web': {
            'client_id':     _GCAL_CLIENT_ID,
            'client_secret': _GCAL_CLIENT_SECRET,
            'redirect_uris': [redirect_uri],
            'auth_uri':      'https://accounts.google.com/o/oauth2/auth',
            'token_uri':     'https://oauth2.googleapis.com/token',
        }
    }

_APPT_PREP: dict = {
    'Dashboard Feature Training': [
        'No preparation needed — bring any questions you have about the platform.',
        'Consider noting which features or sections you find most confusing.',
        'Have your account credentials ready to log in during the session.',
    ],
    'On-Site Cyber Security Audit': [
        'Prepare a list of all systems, networks, and physical access points on-site.',
        'Ensure a technical contact is available to assist during the audit.',
        'Have network diagrams or asset inventories ready if available.',
    ],
    'Website Security Analysis': [
        'Provide the full URL(s) of the website(s) to be assessed.',
        'Have admin access to your web hosting panel or CMS ready.',
        'Prepare a list of any known vulnerabilities or recent security incidents.',
    ],
    'Manual Report Writing': [
        'Gather any existing scan results, audit logs, or findings you want included.',
        'Prepare a brief overview of the systems or scope the report should cover.',
        'Have stakeholder names and titles ready for the executive summary.',
    ],
}


def _parse_duration_minutes(s: str) -> int:
    """Parse a human duration string to minutes. Returns 60 if unparseable."""
    import re as _re
    if not s:
        return 60
    s = s.lower().strip()
    if 'full day' in s or 'whole day' in s:
        return 480
    if 'half day' in s or 'half-day' in s:
        return 240
    total = 0
    m = _re.search(r'(\d+(?:\.\d+)?)\s*h(?:our|r)?s?', s)
    if m:
        total += int(float(m.group(1)) * 60)
    m = _re.search(r'(\d+)\s*m(?:in(?:ute)?s?)?', s)
    if m:
        total += int(m.group(1))
    return total if total > 0 else 60


def _get_gcal_service():
    if not _GCAL_AVAILABLE or not os.path.exists(_GCAL_TOKEN_FILE):
        return None
    try:
        creds = _GCreds.from_authorized_user_file(_GCAL_TOKEN_FILE, _GCAL_SCOPES)
        if creds.expired and creds.refresh_token:
            creds.refresh(_GRequest())
            with open(_GCAL_TOKEN_FILE, 'w') as fh:
                fh.write(creds.to_json())
        return _gbuild('calendar', 'v3', credentials=creds, cache_discovery=False)
    except Exception:
        return None


def _create_gcal_appointment(service_type: str, date_str: str, end_date_str: str,
                              time_str: str, duration_str: str, timezone: str,
                              full_name: str, user_email: str,
                              company: str, notes: str) -> dict | None:
    svc = _get_gcal_service()
    if not svc:
        return None
    try:
        import datetime as _dt
        duration_min = _parse_duration_minutes(duration_str)
        start_naive  = _dt.datetime.strptime(f'{date_str} {time_str}', '%Y-%m-%d %H:%M')
        if end_date_str and end_date_str != date_str:
            end_naive = _dt.datetime.strptime(f'{end_date_str} {time_str}', '%Y-%m-%d %H:%M')
        else:
            end_naive = start_naive + _dt.timedelta(minutes=duration_min)
        start_iso   = start_naive.strftime('%Y-%m-%dT%H:%M:%S')
        end_iso     = end_naive.strftime('%Y-%m-%dT%H:%M:%S')

        svc_descs = {
            'Dashboard Feature Training':   'Personalised walkthrough of CyberINK Security Intelligence platform features.',
            'On-Site Cyber Security Audit': 'Comprehensive on-premises security assessment.',
            'Website Security Analysis':    'In-depth vulnerability assessment and security analysis of web properties.',
            'Manual Report Writing':        'Collaborative session to produce a customised security report with findings, risk ratings, and remediation recommendations.',
        }
        desc_lines = [f'Service: {service_type}', svc_descs.get(service_type, ''), '',
                      f'Client: {full_name}']
        if company:
            desc_lines.append(f'Company: {company}')
        desc_lines.append(f'Email: {user_email}')
        if notes:
            desc_lines.extend(['', 'Notes:', notes])

        event = {
            'summary':     f'[CyberINK] {service_type} — {full_name}',
            'description': '\n'.join(desc_lines),
            'start': {'dateTime': start_iso, 'timeZone': timezone},
            'end':   {'dateTime': end_iso,   'timeZone': timezone},
            'attendees': [
                {'email': user_email,   'displayName': full_name},
                {'email': _SUPPORT_EMAIL, 'displayName': 'CyberINK Support'},
            ],
            'conferenceData': {
                'createRequest': {
                    'requestId': str(_uuid.uuid4()),
                    'conferenceSolutionKey': {'type': 'hangoutsMeet'},
                }
            },
            'reminders': {
                'useDefault': False,
                'overrides': [
                    {'method': 'email',  'minutes': 1440},
                    {'method': 'popup',  'minutes': 30},
                ],
            },
        }
        result = svc.events().insert(
            calendarId='primary',
            body=event,
            conferenceDataVersion=1,
            sendUpdates='all',
        ).execute()
        return {
            'meet_link':  result.get('hangoutLink', ''),
            'event_link': result.get('htmlLink', ''),
            'event_id':   result.get('id', ''),
        }
    except Exception as exc:
        app.logger.error(f'[GCAL] Event creation failed: {exc}')
        return None


def _format_appt_datetime(date_str: str, end_date_str: str, time_str: str,
                          duration_str: str, timezone: str) -> str:
    try:
        import datetime as _dt
        dt = _dt.datetime.strptime(f'{date_str} {time_str}', '%Y-%m-%d %H:%M')
        result = f'{dt.strftime("%A, %d %B %Y")} at {dt.strftime("%H:%M")}'
        if end_date_str and end_date_str != date_str:
            edt = _dt.datetime.strptime(end_date_str, '%Y-%m-%d')
            result += f' — {edt.strftime("%A, %d %B %Y")}'
        if duration_str:
            result += f' · {duration_str}'
        result += f' ({timezone})'
        return result
    except Exception:
        return f'{date_str} {time_str} ({timezone})'


def _send_appointment_admin_email(service_type: str, date_str: str, end_date_str: str,
                                   time_str: str, duration_str: str, timezone: str,
                                   full_name: str, user_email: str, company: str,
                                   notes: str, meet_link: str, event_link: str,
                                   username: str) -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        appt_dt    = _format_appt_datetime(date_str, end_date_str, time_str, duration_str, timezone)
        subject    = f'[CyberINK Appointment] {service_type} — {full_name}'
        safe_notes = notes.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;').replace('\n', '<br>')

        company_row = (f'<tr style="border-bottom:1px solid #bfdbfe;"><td style="padding:7px 0;'
                       f'font-weight:700;color:#1e3a8a;width:130px;">Company</td>'
                       f'<td style="padding:7px 0;color:#2563eb;">{company}</td></tr>') if company else ''
        meet_row = (f'<tr style="border-bottom:1px solid #bfdbfe;"><td style="padding:7px 0;'
                    f'font-weight:700;color:#1e3a8a;">Google Meet</td><td style="padding:7px 0;">'
                    f'<a href="{meet_link}" style="color:#2563eb;">{meet_link}</a></td></tr>') if meet_link else ''
        event_row = (f'<tr style="border-bottom:1px solid #bfdbfe;"><td style="padding:7px 0;'
                     f'font-weight:700;color:#1e3a8a;">Calendar</td><td style="padding:7px 0;">'
                     f'<a href="{event_link}" style="color:#2563eb;">View in Google Calendar</a>'
                     f'</td></tr>') if event_link else ''
        notes_block = (f'<div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;'
                       f'letter-spacing:.4px;margin:20px 0 8px;">Additional Notes</div>'
                       f'<div style="background:#f8faff;border:1px solid #bfdbfe;border-radius:8px;'
                       f'padding:12px 16px;font-size:13px;color:#1e3a8a;line-height:1.6;">'
                       f'{safe_notes}</div>') if notes else ''

        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 16px;">
            New Appointment Request
          </p>
          <table style="width:100%;border-collapse:collapse;font-size:12px;margin-bottom:16px;">
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:130px;">Service</td>
              <td style="padding:7px 0;color:#2563eb;font-weight:600;">{service_type}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Date and Time</td>
              <td style="padding:7px 0;color:#2563eb;">{appt_dt}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Full Name</td>
              <td style="padding:7px 0;color:#2563eb;">{full_name}</td>
            </tr>
            {company_row}
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Email</td>
              <td style="padding:7px 0;"><a href="mailto:{user_email}" style="color:#2563eb;">{user_email}</a></td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Account</td>
              <td style="padding:7px 0;color:#2563eb;">{username}</td>
            </tr>
            {meet_row}
            {event_row}
          </table>
          {notes_block}
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
            Reply to this email to contact the client at
            <a href="mailto:{user_email}" style="color:#2563eb;">{user_email}</a>.
          </p>"""

        msg = MIMEMultipart('alternative')
        msg['Subject']  = subject
        msg['From']     = f'CyberINK Appointments <{_SMTP_USER}>'
        msg['To']       = _SUPPORT_EMAIL
        msg['Reply-To'] = user_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, _SUPPORT_EMAIL, msg.as_string())
        return True
    except Exception:
        return False


def _send_appointment_user_email(service_type: str, date_str: str, end_date_str: str,
                                  time_str: str, duration_str: str, timezone: str,
                                  full_name: str, user_email: str, meet_link: str) -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        appt_dt = _format_appt_datetime(date_str, end_date_str, time_str, duration_str, timezone)
        subject = f'CyberINK — Appointment Request Received: {service_type}'

        if meet_link:
            meet_block = (f'<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;'
                          f'padding:14px 16px;margin:20px 0;text-align:center;">'
                          f'<div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;'
                          f'letter-spacing:.4px;margin-bottom:8px;">Google Meet Link</div>'
                          f'<a href="{meet_link}" style="color:#2563eb;font-size:13px;font-weight:600;'
                          f'word-break:break-all;">{meet_link}</a>'
                          f'<div style="font-size:11px;color:#3b82f6;margin-top:6px;">'
                          f'This link will be active at your appointment time.</div></div>')
        else:
            meet_block = ('<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;'
                          'padding:14px 16px;margin:20px 0;font-size:12px;color:#1e3a8a;">'
                          'A Google Meet link will be sent to you once the appointment is confirmed by our team.'
                          '</div>')

        prep_items = _APPT_PREP.get(service_type, [])
        prep_block = ''
        if prep_items:
            items_html = ''.join(f'<li style="margin-bottom:6px;color:#1e3a8a;">{i}</li>' for i in prep_items)
            prep_block = (f'<div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;'
                          f'letter-spacing:.4px;margin:20px 0 10px;">How to Prepare</div>'
                          f'<ul style="margin:0;padding-left:18px;font-size:12px;line-height:1.65;">'
                          f'{items_html}</ul>')

        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 8px;">
            Appointment Request Received
          </p>
          <p class="ep" style="color:#3b82f6;font-size:13px;margin:0 0 20px;line-height:1.6;">
            Hello {full_name}, your appointment request has been received. Our team will review
            and confirm within 1 business day.
          </p>
          <table style="width:100%;border-collapse:collapse;font-size:12px;">
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:110px;">Service</td>
              <td style="padding:7px 0;color:#2563eb;font-weight:600;">{service_type}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Date and Time</td>
              <td style="padding:7px 0;color:#2563eb;">{appt_dt}</td>
            </tr>
          </table>
          {meet_block}
          {prep_block}
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
            To cancel or reschedule, reply to this email or contact us at
            <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.
          </p>"""

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Security <{_SMTP_USER}>'
        msg['To']      = user_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, user_email, msg.as_string())
        return True
    except Exception:
        return False


@app.route('/api/book-appointment', methods=['POST'])
def book_appointment():
    if 'user' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    data          = request.get_json(force=True) or {}
    service_type  = (data.get('service_type') or '').strip()
    date_str      = (data.get('date')         or '').strip()
    end_date_str  = (data.get('end_date')     or '').strip()
    time_str      = (data.get('time')         or '').strip()
    duration_str  = (data.get('duration')     or '').strip()
    timezone      = (data.get('timezone')     or 'Asia/Jakarta').strip()
    full_name     = (data.get('full_name')    or '').strip()
    user_email    = (data.get('email')        or '').strip()
    company       = (data.get('company')      or '').strip()
    notes         = (data.get('notes')        or '').strip()

    valid_services = {'Dashboard Feature Training', 'On-Site Cyber Security Audit',
                      'Website Security Analysis', 'Manual Report Writing'}
    if not service_type or service_type not in valid_services:
        return jsonify({'error': 'Please select a valid service type.'}), 400
    if not date_str or not time_str:
        return jsonify({'error': 'Date and time are required.'}), 400
    if not full_name:
        return jsonify({'error': 'Full name is required.'}), 400
    if not user_email or '@' not in user_email:
        return jsonify({'error': 'A valid email address is required.'}), 400

    try:
        import datetime as _dt
        appt_date = _dt.datetime.strptime(date_str, '%Y-%m-%d').date()
        if appt_date <= _dt.date.today():
            return jsonify({'error': 'Please select a future start date.'}), 400
        if end_date_str:
            end_appt = _dt.datetime.strptime(end_date_str, '%Y-%m-%d').date()
            if end_appt < appt_date:
                return jsonify({'error': 'End date cannot be before the start date.'}), 400
    except ValueError:
        return jsonify({'error': 'Invalid date format.'}), 400

    import datetime as _dt
    now_iso  = _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
    appt_id  = 'appt-' + str(_uuid.uuid4())[:8]
    appts    = _load_appointments()
    appts[appt_id] = {
        'id': appt_id, 'service_type': service_type,
        'date': date_str, 'end_date': end_date_str, 'time': time_str,
        'duration': duration_str, 'timezone': timezone,
        'full_name': full_name, 'email': user_email,
        'company': company, 'notes': notes,
        'username': session['user']['username'],
        'status': 'pending', 'meet_link': '', 'event_link': '', 'event_id': '',
        'admin_notes': '', 'created_at': now_iso, 'updated_at': now_iso,
    }
    _save_appointments(appts)

    _send_appointment_admin_email(
        service_type, date_str, end_date_str, time_str, duration_str, timezone,
        full_name, user_email, company, notes,
        '', '', session['user']['username']
    )
    _send_appointment_user_email(
        service_type, date_str, end_date_str, time_str, duration_str, timezone,
        full_name, user_email, ''
    )
    return jsonify({'ok': True,
                    'message': f'Appointment request submitted. Our team will review and confirm at {user_email} within 1 business day.'})


# ── Appointment Management ────────────────────────────────────────────────────
_APPTS_FILE = os.path.join(os.path.dirname(__file__), '..', 'data', 'appointments.json')

def _load_appointments() -> dict:
    try:
        if os.path.exists(_APPTS_FILE):
            with open(_APPTS_FILE) as fh:
                return _json.load(fh)
    except Exception:
        pass
    return {}

def _save_appointments(appts: dict) -> None:
    os.makedirs(os.path.dirname(_APPTS_FILE), exist_ok=True)
    with open(_APPTS_FILE, 'w') as fh:
        _json.dump(appts, fh, indent=2)


def _send_appointment_approved_email(service_type: str, date_str: str, end_date_str: str,
                                      time_str: str, duration_str: str, timezone: str,
                                      full_name: str, user_email: str,
                                      meet_link: str, admin_note: str = '') -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        appt_dt = _format_appt_datetime(date_str, end_date_str, time_str, duration_str, timezone)
        subject = f'CyberINK — Appointment Confirmed: {service_type}'
        if meet_link:
            meet_block = (
                f'<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;'
                f'padding:16px;margin:20px 0;text-align:center;">'
                f'<div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;'
                f'letter-spacing:.4px;margin-bottom:10px;">Your Google Meet Link</div>'
                f'<a href="{meet_link}" style="color:#2563eb;font-size:14px;font-weight:700;'
                f'word-break:break-all;display:block;margin-bottom:8px;">{meet_link}</a>'
                f'<div style="font-size:11px;color:#3b82f6;">Click the link above at your appointment time to join the session.</div>'
                f'</div>'
            )
        else:
            meet_block = ('<div style="background:#eff6ff;border:1px solid #bfdbfe;border-radius:8px;'
                          'padding:14px 16px;margin:20px 0;font-size:12px;color:#1e3a8a;">'
                          'Our team will send you a Google Meet link before the session.</div>')
        note_block = ''
        if admin_note:
            safe_n = admin_note.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('\n','<br>')
            note_block = (f'<div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;'
                          f'letter-spacing:.4px;margin:20px 0 8px;">Note from our team</div>'
                          f'<div style="background:#f8faff;border:1px solid #bfdbfe;border-radius:8px;'
                          f'padding:12px 16px;font-size:13px;color:#1e3a8a;line-height:1.6;">{safe_n}</div>')
        prep_items = _APPT_PREP.get(service_type, [])
        prep_block = ''
        if prep_items:
            items_html = ''.join(f'<li style="margin-bottom:6px;color:#1e3a8a;">{i}</li>' for i in prep_items)
            prep_block = (f'<div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;'
                          f'letter-spacing:.4px;margin:20px 0 10px;">How to Prepare</div>'
                          f'<ul style="margin:0;padding-left:18px;font-size:12px;line-height:1.65;">{items_html}</ul>')
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 8px;">Appointment Confirmed</p>
          <p class="ep" style="color:#3b82f6;font-size:13px;margin:0 0 20px;line-height:1.6;">
            Hello {full_name}, your appointment has been confirmed by our team.</p>
          <table style="width:100%;border-collapse:collapse;font-size:12px;">
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:110px;">Service</td>
              <td style="padding:7px 0;color:#2563eb;font-weight:600;">{service_type}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Date and Time</td>
              <td style="padding:7px 0;color:#2563eb;">{appt_dt}</td>
            </tr>
          </table>
          {meet_block}{note_block}{prep_block}
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
            To cancel or reschedule contact us at <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.</p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Security <{_SMTP_USER}>'
        msg['To']      = user_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, user_email, msg.as_string())
        return True
    except Exception:
        return False


def _send_appointment_cancelled_email(service_type: str, appt_dt: str,
                                       full_name: str, user_email: str, reason: str = '') -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        subject = f'CyberINK — Appointment Cancelled: {service_type}'
        reason_block = ''
        if reason:
            safe_r = reason.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('\n','<br>')
            reason_block = (f'<div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;'
                            f'letter-spacing:.4px;margin:20px 0 8px;">Reason</div>'
                            f'<div style="background:#f8faff;border:1px solid #bfdbfe;border-radius:8px;'
                            f'padding:12px 16px;font-size:13px;color:#1e3a8a;line-height:1.6;">{safe_r}</div>')
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 8px;">Appointment Cancelled</p>
          <p class="ep" style="color:#3b82f6;font-size:13px;margin:0 0 20px;line-height:1.6;">
            Hello {full_name}, your appointment has been cancelled.</p>
          <table style="width:100%;border-collapse:collapse;font-size:12px;">
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:110px;">Service</td>
              <td style="padding:7px 0;color:#2563eb;">{service_type}</td>
            </tr>
            <tr><td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Date and Time</td>
              <td style="padding:7px 0;color:#2563eb;">{appt_dt}</td>
            </tr>
          </table>{reason_block}
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
            To book a new appointment visit the Customer Service Centre or contact
            <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.</p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Security <{_SMTP_USER}>'
        msg['To']      = user_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, user_email, msg.as_string())
        return True
    except Exception:
        return False


def _send_appointment_reschedule_email(service_type: str, old_dt: str, new_dt: str,
                                        full_name: str, user_email: str, message: str = '') -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        subject = f'CyberINK — Appointment Rescheduled: {service_type}'
        msg_block = ''
        if message:
            safe_m = message.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('\n','<br>')
            msg_block = (f'<div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;'
                         f'letter-spacing:.4px;margin:20px 0 8px;">Message from our team</div>'
                         f'<div style="background:#f8faff;border:1px solid #bfdbfe;border-radius:8px;'
                         f'padding:12px 16px;font-size:13px;color:#1e3a8a;line-height:1.6;">{safe_m}</div>')
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 8px;">Appointment Rescheduled</p>
          <p class="ep" style="color:#3b82f6;font-size:13px;margin:0 0 20px;line-height:1.6;">
            Hello {full_name}, your appointment has been moved to a new date and time.</p>
          <table style="width:100%;border-collapse:collapse;font-size:12px;">
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:120px;">Service</td>
              <td style="padding:7px 0;color:#2563eb;">{service_type}</td>
            </tr>
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Previous</td>
              <td style="padding:7px 0;color:#93c5fd;text-decoration:line-through;">{old_dt}</td>
            </tr>
            <tr><td style="padding:7px 0;font-weight:700;color:#1e3a8a;">New Time</td>
              <td style="padding:7px 0;color:#2563eb;font-weight:600;">{new_dt}</td>
            </tr>
          </table>{msg_block}
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
            If this does not work for you contact us at
            <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.</p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Security <{_SMTP_USER}>'
        msg['To']      = user_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, user_email, msg.as_string())
        return True
    except Exception:
        return False


def _send_appointment_completed_email(service_type: str, appt_dt: str,
                                       full_name: str, user_email: str) -> bool:
    if not _SMTP_USER or not _SMTP_PASS:
        return False
    try:
        subject = f'CyberINK — Appointment Completed: {service_type}'
        body = f"""
          <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 8px;">Appointment Completed</p>
          <p class="ep" style="color:#3b82f6;font-size:13px;margin:0 0 20px;line-height:1.6;">
            Hello {full_name}, your appointment has been marked as completed. Thank you for choosing CyberINK Security.</p>
          <table style="width:100%;border-collapse:collapse;font-size:12px;">
            <tr style="border-bottom:1px solid #bfdbfe;">
              <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:110px;">Service</td>
              <td style="padding:7px 0;color:#2563eb;font-weight:600;">{service_type}</td>
            </tr>
            <tr><td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Date and Time</td>
              <td style="padding:7px 0;color:#2563eb;">{appt_dt}</td>
            </tr>
          </table>
          <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
            We hope the session was valuable. To book another appointment visit the Customer Service Centre or
            contact <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.</p>"""
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Security <{_SMTP_USER}>'
        msg['To']      = user_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, user_email, msg.as_string())
        return True
    except Exception:
        return False


def _auto_expire_appointments() -> None:
    """Auto-complete approved appointments whose end date (or date) has passed and email the client."""
    import datetime as _dt
    today = _dt.date.today().isoformat()
    appts = _load_appointments()
    changed = False
    for appt in appts.values():
        if appt.get('status') != 'approved':
            continue
        check_date = appt.get('end_date') or appt.get('date', '')
        if not check_date or check_date >= today:
            continue
        now = _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
        appt.update({'status': 'completed', 'updated_at': now, 'completed_at': now})
        appt_dt = _format_appt_datetime(appt['date'], appt.get('end_date', ''),
                                         appt['time'], appt.get('duration', ''), appt['timezone'])
        _send_appointment_completed_email(appt['service_type'], appt_dt,
                                           appt['full_name'], appt['email'])
        changed = True
    if changed:
        _save_appointments(appts)


@app.route('/api/appointments', methods=['GET'])
def get_appointments():
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    _auto_expire_appointments()
    appts = _load_appointments()
    result = []
    for a in sorted(appts.values(), key=lambda x: x.get('created_at', ''), reverse=True):
        item = dict(a)
        ca = a.get('cost_amount')
        cc = a.get('cost_currency', 'USD')
        if ca is not None:
            item['cost_display'] = _fmt_cost(float(ca), cc)
        result.append(item)
    return jsonify({'appointments': result})


@app.route('/api/appointments/<appt_id>/approve', methods=['POST'])
def approve_appointment(appt_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data       = request.get_json(force=True) or {}
    admin_note = (data.get('admin_note') or '').strip()
    appts = _load_appointments()
    if appt_id not in appts:
        return jsonify({'error': 'Appointment not found'}), 404
    appt = appts[appt_id]
    if appt['status'] not in ('pending', 'rescheduled'):
        return jsonify({'error': f'Cannot approve a {appt["status"]} appointment'}), 400
    gcal_result = _create_gcal_appointment(
        appt['service_type'], appt['date'], appt.get('end_date', ''),
        appt['time'], appt.get('duration', ''), appt['timezone'],
        appt['full_name'], appt['email'], appt.get('company', ''), appt.get('notes', '')
    )
    meet_link  = gcal_result['meet_link']  if gcal_result else ''
    event_link = gcal_result['event_link'] if gcal_result else ''
    event_id   = gcal_result['event_id']   if gcal_result else ''
    import datetime as _dt
    now = _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
    appt.update({'status': 'approved', 'meet_link': meet_link, 'event_link': event_link,
                 'event_id': event_id, 'admin_notes': admin_note,
                 'updated_at': now, 'approved_at': now})
    _save_appointments(appts)
    _send_appointment_approved_email(
        appt['service_type'], appt['date'], appt.get('end_date', ''),
        appt['time'], appt.get('duration', ''), appt['timezone'],
        appt['full_name'], appt['email'], meet_link, admin_note
    )
    msg = 'Appointment approved.'
    if meet_link:
        msg += f' Google Meet link sent to {appt["email"]}.'
    else:
        msg += f' Confirmation sent to {appt["email"]}. (No Meet link — Google Calendar not configured.)'
    return jsonify({'ok': True, 'message': msg, 'meet_link': meet_link})


@app.route('/api/appointments/<appt_id>/cancel', methods=['POST'])
def cancel_appointment(appt_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data   = request.get_json(force=True) or {}
    reason = (data.get('reason') or '').strip()
    appts  = _load_appointments()
    if appt_id not in appts:
        return jsonify({'error': 'Appointment not found'}), 404
    appt = appts[appt_id]
    if appt['status'] == 'completed':
        return jsonify({'error': 'Cannot cancel a completed appointment'}), 400
    import datetime as _dt
    appt.update({'status': 'cancelled', 'admin_notes': reason,
                 'cancel_reason': reason, 'cancelled_by': 'admin',
                 'updated_at': _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')})
    _save_appointments(appts)
    appt_dt = _format_appt_datetime(appt['date'], appt.get('end_date', ''),
                                     appt['time'], appt.get('duration', ''), appt['timezone'])
    _send_appointment_cancelled_email(appt['service_type'], appt_dt,
                                       appt['full_name'], appt['email'], reason)
    return jsonify({'ok': True, 'message': f'Appointment cancelled. User notified at {appt["email"]}.'})


@app.route('/api/appointments/<appt_id>/complete', methods=['POST'])
def complete_appointment(appt_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    appts = _load_appointments()
    if appt_id not in appts:
        return jsonify({'error': 'Appointment not found'}), 404
    appt = appts[appt_id]
    if appt['status'] not in ('approved', 'rescheduled'):
        return jsonify({'error': f'Cannot complete a {appt["status"]} appointment'}), 400
    import datetime as _dt
    now = _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
    appt.update({'status': 'completed', 'updated_at': now, 'completed_at': now})
    _save_appointments(appts)
    appt_dt = _format_appt_datetime(appt['date'], appt.get('end_date', ''),
                                     appt['time'], appt.get('duration', ''), appt['timezone'])
    _send_appointment_completed_email(appt['service_type'], appt_dt,
                                       appt['full_name'], appt['email'])
    return jsonify({'ok': True, 'message': f'Appointment marked as completed. Notification sent to {appt["email"]}.'})


@app.route('/api/appointments/<appt_id>/reschedule', methods=['POST'])
def reschedule_appointment(appt_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data         = request.get_json(force=True) or {}
    new_date     = (data.get('date')     or '').strip()
    new_end_date = (data.get('end_date') or '').strip()
    new_time     = (data.get('time')     or '').strip()
    new_dur      = (data.get('duration') or '').strip()
    message      = (data.get('message')  or '').strip()
    if not new_date or not new_time:
        return jsonify({'error': 'New date and time are required'}), 400
    appts = _load_appointments()
    if appt_id not in appts:
        return jsonify({'error': 'Appointment not found'}), 404
    appt   = appts[appt_id]
    old_dt = _format_appt_datetime(appt['date'], appt.get('end_date', ''),
                                    appt['time'], appt.get('duration', ''), appt['timezone'])
    import datetime as _dt
    appt.update({'date': new_date, 'end_date': new_end_date, 'time': new_time,
                 'duration': new_dur or appt.get('duration', ''),
                 'status': 'rescheduled', 'meet_link': '', 'event_link': '', 'event_id': '',
                 'updated_at': _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')})
    _save_appointments(appts)
    new_dt = _format_appt_datetime(new_date, new_end_date, new_time,
                                    new_dur or appt.get('duration', ''), appt['timezone'])
    _send_appointment_reschedule_email(appt['service_type'], old_dt, new_dt,
                                        appt['full_name'], appt['email'], message)
    return jsonify({'ok': True, 'message': f'Rescheduled. User notified at {appt["email"]}.'})


@app.route('/api/appointments/<appt_id>/notes', methods=['PATCH'])
def update_appointment_notes(appt_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data  = request.get_json(force=True) or {}
    notes = (data.get('notes') or '').strip()
    appts = _load_appointments()
    if appt_id not in appts:
        return jsonify({'error': 'Appointment not found'}), 404
    import datetime as _dt
    appts[appt_id]['admin_notes'] = notes
    appts[appt_id]['updated_at']  = _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
    _save_appointments(appts)
    return jsonify({'ok': True})


# ── Currency conversion (static rates vs USD) ─────────────────────────────────
_CURRENCY_RATES: dict[str, float] = {
    'USD': 1.0,    'EUR': 0.93,   'GBP': 0.79,   'SGD': 1.35,
    'MYR': 4.72,   'IDR': 16500.0,'AUD': 1.55,   'CAD': 1.36,
    'JPY': 155.0,  'CNY': 7.24,   'INR': 84.0,   'KRW': 1380.0,
    'THB': 36.0,   'PHP': 58.0,   'VND': 25000.0,'HKD': 7.82,
    'NZD': 1.68,   'CHF': 0.90,   'SEK': 10.5,   'NOK': 10.7,
    'DKK': 6.9,    'BRL': 5.1,    'MXN': 17.0,   'ZAR': 18.5,
    'NGN': 1600.0, 'AED': 3.67,   'SAR': 3.75,   'PKR': 280.0,
    'BDT': 110.0,  'TWD': 32.0,
}
_CURRENCY_SYMBOLS: dict[str, str] = {
    'USD': '$',   'EUR': '€',   'GBP': '£',   'SGD': 'S$',
    'MYR': 'RM',  'IDR': 'Rp',  'AUD': 'A$',  'CAD': 'C$',
    'JPY': '¥',   'CNY': '¥',   'INR': '₹',   'KRW': '₩',
    'THB': '฿',   'PHP': '₱',   'VND': '₫',   'HKD': 'HK$',
    'NZD': 'NZ$', 'CHF': 'Fr',  'SEK': 'kr',  'NOK': 'kr',
    'DKK': 'kr',  'BRL': 'R$',  'MXN': '$',   'ZAR': 'R',
    'NGN': '₦',   'AED': 'د.إ', 'SAR': '﷼',   'PKR': '₨',
    'BDT': '৳',   'TWD': 'NT$',
}

def _convert_currency(amount: float, from_cur: str, to_cur: str) -> float:
    r_from = _CURRENCY_RATES.get(from_cur.upper(), 1.0)
    r_to   = _CURRENCY_RATES.get(to_cur.upper(), 1.0)
    return round(amount / r_from * r_to, 2)

def _fmt_cost(amount: float, currency: str) -> str:
    sym = _CURRENCY_SYMBOLS.get(currency.upper(), '')
    if currency.upper() in ('JPY', 'KRW', 'IDR', 'VND', 'NGN'):
        return f'{sym}{int(round(amount)):,} {currency}'
    return f'{sym}{amount:,.2f} {currency}'


# ── My Appointments (user-facing) ──────────────────────────────────────────────
@app.route('/api/my-appointments', methods=['GET'])
def my_appointments():
    if 'user' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    _auto_expire_appointments()
    username = session['user']['username']
    users    = _load_users()
    user_obj = users.get(username, {})
    cur_code = user_obj.get('currency_code', 'USD')
    appts    = _load_appointments()
    result   = []
    for a in sorted(appts.values(), key=lambda x: x.get('created_at', ''), reverse=True):
        if a.get('username') == username:
            item = dict(a)
            # Attach cost in user's currency
            cost_amount   = a.get('cost_amount')
            cost_currency = a.get('cost_currency', 'USD')
            if cost_amount is not None:
                converted = _convert_currency(float(cost_amount), cost_currency, cur_code)
                item['cost_display_base']  = _fmt_cost(float(cost_amount), cost_currency)
                item['cost_display_local'] = _fmt_cost(converted, cur_code)
                item['cost_currency_local']= cur_code
            result.append(item)
    return jsonify({'appointments': result})


def _send_user_cancel_email(service_type: str, appt_dt: str, full_name: str,
                             user_email: str, reason: str) -> None:
    if not _SMTP_USER or not _SMTP_PASS:
        return
    safe_r = reason.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('\n','<br>')
    body = f"""
      <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 8px;">Appointment Cancellation Request</p>
      <p class="ep" style="color:#3b82f6;font-size:13px;margin:0 0 20px;line-height:1.6;">
        Hello {full_name}, your cancellation request has been received. Your appointment has been cancelled.</p>
      <table style="width:100%;border-collapse:collapse;font-size:12px;">
        <tr style="border-bottom:1px solid #bfdbfe;">
          <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:120px;">Service</td>
          <td style="padding:7px 0;color:#2563eb;">{service_type}</td>
        </tr>
        <tr style="border-bottom:1px solid #bfdbfe;">
          <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Date and Time</td>
          <td style="padding:7px 0;color:#2563eb;">{appt_dt}</td>
        </tr>
      </table>
      <div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;letter-spacing:.4px;margin:20px 0 8px;">Reason Provided</div>
      <div style="background:#f8faff;border:1px solid #bfdbfe;border-radius:8px;padding:12px 16px;font-size:13px;color:#1e3a8a;line-height:1.6;">{safe_r}</div>
      <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
        To book a new appointment visit the Customer Service Centre or contact
        <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.</p>"""
    try:
        subject = f'CyberINK — Appointment Cancelled: {service_type}'
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Security <{_SMTP_USER}>'
        msg['To']      = user_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, [user_email, _SUPPORT_EMAIL], msg.as_string())
    except Exception:
        pass


def _send_user_reschedule_request_email(service_type: str, old_dt: str, new_dt: str,
                                         full_name: str, user_email: str, reason: str) -> None:
    if not _SMTP_USER or not _SMTP_PASS:
        return
    safe_r = reason.replace('&','&amp;').replace('<','&lt;').replace('>','&gt;').replace('\n','<br>')
    body = f"""
      <p class="eh1" style="color:#0f172a;font-size:16px;font-weight:700;margin:0 0 8px;">Reschedule Request Submitted</p>
      <p class="ep" style="color:#3b82f6;font-size:13px;margin:0 0 20px;line-height:1.6;">
        Hello {full_name}, your reschedule request has been received. Our team will review and re-confirm your appointment.</p>
      <table style="width:100%;border-collapse:collapse;font-size:12px;">
        <tr style="border-bottom:1px solid #bfdbfe;">
          <td style="padding:7px 0;font-weight:700;color:#1e3a8a;width:120px;">Service</td>
          <td style="padding:7px 0;color:#2563eb;">{service_type}</td>
        </tr>
        <tr style="border-bottom:1px solid #bfdbfe;">
          <td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Previous Date</td>
          <td style="padding:7px 0;color:#93c5fd;text-decoration:line-through;">{old_dt}</td>
        </tr>
        <tr><td style="padding:7px 0;font-weight:700;color:#1e3a8a;">Requested Date</td>
          <td style="padding:7px 0;color:#2563eb;font-weight:600;">{new_dt}</td>
        </tr>
      </table>
      <div style="font-size:11px;font-weight:700;color:#1e3a8a;text-transform:uppercase;letter-spacing:.4px;margin:20px 0 8px;">Reason for Reschedule</div>
      <div style="background:#f8faff;border:1px solid #bfdbfe;border-radius:8px;padding:12px 16px;font-size:13px;color:#1e3a8a;line-height:1.6;">{safe_r}</div>
      <p class="ep" style="color:#64748b;font-size:11px;margin-top:20px;line-height:1.6;">
        You will receive a separate confirmation once the team approves your new schedule.
        Questions? Contact <a href="mailto:{_SUPPORT_EMAIL}" style="color:#2563eb;">{_SUPPORT_EMAIL}</a>.</p>"""
    try:
        subject = f'CyberINK — Reschedule Request: {service_type}'
        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From']    = f'CyberINK Security <{_SMTP_USER}>'
        msg['To']      = user_email
        msg.attach(MIMEText(_email_html(subject, body), 'html'))
        with smtplib.SMTP_SSL(_SMTP_HOST, _SMTP_PORT) as srv:
            srv.login(_SMTP_USER, _SMTP_PASS)
            srv.sendmail(_SMTP_USER, [user_email, _SUPPORT_EMAIL], msg.as_string())
    except Exception:
        pass


@app.route('/api/appointments/<appt_id>/user-cancel', methods=['POST'])
def user_cancel_appointment(appt_id):
    if 'user' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    if session['user'].get('role') == 'admin':
        return jsonify({'error': 'Admins use the management panel'}), 400
    data   = request.get_json(force=True) or {}
    reason = (data.get('reason') or '').strip()
    if not reason:
        return jsonify({'error': 'A cancellation reason is required'}), 400
    appts  = _load_appointments()
    if appt_id not in appts:
        return jsonify({'error': 'Appointment not found'}), 404
    appt = appts[appt_id]
    if appt.get('username') != session['user']['username']:
        return jsonify({'error': 'You can only cancel your own appointments'}), 403
    if appt['status'] == 'completed':
        return jsonify({'error': 'Cannot cancel a completed appointment'}), 400
    if appt['status'] == 'cancelled':
        return jsonify({'error': 'Appointment is already cancelled'}), 400
    import datetime as _dt
    appt.update({'status': 'cancelled', 'cancel_reason': reason,
                 'cancelled_by': 'user',
                 'updated_at': _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')})
    _save_appointments(appts)
    appt_dt = _format_appt_datetime(appt['date'], appt.get('end_date', ''),
                                     appt['time'], appt.get('duration', ''), appt['timezone'])
    _send_user_cancel_email(appt['service_type'], appt_dt,
                             appt['full_name'], appt['email'], reason)
    return jsonify({'ok': True, 'message': 'Appointment cancelled. A confirmation has been sent to your email.'})


@app.route('/api/appointments/<appt_id>/user-reschedule', methods=['POST'])
def user_reschedule_appointment(appt_id):
    if 'user' not in session:
        return jsonify({'error': 'Not logged in'}), 401
    if session['user'].get('role') == 'admin':
        return jsonify({'error': 'Admins use the management panel'}), 400
    data         = request.get_json(force=True) or {}
    new_date     = (data.get('date')     or '').strip()
    new_end_date = (data.get('end_date') or '').strip()
    new_time     = (data.get('time')     or '').strip()
    new_dur      = (data.get('duration') or '').strip()
    reason       = (data.get('reason')   or '').strip()
    if not new_date or not new_time:
        return jsonify({'error': 'New date and time are required'}), 400
    if not reason:
        return jsonify({'error': 'A reason for rescheduling is required'}), 400
    appts = _load_appointments()
    if appt_id not in appts:
        return jsonify({'error': 'Appointment not found'}), 404
    appt = appts[appt_id]
    if appt.get('username') != session['user']['username']:
        return jsonify({'error': 'You can only reschedule your own appointments'}), 403
    if appt['status'] in ('completed', 'cancelled'):
        return jsonify({'error': f'Cannot reschedule a {appt["status"]} appointment'}), 400
    old_dt = _format_appt_datetime(appt['date'], appt.get('end_date', ''),
                                    appt['time'], appt.get('duration', ''), appt['timezone'])
    import datetime as _dt
    appt.update({'date': new_date, 'end_date': new_end_date, 'time': new_time,
                 'duration': new_dur or appt.get('duration', ''),
                 'status': 'rescheduled', 'meet_link': '', 'event_link': '', 'event_id': '',
                 'reschedule_reason': reason, 'rescheduled_by': 'user',
                 'updated_at': _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')})
    _save_appointments(appts)
    new_dt = _format_appt_datetime(new_date, new_end_date, new_time,
                                    new_dur or appt.get('duration', ''), appt['timezone'])
    _send_user_reschedule_request_email(appt['service_type'], old_dt, new_dt,
                                         appt['full_name'], appt['email'], reason)
    return jsonify({'ok': True, 'message': 'Reschedule request submitted. You will be notified once the team confirms your new schedule.'})


@app.route('/api/appointments/<appt_id>/set-cost', methods=['POST'])
def set_appointment_cost(appt_id):
    if 'user' not in session or session['user'].get('role') != 'admin':
        return jsonify({'error': 'Unauthorized'}), 403
    data     = request.get_json(force=True) or {}
    amount   = data.get('amount')
    currency = (data.get('currency') or 'USD').strip().upper()
    if amount is None:
        return jsonify({'error': 'Amount is required'}), 400
    try:
        amount = float(amount)
        if amount < 0:
            raise ValueError()
    except (ValueError, TypeError):
        return jsonify({'error': 'Amount must be a non-negative number'}), 400
    if currency not in _CURRENCY_RATES:
        return jsonify({'error': f'Unsupported currency: {currency}'}), 400
    appts = _load_appointments()
    if appt_id not in appts:
        return jsonify({'error': 'Appointment not found'}), 404
    import datetime as _dt
    appts[appt_id]['cost_amount']   = amount
    appts[appt_id]['cost_currency'] = currency
    appts[appt_id]['updated_at']    = _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S')
    _save_appointments(appts)
    return jsonify({'ok': True, 'message': f'Cost set to {_fmt_cost(amount, currency)}.'})


# ── User File Manager ─────────────────────────────────────────────────────────
import uuid as _uuid_fm
from werkzeug.utils import secure_filename as _secure_fn

_FILES_DIR      = os.path.join(os.path.dirname(__file__), '..', 'data', 'user_files')
_FILES_MANIFEST = os.path.join(_FILES_DIR, 'manifest.json')
_FILES_MAX_MB   = 50

_ALLOWED_EXTS = {
    # Documents
    'pdf','doc','docx','xls','xlsx','ppt','pptx','txt','csv','rtf','odt','ods',
    # Code
    'py','js','ts','jsx','tsx','html','htm','css','scss','sass','less',
    'php','rb','go','rs','java','kt','swift','c','cpp','h','hpp',
    'cs','vb','sh','bash','ps1','bat','sql','json','yaml','yml','toml','xml',
    'ini','env','conf','cfg','md','rst',
    # Images / misc
    'png','jpg','jpeg','gif','svg','webp','zip','tar','gz',
}

def _load_fm_manifest() -> list:
    try:
        if os.path.exists(_FILES_MANIFEST):
            with open(_FILES_MANIFEST, encoding='utf-8') as fh:
                return _json.load(fh)
    except Exception:
        pass
    return []

def _save_fm_manifest(entries: list) -> None:
    os.makedirs(_FILES_DIR, exist_ok=True)
    with open(_FILES_MANIFEST, 'w', encoding='utf-8') as fh:
        _json.dump(entries, fh, indent=2)

def _fmt_filesize(n: int) -> str:
    for unit in ('B','KB','MB','GB'):
        if n < 1024:
            return f'{n:.1f} {unit}' if unit != 'B' else f'{n} B'
        n /= 1024
    return f'{n:.1f} TB'


@app.route('/api/files', methods=['GET'])
@login_required
def list_user_files():
    username = session['user']['username']
    is_admin = session['user'].get('role') == 'admin'
    entries  = _load_fm_manifest()
    if is_admin:
        result = entries  # admin sees all
    else:
        result = [e for e in entries if e.get('username') == username]
    return jsonify({'files': sorted(result, key=lambda x: x.get('uploaded_at', ''), reverse=True)})


@app.route('/api/files/upload', methods=['POST'])
@login_required
def upload_user_file():
    username = session['user']['username']
    files    = request.files.getlist('files')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'No files provided'}), 400

    max_bytes = _FILES_MAX_MB * 1024 * 1024
    saved     = []
    errors    = []

    for f in files:
        original = f.filename or 'upload'
        ext      = original.rsplit('.', 1)[-1].lower() if '.' in original else ''
        if ext not in _ALLOWED_EXTS:
            errors.append(f'{original}: file type .{ext} is not allowed')
            continue
        safe     = _secure_fn(original) or 'upload'
        file_id  = 'f-' + str(_uuid_fm.uuid4())[:12]
        dest_dir = os.path.join(_FILES_DIR, username)
        os.makedirs(dest_dir, exist_ok=True)
        disk_name = f'{file_id}_{safe}'
        disk_path = os.path.join(dest_dir, disk_name)
        f.seek(0, 2)
        size = f.tell()
        f.seek(0)
        if size > max_bytes:
            errors.append(f'{original}: exceeds {_FILES_MAX_MB} MB limit')
            continue
        f.save(disk_path)
        entry = {
            'id':          file_id,
            'username':    username,
            'filename':    original,
            'safe_name':   safe,
            'disk_name':   disk_name,
            'size':        size,
            'size_display':_fmt_filesize(size),
            'ext':         ext,
            'uploaded_at': _dt.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S'),
        }
        manifest = _load_fm_manifest()
        manifest.append(entry)
        _save_fm_manifest(manifest)
        saved.append(entry)

    if not saved and errors:
        return jsonify({'error': errors[0], 'errors': errors}), 400
    return jsonify({'ok': True, 'saved': len(saved),
                    'files': saved, 'errors': errors,
                    'message': f'{len(saved)} file(s) uploaded.' + (f' {len(errors)} skipped.' if errors else '')})


@app.route('/api/files/<file_id>/download', methods=['GET'])
@login_required
def download_user_file(file_id):
    username = session['user']['username']
    is_admin = session['user'].get('role') == 'admin'
    manifest = _load_fm_manifest()
    entry    = next((e for e in manifest if e['id'] == file_id), None)
    if not entry:
        return jsonify({'error': 'File not found'}), 404
    if not is_admin and entry.get('username') != username:
        return jsonify({'error': 'Access denied'}), 403
    disk_path = os.path.join(_FILES_DIR, entry['username'], entry['disk_name'])
    if not os.path.exists(disk_path):
        return jsonify({'error': 'File missing from storage'}), 404
    return send_file(disk_path, as_attachment=True, download_name=entry['filename'])


@app.route('/api/files/<file_id>', methods=['DELETE'])
@login_required
def delete_user_file(file_id):
    username = session['user']['username']
    is_admin = session['user'].get('role') == 'admin'
    manifest = _load_fm_manifest()
    entry    = next((e for e in manifest if e['id'] == file_id), None)
    if not entry:
        return jsonify({'error': 'File not found'}), 404
    if not is_admin and entry.get('username') != username:
        return jsonify({'error': 'Access denied'}), 403
    disk_path = os.path.join(_FILES_DIR, entry['username'], entry['disk_name'])
    try:
        if os.path.exists(disk_path):
            os.remove(disk_path)
    except Exception:
        pass
    manifest = [e for e in manifest if e['id'] != file_id]
    _save_fm_manifest(manifest)
    return jsonify({'ok': True, 'message': 'File deleted.'})


# ── Google Calendar OAuth Setup (admin only) ───────────────────────────────────
@app.route('/api/auth/google/setup')
def gcal_setup():
    if 'user' not in session or session['user'].get('role') != 'admin':
        return 'Unauthorized', 403
    if not _GCAL_AVAILABLE:
        return ('Google Calendar API libraries not installed.\n'
                'Run: pip install google-api-python-client google-auth-httplib2 google-auth-oauthlib'), 503
    if not _GCAL_CLIENT_ID or not _GCAL_CLIENT_SECRET:
        return 'GOOGLE_CLIENT_ID / GOOGLE_CLIENT_SECRET not set in .env', 503
    if _get_gcal_service():
        return 'Google Calendar is already connected.', 200
    try:
        from flask import redirect as _redir
        cb = request.host_url.rstrip('/') + '/api/auth/google/callback'
        flow = _GFlow.from_client_config(_gcal_client_config(cb), scopes=_GCAL_SCOPES)
        flow.redirect_uri = cb
        auth_url, state = flow.authorization_url(access_type='offline',
                                                  include_granted_scopes='true', prompt='consent')
        session['gcal_oauth_state'] = state
        return _redir(auth_url)
    except Exception as exc:
        return f'OAuth setup error: {exc}', 500


@app.route('/api/auth/google/callback')
def gcal_callback():
    if 'user' not in session or session['user'].get('role') != 'admin':
        return 'Unauthorized', 403
    if not _GCAL_AVAILABLE:
        return 'Google Calendar libraries not installed.', 503
    try:
        cb = request.host_url.rstrip('/') + '/api/auth/google/callback'
        flow = _GFlow.from_client_config(_gcal_client_config(cb), scopes=_GCAL_SCOPES)
        flow.redirect_uri = cb
        flow.fetch_token(authorization_response=request.url)
        creds = flow.credentials
        os.makedirs(os.path.dirname(_GCAL_TOKEN_FILE), exist_ok=True)
        with open(_GCAL_TOKEN_FILE, 'w') as fh:
            fh.write(creds.to_json())
        return ('Google Calendar connected successfully. '
                'Appointments will now auto-create Google Meet links. '
                'You may close this tab.'), 200
    except Exception as exc:
        return f'OAuth callback error: {exc}', 500


if __name__ == '__main__':
    port = int(os.environ.get('CFAI_DASHBOARD_PORT', 8889))
    print(f'CF_AI Dashboard running on http://0.0.0.0:{port}')
    app.run(host='0.0.0.0', port=port, debug=False, threaded=True)
